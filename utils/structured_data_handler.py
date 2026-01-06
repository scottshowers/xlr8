"""
Structured Data Handler - DuckDB Storage for Excel/CSV v5.14
============================================================

Deploy to: utils/structured_data_handler.py

v5.14 CHANGES (Simple Fast Processing - MAJOR REWRITE):
- REMOVED: All chunked/streaming processing (was SLOWER than pandas!)
- SIMPLIFIED: Direct pd.read_excel(file, sheet_name=X) for each sheet
  - Pandas C-level code is FAST
  - Only reads one sheet at a time (not entire file)
- openpyxl ONLY used for: getting sheet names (instant metadata read)
- Small sheets (<1000 rows): Run detection for multi-table layouts
- Large sheets (>=1000 rows): Skip detection, direct read → DuckDB
- NEW: _store_single_table() helper centralizes all storage logic
- Expected: 15+ min → 2-3 min for large XLSX files

v5.6 CHANGES (Performance Optimization):
- REMOVED: Verbose diagnostics from safe_fetchall() that ran 6 queries per call
- REMOVED: Verbose logging from safe_commit()
- These diagnostics were killing API responsiveness on polling endpoints

v5.5 CHANGES (Junk Column Removal & Observation Classification):
- NEW: _remove_junk_columns() - Auto-removes parsing artifacts before storage
  - Removes: col_X, unnamed, 100% empty columns
  - Removes: junk-named columns with <5% fill rate
  - Tracks removed columns in results for transparency
- NEW: _classify_column_observation() - Classifies findings properly
  - INSIGHT: Optional fields not in use (UDFs, report_category)
  - WARNING: Suspicious patterns that might need attention
  - ERROR: Actual data quality problems
- UPDATED: _validate_upload_quality() - Returns insights AND issues separately
  - Health score only affected by actual issues, not insights
  - Insights tracked for configuration understanding
- NEW: cleanup_junk_columns() - Utility to clean existing tables

v5.4 CHANGES (Universal Header Detection):
- NEW: _find_header_row() - Universal function based on ratios, not heuristics
  - fill_ratio: non-empty cells / total cells (filters sparse title rows)
  - text_ratio: text cells / non-empty cells (identifies headers vs data)
  - Score = (text_ratio * 2) + fill_ratio
- Removed merged cell detection from normal sheet processing (kept in vertical split only)
- Single function used by: normal processing, horizontal split, vertical split
- Works on ANY Excel file, not just UKG-specific formats

v5.3 CHANGES (Smart Header & Vertical Table Detection):
- Added _split_vertical_tables() for stacked tables on same sheet
- Horizontal split now accepts single blank column separators
- Uses row 0 title for table names when available

v5.2 CHANGES (Thread Safety):
- Added threading lock (_db_lock) to prevent concurrent DuckDB access
- Lock applied to query(), query_to_dataframe(), profile_columns_fast()

v5.0 CHANGES (Performance Optimization):
- Added progress_callback parameter to store_excel() and store_csv()
- SQL-based profiling (no dataframe reload) - 3-5x faster
- Sample-based profiling for large tables (>50K rows) - 10x faster
- Row count estimation for progress tracking
- Chunked CSV loading option for huge files
- All existing functionality preserved

v4.9 CHANGES:
- Improved horizontal table detection (skips header rows, uses 90% threshold)
- Better logging for multi-table sheet parsing
- Fixes Change Reasons and similar multi-table sheets

Stores tabular data in DuckDB for fast SQL queries.
Each project gets isolated tables. Cross-sheet joins supported.

SECURITY FEATURES:
- Field-level encryption for PII (SSN, DOB, etc.)
- Encryption at rest using AES-256-GCM (authenticated encryption)

VERSIONING FEATURES:
- Load versioning for comparison
- Diff detection (new, changed, deleted records)

Author: XLR8 Team
"""

import os
import re
import sys
import json
import logging
import threading
import pandas as pd
import duckdb
import time
from typing import Dict, List, Any, Optional, Tuple, Callable
from datetime import datetime
import hashlib
import base64

# Encryption imports
try:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    from cryptography.fernet import Fernet  # Keep for backward compatibility
    import base64
    ENCRYPTION_AVAILABLE = True
except ImportError:
    ENCRYPTION_AVAILABLE = False
    logging.warning("cryptography not installed - PII encryption disabled. Run: pip install cryptography")

# Openpyxl for colored header detection
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not installed - colored header detection disabled")

logger = logging.getLogger(__name__)

# Track module loads for debugging multi-worker issues
import uuid
_MODULE_LOAD_ID = str(uuid.uuid4())[:8]
logger.warning(f"[MODULE] structured_data_handler loaded, module_id={_MODULE_LOAD_ID}, pid={os.getpid()}")

# DuckDB storage location
DUCKDB_PATH = "/data/structured_data.duckdb"
ENCRYPTION_KEY_PATH = "/data/.encryption_key_v2"  # New path for AES-256 key

# PII field patterns to encrypt
PII_PATTERNS = [
    r'ssn', r'social.*sec', r'social_security',
    r'tax.*id', r'tin',
    r'bank.*account', r'routing.*number', r'account.*number',
    r'credit.*card', r'card.*number',
    r'passport', r'license.*number', r'driver.*lic',
    r'salary', r'pay.*rate', r'wage', r'compensation',
    r'dob', r'date.*birth', r'birth.*date', r'birthdate',
]

# =============================================================================
# PERFORMANCE CONSTANTS
# =============================================================================
LARGE_TABLE_THRESHOLD = 50000  # Sample profiling for tables > 50K rows
PROFILE_SAMPLE_SIZE = 50000    # Rows to sample for profiling large tables
CHUNK_SIZE = 50000             # Rows per chunk for chunked loading


# =========================================================================
# INFERENCE JOB QUEUE - processes one file at a time to avoid overload
# =========================================================================
import queue
import threading

_inference_queue = queue.Queue()
_inference_worker_started = False
_inference_lock = threading.Lock()

def _inference_worker():
    """Background worker that processes inference jobs one at a time"""
    logger.info("[INFERENCE_QUEUE] Worker started")
    while True:
        try:
            job = _inference_queue.get(timeout=60)  # Wait up to 60s for job
            if job is None:  # Poison pill
                break
            
            handler, job_id, project, file_name, tables_info = job
            logger.info(f"[INFERENCE_QUEUE] Processing job {job_id} for {file_name} ({len(tables_info)} tables)")
            
            try:
                handler.run_inference_for_file(job_id, project, file_name, tables_info)
            except Exception as e:
                logger.error(f"[INFERENCE_QUEUE] Job {job_id} failed: {e}")
            
            _inference_queue.task_done()
            
        except queue.Empty:
            # No jobs, just keep waiting
            continue
        except Exception as e:
            logger.error(f"[INFERENCE_QUEUE] Worker error: {e}")

def _ensure_worker_started():
    """Start the worker thread if not already running"""
    global _inference_worker_started
    with _inference_lock:
        if not _inference_worker_started:
            worker = threading.Thread(target=_inference_worker, daemon=True)
            worker.start()
            _inference_worker_started = True
            logger.info("[INFERENCE_QUEUE] Worker thread initialized")

def queue_inference_job(handler, job_id: str, project: str, file_name: str, tables_info: list):
    """Add an inference job to the queue"""
    _ensure_worker_started()
    _inference_queue.put((handler, job_id, project, file_name, tables_info))
    queue_size = _inference_queue.qsize()
    logger.info(f"[INFERENCE_QUEUE] Queued job {job_id}, queue size: {queue_size}")


class FieldEncryptor:
    """
    Handles field-level encryption for PII data using AES-256-GCM.
    
    - AES-256-GCM: Authenticated encryption with 256-bit key
    - 96-bit random nonce per encryption
    - Format: "ENC256:" + base64(nonce + ciphertext + tag)
    - Backward compatible: can still decrypt old "ENC:" Fernet data
    
    Future KMS integration point: Replace _get_key() method
    """
    
    def __init__(self, key_path: str = ENCRYPTION_KEY_PATH):
        self.key_path = key_path
        self.aesgcm = None
        self.fernet = None  # For backward compatibility
        
        if ENCRYPTION_AVAILABLE:
            self._init_encryption()
    
    def _init_encryption(self):
        """Initialize or load AES-256 encryption key"""
        try:
            key = self._get_key()
            self.aesgcm = AESGCM(key)
            logger.info("AES-256-GCM encryption initialized successfully")
            
            # Also init Fernet for backward compatibility with old data
            old_key_path = "/data/.encryption_key"
            if os.path.exists(old_key_path):
                try:
                    with open(old_key_path, 'rb') as f:
                        fernet_key = f.read()
                    self.fernet = Fernet(fernet_key)
                    logger.info("Fernet backward compatibility enabled")
                except Exception:
                    pass
                    
        except Exception as e:
            logger.error(f"Encryption initialization failed: {e}")
            self.aesgcm = None
    
    def _get_key(self) -> bytes:
        """
        Get or generate AES-256 key (32 bytes).
        
        Future KMS integration: Override this method to fetch from KMS
        Example:
            def _get_key(self):
                import boto3
                kms = boto3.client('kms')
                response = kms.generate_data_key(KeyId='alias/pii-key', KeySpec='AES_256')
                return response['Plaintext']
        """
        if os.path.exists(self.key_path):
            with open(self.key_path, 'rb') as f:
                key = f.read()
                if len(key) != 32:
                    raise ValueError("Invalid key length, expected 32 bytes")
                return key
        else:
            # Generate new 256-bit key
            key = os.urandom(32)
            os.makedirs(os.path.dirname(self.key_path), exist_ok=True)
            with open(self.key_path, 'wb') as f:
                f.write(key)
            os.chmod(self.key_path, 0o600)  # Restrict permissions
            logger.info("Generated new AES-256 encryption key")
            return key
    
    def is_pii_column(self, column_name: str) -> bool:
        """Check if column likely contains PII"""
        col_lower = column_name.lower()
        for pattern in PII_PATTERNS:
            if re.search(pattern, col_lower):
                return True
        return False
    
    def encrypt(self, value: Any) -> str:
        """Encrypt a value using AES-256-GCM"""
        if not self.aesgcm or value is None or pd.isna(value):
            return value
        
        try:
            # Convert to string and encode
            val_bytes = str(value).encode('utf-8')
            
            # Generate random 96-bit nonce (12 bytes)
            nonce = os.urandom(12)
            
            # Encrypt (returns ciphertext + 16-byte auth tag)
            ciphertext = self.aesgcm.encrypt(nonce, val_bytes, None)
            
            # Combine: nonce + ciphertext (includes tag)
            encrypted = nonce + ciphertext
            
            # Base64 encode and prefix
            return "ENC256:" + base64.b64encode(encrypted).decode('utf-8')
            
        except Exception as e:
            logger.warning(f"Encryption failed: {e}")
            return value
    
    def decrypt(self, value: Any) -> Any:
        """Decrypt a value (handles both AES-256 and legacy Fernet)"""
        if value is None or pd.isna(value):
            return value
        
        str_val = str(value)
        
        # AES-256-GCM format
        if str_val.startswith("ENC256:") and self.aesgcm:
            try:
                encrypted = base64.b64decode(str_val[7:])
                nonce = encrypted[:12]
                ciphertext = encrypted[12:]
                decrypted = self.aesgcm.decrypt(nonce, ciphertext, None)
                return decrypted.decode('utf-8')
            except Exception as e:
                logger.warning(f"AES-256 decryption failed: {e}")
                return value
        
        # Legacy Fernet format (backward compatibility)
        elif str_val.startswith("ENC:") and self.fernet:
            try:
                encrypted = str_val[4:].encode('utf-8')
                decrypted = self.fernet.decrypt(encrypted)
                return decrypted.decode('utf-8')
            except Exception as e:
                logger.warning(f"Fernet decryption failed: {e}")
                return value
        
        return value
    
    def encrypt_dataframe(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
        """Encrypt PII columns in a DataFrame"""
        encrypted_cols = []
        
        for col in df.columns:
            if self.is_pii_column(col):
                df[col] = df[col].apply(self.encrypt)
                encrypted_cols.append(col)
        
        return df, encrypted_cols
    
    def decrypt_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Decrypt all encrypted columns in a DataFrame"""
        for col in df.columns:
            # Check if any values are encrypted
            sample = df[col].head(10).astype(str)
            if any(s.startswith(('ENC:', 'ENC256:')) for s in sample):
                df[col] = df[col].apply(self.decrypt)
        
        return df


class StructuredDataHandler:
    """
    DuckDB-based storage for structured data with encryption, versioning, and profiling.
    
    Thread-safe: Uses a lock to prevent concurrent DuckDB operations which can cause
    segmentation faults.
    """
    
    # Class-level lock for DuckDB operations
    _db_lock = threading.RLock()
    
    def __init__(self, db_path: str = DUCKDB_PATH):
        self.db_path = db_path
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        
        # Connect to DuckDB with WAL corruption recovery
        self.conn = self._connect_with_recovery(db_path)
        
        # Initialize encryption
        self.encryptor = FieldEncryptor()
        
        # Initialize metadata tables
        self._init_metadata_table()
        
        logger.info(f"StructuredDataHandler initialized with DuckDB at {db_path}")
    
    def _connect_with_recovery(self, db_path: str):
        """
        Connect to DuckDB with automatic WAL corruption recovery.
        
        If the WAL file is corrupted (common after migrations/crashes),
        we backup and remove it, then retry the connection.
        """
        wal_path = f"{db_path}.wal"
        
        try:
            return duckdb.connect(db_path)
        except Exception as e:
            error_str = str(e)
            
            # Check for WAL replay failure
            if "replaying WAL" in error_str or "WAL file" in error_str:
                logger.warning(f"[RECOVERY] WAL corruption detected: {error_str[:200]}")
                
                # Backup the corrupted WAL
                if os.path.exists(wal_path):
                    backup_path = f"{wal_path}.corrupted.{int(time.time())}"
                    try:
                        os.rename(wal_path, backup_path)
                        logger.warning(f"[RECOVERY] Backed up corrupted WAL to {backup_path}")
                    except Exception as backup_e:
                        logger.warning(f"[RECOVERY] Could not backup WAL: {backup_e}")
                        # Try to just delete it
                        try:
                            os.remove(wal_path)
                            logger.warning(f"[RECOVERY] Deleted corrupted WAL file")
                        except Exception as del_e:
                            logger.error(f"[RECOVERY] Could not delete WAL: {del_e}")
                            raise e  # Re-raise original error
                
                # Retry connection
                logger.warning("[RECOVERY] Retrying DuckDB connection without WAL...")
                try:
                    conn = duckdb.connect(db_path)
                    logger.warning("[RECOVERY] ✓ DuckDB connection recovered successfully!")
                    return conn
                except Exception as retry_e:
                    logger.error(f"[RECOVERY] Retry failed: {retry_e}")
                    raise retry_e
            else:
                # Not a WAL error, re-raise
                raise e
    
    def _init_metadata_table(self):
        """Create metadata tables if they don't exist"""
        try:
            # Schema metadata with versioning
            self.conn.execute("""
                CREATE SEQUENCE IF NOT EXISTS schema_metadata_seq START 1
            """)
            
            self.conn.execute("""
                CREATE TABLE IF NOT EXISTS _schema_metadata (
                    id INTEGER PRIMARY KEY,
                    project VARCHAR NOT NULL,
                    file_name VARCHAR NOT NULL,
                    sheet_name VARCHAR NOT NULL,
                    table_name VARCHAR NOT NULL,
                    display_name VARCHAR,
                    entity_type VARCHAR,
                    category VARCHAR,
                    columns JSON NOT NULL,
                    column_count INTEGER,
                    row_count INTEGER,
                    likely_keys JSON,
                    encrypted_columns JSON,
                    truth_type VARCHAR,
                    uploaded_by VARCHAR,
                    version INTEGER DEFAULT 1,
                    is_current BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Load version tracking
            self.conn.execute("""
                CREATE SEQUENCE IF NOT EXISTS load_versions_seq START 1
            """)
            
            self.conn.execute("""
                CREATE TABLE IF NOT EXISTS _load_versions (
                    id INTEGER PRIMARY KEY,
                    project VARCHAR NOT NULL,
                    file_name VARCHAR NOT NULL,
                    version INTEGER NOT NULL,
                    checksum VARCHAR,
                    row_counts JSON,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Table relationships (detected cross-sheet joins)
            self.conn.execute("""
                CREATE TABLE IF NOT EXISTS _table_relationships (
                    id INTEGER PRIMARY KEY,
                    project VARCHAR NOT NULL,
                    source_table VARCHAR NOT NULL,
                    source_columns JSON NOT NULL,
                    target_table VARCHAR NOT NULL,
                    target_columns JSON NOT NULL,
                    relationship_type VARCHAR DEFAULT 'foreign_key',
                    confidence FLOAT DEFAULT 1.0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Column semantic mappings (Claude-inferred + human overrides)
            # EXTENDED: Now includes context graph (hub/spoke) information
            self.conn.execute("""
                CREATE TABLE IF NOT EXISTS _column_mappings (
                    id INTEGER,
                    project VARCHAR NOT NULL,
                    file_name VARCHAR NOT NULL,
                    table_name VARCHAR NOT NULL,
                    original_column VARCHAR NOT NULL,
                    semantic_type VARCHAR NOT NULL,
                    confidence FLOAT DEFAULT 0.5,
                    is_override BOOLEAN DEFAULT FALSE,
                    needs_review BOOLEAN DEFAULT FALSE,
                    
                    -- CONTEXT GRAPH: Hub/Spoke relationships
                    -- For each semantic_type, ONE table is the hub (max cardinality)
                    -- All others are spokes that reference the hub
                    is_hub BOOLEAN DEFAULT FALSE,        -- Is this THE hub for its semantic_type?
                    hub_table VARCHAR,                   -- If spoke, which table is the hub?
                    hub_column VARCHAR,                  -- If spoke, which column in hub?
                    hub_cardinality INTEGER,             -- How many values in the hub?
                    spoke_cardinality INTEGER,           -- How many values in this spoke?
                    coverage_pct FLOAT,                  -- spoke_cardinality / hub_cardinality
                    is_subset BOOLEAN,                   -- Are ALL spoke values in hub? (FK integrity)
                    
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Mapping inference job status
            self.conn.execute("""
                CREATE TABLE IF NOT EXISTS _mapping_jobs (
                    id VARCHAR PRIMARY KEY,
                    project VARCHAR NOT NULL,
                    file_name VARCHAR NOT NULL,
                    status VARCHAR DEFAULT 'pending',
                    total_tables INTEGER DEFAULT 0,
                    completed_tables INTEGER DEFAULT 0,
                    mappings_found INTEGER DEFAULT 0,
                    needs_review_count INTEGER DEFAULT 0,
                    error_message VARCHAR,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # =================================================================
            # COLUMN PROFILES TABLE - Phase 1 Data Foundation
            # Stores detailed statistics about each column for intelligent
            # query generation and clarification decisions.
            # =================================================================
            self.conn.execute("""
                CREATE TABLE IF NOT EXISTS _column_profiles (
                    id INTEGER,
                    project VARCHAR NOT NULL,
                    table_name VARCHAR NOT NULL,
                    column_name VARCHAR NOT NULL,
                    
                    -- Data type info
                    inferred_type VARCHAR,           -- 'numeric', 'categorical', 'date', 'text', 'boolean'
                    original_dtype VARCHAR,          -- pandas dtype
                    
                    -- Basic stats
                    total_count INTEGER,             -- total rows
                    null_count INTEGER,              -- nulls/blanks
                    distinct_count INTEGER,          -- unique values
                    
                    -- For numeric columns
                    min_value DOUBLE,
                    max_value DOUBLE,
                    mean_value DOUBLE,
                    
                    -- For categorical columns (distinct_count <= 100)
                    distinct_values JSON,            -- ['A', 'T', 'L', ...]
                    value_distribution JSON,         -- {'A': 1500, 'T': 200, 'L': 50}
                    
                    -- For date columns
                    min_date VARCHAR,
                    max_date VARCHAR,
                    
                    -- Sample values (first 5 non-null)
                    sample_values JSON,
                    
                    -- Metadata
                    is_likely_key BOOLEAN DEFAULT FALSE,
                    is_categorical BOOLEAN DEFAULT FALSE,
                    
                    -- FILTER INTELLIGENCE (Phase 2.5)
                    filter_category VARCHAR,         -- 'status', 'company', 'organization', 'location', 'pay_type', 'employee_type', 'job', NULL
                    filter_priority INTEGER DEFAULT 0,  -- Higher = more likely to need clarification
                    
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    
                    -- Unique constraint
                    UNIQUE(project, table_name, column_name)
                )
            """)
            
            # MIGRATION: Add new columns if they don't exist (for existing databases)
            try:
                # Check if filter_category column exists
                cols = self.conn.execute("PRAGMA table_info(_column_profiles)").fetchall()
                col_names = [c[1] for c in cols]
                
                if 'filter_category' not in col_names:
                    logger.warning("[MIGRATION] Adding filter_category column to _column_profiles")
                    self.conn.execute("ALTER TABLE _column_profiles ADD COLUMN filter_category VARCHAR")
                
                if 'filter_priority' not in col_names:
                    logger.warning("[MIGRATION] Adding filter_priority column to _column_profiles")
                    self.conn.execute("ALTER TABLE _column_profiles ADD COLUMN filter_priority INTEGER DEFAULT 0")
                
                self.conn.commit()
            except Exception as mig_e:
                logger.warning(f"[MIGRATION] Column check/add for _column_profiles: {mig_e}")
            
            # MIGRATION: Add context graph columns to _column_mappings
            try:
                mapping_cols = self.conn.execute("PRAGMA table_info(_column_mappings)").fetchall()
                mapping_col_names = [c[1] for c in mapping_cols]
                
                if 'is_hub' not in mapping_col_names:
                    logger.warning("[MIGRATION] Adding context graph columns to _column_mappings")
                    self.conn.execute("ALTER TABLE _column_mappings ADD COLUMN is_hub BOOLEAN DEFAULT FALSE")
                    self.conn.execute("ALTER TABLE _column_mappings ADD COLUMN hub_table VARCHAR")
                    self.conn.execute("ALTER TABLE _column_mappings ADD COLUMN hub_column VARCHAR")
                    self.conn.execute("ALTER TABLE _column_mappings ADD COLUMN hub_cardinality INTEGER")
                    self.conn.execute("ALTER TABLE _column_mappings ADD COLUMN spoke_cardinality INTEGER")
                    self.conn.execute("ALTER TABLE _column_mappings ADD COLUMN coverage_pct FLOAT")
                    self.conn.execute("ALTER TABLE _column_mappings ADD COLUMN is_subset BOOLEAN")
                    logger.warning("[MIGRATION] Context graph columns added to _column_mappings")
                
                # DATA-DRIVEN: is_discovered for auto-detected types not in vocabulary
                if 'is_discovered' not in mapping_col_names:
                    logger.warning("[MIGRATION] Adding is_discovered column to _column_mappings")
                    self.conn.execute("ALTER TABLE _column_mappings ADD COLUMN is_discovered BOOLEAN DEFAULT FALSE")
                
                self.conn.commit()
            except Exception as mig_e:
                logger.warning(f"[MIGRATION] Column check/add for _column_mappings: {mig_e}")
            
            # MIGRATION: Add display_name, truth_type, column_count, domain, entity_type, category to _schema_metadata
            try:
                schema_cols = self.conn.execute("PRAGMA table_info(_schema_metadata)").fetchall()
                schema_col_names = [c[1] for c in schema_cols]
                
                if 'display_name' not in schema_col_names:
                    logger.warning("[MIGRATION] Adding display_name column to _schema_metadata")
                    self.conn.execute("ALTER TABLE _schema_metadata ADD COLUMN display_name VARCHAR")
                
                if 'truth_type' not in schema_col_names:
                    logger.warning("[MIGRATION] Adding truth_type column to _schema_metadata")
                    self.conn.execute("ALTER TABLE _schema_metadata ADD COLUMN truth_type VARCHAR")
                
                if 'column_count' not in schema_col_names:
                    logger.warning("[MIGRATION] Adding column_count column to _schema_metadata")
                    self.conn.execute("ALTER TABLE _schema_metadata ADD COLUMN column_count INTEGER")
                
                if 'uploaded_by' not in schema_col_names:
                    logger.warning("[MIGRATION] Adding uploaded_by column to _schema_metadata")
                    self.conn.execute("ALTER TABLE _schema_metadata ADD COLUMN uploaded_by VARCHAR")
                
                if 'domain' not in schema_col_names:
                    logger.warning("[MIGRATION] Adding domain column to _schema_metadata")
                    self.conn.execute("ALTER TABLE _schema_metadata ADD COLUMN domain VARCHAR")
                
                # CONTEXT GRAPH: entity_type and category for semantic typing
                if 'entity_type' not in schema_col_names:
                    logger.warning("[MIGRATION] Adding entity_type column to _schema_metadata")
                    self.conn.execute("ALTER TABLE _schema_metadata ADD COLUMN entity_type VARCHAR")
                
                if 'category' not in schema_col_names:
                    logger.warning("[MIGRATION] Adding category column to _schema_metadata")
                    self.conn.execute("ALTER TABLE _schema_metadata ADD COLUMN category VARCHAR")
                
                # Backfill display_name for existing records (use file_name without extension)
                self.conn.execute("""
                    UPDATE _schema_metadata 
                    SET display_name = CASE 
                        WHEN display_name IS NULL THEN 
                            COALESCE(
                                NULLIF(sheet_name, ''),
                                REGEXP_REPLACE(file_name, '\\.[^.]+$', '')
                            )
                        ELSE display_name
                    END
                    WHERE display_name IS NULL
                """)
                
                self.conn.commit()
                logger.info("[MIGRATION] _schema_metadata columns updated successfully")
            except Exception as mig_e:
                logger.warning(f"[MIGRATION] Column check/add for _schema_metadata: {mig_e}")
            
            self.conn.commit()
            logger.info("Metadata tables initialized successfully")
            
        except Exception as e:
            logger.warning(f"Metadata table init issue: {e}, attempting to recreate...")
            try:
                # Drop and recreate if there's a conflict
                self.conn.execute("DROP TABLE IF EXISTS _schema_metadata")
                self.conn.execute("DROP TABLE IF EXISTS _load_versions")
                self.conn.execute("DROP SEQUENCE IF EXISTS schema_metadata_seq")
                self.conn.execute("DROP SEQUENCE IF EXISTS load_versions_seq")
                self.conn.commit()
                
                # Recursive call to create fresh
                self._init_metadata_table()
            except Exception as e2:
                logger.error(f"Failed to recreate metadata tables: {e2}")
    
    def _sanitize_name(self, name: str) -> str:
        """Sanitize table/column names for SQL"""
        # Handle nan/None values that come from pandas
        name_str = str(name).strip()
        if name_str.lower() in ['nan', 'none', 'nat', ''] or pd.isna(name):
            return 'unnamed'
        
        # Handle numeric column names (0, 1, 2, etc.) 
        if name_str.replace('.', '').replace('-', '').isdigit():
            return f'col_{name_str}'
        
        # Remove special chars, replace spaces with underscores
        sanitized = re.sub(r'[^\w\s]', '', name_str)
        sanitized = re.sub(r'\s+', '_', sanitized.strip())
        sanitized = sanitized.lower()
        # Ensure it doesn't start with a number
        if sanitized and sanitized[0].isdigit():
            sanitized = 'col_' + sanitized
        return sanitized or 'unnamed'
    
    def _derive_entity_metadata(
        self, 
        file_name: str, 
        sheet_name: str, 
        sub_table_title: str = None
    ) -> Dict[str, Optional[str]]:
        """
        Derive entity_type and category from file/sheet/sub-table context.
        
        This enables semantic inference to correctly type generic columns.
        For example, a "code" column in a "Termination Reasons" sub-table
        becomes "termination_reason_code" instead of unknown.
        
        Priority:
        1. Sub-table title → entity_type (e.g., "Termination Reasons" → "termination_reasons")
        2. Sheet name → entity_type if no sub-table, or category if sub-table exists
        3. File name → entity_type for simple single-table files
        
        Args:
            file_name: Original filename (e.g., "Change Reasons.xlsx")
            sheet_name: Sheet/tab name (e.g., "Change Reasons" or "Change Reasons - Termination Reasons")
            sub_table_title: Title of sub-table within sheet if detected (e.g., "Termination Reasons")
            
        Returns:
            Dict with 'entity_type' and 'category' keys (values may be None)
        """
        result = {'entity_type': None, 'category': None}
        
        # CONTEXT GRAPH FIX: Parse combined sheet_name when it contains " - "
        # This happens when sub-tables are detected and stored as "Sheet - SubTable"
        if sheet_name and ' - ' in sheet_name and not sub_table_title:
            parts = sheet_name.split(' - ', 1)
            sheet_name = parts[0].strip()
            sub_table_title = parts[1].strip()
            logger.info(f"[ENTITY-META] Parsed combined name: sheet='{sheet_name}', sub='{sub_table_title}'")
        
        def normalize_to_entity_type(name: str) -> str:
            """Convert a name to entity_type format: lowercase, underscores, singular-ish"""
            if not name:
                return None
            # Remove file extensions
            name = re.sub(r'\.[^.]+$', '', name)
            # Remove common prefixes that don't add meaning (only as whole words/prefixes)
            # Match: "Component_Company" -> "Company", "Component Company" -> "Company"
            # Don't match: "Configuration" (should stay as is)
            name = re.sub(r'^(component|config|setup|master)[_\s]+', '', name, flags=re.IGNORECASE)
            # Convert to lowercase with underscores
            name = re.sub(r'[^\w\s]', '', name)
            name = re.sub(r'\s+', '_', name.strip())
            name = name.lower()
            # Remove trailing 's' for simple plurals (but not 'ss' like 'address')
            if name.endswith('s') and not name.endswith('ss'):
                # Keep it plural for now - more natural for table names
                pass
            return name if name else None
        
        # Clean inputs
        file_clean = file_name.strip() if file_name else ''
        sheet_clean = sheet_name.strip() if sheet_name else ''
        sub_clean = sub_table_title.strip() if sub_table_title else ''
        
        # Priority 1: Sub-table title is most specific
        if sub_clean:
            result['entity_type'] = normalize_to_entity_type(sub_clean)
            # Sheet becomes category when sub-table exists
            if sheet_clean and sheet_clean.lower() != sub_clean.lower():
                result['category'] = normalize_to_entity_type(sheet_clean)
            logger.info(f"[ENTITY-META] Sub-table '{sub_clean}' → entity_type='{result['entity_type']}', category='{result['category']}'")
            return result
        
        # Priority 2: Sheet name
        if sheet_clean and sheet_clean.lower() not in ['sheet1', 'data', 'sheet 1']:
            result['entity_type'] = normalize_to_entity_type(sheet_clean)
            # File name can be category if different
            file_normalized = normalize_to_entity_type(file_clean)
            if file_normalized and file_normalized != result['entity_type']:
                result['category'] = file_normalized
            logger.info(f"[ENTITY-META] Sheet '{sheet_clean}' → entity_type='{result['entity_type']}', category='{result['category']}'")
            return result
        
        # Priority 3: File name (for CSVs or simple single-sheet files)
        if file_clean:
            result['entity_type'] = normalize_to_entity_type(file_clean)
            logger.info(f"[ENTITY-META] File '{file_clean}' → entity_type='{result['entity_type']}'")
            return result
        
        logger.warning(f"[ENTITY-META] Could not derive entity_type from file='{file_name}', sheet='{sheet_name}', sub='{sub_table_title}'")
        return result
    
    def _remove_junk_columns(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, List[Dict]]:
        """
        Remove junk columns that are parsing artifacts.
        
        Identifies and removes:
        - Columns named 'unnamed', 'col_X' (auto-generated)
        - Columns that are 100% empty
        - Columns that are 95%+ empty AND have junk names
        
        Returns:
            Tuple of (cleaned_dataframe, list_of_removed_columns)
        """
        removed = []
        cols_to_drop = []
        
        for col in df.columns:
            col_str = str(col).lower()
            
            # Check for junk column patterns
            is_junk_name = (
                col_str == 'unnamed' or
                col_str.startswith('unnamed_') or
                col_str.startswith('unnamed:') or
                re.match(r'^col_\d+$', col_str) or  # col_0, col_1, etc.
                col_str in ['nan', 'none', 'nat', '']
            )
            
            # Check fill rate
            if len(df) > 0:
                non_empty = df[col].notna().sum()
                non_empty_str = (df[col].astype(str).str.strip() != '').sum()
                fill_rate = max(non_empty, non_empty_str) / len(df)
            else:
                fill_rate = 0
            
            # Decision logic
            should_remove = False
            reason = None
            
            if fill_rate == 0:
                should_remove = True
                reason = 'completely_empty'
            elif is_junk_name and fill_rate < 0.05:
                should_remove = True
                reason = 'junk_name_mostly_empty'
            elif is_junk_name and fill_rate < 0.10:
                # Borderline - check if it's all the same value (often parsing artifact)
                unique_values = df[col].dropna().astype(str).str.strip().unique()
                unique_values = [v for v in unique_values if v and v.lower() not in ['nan', 'none', '']]
                if len(unique_values) <= 1:
                    should_remove = True
                    reason = 'junk_name_single_value'
            
            if should_remove:
                cols_to_drop.append(col)
                removed.append({
                    'column': str(col),
                    'reason': reason,
                    'fill_rate': round(fill_rate, 3)
                })
                logger.info(f"[JUNK-REMOVAL] Removing column '{col}': {reason} (fill={fill_rate:.1%})")
        
        if cols_to_drop:
            df = df.drop(columns=cols_to_drop)
            logger.info(f"[JUNK-REMOVAL] Removed {len(cols_to_drop)} junk column(s)")
        
        return df, removed
    
    def _classify_column_observation(
        self, 
        column_name: str, 
        fill_rate: float, 
        table_name: str = None
    ) -> Optional[Dict]:
        """
        Classify a column observation as insight, warning, error, or None.
        
        Classification logic:
        - INSIGHT: Optional fields not in use (UDFs, report categories at 0%)
        - WARNING: Partially filled required-looking fields
        - ERROR: Clearly invalid data patterns
        
        Returns:
            Dict with classification info, or None if no observation needed
        """
        col_lower = column_name.lower()
        
        # Pattern matching for known optional fields
        OPTIONAL_FIELD_PATTERNS = [
            r'orgud?field\d+',      # orgudfield1-10, orgufield1-10
            r'udf\d+',              # udf1, udf2, etc.
            r'custom_?\d+',         # custom1, custom_field_2
            r'report_category',     # report_category_code, report_category
            r'user_defined',        # user_defined_*
            r'flex_?\d*',           # flex1, flex_field
            r'attribute_?\d*',      # attribute1, attribute_field
        ]
        
        is_optional_field = any(re.search(pattern, col_lower) for pattern in OPTIONAL_FIELD_PATTERNS)
        
        # Classification logic
        if is_optional_field:
            if fill_rate == 0:
                return {
                    'classification': 'insight',
                    'column': column_name,
                    'fill_rate': fill_rate,
                    'message': f"Not using {column_name}",
                    'affects_health_score': False,
                    'insight_type': 'unused_optional_field'
                }
            elif fill_rate > 0.9:
                return {
                    'classification': 'insight',
                    'column': column_name,
                    'fill_rate': fill_rate,
                    'message': f"Using {column_name} ({fill_rate:.0%} populated)",
                    'affects_health_score': False,
                    'insight_type': 'active_optional_field'
                }
            else:
                return {
                    'classification': 'insight',
                    'column': column_name,
                    'fill_rate': fill_rate,
                    'message': f"Partially using {column_name} ({fill_rate:.0%})",
                    'affects_health_score': False,
                    'insight_type': 'partial_optional_field'
                }
        
        # Check for suspicious columns (might indicate parsing issues)
        SUSPICIOUS_PATTERNS = [
            r'^col_\d+$',
            r'^unnamed',
            r'^\d+$',
        ]
        
        is_suspicious_name = any(re.match(pattern, col_lower) for pattern in SUSPICIOUS_PATTERNS)
        
        if is_suspicious_name:
            return {
                'classification': 'warning',
                'column': column_name,
                'fill_rate': fill_rate,
                'message': f"Suspicious column name: {column_name}",
                'affects_health_score': True,
                'warning_type': 'suspicious_column_name'
            }
        
        # Check for potentially required fields with low fill rates
        LIKELY_REQUIRED_PATTERNS = [
            r'employee_?id',
            r'emp_?id', 
            r'person_?id',
            r'ssn',
            r'social_security',
            r'hire_date',
            r'first_?name',
            r'last_?name',
        ]
        
        is_likely_required = any(re.search(pattern, col_lower) for pattern in LIKELY_REQUIRED_PATTERNS)
        
        if is_likely_required and fill_rate < 0.5:
            return {
                'classification': 'warning',
                'column': column_name,
                'fill_rate': fill_rate,
                'message': f"Required field {column_name} only {fill_rate:.0%} populated",
                'affects_health_score': True,
                'warning_type': 'low_fill_required_field'
            }
        
        # Default: no observation needed
        return None
    
    def _find_header_row(self, df: pd.DataFrame, max_rows: int = 15) -> int:
        """
        Universal header row detection based on simple ratios.
        
        A header row has:
        - High text ratio (mostly strings, not numbers)
        - Good fill ratio (most cells have values)
        
        A title/separator row has:
        - Low fill ratio (0-2 values)
        
        A data row has:
        - Mixed types (some text, some numbers, some blanks)
        
        Returns: 0-indexed row number of the best header candidate
        """
        if df.empty:
            return 0
        
        best_row = 0
        best_score = -1
        
        for row_idx in range(min(max_rows, len(df))):
            row = df.iloc[row_idx]
            total_cells = len(row)
            
            # Count cell types
            non_empty = 0
            text_cells = 0
            
            for val in row:
                if pd.notna(val):
                    val_str = str(val).strip()
                    if val_str and val_str.lower() not in ['nan', 'none', 'nat']:
                        non_empty += 1
                        # Is it text (not purely numeric)?
                        if isinstance(val, str) and not val_str.replace('.', '').replace('-', '').isdigit():
                            text_cells += 1
            
            if non_empty == 0:
                continue  # Skip empty rows
            
            # Calculate ratios
            fill_ratio = non_empty / total_cells
            text_ratio = text_cells / non_empty if non_empty > 0 else 0
            
            # Score: reward high text ratio AND good fill
            # Header: high text (>0.6), decent fill (>0.3)
            # Title row: might have high text but very low fill (<0.1)
            # Data row: lower text ratio (mixed with numbers)
            
            if fill_ratio < 0.1:
                continue  # Skip sparse rows (titles/separators)
            
            score = (text_ratio * 2) + fill_ratio  # Weight text ratio more
            
            if score > best_score:
                best_score = score
                best_row = row_idx
        
        return best_row
    
    # =========================================================================
    # STORE SINGLE TABLE - Helper to store one dataframe to DuckDB
    # =========================================================================
    
    def _store_single_table(
        self,
        df: pd.DataFrame,
        project: str,
        file_name: str,
        sheet_name: str,
        version: int,
        encrypt_pii: bool,
        all_encrypted_cols: list,
        results: dict,
        uploaded_by: str = None
    ) -> bool:
        """
        Store a single dataframe as a DuckDB table.
        
        Handles: column sanitization, deduplication, type coercion, encryption, metadata.
        Returns True on success, False on failure.
        """
        try:
            # Drop completely empty rows/columns
            df = df.dropna(how='all').dropna(axis=1, how='all')
            
            if df.empty:
                logger.warning(f"[STORE] Sheet '{sheet_name}' is empty after cleanup, skipping")
                return False
            
            # Sanitize column names
            df.columns = [self._sanitize_name(str(c) if c else f"col_{i}") for i, c in enumerate(df.columns)]
            
            # Handle duplicate column names
            seen = {}
            new_cols = []
            for col in df.columns:
                if col in seen:
                    seen[col] += 1
                    new_cols.append(f"{col}_{seen[col]}")
                else:
                    seen[col] = 0
                    new_cols.append(col)
            df.columns = new_cols
            
            # Force all columns to string
            for col in df.columns:
                df[col] = df[col].fillna('').astype(str)
                df[col] = df[col].replace({'nan': '', 'None': '', 'NaT': ''})
            
            # Remove junk columns
            df, removed_junk = self._remove_junk_columns(df)
            if removed_junk:
                if 'junk_columns_removed' not in results:
                    results['junk_columns_removed'] = []
                results['junk_columns_removed'].extend([{**j, 'sheet': sheet_name} for j in removed_junk])
            
            # Encrypt PII if requested
            encrypted_cols = []
            if encrypt_pii and self.encryptor.fernet:
                df, encrypted_cols = self.encryptor.encrypt_dataframe(df)
                all_encrypted_cols.extend(encrypted_cols)
            
            # Generate table name and display name
            table_name = self._generate_table_name(project, file_name, sheet_name)
            display_name = self._generate_display_name(file_name, sheet_name)
            
            # Derive entity_type and category from context
            entity_meta = self._derive_entity_metadata(file_name, sheet_name)
            
            # Create table (thread-safe)
            self.safe_create_table_from_df(table_name, df)
            
            # Detect key columns
            likely_keys = self._detect_key_columns(df)
            
            # Store metadata
            columns_info = [
                {'name': col, 'type': 'VARCHAR', 'encrypted': col in encrypted_cols}
                for col in df.columns
            ]
            
            self.safe_execute("""
                INSERT INTO _schema_metadata 
                (id, project, file_name, sheet_name, table_name, display_name, entity_type, category, columns, column_count, row_count, likely_keys, encrypted_columns, truth_type, uploaded_by, version, is_current)
                VALUES (nextval('schema_metadata_seq'), ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, TRUE)
            """, [
                project,
                file_name,
                sheet_name,
                table_name,
                display_name,
                entity_meta.get('entity_type'),
                entity_meta.get('category'),
                json.dumps(columns_info),
                len(df.columns),
                len(df),
                json.dumps(likely_keys),
                json.dumps(encrypted_cols),
                None,  # truth_type - set by smart_router if provided
                uploaded_by,
                version
            ])
            
            # Add to results
            results['sheets'].append({
                'sheet_name': sheet_name,
                'table_name': table_name,
                'display_name': display_name,
                'columns': list(df.columns),
                'column_count': len(df.columns),
                'row_count': len(df),
                'likely_keys': likely_keys,
                'encrypted_columns': encrypted_cols
            })
            results['total_rows'] += len(df)
            results['tables_created'].append(table_name)
            
            return True
            
        except Exception as e:
            logger.error(f"[STORE] Failed to store '{sheet_name}': {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    # =========================================================================
    # HORIZONTAL TABLE DETECTION - Side-by-side tables on same sheet
    # =========================================================================
    
    def _split_horizontal_tables(self, file_path: str, sheet_name: str) -> List[Tuple[str, pd.DataFrame]]:
        """
        Detect and split horizontally arranged tables within a single sheet.
        
        ONLY splits when there are CLEAR blank column separators (2+ consecutive blank cols).
        Returns empty list for normal single-table sheets.
        """
        try:
            logger.warning(f"[HORIZONTAL-DETECT] Checking sheet '{sheet_name}' for side-by-side tables")
            
            raw_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            logger.warning(f"[HORIZONTAL-DETECT] Sheet has {len(raw_df)} rows, {raw_df.shape[1]} columns")
            
            if raw_df.empty or raw_df.shape[1] < 4:
                logger.warning(f"[HORIZONTAL-DETECT] Sheet too small, skipping")
                return []
            
            # Find COMPLETELY blank columns (all NaN or empty in data rows)
            # Skip first 2 rows which might be titles
            data_rows = raw_df.iloc[2:] if len(raw_df) > 2 else raw_df
            
            blank_cols = []
            for col_idx in range(raw_df.shape[1]):
                col_data = data_rows.iloc[:, col_idx]
                is_blank = col_data.isna().all() or (col_data.astype(str).str.strip().replace('', pd.NA).isna().all())
                if is_blank:
                    blank_cols.append(col_idx)
            
            logger.warning(f"[HORIZONTAL-DETECT] Found {len(blank_cols)} blank columns: {blank_cols[:10]}{'...' if len(blank_cols) > 10 else ''}")
            
            # Find blank columns as separators
            # Only trigger horizontal split if there are multiple separators
            # (single blank column could be accidental, multiple = intentional layout)
            separators = blank_cols.copy()
            
            if len(separators) < 2:
                logger.warning(f"[HORIZONTAL-DETECT] Less than 2 separators, treating as single table")
                return []  # Need multiple separators to confirm multi-table layout
            
            logger.warning(f"[HORIZONTAL-DETECT] Sheet '{sheet_name}': Found {len(separators)} separator(s) at columns {separators}")
            
            # Split into regions based on separators
            regions = []
            prev_end = 0
            for sep_col in separators:
                if sep_col > prev_end:
                    regions.append((prev_end, sep_col))
                prev_end = sep_col + 1  # Skip the blank column
            # Add final region
            if prev_end < raw_df.shape[1]:
                regions.append((prev_end, raw_df.shape[1]))
            
            # Filter out tiny regions (less than 2 columns of data)
            regions = [(s, e) for s, e in regions if e - s >= 2]
            
            if len(regions) < 2:
                return []  # Need at least 2 real regions to split
            
            logger.warning(f"[HORIZONTAL-DETECT] Sheet '{sheet_name}': Splitting into {len(regions)} tables")
            
            result = []
            for idx, (start_col, end_col) in enumerate(regions):
                sub_df = raw_df.iloc[:, start_col:end_col].copy()
                
                # Check if row 0 has a title (single value that could be table name)
                row0_vals = [v for v in sub_df.iloc[0] if pd.notna(v) and str(v).strip()]
                title_from_row0 = row0_vals[0] if len(row0_vals) == 1 else None
                
                # Find header row using universal detection
                header_row = self._find_header_row(sub_df)
                
                # Set headers and remove header rows
                sub_df.columns = [str(c).strip() for c in sub_df.iloc[header_row]]
                sub_df = sub_df.iloc[header_row + 1:].reset_index(drop=True)
                
                # Remove empty rows
                sub_df = sub_df.dropna(how='all')
                
                if len(sub_df) > 0 and len(sub_df.columns) > 0:
                    # Use title from row 0 if available, otherwise first column header
                    if title_from_row0:
                        table_name = str(title_from_row0).strip()[:40]
                    else:
                        table_name = str(sub_df.columns[0]).strip()[:40] or f"table_{idx + 1}"
                    result.append((table_name, sub_df))
                    logger.warning(f"[HORIZONTAL-DETECT]   Region {idx + 1}: '{table_name}' ({len(sub_df)} rows, {len(sub_df.columns)} cols)")
            
            return result
            
        except Exception as e:
            logger.warning(f"Horizontal table detection failed for '{sheet_name}': {e}")
            return []
    
    def _split_vertical_tables(self, file_path: str, sheet_name: str) -> List[Tuple[str, pd.DataFrame]]:
        """
        Detect and split vertically stacked tables within a single sheet.
        
        Detection methods:
        1. Merged title rows that span 5+ columns (config reports)
        2. Title rows: text in col A only, rest of row empty (section headers)
        3. Completely blank rows as separators
        
        Returns list of (table_name, dataframe) tuples, or empty list for single-table sheets.
        """
        try:
            logger.warning(f"[VERTICAL-DETECT] Checking sheet '{sheet_name}' for stacked tables")
            
            # Read full sheet first
            raw_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            logger.warning(f"[VERTICAL-DETECT] Sheet has {len(raw_df)} rows, {len(raw_df.columns)} columns")
            
            if len(raw_df) < 3:
                logger.warning(f"[VERTICAL-DETECT] Sheet too small, skipping")
                return []
            
            separator_rows = set()
            title_rows = {}  # row_idx -> title text
            
            # Method 1: Find merged rows using openpyxl
            if OPENPYXL_AVAILABLE:
                try:
                    wb = load_workbook(file_path, data_only=True)
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        
                        merge_count = len(list(ws.merged_cells.ranges))
                        logger.warning(f"[VERTICAL-DETECT] Found {merge_count} merged cell ranges")
                        
                        for merged_range in ws.merged_cells.ranges:
                            col_span = merged_range.max_col - merged_range.min_col
                            # If merge spans 5+ columns, it's a title/separator row
                            if col_span >= 4:
                                for row in range(merged_range.min_row, merged_range.max_row + 1):
                                    row_idx = row - 1  # 0-indexed
                                    separator_rows.add(row_idx)
                                    # Get the title text
                                    cell_val = ws.cell(merged_range.min_row, merged_range.min_col).value
                                    if cell_val and str(cell_val).strip():
                                        title_rows[row_idx] = str(cell_val).strip()
                                        logger.warning(f"[VERTICAL-DETECT] Merged title at row {row_idx}: {str(cell_val)[:50]}")
                        wb.close()
                except Exception as e:
                    logger.warning(f"Merged cell detection failed for vertical split: {e}")
            
            logger.warning(f"[VERTICAL-DETECT] After merged cell check: {len(separator_rows)} separator rows")
            
            # Method 2: Find title rows (text in col A only, rest empty or mostly empty)
            for row_idx in range(len(raw_df)):
                if row_idx in separator_rows:
                    continue  # Already marked
                
                row = raw_df.iloc[row_idx]
                first_val = row.iloc[0] if len(row) > 0 else None
                
                # Check if first cell has text
                if pd.isna(first_val) or not str(first_val).strip():
                    # Completely blank row is also a separator
                    if row.isna().all() or (row.astype(str).str.strip() == '').all():
                        separator_rows.add(row_idx)
                    continue
                
                first_text = str(first_val).strip()
                
                # Check if rest of row is empty (allowing for 1-2 stray values)
                other_values = row.iloc[1:] if len(row) > 1 else pd.Series()
                non_empty_others = sum(1 for v in other_values if pd.notna(v) and str(v).strip())
                
                # STRICT: Only treat as title row if it's COMPLETELY alone (no other values)
                # AND preceded/followed by blank rows. The old 90% heuristic was too aggressive.
                # Merged cells (detected above) are the reliable indicator of section headers.
                # Skip the text-in-col-A heuristic - it causes over-splitting.
                pass  # Removed aggressive title detection
            
            logger.warning(f"[VERTICAL-DETECT] Total separator rows found: {len(separator_rows)}")
            
            if not separator_rows:
                logger.warning(f"[VERTICAL-DETECT] No separators found, treating as single table")
                return []  # No separators = probably single table
            
            # Find table boundaries
            # A header row is: not a separator, has 2+ unique text values
            tables = []
            current_title = None
            current_header = None
            current_start = None
            
            for row_idx in range(len(raw_df)):
                is_separator = row_idx in separator_rows
                
                if is_separator:
                    # End current table if exists
                    if current_header is not None and current_start is not None:
                        if current_start < row_idx:  # Has at least some data rows
                            table_name = current_title or f"Table_{len(tables)+1}"
                            tables.append((table_name, current_header, current_start, row_idx))
                        current_header = None
                        current_start = None
                    
                    # Capture title for next table
                    if row_idx in title_rows:
                        current_title = title_rows[row_idx]
                    else:
                        current_title = None
                else:
                    # Check if this could be a header row (first non-separator after separator)
                    if current_header is None:
                        text_values = set()
                        for val in raw_df.iloc[row_idx]:
                            if pd.notna(val) and isinstance(val, str) and len(str(val).strip()) >= 2:
                                text_values.add(str(val).strip())
                        
                        if len(text_values) >= 2:
                            current_header = row_idx
                            current_start = row_idx + 1
            
            # Don't forget last table
            if current_header is not None and current_start is not None:
                table_name = current_title or f"Table_{len(tables)+1}"
                tables.append((table_name, current_header, current_start, len(raw_df)))
            
            logger.warning(f"[VERTICAL-DETECT] Tables found: {len(tables)}")
            
            # Only return if we found multiple tables
            if len(tables) < 2:
                logger.warning(f"[VERTICAL-DETECT] Less than 2 tables, treating as single table")
                return []
            
            logger.warning(f"[VERTICAL-DETECT] Sheet '{sheet_name}': Found {len(tables)} stacked tables")
            
            result = []
            
            for idx, (title, header_row, data_start, data_end) in enumerate(tables):
                # Extract this table's data
                header_vals = raw_df.iloc[header_row].tolist()
                data_df = raw_df.iloc[data_start:data_end].copy()
                
                # CRITICAL: Trim trailing empty columns BEFORE setting names
                # Each sub-table may have different widths, but raw_df has max width
                last_real_col = 0
                for col_idx in range(len(header_vals)):
                    header_val = header_vals[col_idx]
                    has_header = pd.notna(header_val) and str(header_val).strip()
                    
                    # Check if any data exists in this column
                    col_data = data_df.iloc[:, col_idx] if col_idx < data_df.shape[1] else pd.Series()
                    has_data = not col_data.isna().all() and not (col_data.astype(str).str.strip() == '').all()
                    
                    if has_header or has_data:
                        last_real_col = col_idx
                
                # Trim to only columns with headers or data
                if last_real_col < len(header_vals) - 1:
                    header_vals = header_vals[:last_real_col + 1]
                    data_df = data_df.iloc[:, :last_real_col + 1]
                    logger.warning(f"[VERTICAL-DETECT] Trimmed table to {last_real_col + 1} columns")
                
                # Set column names from header row
                data_df.columns = [str(v).strip() if pd.notna(v) else f'col_{i}' for i, v in enumerate(header_vals)]
                data_df = data_df.reset_index(drop=True)
                
                # Remove completely empty rows
                data_df = data_df.dropna(how='all')
                
                if len(data_df) > 0:
                    # Sanitize table name from title
                    table_name = re.sub(r'[^a-zA-Z0-9_\s]', '', title)[:40].strip()
                    if not table_name:
                        table_name = f"table_{idx + 1}"
                    
                    result.append((table_name, data_df))
                    logger.warning(f"[VERTICAL-DETECT]   Table {idx + 1}: '{table_name}' - header row {header_row}, {len(data_df)} data rows, {len(data_df.columns)} cols")
            
            return result
            
        except Exception as e:
            logger.warning(f"Vertical table detection failed for '{sheet_name}': {e}")
            return []
    
    def _generate_table_name(self, project: str, file_name: str, sheet_name: str) -> str:
        """Generate unique internal table name for DuckDB storage"""
        clean_project = self._sanitize_name(project)[:20]  # Limit project
        clean_file = self._sanitize_name(file_name.split('.')[0])[:40]  # Limit file
        clean_sheet = self._sanitize_name(sheet_name)  # Keep full sheet name - it's what makes tables unique!
        
        # DuckDB handles long names fine - don't truncate the sheet name
        table_name = f"{clean_project}_{clean_file}_{clean_sheet}"
        return table_name
    
    def _generate_display_name(self, file_name: str, sheet_name: str = None) -> str:
        """
        Generate a clean, human-readable display name for UI.
        
        Priority:
        1. If sheet_name exists and is meaningful, use it
        2. Otherwise use the file name without extension
        
        Rules:
        - Remove file extension
        - Replace underscores with spaces
        - Title case
        - Max 60 characters
        """
        # Start with sheet name if it's meaningful
        if sheet_name and sheet_name.lower() not in ('data', 'sheet1', 'sheet 1', '', 'pdf'):
            base_name = sheet_name
        else:
            # Use filename without extension
            base_name = file_name.rsplit('.', 1)[0] if '.' in file_name else file_name
        
        # Clean up the name
        # Remove common suffixes that don't add value
        for suffix in ['_export', '_data', '_report', '_final', '_v1', '_v2', '_copy']:
            if base_name.lower().endswith(suffix):
                base_name = base_name[:-len(suffix)]
        
        # Replace underscores and hyphens with spaces
        display = base_name.replace('_', ' ').replace('-', ' ')
        
        # Title case, but preserve acronyms (all caps words)
        words = display.split()
        titled_words = []
        for word in words:
            if word.isupper() and len(word) <= 5:  # Keep short acronyms like "TAX", "HR"
                titled_words.append(word)
            else:
                titled_words.append(word.title())
        
        display = ' '.join(titled_words)
        
        # Truncate if too long
        if len(display) > 60:
            display = display[:57] + '...'
        
        return display.strip()
    
    def _generate_names(self, project: str, file_name: str, sheet_name: str) -> Dict[str, str]:
        """
        Generate both internal table_name and user-facing display_name.
        
        Returns:
            Dict with 'table_name' (for DuckDB) and 'display_name' (for UI)
        """
        return {
            'table_name': self._generate_table_name(project, file_name, sheet_name),
            'display_name': self._generate_display_name(file_name, sheet_name)
        }
    
    def _detect_key_columns(self, df: pd.DataFrame) -> List[str]:
        """Detect likely primary/foreign key columns based on naming and uniqueness"""
        likely_keys = []
        
        key_patterns = [
            r'_id$', r'^id$', r'_key$', r'^key$',
            r'employee.*num', r'emp.*num', r'ee.*num',
            r'employee.*id', r'emp.*id',
            r'_code$', r'company.*code', r'dept.*code'
        ]
        
        for col in df.columns:
            col_lower = col.lower()
            
            # Check naming patterns
            for pattern in key_patterns:
                if re.search(pattern, col_lower):
                    likely_keys.append(col)
                    break
            
            # Check uniqueness for non-pattern columns
            if col not in likely_keys and len(df) > 0:
                uniqueness = df[col].nunique() / len(df)
                if uniqueness > 0.95:  # 95%+ unique values
                    likely_keys.append(col)
        
        return likely_keys[:5]  # Limit to top 5 candidates
    
    def detect_relationships(self, project: str) -> List[Dict[str, Any]]:
        """
        Detect relationships between tables based on column names and values.
        Looks for foreign key relationships across sheets.
        """
        relationships = []
        
        try:
            # Get all tables for this project (thread-safe)
            tables_info = self.safe_fetchall("""
                SELECT DISTINCT table_name, columns, likely_keys 
                FROM _schema_metadata 
                WHERE project = ? AND is_current = TRUE
            """, [project])
            
            if len(tables_info) < 2:
                logger.warning(f"[RELATIONSHIPS] Project {project} has {len(tables_info)} tables, need at least 2")
                return []  # Need at least 2 tables
            
            logger.warning(f"[RELATIONSHIPS] Analyzing {len(tables_info)} tables for project {project}")
            
            # Parse table schemas
            tables = []
            for table_name, columns_json, keys_json in tables_info:
                columns = json.loads(columns_json) if columns_json else []
                keys = json.loads(keys_json) if keys_json else []
                tables.append({
                    'name': table_name,
                    'columns': [c['name'] if isinstance(c, dict) else c for c in columns],
                    'keys': keys
                })
            
            # Look for matching column names between tables
            for i, source in enumerate(tables):
                for target in tables[i+1:]:
                    # Find common column names
                    common_cols = set(source['columns']) & set(target['columns'])
                    
                    for col in common_cols:
                        # Skip if column is likely not a key (too generic)
                        if col in ['name', 'description', 'notes', 'comments', 'date']:
                            continue
                        
                        # Calculate relationship confidence
                        confidence = 0.5
                        if col in source['keys'] or col in target['keys']:
                            confidence = 0.9
                        elif any(pattern in col.lower() for pattern in ['_id', '_code', '_num', 'employee']):
                            confidence = 0.8
                        
                        relationship = {
                            'source_table': source['name'],
                            'source_columns': [col],
                            'target_table': target['name'],
                            'target_columns': [col],
                            'key_type': 'foreign_key',
                            'confidence': confidence
                        }
                        relationships.append(relationship)
            
            logger.warning(f"[RELATIONSHIPS] Found {len(relationships)} potential relationships")
            
            # Store relationships (thread-safe)
            if relationships:
                # Clear existing relationships for this project
                self.safe_execute("DELETE FROM _table_relationships WHERE project = ?", [project])
                
                for idx, rel in enumerate(relationships):
                    self.safe_execute("""
                        INSERT INTO _table_relationships 
                        (id, project, source_table, source_columns, target_table, target_columns, 
                         relationship_type, confidence)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, [
                        idx + 1,
                        project,
                        rel['source_table'],
                        json.dumps(rel['source_columns']),
                        rel['target_table'],
                        json.dumps(rel['target_columns']),
                        rel['key_type'],
                        rel['confidence']
                    ])
                
                self.conn.commit()
                logger.warning(f"[RELATIONSHIPS] Stored {len(relationships)} relationships for project {project}")
            
            return relationships
            
        except Exception as e:
            logger.error(f"[RELATIONSHIPS] Detection failed: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def get_relationships(self, project: str) -> List[Dict[str, Any]]:
        """Get stored relationships for a project"""
        try:
            result = self.safe_fetchall("""
                SELECT source_table, source_columns, target_table, target_columns, 
                       relationship_type, confidence
                FROM _table_relationships
                WHERE project = ?
            """, [project])
            
            relationships = []
            for row in result:
                relationships.append({
                    'source_table': row[0],
                    'source_columns': json.loads(row[1]) if row[1] else [],
                    'target_table': row[2],
                    'target_columns': json.loads(row[3]) if row[3] else [],
                    'relationship_type': row[4],
                    'confidence': row[5]
                })
            
            return relationships
        except Exception as e:
            logger.warning(f"Failed to get relationships: {e}")
            return []
    
    # =========================================================================
    # COLUMN MAPPING METHODS (Claude-inferred + human override)
    # =========================================================================
    
    def infer_column_mappings(self, project: str, file_name: str, table_name: str, 
                               columns: List[str], sample_data: List[Dict]) -> List[Dict]:
        """
        Use LLM to infer semantic meaning of columns.
        LOCAL FIRST: Uses Mistral via LLMOrchestrator, Claude only as fallback.
        
        CONTEXT GRAPH: Uses entity_type to resolve ambiguous column names like "code".
        
        Args:
            project: Project name
            file_name: Source file name
            table_name: DuckDB table name
            columns: List of column names
            sample_data: First few rows of data
            
        Returns:
            List of mapping dicts with confidence scores
        """
        mappings = []
        
        try:
            # Prepare sample data string
            sample_str = ""
            for i, row in enumerate(sample_data[:5]):
                sample_str += f"Row {i+1}: {row}\n"
            
            # CONTEXT GRAPH: Fetch entity_type for this table
            entity_type = None
            category = None
            try:
                meta = self.safe_fetchone("""
                    SELECT entity_type, category FROM _schema_metadata 
                    WHERE table_name = ? AND is_current = TRUE
                """, [table_name])
                if meta:
                    entity_type = meta[0]
                    category = meta[1]
                    logger.info(f"[MAPPINGS] Using entity_type='{entity_type}', category='{category}' for inference")
            except Exception as meta_e:
                logger.debug(f"[MAPPINGS] Could not fetch entity_type: {meta_e}")
            
            # Get semantic types from vocabulary (ONE SOURCE OF TRUTH)
            try:
                from backend.utils.semantic_vocabulary import get_type_names_for_prompt
            except ImportError:
                from utils.semantic_vocabulary import get_type_names_for_prompt
            
            semantic_types_text = get_type_names_for_prompt()
            
            # Build STRICT prompt for LLM - conservative matching
            # CONTEXT GRAPH: Include entity_type so LLM can resolve generic columns dynamically
            entity_context = ""
            if entity_type:
                entity_context = f"""
ENTITY TYPE: {entity_type}
{f"CATEGORY: {category}" if category else ""}

CRITICAL - USE ENTITY TYPE TO RESOLVE AMBIGUOUS COLUMNS:
When a column has a generic name like "code", "type", "id", "name", "description", 
combine the ENTITY TYPE with the column name to find the semantic type.

THE PATTERN: entity_type (singularized) + column_name = semantic_type
Examples of this pattern:
- entity_type="termination_reasons" + col="code" → Look for "termination_reason_code" in allowed types
- entity_type="employee_types" + col="code" → Look for "employee_type_code" in allowed types
- entity_type="pay_groups" + col="code" → Look for "pay_group_code" in allowed types
- entity_type="earnings" + col="code" → Look for "earning_code" in allowed types

The entity_type tells you WHAT this table contains. Search the allowed semantic types below
for one that matches the pattern: [entity singular]_[column]. If found, use it with HIGH CONFIDENCE.
"""
            
            prompt = f"""Analyze these column names and sample data to identify semantic types.

SOURCE FILE: {file_name}
{entity_context}
COLUMN NAMES:
{columns}

SAMPLE DATA:
{sample_str}

CRITICAL RULES - BE CONSERVATIVE:

1. ONLY tag a column if you are HIGHLY CONFIDENT it actually IS that semantic type.
   - "company_code" column with values like "001", "002" → company_code ✓
   - "code" column in ambiguous context → NONE (not enough info)
   - BUT: "code" column WITH entity_type context → USE THE ENTITY TYPE (see above)

2. DO NOT tag these as semantic types:
   - Boolean/flag columns (is_exempt, is_active, inactive, futa_exempt) → NONE
   - Currency codes like "USD", "CAD" → NONE (not amounts)
   - Generic "code" columns WITHOUT entity_type context → NONE
   - "contact" or "name" fields that aren't specifically employee names → NONE
   - ID numbers that aren't SSNs (tax IDs, FEIN, business numbers) → NONE
   - Dates that aren't effective dates (created_at, modified_at) → NONE
   - Generic counters or limits (fte_work_hours with 2 values) → NONE
   - Columns where cardinality is very low (<5) in reference tables → NONE

3. Use filename context ONLY if entity_type is not available:
   - "code" in "Earnings Codes.pdf" → earning_code
   - "code" in "Deduction Codes.xlsx" → deduction_code
   - "code" in "Job Codes" → job_code

4. Column name must CLOSELY MATCH the semantic type:
   - "company_code" → company_code ✓
   - "employer_type_code" → NONE (not the same as employee_type_code)
   - "gl_account_number" → NONE (not company_code)

5. Match ONLY these exact semantic types:
{semantic_types_text}

6. When in doubt, use NONE. Wrong matches are worse than no matches.

Respond with JSON array only:
[
  {{"column": "column_name", "semantic_type": "exact_type_or_NONE", "confidence": 0.0-1.0}},
  ...
]

Use confidence 0.95+ for exact column name matches AND for "code" columns with clear entity_type context."""

            # Use LLMOrchestrator - LOCAL FIRST (Mistral), Claude fallback
            try:
                from utils.llm_orchestrator import LLMOrchestrator
                orchestrator = LLMOrchestrator()
            except ImportError:
                from backend.utils.llm_orchestrator import LLMOrchestrator
                orchestrator = LLMOrchestrator()
            
            result = orchestrator.generate_json(prompt)
            
            if not result.get('success'):
                logger.warning(f"[MAPPINGS] LLM inference failed: {result.get('error')}, using fallback")
                return self._fallback_column_inference(columns, project, file_name, table_name)
            
            inferred = result.get('response')
            if not isinstance(inferred, list):
                logger.warning(f"[MAPPINGS] LLM returned non-list, using fallback")
                return self._fallback_column_inference(columns, project, file_name, table_name)
            
            logger.info(f"[MAPPINGS] Using {result.get('model_used', 'unknown')} for column inference")
            
            # Get valid semantic types for validation
            try:
                from backend.utils.semantic_vocabulary import get_all_type_names
            except ImportError:
                try:
                    from utils.semantic_vocabulary import get_all_type_names
                except ImportError:
                    get_all_type_names = None
            
            valid_types = get_all_type_names() if get_all_type_names else set()
            
            # Store mappings - ONLY high confidence ones
            MIN_CONFIDENCE_THRESHOLD = 0.80  # Skip mappings below this
            
            for item in inferred:
                col = item.get('column', '')
                sem_type = item.get('semantic_type', 'NONE')
                confidence = item.get('confidence', 0.5)
                
                # SKIP low confidence mappings entirely
                if confidence < MIN_CONFIDENCE_THRESHOLD:
                    logger.debug(f"[MAPPINGS] Skipped low-confidence ({confidence:.2f}) mapping: {col} → {sem_type}")
                    continue
                
                # VALIDATE: Only accept types from the vocabulary
                # Reject anything the LLM made up (e.g., "PAY & COMPENSATION", "Monetary amounts")
                if sem_type and sem_type != 'NONE':
                    sem_type_lower = sem_type.lower().strip()
                    
                    # Check if it's a valid type
                    if valid_types and sem_type_lower not in valid_types:
                        logger.debug(f"[MAPPINGS] Rejected invalid semantic type '{sem_type}' for {col}")
                        continue  # Skip this mapping
                    
                    # Normalize to lowercase
                    sem_type = sem_type_lower
                    
                    needs_review = confidence < 0.90  # Tightened from 0.85
                    
                    mapping = {
                        'project': project,
                        'file_name': file_name,
                        'table_name': table_name,
                        'original_column': col,
                        'semantic_type': sem_type,
                        'confidence': confidence,
                        'is_override': False,
                        'needs_review': needs_review
                    }
                    mappings.append(mapping)
                    
                    # Store in database
                    self._store_column_mapping(mapping)
            
            logger.info(f"[MAPPINGS] Inferred {len(mappings)} semantic mappings for {table_name}")
            return mappings
            
        except Exception as e:
            logger.warning(f"[MAPPINGS] LLM inference failed: {e}, using fallback")
            return self._fallback_column_inference(columns, project, file_name, table_name)
    
    def _fallback_column_inference(self, columns: List[str], project: str = None, 
                                    file_name: str = None, table_name: str = None) -> List[Dict]:
        """
        Pattern-based fallback when LLM is unavailable.
        
        CONTEXT GRAPH: Uses entity_type to dynamically resolve generic columns
        by searching the semantic vocabulary.
        """
        mappings = []
        
        # CONTEXT GRAPH: Fetch entity_type for this table
        entity_type = None
        if table_name:
            try:
                meta = self.safe_fetchone("""
                    SELECT entity_type FROM _schema_metadata 
                    WHERE table_name = ? AND is_current = TRUE
                """, [table_name])
                if meta:
                    entity_type = meta[0]
            except Exception:
                pass
        
        # DYNAMIC: Search vocabulary for semantic type matching entity_type + column
        def find_semantic_type_for_entity(entity_type: str, col_name: str) -> Optional[str]:
            """
            Dynamically find semantic type that matches entity_type + column.
            
            For example:
            - entity_type="termination_reasons", col="code" → termination_reason_code
            - entity_type="employee_types", col="code" → employee_type_code
            - entity_type="pay_groups", col="code" → pay_group_code
            
            This is NOT hardcoded - it searches the vocabulary.
            """
            if not entity_type:
                return None
            
            col_lower = col_name.lower()
            
            # Generic columns that benefit from entity_type context
            generic_cols = {'code', 'type', 'id', 'number', 'num'}
            if col_lower not in generic_cols:
                return None
            
            try:
                try:
                    from backend.utils.semantic_vocabulary import get_all_types
                except ImportError:
                    from utils.semantic_vocabulary import get_all_types
                
                all_types = get_all_types()
                
                # Normalize entity_type for matching
                # "termination_reasons" → "termination_reason" (singularize)
                # "employee_types" → "employee_type"
                entity_normalized = entity_type.lower().replace('_', '')
                if entity_normalized.endswith('s') and not entity_normalized.endswith('ss'):
                    entity_singular = entity_normalized[:-1]
                else:
                    entity_singular = entity_normalized
                
                # Also try without trailing 's' on each word
                # "termination_reasons" → "terminationreason"
                entity_words = entity_type.lower().split('_')
                entity_words_singular = [w[:-1] if w.endswith('s') and not w.endswith('ss') else w for w in entity_words]
                entity_pattern = ''.join(entity_words_singular)
                
                # Search for matching semantic type
                for sem_type in all_types:
                    type_name = sem_type.name.lower().replace('_', '')
                    
                    # Match patterns:
                    # - "terminationreasoncode" contains "terminationreason"
                    # - "employeetypecode" contains "employeetype"
                    if entity_singular in type_name or entity_pattern in type_name:
                        # Prefer types ending in the column type (code, type, etc.)
                        if col_lower in type_name or type_name.endswith(col_lower):
                            logger.info(f"[DYNAMIC-MATCH] entity='{entity_type}' + col='{col_name}' → {sem_type.name}")
                            return sem_type.name
                
                # Second pass: looser match
                for sem_type in all_types:
                    type_name = sem_type.name.lower().replace('_', '')
                    if entity_singular in type_name or entity_pattern in type_name:
                        logger.info(f"[DYNAMIC-MATCH] entity='{entity_type}' + col='{col_name}' → {sem_type.name} (loose)")
                        return sem_type.name
                        
            except Exception as e:
                logger.debug(f"[FALLBACK] Vocabulary search failed: {e}")
            
            return None
        
        patterns = {
            'employee_number': [r'employee.*num', r'emp.*num', r'ee.*num', r'emp.*id', r'employee.*id', r'^emp_no$', r'^ee_id$'],
            'company_code': [r'company.*code', r'co.*code', r'comp.*code', r'home.*company'],
            'employment_status_code': [r'employment.*status', r'emp.*status', r'status.*code'],
            'earning_code': [r'earning.*code', r'earn.*code'],
            'deduction_code': [r'deduction.*code', r'ded.*code', r'benefit.*code'],
            'job_code': [r'job.*code', r'position.*code'],
            'department_code': [r'dept.*code', r'department.*code', r'org.*level'],
            'amount': [r'amount', r'amt$', r'_amt$'],
            'rate': [r'rate$', r'pay.*rate', r'hourly.*rate', r'salary'],
            'effective_date': [r'effective.*date', r'eff.*date'],
            'employee_name': [r'^name$', r'employee.*name', r'emp.*name', r'full.*name']
        }
        
        for col in columns:
            col_lower = col.lower()
            matched = False
            
            # CONTEXT GRAPH: Try dynamic entity_type + column matching first
            if entity_type:
                sem_type = find_semantic_type_for_entity(entity_type, col)
                if sem_type:
                    mapping = {
                        'project': project,
                        'file_name': file_name,
                        'table_name': table_name,
                        'original_column': col,
                        'semantic_type': sem_type,
                        'confidence': 0.90,  # High confidence when entity_type matches
                        'is_override': False,
                        'needs_review': False
                    }
                    mappings.append(mapping)
                    if project and file_name and table_name:
                        self._store_column_mapping(mapping)
                    continue
            
            # Standard pattern matching
            for sem_type, pattern_list in patterns.items():
                for pattern in pattern_list:
                    if re.search(pattern, col_lower):
                        mapping = {
                            'project': project,
                            'file_name': file_name,
                            'table_name': table_name,
                            'original_column': col,
                            'semantic_type': sem_type,
                            'confidence': 0.7,  # Lower confidence for pattern matching
                            'is_override': False,
                            'needs_review': True
                        }
                        mappings.append(mapping)
                        
                        if project and file_name and table_name:
                            self._store_column_mapping(mapping)
                        matched = True
                        break
                if matched:
                    break
        
        return mappings
    
    def _store_column_mapping(self, mapping: Dict):
        """Store a single column mapping in database"""
        try:
            # Check if mapping already exists
            existing = self.conn.execute("""
                SELECT id FROM _column_mappings 
                WHERE project = ? AND table_name = ? AND original_column = ?
            """, [mapping['project'], mapping['table_name'], mapping['original_column']]).fetchone()
            
            if existing:
                # Update existing (only if not an override)
                self.conn.execute("""
                    UPDATE _column_mappings 
                    SET semantic_type = ?, confidence = ?, needs_review = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE project = ? AND table_name = ? AND original_column = ? AND is_override = FALSE
                """, [
                    mapping['semantic_type'],
                    mapping['confidence'],
                    mapping['needs_review'],
                    mapping['project'],
                    mapping['table_name'],
                    mapping['original_column']
                ])
            else:
                # Insert new
                self.conn.execute("""
                    INSERT INTO _column_mappings 
                    (id, project, file_name, table_name, original_column, semantic_type, 
                     confidence, is_override, needs_review)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, [
                    hash(f"{mapping['project']}_{mapping['table_name']}_{mapping['original_column']}") % 2147483647,
                    mapping['project'],
                    mapping['file_name'],
                    mapping['table_name'],
                    mapping['original_column'],
                    mapping['semantic_type'],
                    mapping['confidence'],
                    mapping['is_override'],
                    mapping['needs_review']
                ])
            
            self.conn.commit()
        except Exception as e:
            logger.warning(f"[MAPPINGS] Failed to store mapping: {e}")
    
    def get_column_mappings(self, project: str, file_name: str = None, 
                            table_name: str = None) -> List[Dict]:
        """
        Get column mappings for a project/file/table.
        
        Args:
            project: Project name (required)
            file_name: Optional file filter
            table_name: Optional table filter
            
        Returns:
            List of mapping dicts
        """
        try:
            query = "SELECT * FROM _column_mappings WHERE project = ?"
            params = [project]
            
            if file_name:
                query += " AND file_name = ?"
                params.append(file_name)
            
            if table_name:
                query += " AND table_name = ?"
                params.append(table_name)
            
            query += " ORDER BY table_name, original_column"
            
            result = self.conn.execute(query, params).fetchall()
            
            mappings = []
            for row in result:
                mappings.append({
                    'id': row[0],
                    'project': row[1],
                    'file_name': row[2],
                    'table_name': row[3],
                    'original_column': row[4],
                    'semantic_type': row[5],
                    'confidence': row[6],
                    'is_override': row[7],
                    'needs_review': row[8],
                    'created_at': str(row[9]) if row[9] else None,
                    'updated_at': str(row[10]) if row[10] else None
                })
            
            return mappings
            
        except Exception as e:
            logger.warning(f"[MAPPINGS] Failed to get mappings: {e}")
            return []
    
    def update_column_mapping(self, project: str, table_name: str, column_name: str,
                               semantic_type: str) -> bool:
        """
        Human override of a column mapping.
        Sets is_override=TRUE and needs_review=FALSE.
        """
        try:
            self.conn.execute("""
                UPDATE _column_mappings 
                SET semantic_type = ?, is_override = TRUE, needs_review = FALSE, 
                    confidence = 1.0, updated_at = CURRENT_TIMESTAMP
                WHERE project = ? AND table_name = ? AND original_column = ?
            """, [semantic_type, project, table_name, column_name])
            self.conn.commit()
            
            logger.info(f"[MAPPINGS] Human override: {table_name}.{column_name} -> {semantic_type}")
            return True
            
        except Exception as e:
            logger.error(f"[MAPPINGS] Failed to update mapping: {e}")
            return False
    
    def get_mappings_needing_review(self, project: str) -> List[Dict]:
        """Get all mappings that need human review"""
        try:
            result = self.conn.execute("""
                SELECT table_name, original_column, semantic_type, confidence
                FROM _column_mappings 
                WHERE project = ? AND needs_review = TRUE
                ORDER BY confidence ASC
            """, [project]).fetchall()
            
            return [
                {
                    'table_name': row[0],
                    'original_column': row[1],
                    'semantic_type': row[2],
                    'confidence': row[3]
                }
                for row in result
            ]
            
        except Exception as e:
            logger.warning(f"[MAPPINGS] Failed to get review list: {e}")
            return []
    
    def compute_context_graph(self, project: str) -> Dict:
        """
        Compute hub/spoke relationships for the context graph.
        
        DATA-DRIVEN APPROACH (v2.0):
        ============================
        OLD: Vocabulary → Semantic Type → Hub Candidate → Check Data
        NEW: Data Patterns → Hub Candidate → Value Matching → Vocabulary Labels
        
        The data determines what's a hub. Vocabulary just provides labels.
        
        Algorithm:
        1. Identify hub candidates from DATA PATTERNS (not vocabulary)
        2. Find spokes by VALUE MATCHING across all columns
        3. Apply vocabulary labels (auto-add if not found)
        
        Returns:
            Dict with stats about the computed graph
        """
        logger.warning(f"[CONTEXT-GRAPH] Computing hub/spoke relationships for {project} (DATA-DRIVEN v2.0)")
        
        # Configurable thresholds
        HUB_PATTERN_SCORE_THRESHOLD = 4   # Minimum score to be hub candidate
        MIN_COVERAGE_PCT = 20             # Minimum overlap for relationship  
        MIN_HUB_CARDINALITY = 3           # Skip hubs with fewer values
        MAX_HUB_CARDINALITY = 10000       # Skip if too large (not a lookup)
        
        try:
            from collections import defaultdict
            
            # =================================================================
            # STEP 1: Get all tables and their metadata for this project
            # =================================================================
            tables = self.conn.execute("""
                SELECT 
                    s.table_name,
                    s.file_name,
                    s.entity_type,
                    s.category,
                    s.row_count,
                    COALESCE(s.truth_type, 'unknown') as truth_type
                FROM _schema_metadata s
                WHERE s.project = ? 
                  AND s.is_current = TRUE
                  AND s.table_name NOT LIKE '\\_%' ESCAPE '\\'
            """, [project]).fetchall()
            
            if not tables:
                logger.warning(f"[CONTEXT-GRAPH] No tables found for {project}")
                return {'hubs': 0, 'spokes': 0, 'semantic_types': 0}
            
            logger.info(f"[CONTEXT-GRAPH] Found {len(tables)} tables to analyze")
            
            # =================================================================
            # STEP 2: Score each table as potential hub candidate
            # =================================================================
            hub_candidates = []  # List of (table_name, key_column, score, entity_type, category, truth_type)
            
            for table_name, file_name, entity_type, category, row_count, truth_type in tables:
                try:
                    # Get column info for this table
                    columns = self.conn.execute(f"""
                        PRAGMA table_info('{table_name}')
                    """).fetchall()
                    
                    col_names = [c[1].lower() for c in columns if c[1] and not c[1].startswith('_')]
                    
                    if not col_names:
                        continue
                    
                    # Look for "code" or "id" column (the key column)
                    key_column = None
                    has_description = False
                    
                    for col in col_names:
                        # Key column patterns
                        if col == 'code' or col.endswith('_code') or col.endswith('_id'):
                            if not key_column:  # Take first match
                                key_column = col
                        # Description column patterns
                        if 'description' in col or 'name' in col or 'label' in col or col == 'desc':
                            has_description = True
                    
                    # If no explicit key, check for columns that LOOK like keys
                    if not key_column:
                        for col in col_names:
                            if col in ['id', 'type', 'status']:
                                key_column = col
                                break
                    
                    if not key_column:
                        continue  # Can't be a hub without a key column
                    
                    # Get cardinality of the key column
                    try:
                        card_result = self.conn.execute(f"""
                            SELECT COUNT(DISTINCT "{key_column}") 
                            FROM "{table_name}"
                            WHERE "{key_column}" IS NOT NULL 
                              AND TRIM(CAST("{key_column}" AS VARCHAR)) != ''
                        """).fetchone()
                        cardinality = card_result[0] if card_result else 0
                    except Exception:
                        cardinality = 0
                    
                    if cardinality < MIN_HUB_CARDINALITY or cardinality > MAX_HUB_CARDINALITY:
                        continue
                    
                    # Calculate uniqueness ratio
                    row_count = row_count or 0
                    if row_count > 0:
                        uniqueness = cardinality / row_count
                    else:
                        uniqueness = 0
                    
                    # =================================================
                    # Score the table as a hub candidate
                    # =================================================
                    score = 0
                    
                    # +2: Has code/id column
                    if key_column == 'code' or key_column.endswith('_code'):
                        score += 2
                    elif key_column.endswith('_id') or key_column == 'id':
                        score += 1
                    
                    # +2: Has description/name column
                    if has_description:
                        score += 2
                    
                    # +2: High uniqueness (> 80%)
                    if uniqueness > 0.8:
                        score += 2
                    elif uniqueness > 0.5:
                        score += 1
                    
                    # +1: Reasonable size (< 2000 rows)
                    if row_count < 2000:
                        score += 1
                    
                    # +1: Entity type suggests reference table
                    if entity_type:
                        et_lower = entity_type.lower()
                        if et_lower.endswith('_codes') or et_lower.endswith('_types') or \
                           et_lower.endswith('_reasons') or et_lower.endswith('_groups') or \
                           et_lower.endswith('s'):  # Plural suggests lookup
                            score += 1
                    
                    # +1: truth_type = configuration
                    if truth_type == 'configuration':
                        score += 1
                    
                    # Check threshold
                    if score >= HUB_PATTERN_SCORE_THRESHOLD:
                        hub_candidates.append({
                            'table_name': table_name,
                            'key_column': key_column,
                            'score': score,
                            'cardinality': cardinality,
                            'entity_type': entity_type,
                            'category': category,
                            'truth_type': truth_type,
                            'file_name': file_name
                        })
                        logger.debug(f"[CONTEXT-GRAPH] Hub candidate: {table_name}.{key_column} (score={score}, card={cardinality})")
                    
                except Exception as table_err:
                    logger.debug(f"[CONTEXT-GRAPH] Error analyzing {table_name}: {table_err}")
                    continue
            
            logger.info(f"[CONTEXT-GRAPH] Found {len(hub_candidates)} hub candidates")
            
            if not hub_candidates:
                return {'hubs': 0, 'spokes': 0, 'semantic_types': 0}
            
            # =================================================================
            # STEP 3: Get actual values for each hub candidate
            # =================================================================
            for hub in hub_candidates:
                try:
                    values_result = self.conn.execute(f"""
                        SELECT DISTINCT LOWER(TRIM(CAST("{hub['key_column']}" AS VARCHAR))) as val
                        FROM "{hub['table_name']}"
                        WHERE "{hub['key_column']}" IS NOT NULL 
                          AND TRIM(CAST("{hub['key_column']}" AS VARCHAR)) != ''
                        LIMIT 5000
                    """).fetchall()
                    hub['values'] = {r[0] for r in values_result if r[0]}
                except Exception:
                    hub['values'] = set()
            
            # =================================================================
            # STEP 4: Determine semantic type for each hub
            # =================================================================
            for hub in hub_candidates:
                semantic_type = self._derive_semantic_type(
                    hub['entity_type'], 
                    hub['key_column'],
                    hub['category']
                )
                hub['semantic_type'] = semantic_type
                hub['is_discovered'] = not self._is_known_semantic_type(semantic_type)
            
            # =================================================================
            # STEP 5: Deduplicate hubs (same semantic_type → pick best)
            # =================================================================
            hubs_by_type = defaultdict(list)
            for hub in hub_candidates:
                hubs_by_type[hub['semantic_type']].append(hub)
            
            final_hubs = []
            for sem_type, candidates in hubs_by_type.items():
                # Sort: prefer configuration, then highest cardinality
                candidates.sort(key=lambda x: (
                    x['truth_type'] == 'configuration',  # True sorts after False
                    x['cardinality']
                ), reverse=True)
                
                final_hubs.append(candidates[0])
                # Mark others as secondary
                for secondary in candidates[1:]:
                    secondary['is_secondary'] = True
            
            logger.info(f"[CONTEXT-GRAPH] {len(final_hubs)} unique hubs after deduplication")
            
            # =================================================================
            # STEP 6: Find spokes by value matching
            # =================================================================
            # Get all columns from all tables (from _column_profiles for efficiency)
            all_columns = self.conn.execute("""
                SELECT 
                    p.table_name,
                    p.column_name,
                    p.distinct_count,
                    p.distinct_values,
                    COALESCE(s.truth_type, 'unknown') as truth_type,
                    s.entity_type,
                    s.category,
                    s.file_name
                FROM _column_profiles p
                LEFT JOIN _schema_metadata s 
                    ON p.project = s.project AND p.table_name = s.table_name
                WHERE p.project = ?
                  AND p.distinct_count > 0
                  AND p.distinct_count < 5000
            """, [project]).fetchall()
            
            # Build lookup of column values (only for columns with cached values)
            column_values_cache = {}
            for table_name, col_name, distinct_count, distinct_values, truth_type, entity_type, category, file_name in all_columns:
                if distinct_values:
                    values = self._parse_distinct_values(distinct_values)
                    if values:
                        column_values_cache[(table_name, col_name)] = {
                            'values': values,
                            'cardinality': distinct_count or len(values),
                            'truth_type': truth_type,
                            'entity_type': entity_type,
                            'category': category,
                            'file_name': file_name
                        }
            
            # For each hub, find matching spokes
            relationships = []  # (hub, spoke_table, spoke_column, coverage_pct, is_subset)
            
            for hub in final_hubs:
                hub_values = hub['values']
                hub_cardinality = len(hub_values)
                
                if not hub_values:
                    continue
                
                # Check every column for value overlap
                for (table_name, col_name), col_info in column_values_cache.items():
                    # Skip the hub itself
                    if table_name == hub['table_name'] and col_name == hub['key_column']:
                        continue
                    
                    spoke_values = col_info['values']
                    if not spoke_values:
                        continue
                    
                    # Calculate overlap
                    overlap = len(spoke_values & hub_values)
                    if overlap == 0:
                        continue
                    
                    coverage_pct = (overlap / hub_cardinality * 100) if hub_cardinality > 0 else 0
                    is_subset = spoke_values.issubset(hub_values)
                    
                    # Check thresholds
                    if coverage_pct >= MIN_COVERAGE_PCT or is_subset:
                        relationships.append({
                            'hub': hub,
                            'spoke_table': table_name,
                            'spoke_column': col_name,
                            'spoke_cardinality': col_info['cardinality'],
                            'coverage_pct': round(coverage_pct, 2),
                            'is_subset': is_subset,
                            'truth_type': col_info['truth_type'],
                            'entity_type': col_info['entity_type'],
                            'file_name': col_info['file_name']
                        })
            
            logger.info(f"[CONTEXT-GRAPH] Found {len(relationships)} spoke relationships")
            
            # =================================================================
            # STEP 7: Update _column_mappings with hub/spoke data
            # =================================================================
            hub_count = 0
            spoke_count = 0
            
            # First, mark all hubs
            for hub in final_hubs:
                # Ensure mapping exists (upsert)
                self._upsert_column_mapping(
                    project=project,
                    file_name=hub['file_name'],
                    table_name=hub['table_name'],
                    column_name=hub['key_column'],
                    semantic_type=hub['semantic_type'],
                    confidence=0.95,
                    is_hub=True,
                    hub_cardinality=hub['cardinality'],
                    is_discovered=hub.get('is_discovered', False)
                )
                hub_count += 1
            
            # Then, mark all spokes
            for rel in relationships:
                hub = rel['hub']
                self._upsert_column_mapping(
                    project=project,
                    file_name=rel['file_name'],
                    table_name=rel['spoke_table'],
                    column_name=rel['spoke_column'],
                    semantic_type=hub['semantic_type'],
                    confidence=0.85,
                    is_hub=False,
                    hub_table=hub['table_name'],
                    hub_column=hub['key_column'],
                    hub_cardinality=hub['cardinality'],
                    spoke_cardinality=rel['spoke_cardinality'],
                    coverage_pct=rel['coverage_pct'],
                    is_subset=rel['is_subset'],
                    is_discovered=hub.get('is_discovered', False)
                )
                spoke_count += 1
            
            self.conn.execute("CHECKPOINT")
            
            # =================================================================
            # STEP 8: Auto-add discovered types to vocabulary
            # =================================================================
            discovered_types = [h['semantic_type'] for h in final_hubs if h.get('is_discovered')]
            if discovered_types:
                logger.warning(f"[CONTEXT-GRAPH] Discovered {len(discovered_types)} new semantic types: {discovered_types}")
                # Note: These are tracked via is_discovered flag
                # UI can show them for user to name/confirm
            
            result = {
                'hubs': hub_count,
                'spokes': spoke_count,
                'semantic_types': len(final_hubs),
                'discovered_types': len(discovered_types),
                'total_candidates': len(hub_candidates),
                'total_relationships': len(relationships)
            }
            
            logger.warning(f"[CONTEXT-GRAPH] Complete: {hub_count} hubs ({len(discovered_types)} discovered), {spoke_count} spokes across {len(final_hubs)} semantic types")
            return result
            
        except Exception as e:
            logger.error(f"[CONTEXT-GRAPH] Failed to compute: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return {'hubs': 0, 'spokes': 0, 'semantic_types': 0, 'error': str(e)}
    
    def _derive_semantic_type(self, entity_type: str, column_name: str, category: str = None) -> str:
        """
        Derive semantic type from entity_type and column name.
        
        Pattern: entity_type (singularized) + column suffix = semantic_type
        
        Examples:
        - entity_type="termination_reasons" + col="code" → "termination_reason_code"
        - entity_type="pay_groups" + col="pay_group_code" → "pay_group_code"
        - entity_type="earnings" + col="earnings_code" → "earning_code"
        """
        col = column_name.lower()
        
        # If column already looks like a semantic type, use it
        if col.endswith('_code') and len(col) > 5:
            return col
        
        if entity_type:
            # Singularize entity_type
            et = entity_type.lower().strip()
            if et.endswith('ies'):
                singular = et[:-3] + 'y'  # companies → company
            elif et.endswith('ses'):
                singular = et[:-2]  # statuses → status
            elif et.endswith('s') and not et.endswith('ss'):
                singular = et[:-1]  # earnings → earning
            else:
                singular = et
            
            # Build semantic type
            if col == 'code' or col == 'id' or col == 'type':
                return f"{singular}_code"
            elif col.endswith('_code') or col.endswith('_id'):
                return col
            else:
                return f"{singular}_{col}"
        
        # Fallback: just use column name
        if col.endswith('_code'):
            return col
        return f"{col}_code"
    
    def _is_known_semantic_type(self, semantic_type: str) -> bool:
        """Check if semantic type exists in vocabulary."""
        try:
            try:
                from backend.utils.semantic_vocabulary import get_all_types
            except ImportError:
                from utils.semantic_vocabulary import get_all_types
            
            known_types = {t.name for t in get_all_types()}
            return semantic_type in known_types
        except Exception:
            return False
    
    def _upsert_column_mapping(
        self,
        project: str,
        file_name: str,
        table_name: str,
        column_name: str,
        semantic_type: str,
        confidence: float = 0.85,
        is_hub: bool = False,
        hub_table: str = None,
        hub_column: str = None,
        hub_cardinality: int = None,
        spoke_cardinality: int = None,
        coverage_pct: float = None,
        is_subset: bool = None,
        is_discovered: bool = False
    ):
        """Upsert a column mapping with hub/spoke info."""
        try:
            # Check if exists
            existing = self.safe_fetchone("""
                SELECT id FROM _column_mappings 
                WHERE project = ? AND table_name = ? AND original_column = ?
            """, [project, table_name, column_name])
            
            if existing:
                # Update
                self.conn.execute("""
                    UPDATE _column_mappings SET
                        semantic_type = ?,
                        confidence = ?,
                        is_hub = ?,
                        hub_table = ?,
                        hub_column = ?,
                        hub_cardinality = ?,
                        spoke_cardinality = ?,
                        coverage_pct = ?,
                        is_subset = ?,
                        is_discovered = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE project = ? AND table_name = ? AND original_column = ?
                """, [
                    semantic_type, confidence, is_hub,
                    hub_table, hub_column, hub_cardinality,
                    spoke_cardinality, coverage_pct, is_subset, is_discovered,
                    project, table_name, column_name
                ])
            else:
                # Insert
                self.conn.execute("""
                    INSERT INTO _column_mappings 
                    (project, file_name, table_name, original_column, semantic_type, confidence,
                     is_hub, hub_table, hub_column, hub_cardinality, spoke_cardinality, 
                     coverage_pct, is_subset, is_discovered)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, [
                    project, file_name or '', table_name, column_name, semantic_type, confidence,
                    is_hub, hub_table, hub_column, hub_cardinality, spoke_cardinality,
                    coverage_pct, is_subset, is_discovered
                ])
        except Exception as e:
            logger.debug(f"[CONTEXT-GRAPH] Mapping upsert error: {e}")
    
    def _parse_distinct_values(self, values_json) -> set:
        """Parse distinct_values JSON into a set for comparison."""
        if not values_json:
            return set()
        
        try:
            import json
            values = json.loads(values_json) if isinstance(values_json, str) else values_json
            if isinstance(values, list):
                return {str(v).lower().strip() for v in values if v is not None}
            return set()
        except Exception:
            return set()
    
    def get_context_graph(self, project: str) -> Dict:
        """
        Get the computed context graph for query-time use.
        
        Returns:
            Dict with:
            - hubs: List of hub columns (Config tables = sources of truth)
            - relationships: List of Reality→Config edges with coverage
            - summary: Hub count, spoke count, has_reality_data flag
        """
        with self._db_lock:
            try:
                # Get all hubs with truth_type, entity_type, and is_discovered
                hubs = self.conn.execute("""
                    SELECT 
                        m.semantic_type, 
                        m.table_name, 
                        m.original_column, 
                        m.hub_cardinality,
                        COALESCE(s.truth_type, 'unknown') as truth_type,
                        s.entity_type,
                        s.category,
                        COALESCE(m.is_discovered, FALSE) as is_discovered
                    FROM _column_mappings m
                    LEFT JOIN _schema_metadata s 
                        ON m.project = s.project AND m.table_name = s.table_name
                    WHERE m.project = ? AND m.is_hub = TRUE
                """, [project]).fetchall()
                
                # Get all spokes (both Config and Reality) with their hub references
                spokes = self.conn.execute("""
                    SELECT 
                        m.semantic_type, 
                        m.table_name, 
                        m.original_column,
                        m.hub_table,
                        m.hub_column,
                        m.hub_cardinality,
                        m.spoke_cardinality,
                        m.coverage_pct,
                        m.is_subset,
                        COALESCE(s.truth_type, 'unknown') as truth_type,
                        s.entity_type,
                        s.category,
                        COALESCE(m.is_discovered, FALSE) as is_discovered
                    FROM _column_mappings m
                    LEFT JOIN _schema_metadata s 
                        ON m.project = s.project AND m.table_name = s.table_name
                    WHERE m.project = ? 
                      AND m.is_hub = FALSE 
                      AND m.hub_table IS NOT NULL
                """, [project]).fetchall()
                
                # Separate spokes by truth_type for stats
                reality_spokes = [s for s in spokes if s[9] == 'reality']
                config_spokes = [s for s in spokes if s[9] == 'configuration']
                
                hub_list = [
                    {
                        'semantic_type': h[0],
                        'table': h[1],
                        'column': h[2],
                        'cardinality': h[3],
                        'truth_type': h[4],
                        'entity_type': h[5],
                        'category': h[6],
                        'is_discovered': h[7],
                        'has_reality_spokes': any(s[0] == h[0] for s in reality_spokes)
                    }
                    for h in hubs
                ]
                
                # Count discovered hubs
                discovered_hubs = [h for h in hub_list if h.get('is_discovered')]
                
                return {
                    'hubs': hub_list,
                    'relationships': [  # ALL spokes - Config and Reality
                        {
                            'semantic_type': s[0],
                            'spoke_table': s[1],
                            'spoke_column': s[2],
                            'hub_table': s[3],
                            'hub_column': s[4],
                            'hub_cardinality': s[5],
                            'spoke_cardinality': s[6],
                            'coverage_pct': s[7],
                            'is_valid_fk': s[8],  # is_subset = valid foreign key
                            'truth_type': s[9],
                            'entity_type': s[10],
                            'category': s[11],
                            'is_discovered': s[12]
                        }
                        for s in spokes
                    ],
                    'summary': {
                        'hub_count': len(hubs),
                        'spoke_count': len(spokes),
                        'config_spoke_count': len(config_spokes),
                        'reality_spoke_count': len(reality_spokes),
                        'semantic_types': list(set(h[0] for h in hubs)),
                        'has_reality_data': len(reality_spokes) > 0,
                        'hubs_awaiting_reality': len([h for h in hub_list if not h['has_reality_spokes']]),
                        'discovered_hubs': len(discovered_hubs)
                    }
                }
                
            except Exception as e:
                logger.error(f"[CONTEXT-GRAPH] Failed to get graph: {e}")
                return {'hubs': [], 'relationships': [], 'secondary_hubs': [], 'summary': {}}

    def create_mapping_job(self, job_id: str, project: str, file_name: str, total_tables: int):
        """Create a mapping job record"""
        try:
            self.conn.execute("""
                INSERT INTO _mapping_jobs (id, project, file_name, status, total_tables)
                VALUES (?, ?, ?, 'pending', ?)
            """, [job_id, project, file_name, total_tables])
            self.conn.commit()
            logger.info(f"[MAPPING_JOB] Created job {job_id} for {file_name}")
        except Exception as e:
            logger.warning(f"[MAPPING_JOB] Failed to create job: {e}")
    
    def get_mapping_job_status(self, job_id: str = None, project: str = None, 
                                file_name: str = None) -> Optional[Dict]:
        """Get mapping job status by ID or by project+file"""
        try:
            if job_id:
                result = self.conn.execute(
                    "SELECT * FROM _mapping_jobs WHERE id = ?", [job_id]
                ).fetchone()
            elif project and file_name:
                result = self.conn.execute("""
                    SELECT * FROM _mapping_jobs 
                    WHERE project = ? AND file_name = ?
                    ORDER BY created_at DESC LIMIT 1
                """, [project, file_name]).fetchone()
            else:
                return None
            
            if result:
                return {
                    'id': result[0],
                    'project': result[1],
                    'file_name': result[2],
                    'status': result[3],
                    'total_tables': result[4],
                    'completed_tables': result[5],
                    'mappings_found': result[6],
                    'needs_review_count': result[7],
                    'error_message': result[8],
                    'created_at': str(result[9]) if result[9] else None,
                    'updated_at': str(result[10]) if result[10] else None
                }
            return None
            
        except Exception as e:
            logger.warning(f"[MAPPING_JOB] Failed to get status: {e}")
            return None
    
    def get_file_mapping_summary(self, project: str, file_name: str) -> Dict:
        """Get mapping summary for a file (for UI display)"""
        try:
            # Get job status
            job = self.get_mapping_job_status(project=project, file_name=file_name)
            
            # Get actual mapping counts
            mapping_result = self.conn.execute("""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN needs_review = TRUE THEN 1 ELSE 0 END) as needs_review
                FROM _column_mappings 
                WHERE project = ? AND file_name = ?
            """, [project, file_name]).fetchone()
            
            total_mappings = mapping_result[0] if mapping_result else 0
            needs_review_count = mapping_result[1] if mapping_result else 0
            
            # Get all mappings for this file
            mappings = self.get_column_mappings(project, file_name=file_name)
            
            return {
                'status': job['status'] if job else ('complete' if total_mappings > 0 else 'none'),
                'total_mappings': total_mappings,
                'needs_review_count': needs_review_count,
                'mappings': mappings,
                'job': job
            }
            
        except Exception as e:
            logger.warning(f"[MAPPING_JOB] Failed to get summary: {e}")
            return {'status': 'error', 'total_mappings': 0, 'needs_review_count': 0, 'mappings': []}
    
    def run_inference_for_file(self, job_id: str, project: str, file_name: str,
                                tables_info: List[Dict]):
        """
        Run column inference for all tables in a file.
        Called in background thread - uses separate DB connection for thread safety.
        """
        # Create separate connection for this thread
        thread_conn = None
        try:
            thread_conn = duckdb.connect(self.db_path)
            logger.info(f"[MAPPING_JOB] Started background inference with separate connection")
            
            total_mappings = 0
            total_needs_review = 0
            
            for i, table_info in enumerate(tables_info):
                table_name = table_info.get('table_name', '')
                columns = table_info.get('columns', [])
                sample_data = table_info.get('sample_data', [])
                
                logger.info(f"[MAPPING_JOB] Inferring {table_name} ({i+1}/{len(tables_info)})")
                
                # Run inference (uses thread connection)
                mappings = self._infer_column_mappings_threaded(
                    thread_conn=thread_conn,
                    project=project,
                    file_name=file_name,
                    table_name=table_name,
                    columns=columns,
                    sample_data=sample_data
                )
                
                # Count results
                total_mappings += len(mappings)
                total_needs_review += sum(1 for m in mappings if m.get('needs_review'))
                
                # Update progress (uses thread connection)
                self._update_mapping_job_threaded(
                    thread_conn,
                    job_id,
                    completed_tables=i + 1,
                    mappings_found=total_mappings,
                    needs_review_count=total_needs_review
                )
            
            # Mark complete
            self._update_mapping_job_threaded(thread_conn, job_id, status='complete')
            logger.info(f"[MAPPING_JOB] Complete: {total_mappings} mappings, {total_needs_review} need review")
            
        except Exception as e:
            logger.error(f"[MAPPING_JOB] Inference failed: {e}")
            if thread_conn:
                try:
                    self._update_mapping_job_threaded(thread_conn, job_id, status='error', error_message=str(e))
                except Exception:
                    pass
        finally:
            if thread_conn:
                try:
                    thread_conn.close()
                    logger.info("[MAPPING_JOB] Closed background thread connection")
                except Exception:
                    pass
    
    def _infer_column_mappings_threaded(self, thread_conn, project: str, file_name: str, 
                                         table_name: str, columns: List[str], 
                                         sample_data: List[Dict]) -> List[Dict]:
        """Thread-safe version of infer_column_mappings using provided connection.
        LOCAL FIRST: Uses Mistral via LLMOrchestrator, Claude only as fallback."""
        mappings = []
        
        try:
            # Prepare sample data string
            sample_str = ""
            for i, row in enumerate(sample_data[:5]):
                sample_str += f"Row {i+1}: {row}\n"
            
            # Get semantic types from vocabulary (ONE SOURCE OF TRUTH)
            try:
                from backend.utils.semantic_vocabulary import get_type_names_for_prompt
            except ImportError:
                from utils.semantic_vocabulary import get_type_names_for_prompt
            
            semantic_types_text = get_type_names_for_prompt()
            
            # Build STRICT prompt for LLM - conservative matching (threaded version)
            # CONTEXT GRAPH: Include filename so "code" in "Earnings Codes.pdf" becomes "earnings_code"
            prompt = f"""Analyze these column names and sample data to identify semantic types.

SOURCE FILE: {file_name}

COLUMN NAMES:
{columns}

SAMPLE DATA:
{sample_str}

CRITICAL RULES - BE CONSERVATIVE:

1. ONLY tag a column if you are HIGHLY CONFIDENT it actually IS that semantic type.
   - "company_code" column with values like "001", "002" → company_code ✓
   - "code" column in ambiguous context → NONE (not enough info)

2. DO NOT tag these as semantic types:
   - Boolean/flag columns (is_exempt, is_active, inactive, futa_exempt) → NONE
   - Currency codes like "USD", "CAD" → NONE (not amounts)
   - Generic "code" columns without clear context → NONE
   - "contact" or "name" fields that aren't specifically employee names → NONE
   - ID numbers that aren't SSNs (tax IDs, FEIN, business numbers) → NONE
   - Dates that aren't effective dates (created_at, modified_at) → NONE
   - Generic counters or limits (fte_work_hours with 2 values) → NONE
   - Columns where cardinality is very low (<5) in reference tables → NONE

3. Use filename context ONLY for "code" columns:
   - "code" in "Earnings Codes.pdf" → earning_code
   - "code" in "Deduction Codes.xlsx" → deduction_code
   - "code" in "Job Codes" → job_code

4. Column name must CLOSELY MATCH the semantic type:
   - "company_code" → company_code ✓
   - "employer_type_code" → NONE (not the same as employee_type_code)
   - "gl_account_number" → NONE (not company_code)

5. Match ONLY these exact semantic types:
{semantic_types_text}

6. When in doubt, use NONE. Wrong matches are worse than no matches.

Respond with JSON array only:
[
  {{"column": "column_name", "semantic_type": "exact_type_or_NONE", "confidence": 0.0-1.0}},
  ...
]

Use confidence 0.95+ ONLY for exact column name matches like "company_code"→company_code."""

            # Use LLMOrchestrator - LOCAL FIRST (Mistral), Claude fallback
            try:
                from utils.llm_orchestrator import LLMOrchestrator
                orchestrator = LLMOrchestrator()
            except ImportError:
                from backend.utils.llm_orchestrator import LLMOrchestrator
                orchestrator = LLMOrchestrator()
            
            result = orchestrator.generate_json(prompt)
            
            if not result.get('success'):
                logger.warning(f"[MAPPINGS] LLM inference failed: {result.get('error')}, using fallback")
                return self._fallback_column_inference_threaded(thread_conn, columns, project, file_name, table_name)
            
            inferred = result.get('response')
            if not isinstance(inferred, list):
                logger.warning(f"[MAPPINGS] LLM returned non-list, using fallback")
                return self._fallback_column_inference_threaded(thread_conn, columns, project, file_name, table_name)
            
            logger.info(f"[MAPPINGS] Using {result.get('model_used', 'unknown')} for column inference (threaded)")
            
            # Get valid semantic types for validation
            try:
                from backend.utils.semantic_vocabulary import get_all_type_names
            except ImportError:
                try:
                    from utils.semantic_vocabulary import get_all_type_names
                except ImportError:
                    get_all_type_names = None
            
            valid_types = get_all_type_names() if get_all_type_names else set()
            
            # Store mappings - ONLY high confidence ones
            MIN_CONFIDENCE_THRESHOLD = 0.80  # Skip mappings below this
            
            for item in inferred:
                col = item.get('column', '')
                sem_type = item.get('semantic_type', 'NONE')
                confidence = item.get('confidence', 0.5)
                
                # SKIP low confidence mappings entirely
                if confidence < MIN_CONFIDENCE_THRESHOLD:
                    logger.debug(f"[MAPPINGS] Skipped low-confidence ({confidence:.2f}) mapping: {col} → {sem_type}")
                    continue
                
                if sem_type and sem_type != 'NONE':
                    sem_type_lower = sem_type.lower().strip()
                    
                    # Check if it's a valid type
                    if valid_types and sem_type_lower not in valid_types:
                        logger.debug(f"[MAPPINGS] Rejected invalid semantic type '{sem_type}' for {col}")
                        continue
                    
                    # Normalize to lowercase
                    sem_type = sem_type_lower
                    
                    needs_review = confidence < 0.90  # Tightened from 0.85
                    
                    mapping = {
                        'project': project,
                        'file_name': file_name,
                        'table_name': table_name,
                        'original_column': col,
                        'semantic_type': sem_type,
                        'confidence': confidence,
                        'is_override': False,
                        'needs_review': needs_review
                    }
                    mappings.append(mapping)
                    
                    # Store using thread connection
                    self._store_column_mapping_threaded(thread_conn, mapping)
            
            logger.info(f"[MAPPINGS] Inferred {len(mappings)} semantic mappings for {table_name}")
            return mappings
            
        except Exception as e:
            logger.warning(f"[MAPPINGS] LLM inference failed: {e}, using fallback")
            return self._fallback_column_inference_threaded(thread_conn, columns, project, file_name, table_name)
    
    def _fallback_column_inference_threaded(self, thread_conn, columns: List[str], 
                                             project: str, file_name: str, table_name: str) -> List[Dict]:
        """Thread-safe pattern-based fallback with dynamic entity_type resolution."""
        mappings = []
        
        # CONTEXT GRAPH: Fetch entity_type for this table
        entity_type = None
        try:
            meta = thread_conn.execute("""
                SELECT entity_type FROM _schema_metadata 
                WHERE table_name = ? AND is_current = TRUE
            """, [table_name]).fetchone()
            if meta:
                entity_type = meta[0]
        except Exception:
            pass
        
        # DYNAMIC: Search vocabulary for semantic type matching entity_type + column
        def find_semantic_type_for_entity(entity_type: str, col_name: str) -> Optional[str]:
            """Dynamically find semantic type that matches entity_type + column."""
            if not entity_type:
                return None
            
            col_lower = col_name.lower()
            generic_cols = {'code', 'type', 'id', 'number', 'num'}
            if col_lower not in generic_cols:
                return None
            
            try:
                try:
                    from backend.utils.semantic_vocabulary import get_all_types
                except ImportError:
                    from utils.semantic_vocabulary import get_all_types
                
                all_types = get_all_types()
                entity_normalized = entity_type.lower().replace('_', '')
                if entity_normalized.endswith('s') and not entity_normalized.endswith('ss'):
                    entity_singular = entity_normalized[:-1]
                else:
                    entity_singular = entity_normalized
                
                entity_words = entity_type.lower().split('_')
                entity_words_singular = [w[:-1] if w.endswith('s') and not w.endswith('ss') else w for w in entity_words]
                entity_pattern = ''.join(entity_words_singular)
                
                for sem_type in all_types:
                    type_name = sem_type.name.lower().replace('_', '')
                    if entity_singular in type_name or entity_pattern in type_name:
                        if col_lower in type_name or type_name.endswith(col_lower):
                            return sem_type.name
                
                for sem_type in all_types:
                    type_name = sem_type.name.lower().replace('_', '')
                    if entity_singular in type_name or entity_pattern in type_name:
                        return sem_type.name
                        
            except Exception:
                pass
            return None
        
        patterns = {
            'employee_number': [r'employee.*num', r'emp.*num', r'ee.*num', r'emp.*id', r'employee.*id', r'^emp_no$', r'^ee_id$'],
            'company_code': [r'company.*code', r'co.*code', r'comp.*code', r'home.*company'],
            'employment_status_code': [r'employment.*status', r'emp.*status', r'status.*code'],
            'earning_code': [r'earning.*code', r'earn.*code'],
            'deduction_code': [r'deduction.*code', r'ded.*code', r'benefit.*code'],
            'job_code': [r'job.*code', r'position.*code'],
            'department_code': [r'dept.*code', r'department.*code', r'org.*level'],
            'amount': [r'amount', r'amt$', r'_amt$'],
            'rate': [r'rate$', r'pay.*rate', r'hourly.*rate', r'salary'],
            'effective_date': [r'effective.*date', r'eff.*date'],
            'employee_name': [r'^name$', r'employee.*name', r'emp.*name', r'full.*name']
        }
        
        for col in columns:
            col_lower = col.lower()
            
            # CONTEXT GRAPH: Try dynamic entity_type + column matching first
            if entity_type:
                sem_type = find_semantic_type_for_entity(entity_type, col)
                if sem_type:
                    mapping = {
                        'project': project,
                        'file_name': file_name,
                        'table_name': table_name,
                        'original_column': col,
                        'semantic_type': sem_type,
                        'confidence': 0.90,
                        'is_override': False,
                        'needs_review': False
                    }
                    mappings.append(mapping)
                    self._store_column_mapping_threaded(thread_conn, mapping)
                    continue
            
            # Standard pattern matching
            for sem_type, pattern_list in patterns.items():
                for pattern in pattern_list:
                    if re.search(pattern, col_lower):
                        mapping = {
                            'project': project,
                            'file_name': file_name,
                            'table_name': table_name,
                            'original_column': col,
                            'semantic_type': sem_type,
                            'confidence': 0.7,
                            'is_override': False,
                            'needs_review': True
                        }
                        mappings.append(mapping)
                        self._store_column_mapping_threaded(thread_conn, mapping)
                        break
                else:
                    continue
                break
        
        return mappings
    
    def _store_column_mapping_threaded(self, thread_conn, mapping: Dict):
        """Thread-safe mapping storage"""
        try:
            # Check if mapping already exists
            existing = thread_conn.execute("""
                SELECT id FROM _column_mappings 
                WHERE project = ? AND table_name = ? AND original_column = ?
            """, [mapping['project'], mapping['table_name'], mapping['original_column']]).fetchone()
            
            if existing:
                thread_conn.execute("""
                    UPDATE _column_mappings 
                    SET semantic_type = ?, confidence = ?, needs_review = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE project = ? AND table_name = ? AND original_column = ? AND is_override = FALSE
                """, [
                    mapping['semantic_type'],
                    mapping['confidence'],
                    mapping['needs_review'],
                    mapping['project'],
                    mapping['table_name'],
                    mapping['original_column']
                ])
            else:
                thread_conn.execute("""
                    INSERT INTO _column_mappings 
                    (id, project, file_name, table_name, original_column, semantic_type, 
                     confidence, is_override, needs_review)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, [
                    hash(f"{mapping['project']}_{mapping['table_name']}_{mapping['original_column']}") % 2147483647,
                    mapping['project'],
                    mapping['file_name'],
                    mapping['table_name'],
                    mapping['original_column'],
                    mapping['semantic_type'],
                    mapping['confidence'],
                    mapping['is_override'],
                    mapping['needs_review']
                ])
            
            thread_conn.commit()
        except Exception as e:
            logger.warning(f"[MAPPINGS] Failed to store mapping: {e}")
    
    def _update_mapping_job_threaded(self, thread_conn, job_id: str, 
                                      completed_tables: int = None,
                                      mappings_found: int = None, 
                                      needs_review_count: int = None,
                                      status: str = None, 
                                      error_message: str = None):
        """Thread-safe job status update"""
        try:
            updates = ["updated_at = CURRENT_TIMESTAMP"]
            params = []
            
            if completed_tables is not None:
                updates.append("completed_tables = ?")
                params.append(completed_tables)
            if mappings_found is not None:
                updates.append("mappings_found = ?")
                params.append(mappings_found)
            if needs_review_count is not None:
                updates.append("needs_review_count = ?")
                params.append(needs_review_count)
            if status is not None:
                updates.append("status = ?")
                params.append(status)
            if error_message is not None:
                updates.append("error_message = ?")
                params.append(error_message)
            
            params.append(job_id)
            
            thread_conn.execute(f"""
                UPDATE _mapping_jobs 
                SET {', '.join(updates)}
                WHERE id = ?
            """, params)
            thread_conn.commit()
        except Exception as e:
            logger.warning(f"[MAPPING_JOB] Failed to update job: {e}")
    
    def _get_next_version(self, project: str, file_name: str) -> int:
        """Get next version number for a file"""
        result = self.conn.execute("""
            SELECT COALESCE(MAX(version), 0) + 1 
            FROM _load_versions 
            WHERE project = ? AND file_name = ?
        """, [project, file_name]).fetchone()
        return result[0] if result else 1
    
    def _compute_checksum(self, df: pd.DataFrame) -> str:
        """Compute checksum of DataFrame for change detection"""
        # Use first 1000 rows for checksum (performance)
        sample = df.head(1000)
        data_str = sample.to_json()
        return hashlib.md5(data_str.encode()).hexdigest()
    
    def _archive_previous_version(self, project: str, file_name: str, version: int):
        """Archive previous version of tables (rename with version suffix)"""
        try:
            tables = self.conn.execute("""
                SELECT table_name FROM _schema_metadata 
                WHERE project = ? AND file_name = ? AND version = ? AND is_current = TRUE
            """, [project, file_name, version]).fetchall()
            
            for (table_name,) in tables:
                archived_name = f"{table_name}_v{version}"
                try:
                    self.conn.execute(f"ALTER TABLE {table_name} RENAME TO {archived_name}")
                    logger.info(f"Archived {table_name} -> {archived_name}")
                except Exception as e:
                    logger.warning(f"Could not archive {table_name}: {e}")
            
            # Mark as not current
            self.conn.execute("""
                UPDATE _schema_metadata 
                SET is_current = FALSE 
                WHERE project = ? AND file_name = ? AND version = ?
            """, [project, file_name, version])
            
            self.conn.commit()
            
        except Exception as e:
            logger.warning(f"Archive failed: {e}")
    
    # =========================================================================
    # STORE EXCEL - v5.0 with Progress Callback
    # =========================================================================
    
    def store_excel(
        self,
        file_path: str,
        project: str,
        file_name: str,
        encrypt_pii: bool = False,  # Disabled - security at perimeter (Railway, API auth, HTTPS)
        keep_previous_version: bool = True,
        progress_callback: Optional[Callable[[int, str], None]] = None,
        uploaded_by: str = None
    ) -> Dict[str, Any]:
        """
        Store Excel file in DuckDB with encryption and versioning.
        Each sheet becomes a table.
        
        v5.0: Added progress_callback for real-time progress updates.
        v5.21: Added uploaded_by for user tracking.
        
        Args:
            file_path: Path to Excel file
            project: Project name
            file_name: Original filename
            encrypt_pii: Whether to encrypt PII columns
            keep_previous_version: Whether to keep previous version for comparison
            progress_callback: Optional callback function(percent: int, message: str)
            uploaded_by: Email of user who uploaded the file
        
        Returns schema info for Claude to use in queries.
        """
        results = {
            'project': project,
            'file_name': file_name,
            'sheets': [],
            'total_rows': 0,
            'tables_created': [],
            'encrypted_columns': [],
            'version': 1
        }
        
        def report_progress(percent: int, message: str):
            """Helper to safely report progress"""
            if progress_callback:
                try:
                    progress_callback(percent, message)
                except Exception as e:
                    logger.warning(f"Progress callback error: {e}")
        
        try:
            report_progress(5, "Analyzing file structure...")
            
            # =================================================================
            # CLEANUP: Drop ALL existing tables for this file before re-upload
            # This handles cases where split detection names things differently
            # =================================================================
            try:
                existing_tables = self.safe_fetchall("""
                    SELECT table_name FROM _schema_metadata 
                    WHERE project = ? AND file_name = ?
                """, [project, file_name])
                
                if existing_tables:
                    logger.warning(f"[STORE_EXCEL] Cleaning up {len(existing_tables)} existing tables for {file_name}")
                    for (table_name,) in existing_tables:
                        try:
                            self.safe_execute(f'DROP TABLE IF EXISTS "{table_name}"')
                            logger.warning(f"[STORE_EXCEL] Dropped existing table: {table_name}")
                        except Exception as drop_e:
                            logger.warning(f"[STORE_EXCEL] Could not drop {table_name}: {drop_e}")
                    
                    # Clean up metadata
                    self.safe_execute("""
                        DELETE FROM _schema_metadata WHERE project = ? AND file_name = ?
                    """, [project, file_name])
                    
                    # Clean up column profiles
                    self.safe_execute("""
                        DELETE FROM _column_profiles WHERE project = ? AND file_name = ?
                    """, [project, file_name])
                    
                    logger.warning(f"[STORE_EXCEL] Cleanup complete for {file_name}")
            except Exception as cleanup_e:
                logger.warning(f"[STORE_EXCEL] Cleanup step failed (continuing): {cleanup_e}")
            
            # Get version number
            version = self._get_next_version(project, file_name)
            results['version'] = version
            
            # =================================================================
            # SPECIAL HANDLING FOR GLOBAL YEAR-END FILES
            # When uploading a new Year-End file to GLOBAL, mark ALL other
            # year-end files as not current (even with different filenames)
            # =================================================================
            if project.lower() == 'global':
                file_lower = file_name.lower()
                is_year_end = any(kw in file_lower for kw in ['year-end', 'year_end', 'yearend', 'checklist'])
                
                if is_year_end:
                    logger.info(f"[STORE_EXCEL] Detected GLOBAL Year-End file: {file_name}")
                    # Mark ALL other year-end files as not current
                    self.conn.execute("""
                        UPDATE _schema_metadata 
                        SET is_current = FALSE 
                        WHERE LOWER(project) = 'global'
                        AND file_name != ?
                        AND (
                            LOWER(file_name) LIKE '%year-end%'
                            OR LOWER(file_name) LIKE '%year_end%'
                            OR LOWER(file_name) LIKE '%yearend%'
                            OR LOWER(file_name) LIKE '%checklist%'
                        )
                    """, [file_name])
                    logger.info(f"[STORE_EXCEL] Marked other Year-End files as not current")
            
            # If keeping previous version, rename old tables
            if keep_previous_version and version > 1:
                self._archive_previous_version(project, file_name, version - 1)
            
            report_progress(10, "Reading Excel file...")
            logger.warning(f"[STORE_EXCEL] ===== v5.14 SIMPLE FAST PROCESSING =====")
            
            # =================================================================
            # STEP 1: Get sheet names using openpyxl (instant - just reads XML index)
            # =================================================================
            sheet_names_ordered = []
            
            if OPENPYXL_AVAILABLE:
                try:
                    wb_readonly = load_workbook(file_path, read_only=True, data_only=True)
                    sheet_names_ordered = wb_readonly.sheetnames
                    wb_readonly.close()
                    logger.warning(f"[STORE_EXCEL] Found {len(sheet_names_ordered)} sheets: {sheet_names_ordered}")
                except Exception as op_e:
                    logger.warning(f"[STORE_EXCEL] openpyxl failed: {op_e}, using pandas for sheet names")
            
            # Fallback to pandas if openpyxl failed
            if not sheet_names_ordered:
                excel_file = pd.ExcelFile(file_path)
                sheet_names_ordered = excel_file.sheet_names
                logger.warning(f"[STORE_EXCEL] Got {len(sheet_names_ordered)} sheets via pandas fallback")
            
            total_sheets = len(sheet_names_ordered)
            all_encrypted_cols = []
            
            # Detection threshold - only run multi-table detection for small sheets
            DETECTION_THRESHOLD = 1000
            
            # =================================================================
            # STEP 2: Process each sheet
            # =================================================================
            for sheet_idx, sheet_name in enumerate(sheet_names_ordered):
                sheet_start_time = datetime.now()
                sheet_progress = 10 + int((sheet_idx / total_sheets) * 60)
                report_progress(sheet_progress, f"Processing sheet {sheet_idx + 1}/{total_sheets}: {sheet_name}")
                
                try:
                    # ---------------------------------------------------------
                    # Read sheet with pandas (C-level speed, single sheet only)
                    # ---------------------------------------------------------
                    logger.warning(f"[STORE_EXCEL] Reading sheet '{sheet_name}'...")
                    
                    # Quick peek to check size and find header
                    peek_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=50)
                    
                    if peek_df.empty:
                        logger.warning(f"[STORE_EXCEL] Sheet '{sheet_name}' is empty, skipping")
                        continue
                    
                    # Find header row
                    header_row_idx = self._find_header_row(peek_df, max_rows=15)
                    
                    # Read full sheet with detected header
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx)
                    
                    if df.empty:
                        logger.warning(f"[STORE_EXCEL] Sheet '{sheet_name}' is empty after header detection, skipping")
                        continue
                    
                    row_count = len(df)
                    logger.warning(f"[STORE_EXCEL] Sheet '{sheet_name}': {row_count:,} rows x {len(df.columns)} cols")
                    
                    # ---------------------------------------------------------
                    # SMALL SHEETS: Check for multi-table layouts
                    # ---------------------------------------------------------
                    if row_count < DETECTION_THRESHOLD:
                        # Try horizontal split (side-by-side tables)
                        horizontal_tables = self._split_horizontal_tables(file_path, sheet_name)
                        if horizontal_tables:
                            for sub_table_name, sub_df in horizontal_tables:
                                self._store_single_table(
                                    sub_df, project, file_name, 
                                    f"{sheet_name} - {sub_table_name}",
                                    version, encrypt_pii, all_encrypted_cols, results,
                                    uploaded_by=uploaded_by
                                )
                            elapsed = (datetime.now() - sheet_start_time).total_seconds()
                            logger.warning(f"[STORE_EXCEL] Sheet '{sheet_name}' -> {len(horizontal_tables)} horizontal tables ({elapsed:.1f}s)")
                            continue
                        
                        # Try vertical split (stacked tables)
                        vertical_tables = self._split_vertical_tables(file_path, sheet_name)
                        if vertical_tables:
                            for sub_table_name, sub_df in vertical_tables:
                                self._store_single_table(
                                    sub_df, project, file_name,
                                    f"{sheet_name} - {sub_table_name}",
                                    version, encrypt_pii, all_encrypted_cols, results,
                                    uploaded_by=uploaded_by
                                )
                            elapsed = (datetime.now() - sheet_start_time).total_seconds()
                            logger.warning(f"[STORE_EXCEL] Sheet '{sheet_name}' -> {len(vertical_tables)} vertical tables ({elapsed:.1f}s)")
                            continue
                    
                    # ---------------------------------------------------------
                    # Store as single table (standard path for large sheets)
                    # ---------------------------------------------------------
                    self._store_single_table(
                        df, project, file_name, sheet_name,
                        version, encrypt_pii, all_encrypted_cols, results,
                        uploaded_by=uploaded_by
                    )
                    
                    elapsed = (datetime.now() - sheet_start_time).total_seconds()
                    logger.warning(f"[STORE_EXCEL] Sheet '{sheet_name}': {row_count:,} rows stored ({elapsed:.1f}s)")
                    
                except Exception as e:
                    logger.error(f"[STORE_EXCEL] Error processing sheet '{sheet_name}': {e}")
                    import traceback
                    logger.error(traceback.format_exc())
                    continue
            
            self.conn.commit()
            logger.info(f"Stored {len(results['tables_created'])} tables from {file_name}")
            
            # Profile columns (70-85%)
            report_progress(70, "Profiling columns...")
            for i, sheet_info in enumerate(results['sheets']):
                table_name = sheet_info.get('table_name')
                if table_name:
                    profile_progress = 70 + int((i / len(results['sheets'])) * 15)
                    report_progress(profile_progress, f"Profiling {table_name}...")
                    try:
                        profile_result = self.profile_columns_fast(project, table_name, progress_callback)
                        sheet_info['column_profiles'] = profile_result.get('profiles', {})
                        sheet_info['categorical_columns'] = profile_result.get('categorical_columns', [])
                        logger.info(f"[PROFILING] {table_name}: {profile_result.get('columns_profiled', 0)} columns profiled")
                    except Exception as profile_e:
                        logger.warning(f"[PROFILING] Failed for {table_name}: {profile_e}")
                        sheet_info['column_profiles'] = {}
                        sheet_info['categorical_columns'] = []
            
            # Note: Relationship detection is now handled by project_intelligence
            # which runs after upload and stores to Supabase project_relationships
            report_progress(85, "Preparing for intelligence analysis...")
            results['relationships'] = []  # Will be populated by project_intelligence
            
            # Start background column inference (90-95%)
            report_progress(90, "Queueing column inference...")
            # NOTE: Inference queuing disabled to prevent DuckDB thread collision
            # Background inference was colliding with profile_columns_fast, causing segfaults
            # Inference can be triggered manually via the UI after upload completes
            # TODO: Re-enable once proper connection serialization is implemented
            try:
                import uuid
                
                job_id = str(uuid.uuid4())[:8]
                tables_info = results.get('sheets', [])
                
                if tables_info:
                    # Create job record for tracking, but DON'T queue inference
                    self.create_mapping_job(job_id, project, file_name, len(tables_info))
                    results['mapping_job_id'] = job_id
                    
                    # DISABLED: queue_inference_job(self, job_id, project, file_name, tables_info)
                    logger.info(f"[MAPPING] Inference disabled during upload - job {job_id} created for {file_name}")
            except Exception as map_e:
                logger.warning(f"[MAPPING] Failed to create mapping job: {map_e}")
            except Exception as map_e:
                logger.warning(f"[MAPPING] Failed to start inference: {map_e}")
            
            report_progress(95, "Finalizing...")
            
            # =====================================================
            # UPLOAD-TIME VALIDATION
            # Check for bad column names, header detection failures
            # =====================================================
            validation = self._validate_upload_quality(results)
            results['validation'] = validation
            
            if validation['issues']:
                logger.warning(f"[UPLOAD-VALIDATION] {len(validation['issues'])} table(s) have issues")
            
        except Exception as e:
            logger.error(f"Error storing Excel file: {e}")
            raise
        
        return results
    
    def _validate_upload_quality(self, results: Dict) -> Dict:
        """
        Validate uploaded data quality and classify observations.
        
        Classifications:
        - INSIGHT: Configuration observations (not issues) - e.g., optional fields not in use
        - WARNING: Potential concerns worth noting - e.g., suspicious column names
        - ERROR: Actual data quality problems - e.g., header detection failure
        
        Only ERRORS and WARNINGS affect health score.
        INSIGHTS are tracked but don't reduce score.
        
        Returns:
            {
                'status': 'healthy' | 'warning' | 'critical',
                'health_score': int (0-100),
                'tables_checked': int,
                'tables_with_issues': int,
                'total_issues': int,
                'total_insights': int,
                'issues': [...],      # Only actual issues (warnings/errors)
                'insights': [...],    # Configuration observations
                'junk_removed': [...] # Columns auto-removed
            }
        """
        issues = []
        insights = []
        junk_removed = results.get('junk_columns_removed', [])
        
        # Bad patterns that indicate parsing problems (NOT optional fields)
        parsing_problem_patterns = ['nan', 'none']
        
        for sheet_info in results.get('sheets', []):
            table_name = sheet_info.get('table_name', '')
            columns = sheet_info.get('columns', [])
            
            if not columns:
                continue
            
            table_issues = []
            table_insights = []
            
            # Analyze each column
            for col in columns:
                col_name = col.get('name', col) if isinstance(col, dict) else str(col)
                col_lower = col_name.lower()
                
                # Get fill rate if available (default to 1.0 if not tracked)
                fill_rate = col.get('fill_rate', 1.0) if isinstance(col, dict) else 1.0
                
                # Classify the observation
                observation = self._classify_column_observation(col_name, fill_rate, table_name)
                
                if observation:
                    if observation['classification'] == 'insight':
                        table_insights.append(observation)
                    elif observation['classification'] in ['warning', 'error']:
                        table_issues.append({
                            'type': observation.get('warning_type', 'data_quality'),
                            'severity': 'high' if observation['classification'] == 'error' else 'medium',
                            'message': observation['message'],
                            'column': col_name,
                            'fill_rate': fill_rate
                        })
                
                # Check for actual parsing problems (not optional fields)
                if any(pattern == col_lower for pattern in parsing_problem_patterns):
                    table_issues.append({
                        'type': 'parsing_artifact',
                        'severity': 'high',
                        'message': f"Column appears to be parsing artifact: {col_name}",
                        'column': col_name
                    })
            
            # Check for header detection failure (first few columns are numeric)
            col_names = [c.get('name', c) if isinstance(c, dict) else str(c) for c in columns[:5]]
            numeric_first_cols = sum(1 for c in col_names if str(c).replace('.', '').replace('_', '').replace('-', '').isdigit())
            
            if numeric_first_cols >= 3:
                table_issues.append({
                    'type': 'likely_header_failure',
                    'severity': 'high',
                    'message': f"First columns appear numeric - header row may be wrong: {col_names}"
                })
            
            if table_issues:
                issues.append({
                    'table': table_name,
                    'sheet': sheet_info.get('sheet_name', sheet_info.get('name', '')),
                    'problems': table_issues
                })
            
            if table_insights:
                insights.append({
                    'table': table_name,
                    'sheet': sheet_info.get('sheet_name', sheet_info.get('name', '')),
                    'observations': table_insights
                })
        
        tables_checked = len(results.get('sheets', []))
        tables_with_issues = len(issues)
        
        # Calculate health score (only issues affect it, not insights)
        total_issues = sum(len(i['problems']) for i in issues)
        
        if total_issues == 0:
            health_score = 100
            status = 'healthy'
        elif total_issues <= 2:
            health_score = 90
            status = 'healthy'
        elif total_issues <= 5:
            health_score = 75
            status = 'warning'
        elif total_issues <= 10:
            health_score = 60
            status = 'warning'
        else:
            health_score = max(40, 100 - (total_issues * 5))
            status = 'critical'
        
        return {
            'status': status,
            'health_score': health_score,
            'tables_checked': tables_checked,
            'tables_with_issues': tables_with_issues,
            'total_issues': total_issues,
            'total_insights': sum(len(i['observations']) for i in insights),
            'issues': issues,
            'insights': insights,
            'junk_removed': junk_removed
        }
    
    # =========================================================================
    # STORE CSV - v5.0 with Progress Callback
    # =========================================================================
    
    def store_csv(
        self,
        file_path: str,
        project: str,
        file_name: str,
        progress_callback: Optional[Callable[[int, str], None]] = None,
        uploaded_by: str = None
    ) -> Dict[str, Any]:
        """
        Store CSV file in DuckDB.
        
        v5.0: Added progress_callback for real-time progress updates.
        v5.21: Added uploaded_by for user tracking.
        
        Args:
            file_path: Path to CSV file
            project: Project name
            file_name: Original filename
            progress_callback: Optional callback function(percent: int, message: str)
            uploaded_by: Email of user who uploaded the file
            
        Returns:
            Dict with storage results
        """
        def report_progress(percent: int, message: str):
            """Helper to safely report progress"""
            if progress_callback:
                try:
                    progress_callback(percent, message)
                except Exception as e:
                    logger.warning(f"Progress callback error: {e}")
        
        try:
            report_progress(5, "Reading CSV file...")
            
            # Quick row count estimate for progress tracking
            estimated_rows = 0
            try:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    estimated_rows = sum(1 for _ in f) - 1  # Subtract header
                logger.info(f"[CSV] Estimated {estimated_rows:,} rows")
            except Exception:
                pass
            
            report_progress(10, f"Loading {estimated_rows:,} rows..." if estimated_rows else "Loading data...")
            
            df = pd.read_csv(file_path)
            df = df.dropna(how='all').dropna(axis=1, how='all')
            
            report_progress(30, "Sanitizing column names...")
            
            # Sanitize column names
            df.columns = [self._sanitize_name(str(c)) for c in df.columns]
            
            # FORCE ALL COLUMNS TO STRING to avoid type inference issues
            for col in df.columns:
                df[col] = df[col].fillna('').astype(str)
                df[col] = df[col].replace({'nan': '', 'None': '', 'NaT': ''})
            
            # Remove junk columns before storage
            df, removed_junk = self._remove_junk_columns(df)
            junk_columns_removed = removed_junk if removed_junk else []
            
            report_progress(40, "Creating DuckDB table...")
            
            table_name = self._generate_table_name(project, file_name, 'data')
            display_name = self._generate_display_name(file_name, 'data')
            
            # Derive entity_type and category from context (CSVs use file_name)
            entity_meta = self._derive_entity_metadata(file_name, 'data')
            
            # Create table (thread-safe)
            self.safe_create_table_from_df(table_name, df)
            
            report_progress(50, "Detecting key columns...")
            
            # Detect keys
            likely_keys = self._detect_key_columns(df)
            
            # Store metadata
            columns_info = [
                {'name': col, 'type': str(df[col].dtype)}
                for col in df.columns
            ]
            
            self.safe_execute("""
                INSERT INTO _schema_metadata 
                (id, project, file_name, sheet_name, table_name, display_name, entity_type, category, columns, column_count, row_count, likely_keys, truth_type, uploaded_by, is_current)
                VALUES (nextval('schema_metadata_seq'), ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, TRUE)
            """, [
                project, file_name, 'data', table_name, display_name,
                entity_meta.get('entity_type'), entity_meta.get('category'),
                json.dumps(columns_info), len(df.columns), len(df), json.dumps(likely_keys),
                None,  # truth_type - set by smart_router if provided
                uploaded_by
            ])
            
            self.conn.commit()
            
            # Profile columns for CSV
            report_progress(60, "Profiling columns...")
            profile_result = {}
            try:
                profile_result = self.profile_columns_fast(project, table_name, progress_callback)
                logger.info(f"[PROFILING] CSV {table_name}: {profile_result.get('columns_profiled', 0)} columns profiled")
            except Exception as profile_e:
                logger.warning(f"[PROFILING] Failed for CSV {table_name}: {profile_e}")
            
            report_progress(90, "Finalizing...")
            
            return {
                'project': project,
                'file_name': file_name,
                'table_name': table_name,
                'display_name': display_name,
                'columns': list(df.columns),
                'column_count': len(df.columns),
                'row_count': len(df),
                'likely_keys': likely_keys,
                'column_profiles': profile_result.get('profiles', {}),
                'categorical_columns': profile_result.get('categorical_columns', []),
                'junk_columns_removed': junk_columns_removed
            }
            
        except Exception as e:
            logger.error(f"Error storing CSV: {e}")
            raise
    
    # =========================================================================
    # STORE DATAFRAME - For PDF and other sources that provide DataFrames
    # =========================================================================
    
    def store_dataframe(
        self,
        df: pd.DataFrame,
        project: str,
        file_name: str,
        sheet_name: str = 'data',
        source_type: str = 'pdf',
        uploaded_by: str = None
    ) -> Dict[str, Any]:
        """
        Store a DataFrame in DuckDB with proper metadata.
        
        Used by smart_pdf_analyzer and other sources that produce DataFrames
        directly instead of files.
        
        v5.21: Added uploaded_by for user tracking.
        
        Args:
            df: DataFrame to store
            project: Project name
            file_name: Original source filename
            sheet_name: Sheet/section name (default 'data')
            source_type: Source type for metadata (default 'pdf')
            uploaded_by: Email of user who uploaded the file
            
        Returns:
            Dict with storage results including table_name, row_count, etc.
        """
        results = {
            'success': False,
            'project': project,
            'file_name': file_name,
            'table_name': None,
            'tables_created': [],
            'row_count': 0,
            'total_rows': 0,
            'columns': [],
            'source_type': source_type
        }
        
        try:
            # CRITICAL: Clear any pending uncommitted transaction
            # This ensures we're not affected by uncommitted deletes/changes from other operations
            try:
                with self._db_lock:
                    self.conn.rollback()
                    logger.warning("[STORE_DF] Cleared any pending transaction via rollback")
            except Exception as rb_e:
                logger.warning(f"[STORE_DF] Rollback note (OK if no transaction): {rb_e}")
            
            if df is None or df.empty:
                results['error'] = 'DataFrame is empty'
                return results
            
            # Clean up DataFrame
            df = df.dropna(how='all').dropna(axis=1, how='all')
            
            if df.empty:
                results['error'] = 'DataFrame is empty after cleanup'
                return results
            
            # Sanitize column names
            df.columns = [self._sanitize_name(str(c)) for c in df.columns]
            
            # Force all columns to string for consistency
            for col in df.columns:
                df[col] = df[col].fillna('').astype(str)
                df[col] = df[col].replace({'nan': '', 'None': '', 'NaT': ''})
            
            # Generate table name and display name
            table_name = self._generate_table_name(project, file_name, sheet_name)
            display_name = self._generate_display_name(file_name, sheet_name)
            
            # Derive entity_type and category from context
            entity_meta = self._derive_entity_metadata(file_name, sheet_name)
            
            # Clean up any existing data for this file
            try:
                logger.warning(f"[STORE_DF] CLEANUP: Looking for existing data for {project}/{file_name}")
                existing = self.safe_fetchall("""
                    SELECT table_name FROM _schema_metadata 
                    WHERE project = ? AND file_name = ?
                """, [project, file_name])
                
                logger.warning(f"[STORE_DF] CLEANUP: Found {len(existing)} existing tables")
                
                for (old_table,) in existing:
                    try:
                        self.safe_execute(f'DROP TABLE IF EXISTS "{old_table}"')
                        logger.warning(f"[STORE_DF] CLEANUP: Dropped existing table: {old_table}")
                    except Exception:
                        pass
                
                self.safe_execute("""
                    DELETE FROM _schema_metadata WHERE project = ? AND file_name = ?
                """, [project, file_name])
                logger.warning(f"[STORE_DF] CLEANUP: Deleted metadata for {project}/{file_name}")
            except Exception as cleanup_e:
                logger.warning(f"[STORE_DF] Cleanup warning: {cleanup_e}")
            
            # Get version number
            version = self._get_next_version(project, file_name)
            
            # Create table (thread-safe)
            self.safe_create_table_from_df(table_name, df)
            
            # Build columns info
            columns_info = [
                {'name': col, 'type': 'VARCHAR', 'encrypted': False}
                for col in df.columns
            ]
            
            # Store metadata - DuckDB is in autocommit mode, so INSERT commits immediately
            logger.warning(f"[STORE_DF] Inserting _schema_metadata for {table_name}, conn_id={id(self.conn)}")
            
            with self._db_lock:
                self.conn.execute("""
                    INSERT INTO _schema_metadata 
                    (id, project, file_name, sheet_name, table_name, display_name, entity_type, category, columns, column_count, row_count, likely_keys, encrypted_columns, truth_type, uploaded_by, version, is_current)
                    VALUES (nextval('schema_metadata_seq'), ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, TRUE)
                """, [
                    project,
                    file_name,
                    sheet_name,
                    table_name,
                    display_name,
                    entity_meta.get('entity_type'),
                    entity_meta.get('category'),
                    json.dumps(columns_info),
                    len(df.columns),
                    len(df),
                    json.dumps([]),  # likely_keys
                    json.dumps([]),  # encrypted_columns
                    None,  # truth_type - set by smart_router if provided
                    uploaded_by,
                    version
                ])
                logger.warning("[STORE_DF] INSERT executed")
                
                # Immediate verification BEFORE leaving the lock
                immed_count = self.conn.execute("SELECT COUNT(*) FROM _schema_metadata").fetchone()[0]
                logger.warning(f"[STORE_DF] IMMEDIATE CHECK: {immed_count} rows in _schema_metadata")
                
                # Force checkpoint to persist
                self.conn.execute("CHECKPOINT")
                logger.warning("[STORE_DF] CHECKPOINT complete")
                
                # Verify after checkpoint
                post_count = self.conn.execute("SELECT COUNT(*) FROM _schema_metadata").fetchone()[0]
                logger.warning(f"[STORE_DF] POST-CHECKPOINT: {post_count} rows in _schema_metadata")
            
            # Force DuckDB to persist by closing and reopening connection
            with self._db_lock:
                import os
                pre_size = os.path.getsize(self.db_path) if os.path.exists(self.db_path) else 0
                logger.warning(f"[STORE_DF] PRE-RECONNECT: db_path={self.db_path}, file_size={pre_size} bytes")
                
                # Check row count before close
                pre_close_count = self.conn.execute("SELECT COUNT(*) FROM _schema_metadata").fetchone()[0]
                logger.warning(f"[STORE_DF] PRE-CLOSE: _schema_metadata has {pre_close_count} rows")
                
                # List tables before close
                pre_tables = [t[0] for t in self.conn.execute("SHOW TABLES").fetchall()]
                logger.warning(f"[STORE_DF] PRE-CLOSE tables: {pre_tables[:10]}")
                
                old_conn = self.conn
                try:
                    old_conn.close()
                    logger.warning("[STORE_DF] Closed old connection")
                except Exception as close_e:
                    logger.warning(f"[STORE_DF] Close error: {close_e}")
                
                # Check file size after close
                post_close_size = os.path.getsize(self.db_path) if os.path.exists(self.db_path) else 0
                logger.warning(f"[STORE_DF] POST-CLOSE: file_size={post_close_size} bytes")
                    
                self.conn = duckdb.connect(self.db_path)
                logger.warning(f"[STORE_DF] Opened new connection, conn_id={id(self.conn)}")
                
                # Check row count after reconnect
                post_reconnect_count = self.conn.execute("SELECT COUNT(*) FROM _schema_metadata").fetchone()[0]
                logger.warning(f"[STORE_DF] POST-RECONNECT: _schema_metadata has {post_reconnect_count} rows")
                
                # List tables after reconnect
                post_tables = [t[0] for t in self.conn.execute("SHOW TABLES").fetchall()]
                logger.warning(f"[STORE_DF] POST-RECONNECT tables: {post_tables[:10]}")
            
            # Final verification using safe_fetchall (same path as status.py)
            try:
                test_result = self.safe_fetchall("SELECT COUNT(*) FROM _schema_metadata WHERE is_current = TRUE")
                logger.warning(f"[STORE_DF] FINAL CHECK via safe_fetchall: {test_result[0][0] if test_result else 'None'} rows with is_current=TRUE")
            except Exception as sf_e:
                logger.warning(f"[STORE_DF] FINAL CHECK failed: {sf_e}")
            
            logger.warning(f"[STORE_DF] Stored {len(df)} rows to {table_name} from {source_type}")
            
            results['success'] = True
            results['table_name'] = table_name
            results['display_name'] = display_name
            results['tables_created'] = [table_name]
            results['row_count'] = len(df)
            results['total_rows'] = len(df)
            results['columns'] = list(df.columns)
            results['column_count'] = len(df.columns)
            results['version'] = version
            
            return results
            
        except Exception as e:
            logger.error(f"[STORE_DF] Error storing DataFrame: {e}")
            import traceback
            logger.error(traceback.format_exc())
            results['error'] = str(e)
            return results
    
    # =========================================================================
    # COLUMN PROFILING - v5.0 SQL-Based (OPTIMIZED)
    # =========================================================================
    
    def profile_columns_fast(
        self, 
        project: str, 
        table_name: str,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> Dict[str, Any]:
        """
        v5.0 OPTIMIZED: Profile columns using SQL aggregates instead of loading DataFrame.
        
        For large tables (>50K rows), uses sampling for type inference.
        Basic stats (count, distinct, min, max) always use full table via SQL.
        
        Args:
            project: Project name
            table_name: DuckDB table name to profile
            progress_callback: Optional callback for progress updates
            
        Returns:
            Dict with profiling results and summary
        """
        logger.info(f"[PROFILING-FAST] Starting SQL-based profiling for {table_name}")
        
        result = {
            'table_name': table_name,
            'columns_profiled': 0,
            'categorical_columns': [],
            'numeric_columns': [],
            'date_columns': [],
            'profiles': {},
            'method': 'sql_optimized'
        }
        
        try:
            # Use context manager for thread-safe DuckDB access
            with self._db_lock:
                # Get row count first
                row_count = self.conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
                
                if row_count == 0:
                    logger.warning(f"[PROFILING-FAST] Table {table_name} is empty")
                    return result
                
                # Get column names
                columns_result = self.conn.execute(f"DESCRIBE {table_name}").fetchall()
                columns = [col[0] for col in columns_result]
                
                logger.info(f"[PROFILING-FAST] Table {table_name}: {row_count:,} rows, {len(columns)} columns")
                
                # Determine if we need sampling for type inference
                use_sampling = row_count > LARGE_TABLE_THRESHOLD
                if use_sampling:
                    logger.info(f"[PROFILING-FAST] Large table - will sample {PROFILE_SAMPLE_SIZE:,} rows for type inference")
                
                # Profile each column using SQL
                for col_idx, col in enumerate(columns):
                    try:
                        profile = self._profile_column_sql(
                            table_name, col, project, row_count, use_sampling
                        )
                        
                        # Store in database
                        self._store_column_profile(profile)
                        
                        # Track by type
                        result['profiles'][col] = profile
                        result['columns_profiled'] += 1
                        
                        if profile['inferred_type'] == 'categorical' or profile.get('is_categorical'):
                            result['categorical_columns'].append({
                                'name': col,
                                'distinct_count': profile['distinct_count'],
                                'values': profile.get('distinct_values', [])
                            })
                        elif profile['inferred_type'] == 'numeric':
                            result['numeric_columns'].append({
                                'name': col,
                                'min': profile.get('min_value'),
                                'max': profile.get('max_value'),
                                'mean': profile.get('mean_value')
                            })
                        elif profile['inferred_type'] == 'date':
                            result['date_columns'].append({
                                'name': col,
                                'min_date': profile.get('min_date'),
                                'max_date': profile.get('max_date')
                            })
                            
                    except Exception as col_e:
                        logger.warning(f"[PROFILING-FAST] Failed to profile column {col}: {col_e}")
                        continue
                
                logger.info(f"[PROFILING-FAST] Completed {table_name}: {result['columns_profiled']} columns, "
                           f"{len(result['categorical_columns'])} categorical, "
                           f"{len(result['numeric_columns'])} numeric")
            
            return result
            
        except Exception as e:
            logger.error(f"[PROFILING-FAST] Failed for {table_name}: {e}")
            result['error'] = str(e)
            return result
    
    def _profile_column_sql(
        self, 
        table_name: str, 
        col: str, 
        project: str,
        total_rows: int,
        use_sampling: bool = False
    ) -> Dict[str, Any]:
        """
        Profile a single column using SQL aggregates.
        
        This is much faster than loading the entire DataFrame because:
        1. DuckDB computes aggregates directly on disk
        2. Only metadata comes back over the wire
        3. For large tables, type inference uses sampling
        """
        profile = {
            'project': project,
            'table_name': table_name,
            'column_name': col,
            'original_dtype': 'VARCHAR',
            'total_count': total_rows,
            'null_count': 0,
            'distinct_count': 0,
            'inferred_type': 'text',
            'is_likely_key': False,
            'is_categorical': False,
            'distinct_values': None,
            'value_distribution': None,
            'min_value': None,
            'max_value': None,
            'mean_value': None,
            'min_date': None,
            'max_date': None,
            'sample_values': [],
            'filter_category': None,
            'filter_priority': 0
        }
        
        try:
            # First, get the column's actual data type
            col_type_result = self.conn.execute(f"""
                SELECT data_type FROM information_schema.columns 
                WHERE table_name = '{table_name}' AND column_name = '{col}'
            """).fetchone()
            
            col_type = col_type_result[0].upper() if col_type_result else 'VARCHAR'
            profile['original_dtype'] = col_type
            
            # Determine if column is numeric (no TRIM needed)
            is_numeric = any(t in col_type for t in ['INT', 'DOUBLE', 'FLOAT', 'DECIMAL', 'NUMERIC', 'BIGINT', 'SMALLINT', 'REAL'])
            
            # Get basic stats via SQL - different query for numeric vs string
            if is_numeric:
                # Numeric columns - just check NULL
                stats = self.conn.execute(f"""
                    SELECT 
                        COUNT(*) as total,
                        SUM(CASE WHEN "{col}" IS NULL THEN 1 ELSE 0 END) as null_count,
                        COUNT(DISTINCT "{col}") as distinct_count
                    FROM {table_name}
                """).fetchone()
            else:
                # String columns - check NULL, empty, and 'nan'
                stats = self.conn.execute(f"""
                    SELECT 
                        COUNT(*) as total,
                        SUM(CASE WHEN "{col}" IS NULL OR TRIM(CAST("{col}" AS VARCHAR)) = '' OR CAST("{col}" AS VARCHAR) = 'nan' THEN 1 ELSE 0 END) as null_count,
                        COUNT(DISTINCT "{col}") as distinct_count
                    FROM {table_name}
                """).fetchone()
            
            # Guard against None result
            if stats is None:
                logger.warning(f"[PROFILING-SQL] No stats returned for {table_name}.{col}")
                return profile
            
            profile['total_count'] = stats[0] if stats[0] is not None else 0
            profile['null_count'] = stats[1] or 0
            profile['distinct_count'] = stats[2] or 0
            
            # Check if likely key (>95% unique)
            non_null_count = profile['total_count'] - profile['null_count']
            if non_null_count > 0:
                uniqueness = profile['distinct_count'] / non_null_count
                profile['is_likely_key'] = uniqueness > 0.95 and profile['distinct_count'] > 10
            
            # Get sample values - different query for numeric vs string
            if is_numeric:
                samples = self.conn.execute(f"""
                    SELECT DISTINCT "{col}" 
                    FROM {table_name} 
                    WHERE "{col}" IS NOT NULL
                    LIMIT 5
                """).fetchall()
            else:
                samples = self.conn.execute(f"""
                    SELECT DISTINCT "{col}" 
                    FROM {table_name} 
                    WHERE "{col}" IS NOT NULL AND TRIM(CAST("{col}" AS VARCHAR)) != '' AND CAST("{col}" AS VARCHAR) != 'nan'
                    LIMIT 5
                """).fetchall()
            profile['sample_values'] = [str(s[0]) for s in samples]
            
            # If already numeric type, set profile accordingly
            if is_numeric:
                profile['inferred_type'] = 'numeric'
                try:
                    numeric_stats = self.conn.execute(f"""
                        SELECT 
                            MIN("{col}") as min_val,
                            MAX("{col}") as max_val,
                            AVG("{col}") as mean_val
                        FROM {table_name}
                        WHERE "{col}" IS NOT NULL
                    """).fetchone()
                    profile['min_value'] = float(numeric_stats[0]) if numeric_stats[0] is not None else None
                    profile['max_value'] = float(numeric_stats[1]) if numeric_stats[1] is not None else None
                    profile['mean_value'] = float(numeric_stats[2]) if numeric_stats[2] is not None else None
                except Exception:
                    pass
            
            # For categorical (low cardinality), get distinct values and distribution
            elif profile['distinct_count'] <= 100:
                profile['is_categorical'] = True
                profile['inferred_type'] = 'categorical'
                
                # Get all distinct values
                distinct_result = self.conn.execute(f"""
                    SELECT DISTINCT "{col}"
                    FROM {table_name}
                    WHERE "{col}" IS NOT NULL AND TRIM(CAST("{col}" AS VARCHAR)) != '' AND CAST("{col}" AS VARCHAR) != 'nan'
                    ORDER BY "{col}"
                """).fetchall()
                profile['distinct_values'] = sorted([str(d[0]) for d in distinct_result if d[0]])
                
                # Get value distribution
                dist_result = self.conn.execute(f"""
                    SELECT "{col}", COUNT(*) as cnt
                    FROM {table_name}
                    WHERE "{col}" IS NOT NULL
                    GROUP BY "{col}"
                    ORDER BY cnt DESC
                    LIMIT 100
                """).fetchall()
                profile['value_distribution'] = {str(d[0]): d[1] for d in dist_result}
                
                # Check for boolean-like values
                values_set = set(profile['distinct_values'])
                values_upper = set(v.upper() for v in values_set if v)
                bool_patterns = [
                    {'Y', 'N'}, {'YES', 'NO'}, {'TRUE', 'FALSE'}, {'1', '0'},
                    {'T', 'F'}, {'ACTIVE', 'INACTIVE'}
                ]
                for pattern in bool_patterns:
                    if values_upper == pattern or values_upper <= pattern:
                        profile['inferred_type'] = 'boolean'
                        break
            
            # Try numeric detection via SQL (for string columns that might be numeric)
            if profile['inferred_type'] not in ['categorical', 'boolean', 'numeric']:
                try:
                    # Try to cast and get numeric stats
                    numeric_stats = self.conn.execute(f"""
                        SELECT 
                            MIN(TRY_CAST(REPLACE(REPLACE(CAST("{col}" AS VARCHAR), ',', ''), '$', '') AS DOUBLE)) as min_val,
                            MAX(TRY_CAST(REPLACE(REPLACE(CAST("{col}" AS VARCHAR), ',', ''), '$', '') AS DOUBLE)) as max_val,
                            AVG(TRY_CAST(REPLACE(REPLACE(CAST("{col}" AS VARCHAR), ',', ''), '$', '') AS DOUBLE)) as mean_val,
                            COUNT(TRY_CAST(REPLACE(REPLACE(CAST("{col}" AS VARCHAR), ',', ''), '$', '') AS DOUBLE)) as numeric_count
                        FROM {table_name}
                        WHERE "{col}" IS NOT NULL AND TRIM(CAST("{col}" AS VARCHAR)) != ''
                    """).fetchone()
                    
                    if numeric_stats and numeric_stats[3] and numeric_stats[3] > non_null_count * 0.8:
                        profile['inferred_type'] = 'numeric'
                        profile['min_value'] = float(numeric_stats[0]) if numeric_stats[0] is not None else None
                        profile['max_value'] = float(numeric_stats[1]) if numeric_stats[1] is not None else None
                        profile['mean_value'] = float(numeric_stats[2]) if numeric_stats[2] is not None else None
                except Exception:
                    pass
            
            # Try date detection if not already typed
            if profile['inferred_type'] == 'text':
                try:
                    date_stats = self.conn.execute(f"""
                        SELECT 
                            MIN(TRY_CAST("{col}" AS DATE)) as min_date,
                            MAX(TRY_CAST("{col}" AS DATE)) as max_date,
                            COUNT(TRY_CAST("{col}" AS DATE)) as date_count
                        FROM {table_name}
                        WHERE "{col}" IS NOT NULL AND TRIM(CAST("{col}" AS VARCHAR)) != ''
                    """).fetchone()
                    
                    if date_stats and date_stats[2] and date_stats[2] > non_null_count * 0.8:
                        profile['inferred_type'] = 'date'
                        profile['min_date'] = str(date_stats[0]) if date_stats[0] else None
                        profile['max_date'] = str(date_stats[1]) if date_stats[1] else None
                except Exception:
                    pass
            
            # Detect filter category
            if profile.get('distinct_values'):
                values_set = set(str(v).upper().strip() for v in profile['distinct_values'] if v)
                profile = self._detect_filter_category(col, profile, profile['distinct_values'], project)
            
        except Exception as e:
            logger.warning(f"[PROFILING-SQL] Error profiling {col}: {e}")
        
        return profile
    
    # Keep the original profile_columns for backward compatibility
    def profile_columns(self, project: str, table_name: str) -> Dict[str, Any]:
        """
        Original profile_columns method - now delegates to fast version.
        Kept for backward compatibility.
        """
        return self.profile_columns_fast(project, table_name)
    
    def _profile_single_column(self, df: pd.DataFrame, col: str, project: str = None) -> Dict[str, Any]:
        """
        Profile a single column and return statistics.
        LEGACY METHOD - kept for compatibility, use _profile_column_sql for new code.
        """
        series = df[col]
        
        profile = {
            'column_name': col,
            'original_dtype': str(series.dtype),
            'total_count': len(series),
            'null_count': int(series.isna().sum() + (series == '').sum()),
            'distinct_count': 0,
            'inferred_type': 'text',
            'is_likely_key': False,
            'is_categorical': False,
            'distinct_values': None,
            'value_distribution': None,
            'min_value': None,
            'max_value': None,
            'mean_value': None,
            'min_date': None,
            'max_date': None,
            'sample_values': [],
            'filter_category': None,
            'filter_priority': 0
        }
        
        # Get non-null, non-empty values
        non_null = series[series.notna() & (series != '') & (series.astype(str) != 'nan')]
        
        if len(non_null) == 0:
            profile['inferred_type'] = 'empty'
            return profile
        
        # Get distinct values
        distinct_values = non_null.unique()
        profile['distinct_count'] = len(distinct_values)
        
        # Sample values (first 5 unique)
        profile['sample_values'] = [str(v) for v in distinct_values[:5]]
        
        # Check if this is a likely key column (high uniqueness)
        uniqueness_ratio = profile['distinct_count'] / len(non_null) if len(non_null) > 0 else 0
        profile['is_likely_key'] = uniqueness_ratio > 0.95 and profile['distinct_count'] > 10
        
        # Try to infer type and get type-specific stats
        profile = self._infer_column_type(non_null, distinct_values, profile)
        
        # DETECT FILTER CATEGORY (data-driven with lookup matching)
        profile = self._detect_filter_category(col, profile, distinct_values, project)
        
        return profile
    
    def _detect_filter_category(self, col_name: str, profile: Dict, distinct_values, project: str = None) -> Dict:
        """
        Detect if this column is a common filter dimension using INTELLIGENT LOOKUP MATCHING.
        
        Data-driven approach:
        1. Load existing lookups from _intelligence_lookups (from config validation reports)
        2. Compare column values against lookup codes
        3. If significant overlap (50%+), use that lookup's type as filter_category
        4. Column name is secondary hint, VALUES are source of truth
        
        Categories: status, company, organization, location, pay_type, employee_type, job
        """
        col_lower = col_name.lower()
        distinct_count = profile.get('distinct_count', 0)
        inferred_type = profile.get('inferred_type', 'text')
        
        # Skip if too many values (not a filter dimension) or no values
        if distinct_count > 500 or distinct_count == 0 or distinct_values is None:
            return profile
        
        # Handle both list and array-like distinct_values
        try:
            if hasattr(distinct_values, 'tolist'):
                distinct_values = distinct_values.tolist()
            values_set = set(str(v).upper().strip() for v in distinct_values if v is not None and str(v).strip())
        except Exception:
            return profile
        
        if not values_set:
            return profile
        
        # =================================================================
        # STEP 1: Try to match against existing lookups (DATA-DRIVEN)
        # =================================================================
        lookup_match = self._match_column_to_lookup(project, col_name, values_set)
        if lookup_match:
            profile['filter_category'] = lookup_match['category']
            profile['filter_priority'] = lookup_match['priority']
            profile['matched_lookup'] = lookup_match['lookup_name']
            logger.info(f"[PROFILING] Column '{col_name}' matched lookup '{lookup_match['lookup_name']}' → category '{lookup_match['category']}' ({lookup_match['match_pct']:.0%} overlap)")
            return profile
        
        # =================================================================
        # STEP 2: Column name + value pattern hints (fallback only)
        # These are broad semantic hints, not hardcoded values
        # =================================================================
        
        # STATUS: Look for status-related column names with low cardinality
        if distinct_count <= 10:
            status_name_hints = ['employment_status', 'emp_status', 'employee_status', 'status_code', 
                                 'active_status', 'work_status', 'termination']
            # Negative patterns - NOT employment status
            status_negative = ['marital', 'tax_status', 'benefit_status', 'visa', 'citizenship',
                              'military', 'insurance', 'aca', 'cobra', 'union']
            
            if any(h in col_lower for h in status_name_hints) and not any(n in col_lower for n in status_negative):
                profile['filter_category'] = 'status'
                profile['filter_priority'] = 90
                return profile
        
        # COMPANY: Company/entity columns
        if 2 <= distinct_count <= 50:
            company_hints = ['company_code', 'company_name', 'company_id', 'legal_entity', 
                            'home_company', 'employer', 'entity_code', 'payroll_company']
            company_negative = ['routing', 'bank', 'account', 'aba', 'swift', 'insurance', 'benefit', 'employee']
            
            if any(h in col_lower for h in company_hints) and not any(n in col_lower for n in company_negative):
                profile['filter_category'] = 'company'
                profile['filter_priority'] = 80
                return profile
        
        # ORGANIZATION: Dept/division columns
        if 2 <= distinct_count <= 200:
            org_hints = ['department', 'dept', 'division', 'cost_center', 'business_unit',
                        'org_level', 'org_code', 'segment', 'team', 'group', 'profit_center']
            org_negative = ['routing', 'bank', 'account', 'employee', 'deduction']
            
            if any(h in col_lower for h in org_hints) and not any(n in col_lower for n in org_negative):
                profile['filter_category'] = 'organization'
                profile['filter_priority'] = 70
                return profile
        
        # LOCATION: State/site columns - MUST have location-specific name
        if 2 <= distinct_count <= 500:
            location_hints = ['state', 'province', 'stateprovince', 'work_state', 'home_state',
                             'work_location', 'location_code', 'country', 'city', 'county', 
                             'region', 'site', 'facility', 'office', 'address_state', 
                             'mail_state', 'tax_state', 'sui_state', 'work_site', 'geo']
            location_negative = ['routing', 'bank', 'account', 'aba', 'swift', 'transit', 
                                'sort_code', 'bsb', 'branch_number', 'employee', 'deduction', 'benefit']
            
            if any(h in col_lower for h in location_hints) and not any(n in col_lower for n in location_negative):
                profile['filter_category'] = 'location'
                profile['filter_priority'] = 60
                return profile
        
        # PAY TYPE: Hourly/salary type columns
        if distinct_count <= 10:
            pay_hints = ['hourly_salary', 'pay_type', 'flsa', 'exempt', 'fullpart', 
                        'full_part', 'ft_pt', 'salary_hourly', 'pay_class', 'wage_type']
            
            if any(h in col_lower for h in pay_hints):
                profile['filter_category'] = 'pay_type'
                profile['filter_priority'] = 50
                return profile
        
        # EMPLOYEE TYPE: Worker type columns
        if distinct_count <= 20:
            emp_type_hints = ['employee_type', 'worker_type', 'emp_type', 'employment_type',
                            'worker_category', 'contingent', 'staff_type']
            
            if any(h in col_lower for h in emp_type_hints):
                profile['filter_category'] = 'employee_type'
                profile['filter_priority'] = 55
                return profile
        
        # JOB: Job code/family columns
        if 2 <= distinct_count <= 500:
            job_hints = ['job_code', 'job_title', 'job_family', 'job_grade', 'position_code',
                        'occupation', 'job_class', 'pay_grade', 'position_title']
            
            if any(h in col_lower for h in job_hints):
                profile['filter_category'] = 'job'
                profile['filter_priority'] = 40
                return profile
        
        return profile
    
    def _match_column_to_lookup(self, project: str, col_name: str, values_set: set) -> Optional[Dict]:
        """
        Match column values against existing lookups from _intelligence_lookups.
        
        Returns match info if 50%+ of column values exist in a lookup's codes.
        This is the DATA-DRIVEN approach - we use actual uploaded configuration data.
        """
        if not project or not values_set:
            return None
        
        try:
            # Check if _intelligence_lookups table exists
            tables = self.conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='_intelligence_lookups'").fetchall()
            if not tables:
                # Try DuckDB syntax
                try:
                    tables = self.conn.execute("SELECT table_name FROM information_schema.tables WHERE table_name = '_intelligence_lookups'").fetchall()
                except Exception:
                    pass
            
            if not tables:
                return None
            
            # Load all lookups for this project
            lookups = self.conn.execute("""
                SELECT table_name, code_column, lookup_type, lookup_data_json, entry_count
                FROM _intelligence_lookups
                WHERE project_name = ?
            """, [project]).fetchall()
            
            if not lookups:
                return None
            
            best_match = None
            best_match_pct = 0.5  # Minimum 50% overlap required
            
            for lookup_row in lookups:
                table_name, code_col, lookup_type, lookup_data_json, entry_count = lookup_row
                
                if not lookup_data_json:
                    continue
                
                try:
                    lookup_data = json.loads(lookup_data_json)
                    lookup_codes = set(str(k).upper().strip() for k in lookup_data.keys())
                    
                    if not lookup_codes:
                        continue
                    
                    # Calculate overlap
                    overlap = values_set & lookup_codes
                    match_pct = len(overlap) / len(values_set) if values_set else 0
                    
                    if match_pct > best_match_pct:
                        best_match_pct = match_pct
                        
                        # Map lookup_type to filter_category
                        category = self._lookup_type_to_filter_category(lookup_type)
                        priority = self._get_category_priority(category)
                        
                        best_match = {
                            'category': category,
                            'priority': priority,
                            'lookup_name': f"{table_name}.{code_col}",
                            'lookup_type': lookup_type,
                            'match_pct': match_pct,
                            'matched_values': len(overlap)
                        }
                
                except (json.JSONDecodeError, Exception) as e:
                    logger.debug(f"[PROFILING] Failed to parse lookup {table_name}: {e}")
                    continue
            
            return best_match
            
        except Exception as e:
            logger.debug(f"[PROFILING] Lookup matching failed: {e}")
            return None
    
    def _lookup_type_to_filter_category(self, lookup_type: str) -> str:
        """Map lookup_type from intelligence to filter_category."""
        mapping = {
            'location': 'location',
            'site': 'location',
            'state': 'location',
            'country': 'location',
            'region': 'location',
            
            'department': 'organization',
            'division': 'organization',
            'cost_center': 'organization',
            'org': 'organization',
            
            'company': 'company',
            'entity': 'company',
            'employer': 'company',
            
            'status': 'status',
            'employment_status': 'status',
            
            'pay_group': 'pay_type',
            'pay_type': 'pay_type',
            'flsa': 'pay_type',
            
            'job': 'job',
            'position': 'job',
            'occupation': 'job',
            
            'employee_type': 'employee_type',
            'worker_type': 'employee_type',
        }
        return mapping.get(lookup_type.lower(), lookup_type)
    
    def _get_category_priority(self, category: str) -> int:
        """Get filter priority for a category."""
        priorities = {
            'status': 100,
            'company': 80,
            'organization': 70,
            'location': 60,
            'employee_type': 55,
            'pay_type': 50,
            'job': 40
        }
        return priorities.get(category, 30)
    
    def _infer_column_type(self, series: pd.Series, distinct_values, profile: Dict) -> Dict:
        """
        Infer the semantic type of a column and compute type-specific statistics.
        """
        # Try numeric first
        try:
            # Filter out obvious non-numeric
            numeric_series = pd.to_numeric(series.astype(str).str.replace(',', '').str.replace('$', '').str.strip(), errors='coerce')
            numeric_valid = numeric_series.dropna()
            
            if len(numeric_valid) >= len(series) * 0.8:  # 80% numeric
                profile['inferred_type'] = 'numeric'
                profile['min_value'] = float(numeric_valid.min())
                profile['max_value'] = float(numeric_valid.max())
                profile['mean_value'] = float(numeric_valid.mean())
                return profile
        except Exception:
            pass
        
        # Try date
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", UserWarning)
                date_series = pd.to_datetime(series, errors='coerce')
            date_valid = date_series.dropna()
            
            if len(date_valid) >= len(series) * 0.8:  # 80% valid dates
                profile['inferred_type'] = 'date'
                profile['min_date'] = str(date_valid.min())
                profile['max_date'] = str(date_valid.max())
                return profile
        except Exception:
            pass
        
        # Check for boolean-like
        str_values = set(str(v).upper().strip() for v in distinct_values)
        bool_patterns = [
            {'Y', 'N'}, {'YES', 'NO'}, {'TRUE', 'FALSE'}, {'1', '0'},
            {'T', 'F'}, {'ACTIVE', 'INACTIVE'}, {'Y', 'N', ''}
        ]
        
        str_values_cleaned = str_values - {''}
        for pattern in bool_patterns:
            if str_values == pattern or str_values_cleaned == pattern or str_values_cleaned == pattern - {''}:
                profile['inferred_type'] = 'boolean'
                profile['is_categorical'] = True
                profile['distinct_values'] = sorted([str(v) for v in distinct_values])
                profile['value_distribution'] = series.value_counts().to_dict()
                # Convert keys to strings
                profile['value_distribution'] = {str(k): int(v) for k, v in profile['value_distribution'].items()}
                return profile
        
        # Categorical: distinct count <= 100 and not high cardinality
        if profile['distinct_count'] <= 100:
            profile['inferred_type'] = 'categorical'
            profile['is_categorical'] = True
            profile['distinct_values'] = sorted([str(v) for v in distinct_values if str(v).strip()])
            
            # Get value distribution
            value_counts = series.value_counts()
            profile['value_distribution'] = {str(k): int(v) for k, v in value_counts.items()}
            return profile
        
        # Default to text
        profile['inferred_type'] = 'text'
        return profile
    
    def _store_column_profile(self, profile: Dict):
        """Store or update a column profile in the database."""
        try:
            with self._db_lock:
                # Delete existing profile for this column
                self.conn.execute("""
                    DELETE FROM _column_profiles 
                    WHERE project = ? AND table_name = ? AND column_name = ?
                """, [profile['project'], profile['table_name'], profile['column_name']])
                
                # Insert new profile
                self.conn.execute("""
                    INSERT INTO _column_profiles (
                        id, project, table_name, column_name,
                        inferred_type, original_dtype,
                        total_count, null_count, distinct_count,
                        min_value, max_value, mean_value,
                        distinct_values, value_distribution,
                        min_date, max_date,
                        sample_values, is_likely_key, is_categorical,
                        filter_category, filter_priority
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, [
                    hash(f"{profile['project']}_{profile['table_name']}_{profile['column_name']}") % 2147483647,
                    profile['project'],
                    profile['table_name'],
                    profile['column_name'],
                    profile.get('inferred_type'),
                    profile.get('original_dtype'),
                    profile.get('total_count'),
                    profile.get('null_count'),
                    profile.get('distinct_count'),
                    profile.get('min_value'),
                    profile.get('max_value'),
                    profile.get('mean_value'),
                    json.dumps(profile.get('distinct_values')) if profile.get('distinct_values') else None,
                    json.dumps(profile.get('value_distribution')) if profile.get('value_distribution') else None,
                    profile.get('min_date'),
                    profile.get('max_date'),
                    json.dumps(profile.get('sample_values')) if profile.get('sample_values') else None,
                    profile.get('is_likely_key', False),
                    profile.get('is_categorical', False),
                    profile.get('filter_category'),
                    profile.get('filter_priority', 0)
                ])
                self.conn.commit()
            
        except Exception as e:
            logger.warning(f"[PROFILING] Failed to store profile for {profile.get('column_name')}: {e}")
    
    def get_column_profile(self, project: str, table_name: str = None, 
                           column_name: str = None) -> List[Dict]:
        """
        Retrieve column profiles from the database.
        
        Args:
            project: Project name
            table_name: Optional - filter by table
            column_name: Optional - filter by column
            
        Returns:
            List of profile dicts
        """
        try:
            query = "SELECT * FROM _column_profiles WHERE project = ?"
            params = [project]
            
            if table_name:
                query += " AND table_name = ?"
                params.append(table_name)
            
            if column_name:
                query += " AND column_name = ?"
                params.append(column_name)
            
            with self._db_lock:
                result = self.conn.execute(query, params).fetchall()
                columns = [desc[0] for desc in self.conn.execute(query, params).description]
            
            profiles = []
            for row in result:
                profile = dict(zip(columns, row))
                # Parse JSON fields
                for json_field in ['distinct_values', 'value_distribution', 'sample_values']:
                    if profile.get(json_field):
                        try:
                            profile[json_field] = json.loads(profile[json_field])
                        except Exception:
                            pass
                profiles.append(profile)
            
            return profiles
            
        except Exception as e:
            logger.warning(f"[PROFILING] Failed to get profiles: {e}")
            return []
    
    def get_categorical_columns(self, project: str, table_name: str = None) -> List[Dict]:
        """
        Get all categorical columns with their distinct values.
        This is the key method for intelligent clarification.
        
        Returns columns where is_categorical = TRUE with their value distributions.
        """
        try:
            query = """
                SELECT table_name, column_name, distinct_count, 
                       distinct_values, value_distribution
                FROM _column_profiles 
                WHERE project = ? AND is_categorical = TRUE
            """
            params = [project]
            
            if table_name:
                query += " AND table_name = ?"
                params.append(table_name)
            
            query += " ORDER BY table_name, column_name"
            
            result = self.conn.execute(query, params).fetchall()
            
            categorical = []
            for row in result:
                table, col, distinct_count, distinct_values, distribution = row
                categorical.append({
                    'table_name': table,
                    'column_name': col,
                    'distinct_count': distinct_count,
                    'distinct_values': json.loads(distinct_values) if distinct_values else [],
                    'value_distribution': json.loads(distribution) if distribution else {}
                })
            
            return categorical
            
        except Exception as e:
            logger.warning(f"[PROFILING] Failed to get categorical columns: {e}")
            return []
    
    def get_filter_candidates(self, project: str) -> Dict[str, List[Dict]]:
        """
        Get all columns identified as filter candidates, grouped by category.
        
        Categories: status, company, organization, location, pay_type, employee_type, job
        
        Returns:
            Dict with category keys: {
                'status': [{'column': 'termination_date', 'table': '...', 'values': [...], ...}],
                'company': [...],
                ...
            }
        """
        try:
            result = self.conn.execute("""
                SELECT 
                    filter_category,
                    table_name,
                    column_name,
                    inferred_type,
                    distinct_count,
                    distinct_values,
                    value_distribution,
                    filter_priority
                FROM _column_profiles
                WHERE project = ? AND filter_category IS NOT NULL
                ORDER BY filter_priority DESC, filter_category, table_name
            """, [project]).fetchall()
            
            candidates = {}
            for row in result:
                category = row[0]
                if category not in candidates:
                    candidates[category] = []
                
                candidates[category].append({
                    'table_name': row[1],
                    'column_name': row[2],
                    'inferred_type': row[3],
                    'distinct_count': row[4],
                    'distinct_values': json.loads(row[5]) if row[5] else [],
                    'value_distribution': json.loads(row[6]) if row[6] else {},
                    'filter_priority': row[7]
                })
            
            return candidates
            
        except Exception as e:
            logger.warning(f"[PROFILING] Failed to get filter candidates: {e}")
            return {}
    
    # =========================================================================
    # QUERY METHODS
    # =========================================================================
    
    def query(self, sql: str) -> List[Dict]:
        """Execute SQL query and return results as list of dicts"""
        with self._db_lock:
            try:
                result = self.conn.execute(sql)
                columns = [desc[0] for desc in result.description]
                rows = result.fetchall()
                return [dict(zip(columns, row)) for row in rows]
            except Exception as e:
                logger.error(f"Query error: {e}")
                raise
    
    def query_to_dataframe(self, sql: str) -> pd.DataFrame:
        """Execute SQL query and return as DataFrame"""
        with self._db_lock:
            try:
                return self.conn.execute(sql).fetchdf()
            except Exception as e:
                logger.error(f"Query error: {e}")
                raise
    
    def safe_execute(self, sql: str, params: list = None):
        """
        Thread-safe SQL execution wrapper.
        Use this for any external/concurrent access to DuckDB.
        """
        with self._db_lock:
            try:
                # Log INSERT operations for debugging
                if 'INSERT' in sql.upper() and '_schema_metadata' in sql:
                    logger.warning(f"[SAFE_EXECUTE] Running INSERT into _schema_metadata")
                
                if params:
                    result = self.conn.execute(sql, params)
                else:
                    result = self.conn.execute(sql)
                
                # Verify INSERT worked
                if 'INSERT' in sql.upper() and '_schema_metadata' in sql:
                    logger.warning(f"[SAFE_EXECUTE] INSERT completed, result type: {type(result)}")
                
                return result
            except Exception as e:
                logger.error(f"[SAFE_EXECUTE] Error: {e}")
                raise
    
    def safe_fetchall(self, sql: str, params: list = None) -> list:
        """
        Thread-safe SQL query that returns all rows.
        Commits any pending transaction first to ensure we see latest data.
        
        v2.1: Removed verbose diagnostics that were killing performance.
        """
        with self._db_lock:
            try:
                # Commit any pending changes to ensure we see them
                try:
                    self.conn.commit()
                except Exception:
                    pass  # OK if nothing to commit
                
                if params:
                    result = self.conn.execute(sql, params).fetchall()
                else:
                    result = self.conn.execute(sql).fetchall()
                
                return result
            except Exception as e:
                logger.error(f"[SAFE_FETCHALL] Error: {e}")
                raise
    
    def safe_fetchone(self, sql: str, params: list = None):
        """
        Thread-safe SQL query that returns one row.
        """
        with self._db_lock:
            try:
                if params:
                    return self.conn.execute(sql, params).fetchone()
                return self.conn.execute(sql).fetchone()
            except Exception as e:
                logger.error(f"[SAFE_FETCHONE] Error: {e}")
                raise
    
    def safe_commit(self):
        """
        Thread-safe commit wrapper.
        """
        with self._db_lock:
            try:
                self.conn.commit()
                
                # Force DuckDB to flush by running a checkpoint
                try:
                    self.conn.execute("CHECKPOINT")
                except Exception as cp_e:
                    logger.debug(f"[SAFE_COMMIT] Checkpoint note: {cp_e}")
                    
            except Exception as e:
                logger.error(f"[SAFE_COMMIT] Error: {e}")
                raise
    
    def safe_create_table_from_df(self, table_name: str, df: pd.DataFrame, drop_existing: bool = True):
        """
        Thread-safe table creation from DataFrame.
        
        This wraps DROP/register/CREATE/unregister in a single lock to prevent
        race conditions with concurrent status polling.
        """
        with self._db_lock:
            try:
                if drop_existing:
                    self.conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                
                temp_name = f"temp_{table_name[:30]}_{id(df)}"
                self.conn.register(temp_name, df)
                self.conn.execute(f'CREATE TABLE "{table_name}" AS SELECT * FROM {temp_name}')
                self.conn.unregister(temp_name)
                
                logger.debug(f"[SAFE_CREATE] Created table {table_name} with {len(df)} rows")
                return True
            except Exception as e:
                logger.error(f"[SAFE_CREATE] Failed to create {table_name}: {e}")
                raise
    
    def get_schema(self, project: str = None) -> Dict[str, Any]:
        """Get schema information for all or specific project"""
        with self._db_lock:
            try:
                if project:
                    result = self.conn.execute("""
                        SELECT table_name, display_name, file_name, sheet_name, columns, column_count, row_count, likely_keys, encrypted_columns, truth_type, uploaded_by, created_at
                        FROM _schema_metadata 
                        WHERE project = ? AND is_current = TRUE
                    """, [project]).fetchall()
                else:
                    result = self.conn.execute("""
                        SELECT project, table_name, display_name, file_name, sheet_name, columns, column_count, row_count, likely_keys, truth_type, uploaded_by, created_at
                        FROM _schema_metadata 
                        WHERE is_current = TRUE
                    """).fetchall()
                
                schema = {}
                for row in result:
                    if project:
                        table_name, display_name, file_name, sheet_name, columns, column_count, row_count, keys, encrypted, truth_type, uploaded_by, created_at = row
                        
                        # Generate display_name if not set (migration fallback)
                        if not display_name:
                            display_name = self._generate_display_name(file_name, sheet_name)
                        
                        schema[table_name] = {
                            'display_name': display_name,
                            'file_name': file_name,
                            'sheet_name': sheet_name,
                            'columns': json.loads(columns) if columns else [],
                            'column_count': column_count or (len(json.loads(columns)) if columns else 0),
                            'row_count': row_count,
                            'likely_keys': json.loads(keys) if keys else [],
                            'encrypted_columns': json.loads(encrypted) if encrypted else [],
                            'truth_type': truth_type,
                            'uploaded_by': uploaded_by,
                            'created_at': str(created_at) if created_at else None
                        }
                    else:
                        proj, table_name, display_name, file_name, sheet_name, columns, column_count, row_count, keys, truth_type, uploaded_by, created_at = row
                        
                        # Generate display_name if not set (migration fallback)
                        if not display_name:
                            display_name = self._generate_display_name(file_name, sheet_name)
                        
                        if proj not in schema:
                            schema[proj] = {}
                        schema[proj][table_name] = {
                            'display_name': display_name,
                            'file_name': file_name,
                            'sheet_name': sheet_name,
                            'columns': json.loads(columns) if columns else [],
                            'column_count': column_count or (len(json.loads(columns)) if columns else 0),
                            'row_count': row_count,
                            'likely_keys': json.loads(keys) if keys else [],
                            'truth_type': truth_type,
                            'uploaded_by': uploaded_by,
                            'created_at': str(created_at) if created_at else None
                        }
                
                return schema
                
            except Exception as e:
                logger.error(f"Error getting schema: {e}")
                return {}
    
    def get_tables(self, project: str) -> List[Dict[str, Any]]:
        """Get list of tables for a project with detailed info"""
        with self._db_lock:
            try:
                result = self.conn.execute("""
                    SELECT table_name, display_name, sheet_name, file_name, columns, column_count, row_count, likely_keys, 
                           encrypted_columns, truth_type, version, created_at, uploaded_by
                    FROM _schema_metadata 
                    WHERE project = ? AND is_current = TRUE
                    ORDER BY table_name
                """, [project]).fetchall()
                
                tables = []
                for row in result:
                    table_name = row[0]
                    display_name = row[1]
                    
                    # Generate display_name if not set (migration fallback)
                    if not display_name:
                        display_name = self._generate_display_name(row[3], row[2])  # file_name, sheet_name
                    
                    tables.append({
                        'table_name': table_name,
                        'display_name': display_name,
                        'sheet_name': row[2],
                        'file_name': row[3],
                        'columns': json.loads(row[4]) if row[4] else [],
                        'column_count': row[5] or (len(json.loads(row[4])) if row[4] else 0),
                        'row_count': row[6],
                        'likely_keys': json.loads(row[7]) if row[7] else [],
                        'encrypted_columns': json.loads(row[8]) if row[8] else [],
                        'truth_type': row[9],
                        'version': row[10],
                        'created_at': str(row[11]) if row[11] else None,
                        'uploaded_by': row[12]
                    })
                
                return tables
                
            except Exception as e:
                logger.error(f"Error getting tables: {e}")
                return []
    
    def get_sample_data(self, table_name: str, limit: int = 10, 
                        decrypt: bool = True) -> List[Dict]:
        """Get sample data from a table"""
        try:
            df = self.query_to_dataframe(f"SELECT * FROM {table_name} LIMIT {limit}")
            
            if decrypt:
                df = self.encryptor.decrypt_dataframe(df)
            
            return df.to_dict('records')
            
        except Exception as e:
            logger.error(f"Error getting sample data: {e}")
            return []
    
    def list_projects(self) -> List[str]:
        """List all projects in the database"""
        with self._db_lock:
            try:
                result = self.conn.execute("""
                    SELECT DISTINCT project FROM _schema_metadata ORDER BY project
                """).fetchall()
                return [r[0] for r in result]
            except Exception as e:
                logger.error(f"Error listing projects: {e}")
                return []
    
    def delete_project(self, project: str) -> Dict[str, Any]:
        """Delete all data for a project"""
        result = {'tables_deleted': [], 'success': False}
        
        try:
            # Get all tables for this project
            tables = self.conn.execute("""
                SELECT table_name FROM _schema_metadata WHERE project = ?
            """, [project]).fetchall()
            
            for (table_name,) in tables:
                try:
                    self.conn.execute(f"DROP TABLE IF EXISTS {table_name}")
                    result['tables_deleted'].append(table_name)
                except Exception as e:
                    logger.warning(f"Could not drop {table_name}: {e}")
            
            # Delete metadata
            self.conn.execute("DELETE FROM _schema_metadata WHERE project = ?", [project])
            self.conn.execute("DELETE FROM _load_versions WHERE project = ?", [project])
            self.conn.execute("DELETE FROM _table_relationships WHERE project = ?", [project])
            self.conn.execute("DELETE FROM _column_mappings WHERE project = ?", [project])
            self.conn.execute("DELETE FROM _column_profiles WHERE project = ?", [project])
            self.conn.commit()
            
            result['success'] = True
            logger.info(f"Deleted project {project}: {len(result['tables_deleted'])} tables")
            
        except Exception as e:
            logger.error(f"Error deleting project: {e}")
            result['error'] = str(e)
        
        return result
    
    # =========================================================================
    # VERSION COMPARISON
    # =========================================================================
    
    def compare_versions(
        self, 
        project: str, 
        file_name: str, 
        sheet_name: str,
        version1: int = None,
        version2: int = None,
        key_column: str = None
    ) -> Dict[str, Any]:
        """
        Compare two versions of a sheet to find differences.
        
        Args:
            project: Project name
            file_name: File name
            sheet_name: Sheet name to compare
            version1: First version (default: previous)
            version2: Second version (default: current/latest)
            key_column: Column to use as row identifier
            
        Returns:
            Dict with added, removed, and changed records
        """
        # Get available versions
        versions = self.conn.execute("""
            SELECT DISTINCT version FROM _schema_metadata 
            WHERE project = ? AND file_name = ? AND sheet_name = ?
            ORDER BY version DESC
        """, [project, file_name, sheet_name]).fetchall()
        
        if len(versions) < 2:
            return {
                'error': 'Need at least 2 versions to compare',
                'available_versions': [v[0] for v in versions]
            }
        
        v2 = version2 or versions[0][0]  # Latest
        v1 = version1 or versions[1][0]  # Previous
        
        # Get table names
        table1_result = self.conn.execute("""
            SELECT table_name FROM _schema_metadata 
            WHERE project = ? AND file_name = ? AND sheet_name = ? AND version = ?
        """, [project, file_name, sheet_name, v1]).fetchone()
        
        table2_result = self.conn.execute("""
            SELECT table_name FROM _schema_metadata 
            WHERE project = ? AND file_name = ? AND sheet_name = ? AND version = ?
        """, [project, file_name, sheet_name, v2]).fetchone()
        
        if not table1_result or not table2_result:
            # Try archived table names
            base_table = self._generate_table_name(project, file_name, sheet_name)
            table1 = f"{base_table}_v{v1}"
            table2 = base_table  # Current is without version suffix
        else:
            table1 = table1_result[0] if v1 != versions[0][0] else f"{table1_result[0]}_v{v1}"
            table2 = table2_result[0]
        
        result = {
            'project': project,
            'file_name': file_name,
            'sheet_name': sheet_name,
            'version1': v1,
            'version2': v2,
            'key_column': key_column,
            'added': [],
            'removed': [],
            'changed': [],
            'unchanged_count': 0
        }
        
        try:
            # Get data from both versions
            df1 = self.query_to_dataframe(f"SELECT * FROM {table1}")
            df2 = self.query_to_dataframe(f"SELECT * FROM {table2}")
            
            # Decrypt for comparison
            df1 = self.encryptor.decrypt_dataframe(df1)
            df2 = self.encryptor.decrypt_dataframe(df2)
            
            # Ensure key column exists
            if key_column not in df1.columns or key_column not in df2.columns:
                return {'error': f"Key column '{key_column}' not found in both versions"}
            
            # Set key as index
            df1 = df1.set_index(key_column)
            df2 = df2.set_index(key_column)
            
            # Find added (in v2 but not v1)
            added_keys = set(df2.index) - set(df1.index)
            result['added'] = df2.loc[list(added_keys)].reset_index().to_dict('records')
            
            # Find removed (in v1 but not v2)
            removed_keys = set(df1.index) - set(df2.index)
            result['removed'] = df1.loc[list(removed_keys)].reset_index().to_dict('records')
            
            # Find changed (in both but different)
            common_keys = set(df1.index) & set(df2.index)
            changed = []
            unchanged = 0
            
            for key in common_keys:
                row1 = df1.loc[key]
                row2 = df2.loc[key]
                
                # Compare each column
                differences = {}
                for col in df1.columns:
                    if col in df2.columns:
                        val1 = row1[col] if not pd.isna(row1[col]) else None
                        val2 = row2[col] if not pd.isna(row2[col]) else None
                        
                        if str(val1) != str(val2):
                            differences[col] = {'old': val1, 'new': val2}
                
                if differences:
                    changed.append({
                        key_column: key,
                        'changes': differences
                    })
                else:
                    unchanged += 1
            
            result['changed'] = changed
            result['unchanged_count'] = unchanged
            
            # Summary
            result['summary'] = {
                'total_v1': len(df1),
                'total_v2': len(df2),
                'added_count': len(result['added']),
                'removed_count': len(result['removed']),
                'changed_count': len(result['changed']),
                'unchanged_count': unchanged
            }
            
            logger.info(f"Compared {file_name}/{sheet_name}: +{len(result['added'])} -{len(result['removed'])} ~{len(result['changed'])}")
            
        except Exception as e:
            logger.error(f"Comparison error: {e}")
            result['error'] = str(e)
        
        return result
    
    def get_file_versions(self, project: str, file_name: str) -> List[Dict]:
        """Get all versions of a file"""
        versions = self.conn.execute("""
            SELECT DISTINCT version, created_at, row_count
            FROM _schema_metadata 
            WHERE project = ? AND file_name = ?
            ORDER BY version DESC
        """, [project, file_name]).fetchall()
        
        return [
            {'version': v[0], 'created_at': v[1], 'row_count': v[2]}
            for v in versions
        ]
    
    def list_files(self, project: str) -> List[Dict]:
        """List all files in a project with version info"""
        files = self.conn.execute("""
            SELECT DISTINCT file_name, 
                   MAX(version) as latest_version,
                   SUM(row_count) as total_rows,
                   MAX(created_at) as last_updated
            FROM _schema_metadata 
            WHERE project = ? AND is_current = TRUE
            GROUP BY file_name
            ORDER BY file_name
        """, [project]).fetchall()
        
        return [
            {
                'file_name': f[0],
                'latest_version': f[1],
                'total_rows': f[2],
                'last_updated': f[3]
            }
            for f in files
        ]
    
    def reset_database(self) -> Dict[str, Any]:
        """
        Completely reset the database - drop all tables and recreate metadata.
        USE WITH CAUTION - this deletes all data!
        """
        result = {
            'tables_dropped': [],
            'success': False
        }
        
        try:
            # Get all user tables (not metadata)
            tables = self.conn.execute("""
                SELECT table_name FROM information_schema.tables 
                WHERE table_schema = 'main' 
                AND table_name NOT LIKE '\\_%' ESCAPE '\\'
            """).fetchall()
            
            for (table_name,) in tables:
                try:
                    self.conn.execute(f"DROP TABLE IF EXISTS {table_name}")
                    result['tables_dropped'].append(table_name)
                except Exception as e:
                    logger.warning(f"Could not drop {table_name}: {e}")
            
            # Drop metadata tables
            self.conn.execute("DROP TABLE IF EXISTS _schema_metadata")
            self.conn.execute("DROP TABLE IF EXISTS _load_versions")
            self.conn.execute("DROP SEQUENCE IF EXISTS schema_metadata_seq")
            self.conn.execute("DROP SEQUENCE IF EXISTS load_versions_seq")
            self.conn.commit()
            
            # Recreate metadata tables
            self._init_metadata_table()
            
            result['success'] = True
            logger.info(f"Database reset: dropped {len(result['tables_dropped'])} tables")
            
        except Exception as e:
            logger.error(f"Database reset failed: {e}")
            result['error'] = str(e)
        
        return result
    
    def cleanup_junk_columns(self, project: str = None, dry_run: bool = True) -> Dict:
        """
        Clean up junk columns from existing DuckDB tables.
        
        Identifies and optionally removes:
        - Columns named 'unnamed', 'col_X' (auto-generated)
        - Columns that are 100% empty
        - Columns that are mostly empty AND have junk names
        
        Args:
            project: Optional project filter (None = all projects)
            dry_run: If True, report what would be cleaned without making changes
        
        Returns:
            {
                'tables_scanned': int,
                'tables_with_junk': int,
                'columns_identified': [...],
                'columns_removed': [...] (empty if dry_run)
            }
        """
        result = {
            'tables_scanned': 0,
            'tables_with_junk': 0,
            'columns_identified': [],
            'columns_removed': [],
            'dry_run': dry_run,
            'errors': []
        }
        
        try:
            # Get all tables
            tables_query = """
                SELECT DISTINCT table_name, project, file_name
                FROM _schema_metadata
                WHERE is_current = TRUE
            """
            if project:
                tables_query += f" AND project = '{project}'"
            
            tables = self.conn.execute(tables_query).fetchall()
            
            for table_name, proj, file_name in tables:
                result['tables_scanned'] += 1
                
                try:
                    # Get column info
                    col_info = self.conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
                    
                    # Read sample to check fill rates
                    sample = self.conn.execute(f'SELECT * FROM "{table_name}" LIMIT 1000').fetchdf()
                    row_count = self.conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
                    
                    junk_cols = []
                    for col_row in col_info:
                        col_name = col_row[1]  # Column name is at index 1
                        col_lower = str(col_name).lower()
                        
                        # Check if junk name
                        is_junk_name = (
                            col_lower == 'unnamed' or
                            col_lower.startswith('unnamed_') or
                            re.match(r'^col_\d+$', col_lower) or
                            col_lower in ['nan', 'none', '']
                        )
                        
                        # Check fill rate
                        if len(sample) > 0 and col_name in sample.columns:
                            non_empty = sample[col_name].notna().sum()
                            non_empty_str = (sample[col_name].astype(str).str.strip() != '').sum()
                            fill_rate = max(non_empty, non_empty_str) / len(sample)
                        else:
                            fill_rate = 0
                        
                        # Identify as junk
                        if (fill_rate == 0) or (is_junk_name and fill_rate < 0.05):
                            junk_cols.append({
                                'column': col_name,
                                'fill_rate': round(fill_rate, 4),
                                'is_junk_name': is_junk_name,
                                'reason': 'empty' if fill_rate == 0 else 'junk_name_sparse'
                            })
                    
                    if junk_cols:
                        result['tables_with_junk'] += 1
                        
                        for junk in junk_cols:
                            result['columns_identified'].append({
                                'table': table_name,
                                'project': proj,
                                'file': file_name,
                                'column': junk['column'],
                                'fill_rate': junk['fill_rate'],
                                'reason': junk['reason']
                            })
                            
                            if not dry_run:
                                try:
                                    self.conn.execute(f'ALTER TABLE "{table_name}" DROP COLUMN "{junk["column"]}"')
                                    result['columns_removed'].append({
                                        'table': table_name,
                                        'column': junk['column']
                                    })
                                    logger.info(f"[CLEANUP] Removed {table_name}.{junk['column']}")
                                except Exception as drop_e:
                                    result['errors'].append({
                                        'table': table_name,
                                        'column': junk['column'],
                                        'error': str(drop_e)
                                    })
                                    logger.warning(f"[CLEANUP] Could not drop {table_name}.{junk['column']}: {drop_e}")
                
                except Exception as table_e:
                    result['errors'].append({
                        'table': table_name,
                        'error': str(table_e)
                    })
                    logger.warning(f"[CLEANUP] Error processing {table_name}: {table_e}")
            
            if not dry_run:
                self.conn.commit()
            
            logger.info(f"[CLEANUP] Scanned {result['tables_scanned']} tables, found {len(result['columns_identified'])} junk columns")
            return result
            
        except Exception as e:
            logger.error(f"[CLEANUP] Error: {e}")
            result['errors'].append({'general': str(e)})
            return result
    
    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()


# =============================================================================
# PROCESS-LEVEL SINGLETON
# =============================================================================
# 
# Python treats different import paths as different modules:
#   - "utils.structured_data_handler" 
#   - "backend.utils.structured_data_handler"
# 
# Each would get its own _handler variable and thus its own DuckDB connection.
# Commits on one connection aren't visible to the other.
#
# Solution: Store the singleton in sys.modules under a canonical key.
# This guarantees ONE instance per process regardless of import path.
# Thread safety via lock ensures only one instance created even with parallel calls.
# =============================================================================

_SINGLETON_KEY = '_xlr8_structured_data_handler_instance'
_SINGLETON_LOCK = threading.Lock()

def get_structured_handler() -> StructuredDataHandler:
    """
    Get or create the singleton handler.
    
    WARNING: This returns the WRITE handler. Only use for upload operations.
    For API endpoints that only READ data, use get_read_handler() instead.
    
    Uses sys.modules to ensure ONE instance per process, regardless of
    whether this module is imported as 'utils.structured_data_handler'
    or 'backend.utils.structured_data_handler'.
    
    Thread-safe: uses lock to prevent race conditions.
    """
    # Fast path - if already exists, return immediately (no lock needed for reads)
    if _SINGLETON_KEY in sys.modules:
        handler = sys.modules[_SINGLETON_KEY]
        if isinstance(handler, StructuredDataHandler):
            return handler
    
    # Slow path - need to create, use lock
    with _SINGLETON_LOCK:
        # Double-check after acquiring lock
        if _SINGLETON_KEY not in sys.modules:
            handler = StructuredDataHandler()
            sys.modules[_SINGLETON_KEY] = handler
            logging.getLogger(__name__).warning(
                f"[HANDLER] Created singleton (id={id(handler)}, conn_id={id(handler.conn)}, db={handler.db_path}, module={_MODULE_LOAD_ID}, pid={os.getpid()})"
            )
        else:
            handler = sys.modules[_SINGLETON_KEY]
        
        # Verify it's actually a handler instance
        if not isinstance(handler, StructuredDataHandler):
            logging.getLogger(__name__).error(f"[HANDLER] SINGLETON IS WRONG TYPE: {type(handler)}")
            handler = StructuredDataHandler()
            sys.modules[_SINGLETON_KEY] = handler
            
        return handler


class ReadOnlyDuckDBHandler:
    """
    Lightweight handler for API endpoints that only read data.
    
    Creates a new connection each time - safe for concurrent access.
    NOTE: We don't use read_only=True because DuckDB doesn't allow
    mixing read-only and read-write connections to the same file.
    DuckDB handles concurrent reads safely via internal locking.
    """
    def __init__(self, db_path: str):
        self.db_path = db_path
        # Don't use read_only=True - it conflicts with write connections
        self.conn = duckdb.connect(db_path)
    
    def close(self):
        if self.conn:
            self.conn.close()
            self.conn = None
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def get_read_handler() -> ReadOnlyDuckDBHandler:
    """
    Get a read-only handler for API endpoints.
    
    ALWAYS use this for API endpoints that only READ data.
    This creates a new read-only connection that won't conflict with
    upload write operations.
    
    Usage:
        handler = get_read_handler()
        result = handler.conn.execute("SELECT * FROM ...").fetchall()
        handler.close()  # Always close when done
        
    Or with context manager:
        with get_read_handler() as handler:
            result = handler.conn.execute("SELECT * FROM ...").fetchall()
    
    Returns:
        ReadOnlyDuckDBHandler with .conn attribute for SQL execution
    """
    db_path = os.environ.get('DUCKDB_PATH', '/data/structured_data.duckdb')
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"DuckDB database not found: {db_path}")
    return ReadOnlyDuckDBHandler(db_path)


def reset_structured_handler():
    """
    Reset the singleton handler (forces reconnection).
    
    Use this if you need to force a fresh connection, e.g., after
    database file corruption or for testing.
    """
    if _SINGLETON_KEY in sys.modules:
        handler = sys.modules[_SINGLETON_KEY]
        if hasattr(handler, 'close'):
            handler.close()
        del sys.modules[_SINGLETON_KEY]
        logging.getLogger(__name__).info("[HANDLER] Singleton instance reset")
