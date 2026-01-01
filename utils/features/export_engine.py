"""
Export Engine - Template-Based Data Export
==========================================

Exports structured data to various formats using templates.
Used by Playbooks, BI, Chat, Reports.

Supported formats:
- xlsx: Excel with formatting
- csv: Simple CSV
- json: JSON export

Template sources:
- Built-in: /data/export_library/templates/
- User-uploaded: document_registry with truth_type='template'
"""

import logging
import io
import json
import csv
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional, Union
from pathlib import Path

logger = logging.getLogger(__name__)

# Template directory (built-in templates)
TEMPLATE_DIR = Path(__file__).parent.parent.parent / "data" / "export_library" / "templates"


class ExportEngine:
    """
    Exports data to various formats with optional templates.
    
    Supports:
    - ComparisonResult objects
    - List[Dict] data
    - Dict with nested data
    
    All exports include provenance metadata.
    """
    
    def __init__(self):
        self.template_dir = TEMPLATE_DIR
    
    def export_comparison(
        self,
        result,  # ComparisonResult
        format: str = "xlsx",
        include_provenance: bool = True
    ) -> io.BytesIO:
        """
        Export a ComparisonResult to file.
        
        Args:
            result: ComparisonResult from comparison_engine
            format: 'xlsx', 'csv', or 'json'
            include_provenance: Include metadata sheet/section
            
        Returns:
            BytesIO buffer with file content
        """
        if format == "xlsx":
            return self._export_comparison_xlsx(result, include_provenance)
        elif format == "csv":
            return self._export_comparison_csv(result)
        elif format == "json":
            return self._export_comparison_json(result)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _export_comparison_xlsx(self, result, include_provenance: bool) -> io.BytesIO:
        """Export comparison to Excel with multiple sheets."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="285390", end_color="285390", fill_type="solid")
        mismatch_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def style_header(ws, row=1):
            for cell in ws[row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
        
        def auto_width(ws):
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ["Comparison Summary", ""],
            ["", ""],
            ["Source A", result.source_a],
            ["Source A Rows", result.source_a_rows],
            ["Source B", result.source_b],
            ["Source B Rows", result.source_b_rows],
            ["", ""],
            ["Results", ""],
            ["Matched Rows", result.matches],
            ["Mismatches", len(result.mismatches)],
            ["Only in Source A", len(result.only_in_a)],
            ["Only in Source B", len(result.only_in_b)],
            ["Match Rate", f"{result.match_rate:.1%}"],
            ["", ""],
            ["Join Keys", ", ".join(result.join_keys)],
            ["Compared Columns", ", ".join(result.compared_columns[:10])],
            ["Executed At", result.executed_at],
            ["Comparison ID", result.comparison_id],
        ]
        
        for row_data in summary_data:
            ws_summary.append(row_data)
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 60
        
        # Sheet 2: Mismatches
        if result.mismatches:
            ws_mismatch = wb.create_sheet("Mismatches")
            
            # Build headers from first mismatch
            first = result.mismatches[0]
            headers = list(first['keys'].keys()) + ['Column', 'Value in A', 'Value in B']
            ws_mismatch.append(headers)
            style_header(ws_mismatch)
            
            # Add data
            for mismatch in result.mismatches:
                for diff in mismatch['differences']:
                    row = list(mismatch['keys'].values()) + [
                        diff['column'],
                        str(diff['value_a'])[:100],
                        str(diff['value_b'])[:100]
                    ]
                    ws_mismatch.append(row)
            
            auto_width(ws_mismatch)
        
        # Sheet 3: Only in A
        if result.only_in_a:
            ws_a = wb.create_sheet("Only in Source A")
            
            headers = list(result.only_in_a[0].keys())
            ws_a.append(headers)
            style_header(ws_a)
            
            for row_data in result.only_in_a:
                ws_a.append([str(v)[:100] if v is not None else "" for v in row_data.values()])
            
            auto_width(ws_a)
        
        # Sheet 4: Only in B
        if result.only_in_b:
            ws_b = wb.create_sheet("Only in Source B")
            
            headers = list(result.only_in_b[0].keys())
            ws_b.append(headers)
            style_header(ws_b)
            
            for row_data in result.only_in_b:
                ws_b.append([str(v)[:100] if v is not None else "" for v in row_data.values()])
            
            auto_width(ws_b)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _export_comparison_csv(self, result) -> io.BytesIO:
        """Export comparison mismatches to CSV."""
        buffer = io.StringIO()
        
        if result.mismatches:
            first = result.mismatches[0]
            fieldnames = list(first['keys'].keys()) + ['column', 'value_a', 'value_b']
            
            writer = csv.DictWriter(buffer, fieldnames=fieldnames)
            writer.writeheader()
            
            for mismatch in result.mismatches:
                for diff in mismatch['differences']:
                    row = dict(mismatch['keys'])
                    row['column'] = diff['column']
                    row['value_a'] = diff['value_a']
                    row['value_b'] = diff['value_b']
                    writer.writerow(row)
        
        # Convert to bytes
        bytes_buffer = io.BytesIO(buffer.getvalue().encode('utf-8'))
        bytes_buffer.seek(0)
        return bytes_buffer
    
    def _export_comparison_json(self, result) -> io.BytesIO:
        """Export comparison to JSON with full data."""
        data = result.to_dict()
        json_str = json.dumps(data, indent=2, default=str)
        buffer = io.BytesIO(json_str.encode('utf-8'))
        buffer.seek(0)
        return buffer
    
    def export_data(
        self,
        data: List[Dict[str, Any]],
        format: str = "xlsx",
        sheet_name: str = "Data",
        title: str = None,
        provenance: Dict[str, Any] = None
    ) -> io.BytesIO:
        """
        Export generic list of dicts to file.
        
        Args:
            data: List of dictionaries to export
            format: 'xlsx', 'csv', or 'json'
            sheet_name: Name for Excel sheet
            title: Optional title for report
            provenance: Optional metadata dict
            
        Returns:
            BytesIO buffer with file content
        """
        if not data:
            raise ValueError("No data to export")
        
        if format == "xlsx":
            return self._export_data_xlsx(data, sheet_name, title, provenance)
        elif format == "csv":
            return self._export_data_csv(data)
        elif format == "json":
            return self._export_data_json(data, provenance)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _export_data_xlsx(
        self,
        data: List[Dict],
        sheet_name: str,
        title: str,
        provenance: Dict
    ) -> io.BytesIO:
        """Export data list to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="285390", end_color="285390", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title row if provided
        start_row = 1
        if title:
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(data[0]), 5))
            start_row = 3
        
        # Headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data
        for row_idx, row_data in enumerate(data, start_row + 1):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                if value is not None:
                    value = str(value)[:500]  # Limit cell content
                ws.cell(row=row_idx, column=col_idx, value=value).border = thin_border
        
        # Auto-width columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Provenance sheet if provided
        if provenance:
            ws_prov = wb.create_sheet("Provenance")
            ws_prov.append(["Metadata", "Value"])
            for key, value in provenance.items():
                ws_prov.append([str(key), str(value)])
            ws_prov.column_dimensions['A'].width = 25
            ws_prov.column_dimensions['B'].width = 60
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _export_data_csv(self, data: List[Dict]) -> io.BytesIO:
        """Export data to CSV."""
        buffer = io.StringIO()
        
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(buffer, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in data:
            writer.writerow({k: str(v)[:500] if v is not None else "" for k, v in row.items()})
        
        bytes_buffer = io.BytesIO(buffer.getvalue().encode('utf-8'))
        bytes_buffer.seek(0)
        return bytes_buffer
    
    def _export_data_json(self, data: List[Dict], provenance: Dict = None) -> io.BytesIO:
        """Export data to JSON."""
        output = {
            "data": data,
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "row_count": len(data)
        }
        if provenance:
            output["provenance"] = provenance
        
        json_str = json.dumps(output, indent=2, default=str)
        buffer = io.BytesIO(json_str.encode('utf-8'))
        buffer.seek(0)
        return buffer


# Module-level instance
_engine = None

def get_export_engine() -> ExportEngine:
    """Get or create an ExportEngine instance."""
    global _engine
    if _engine is None:
        _engine = ExportEngine()
    return _engine


def export_comparison(result, format: str = "xlsx", **kwargs) -> io.BytesIO:
    """Convenience function to export a comparison result."""
    return get_export_engine().export_comparison(result, format, **kwargs)


def export_data(data: List[Dict], format: str = "xlsx", **kwargs) -> io.BytesIO:
    """Convenience function to export generic data."""
    return get_export_engine().export_data(data, format, **kwargs)
