"""
XLR8 Excel Report Generator
Generates professional, formatted Excel implementation plans
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict
import logging

logger = logging.getLogger(__name__)


class SECURE20ReportGenerator:
    """
    Generates comprehensive SECURE 2.0 implementation plan Excel workbook
    with 8 professionally formatted worksheets.
    """
    
    # Professional color scheme (corporate blue)
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    TITLE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
    
    # Status colors
    HIGH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
    MED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Yellow
    LOW_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # Green
    INFO_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue
    
    BOLD_FONT = Font(bold=True)
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, company_name: str, analysis_results: Dict):
        """
        Initialize report generator.
        
        Args:
            company_name: Customer company name
            analysis_results: Output from SECURE20Analyzer
        """
        self.company_name = company_name.upper()
        self.results = analysis_results
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
    def generate(self, output_path: str):
        """Generate complete 8-worksheet implementation plan"""
        logger.info(f"Generating implementation plan for {self.company_name}")
        
        # Create all 8 worksheets
        self._create_implementation_steps()
        self._create_executive_summary()
        self._create_configuration_guide()
        self._create_high_priority()
        self._create_low_priority()
        self._create_all_rcr()
        self._create_import_template()
        self._create_observations()
        
        # Save
        self.wb.save(output_path)
        logger.info(f"Report saved to {output_path}")
        
    def _create_implementation_steps(self):
        """Tab 1: Implementation Steps"""
        ws = self.wb.create_sheet("Implementation Steps")
        
        # Title
        ws['A1'] = f"{self.company_name} - SECURE 2.0 IMPLEMENTATION STEPS"
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].fill = self.TITLE_FILL
        ws.merge_cells('A1:C1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        # Steps
        steps = [
            ("Step", "Action Item", "Status"),
            ("1", "Run SECURE 2.0 Act Catch-Up Contribution Eligibility Report", "✓ Complete"),
            ("2", "Identify RCR Employees (Age 50+, Wages >$145K)", f"✓ Complete - {self.results['statistics']['rcr_employees']} found"),
            ("3", f"Add ROTH Codes to {self.results['statistics']['high_priority']} High-Priority Employees", "⚠ ACTION NEEDED"),
            ("4", "Configure Spillover in Deduction/Benefit Plans", "📋 Pending"),
            ("5", "Notify Affected Employees", "📋 Pending"),
            ("6", "Test Payroll with RCR Employees", "📋 Pending"),
        ]
        
        for idx, row in enumerate(steps, start=4):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=idx, column=col_idx, value=value)
                if idx == 4:  # Header row
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                cell.border = self.THIN_BORDER
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 25
        
    def _create_executive_summary(self):
        """Tab 2: Executive Summary"""
        ws = self.wb.create_sheet("Executive Summary")
        
        stats = self.results['statistics']
        
        # Title
        ws['A1'] = f"{self.company_name} - SECURE 2.0 Implementation Plan"
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].fill = self.TITLE_FILL
        ws.merge_cells('A1:B1')
        
        # Company info
        info = [
            ("System:", "UKG Pro (Payroll Engine)"),
            ("Decision:", "ADOPTING Deemed ROTH Elections with Automatic Spillover"),
            ("Effective Date:", "January 1, 2026"),
            ("", ""),
            ("Critical Findings", ""),
            ("Total Employees:", f"{stats['total_employees']:,}"),
            ("RCR Employees:", f"{stats['rcr_employees']}"),
            ("High Priority Actions:", f"{stats['high_priority']} employees need ROTH codes"),
            ("Low Priority:", f"{stats['low_priority']} employees to monitor"),
            ("Average RCR Wages:", f"${stats['avg_rcr_wages']:,.2f}"),
        ]
        
        for idx, (label, value) in enumerate(info, start=3):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            ws[f'A{idx}'].font = self.BOLD_FONT
            if "Critical Findings" in label:
                ws[f'A{idx}'].fill = self.INFO_FILL
                ws.merge_cells(f'A{idx}:B{idx}')
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        
    def _create_configuration_guide(self):
        """Tab 3: Configuration Guide"""
        ws = self.wb.create_sheet("Configuration Guide")
        
        ws['A1'] = "UKG PRO CONFIGURATION GUIDE"
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].fill = self.TITLE_FILL
        ws.merge_cells('A1:B1')
        
        guide = [
            ("Step 1:", "Navigate to Deduction/Benefit Plans"),
            ("", "Menu > System Configuration > Business Rules > Deduction/Benefit Plans"),
            ("Step 2:", "Select Pre-Tax Catch-Up Deduction Code"),
            ("", "Choose code used for catch-up contributions (e.g., 401CP)"),
            ("Step 3:", "Enable SECURE 2.0 Act Spillover Configuration"),
            ("", "In Deduction setup > Next > Calculations page"),
            ("Step 4:", "Select ROTH Spillover Plan"),
            ("", "From dropdown, select corresponding ROTH catch-up plan"),
            ("Step 5:", "Save Configuration"),
            ("", "Review Summary page and Save"),
            ("Step 6:", "Add ROTH Codes to RCR Employees"),
            ("", "Use bulk import tool or add individually"),
        ]
        
        for idx, (step, detail) in enumerate(guide, start=3):
            ws[f'A{idx}'] = step
            ws[f'B{idx}'] = detail
            if step:
                ws[f'A{idx}'].font = self.BOLD_FONT
                ws[f'A{idx}'].fill = self.INFO_FILL
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 70
        
    def _create_high_priority(self):
        """Tab 4: HIGH Priority - Action Needed"""
        ws = self.wb.create_sheet("HIGH Priority - Action Needed")
        
        ws['A1'] = f"HIGH PRIORITY: {self.results['statistics']['high_priority']} RCR Employees Need ROTH Codes Added"
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].fill = self.HIGH_FILL
        ws.merge_cells('A1:J1')
        
        # Get high priority employees
        high_df = self.results['high_priority'][[
            'Employee Number',
            'Age_12_31_2025',
            '2024_SOC_MED_Wages',
            '2025_YTD_SOC_MED_Wages',
            'Active_401K_Codes',
            'Amount',
            'Percent',
            'Action_Reason'
        ]].copy()
        
        # Add to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(high_df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:  # Header
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                cell.border = self.THIN_BORDER
        
        # Set column widths
        col_widths = [18, 12, 18, 18, 20, 12, 10, 40]
        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width
        
    def _create_low_priority(self):
        """Tab 5: LOW Priority - Monitor"""
        ws = self.wb.create_sheet("LOW Priority - Monitor")
        
        low_count = len(self.results['low_priority'])
        ws['A1'] = f"LOW PRIORITY: {low_count} RCR Employees to Monitor"
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].fill = self.LOW_FILL
        ws.merge_cells('A1:I1')
        
        if low_count > 0:
            low_df = self.results['low_priority'][[
                'Employee Number',
                'Age_12_31_2025',
                '2024_SOC_MED_Wages',
                'Active_401K_Codes',
                'Action_Reason'
            ]]
            
            for r_idx, row in enumerate(dataframe_to_rows(low_df, index=False, header=True), start=3):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 3:
                        cell.font = self.HEADER_FONT
                        cell.fill = self.HEADER_FILL
                    cell.border = self.THIN_BORDER
        else:
            ws['A3'] = "No employees in this category"
        
    def _create_all_rcr(self):
        """Tab 6: All RCR Employees"""
        ws = self.wb.create_sheet("All RCR Employees")
        
        ws['A1'] = f"ALL RCR EMPLOYEES: Complete List (Total: {self.results['statistics']['rcr_employees']})"
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].fill = self.INFO_FILL
        ws.merge_cells('A1:K1')
        
        all_rcr = self.results['rcr_employees'][[
            'Employee Number',
            'Age_12_31_2025',
            '2024_SOC_MED_Wages',
            '2025_YTD_SOC_MED_Wages',
            'Active_401K_Codes',
            'Has_ROTH_Code',
            'Action',
            'Action_Reason'
        ]]
        
        for r_idx, row in enumerate(dataframe_to_rows(all_rcr, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                cell.border = self.THIN_BORDER
        
    def _create_import_template(self):
        """Tab 7: Import Template - HIGH Priority"""
        ws = self.wb.create_sheet("Import Template - HIGH Priority")
        
        ws['A1'] = f"BULK IMPORT TEMPLATE: Add ROTH Codes to {self.results['statistics']['high_priority']} RCR Employees"
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].fill = self.HIGH_FILL
        ws.merge_cells('A1:E1')
        
        # Create import template
        high_df = self.results['high_priority']
        import_data = pd.DataFrame({
            'Employee Number': high_df['Employee Number'],
            'Deduction Code': 'R401CU',  # ROTH 401k Catch-up
            'Amount': '',
            'Percent': high_df['Percent'].fillna(''),
            'Start Date': '01/01/2026'
        })
        
        for r_idx, row in enumerate(dataframe_to_rows(import_data, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                cell.border = self.THIN_BORDER
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
    def _create_observations(self):
        """Tab 8: Observations & Notes"""
        ws = self.wb.create_sheet("Payroll Observations & Notes")
        
        ws['A1'] = "PAYROLL OBSERVATIONS & HOUSEKEEPING NOTES"
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].fill = self.INFO_FILL
        ws.merge_cells('A1:B1')
        
        notes = [
            ("Configuration Status:", "Spillover configuration NOT YET configured"),
            ("Employee Communications:", "Affected employees must receive 'effective opportunity' notice"),
            ("Testing Required:", "Test payroll with RCR employees before 01/01/2026 go-live"),
            ("Record Keeper:", "Contact record keeper for file upload instructions"),
            ("", ""),
            ("Key Dates:", ""),
            ("Compliance Deadline:", "January 1, 2026"),
            ("Employee Notice Deadline:", "December 2025"),
            ("Testing Deadline:", "December 2025"),
        ]
        
        for idx, (label, value) in enumerate(notes, start=3):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            ws[f'A{idx}'].font = self.BOLD_FONT
            if "Key Dates" in label:
                ws[f'A{idx}'].fill = self.MED_FILL
                ws.merge_cells(f'A{idx}:B{idx}')
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60


def generate_report(company_name: str, analysis_results: Dict, output_path: str):
    """
    Convenience function to generate complete implementation plan.
    
    Args:
        company_name: Customer company name
        analysis_results: Output from SECURE20Analyzer
        output_path: Where to save the Excel file
    """
    generator = SECURE20ReportGenerator(company_name, analysis_results)
    generator.generate(output_path)


if __name__ == "__main__":
    # Test with Meyer analysis
    from secure_20_analyzer import run_analysis
    
    print("Running analysis...")
    results = run_analysis('/mnt/user-data/uploads/Meyer_Secure_2_0.xlsx')
    
    print("Generating report...")
    generate_report(
        company_name="MEYER COMPANY",
        analysis_results=results,
        output_path="/mnt/user-data/outputs/Meyer_XLR8_Implementation_Plan.xlsx"
    )
    
    print("\n✅ DONE! Report generated at /mnt/user-data/outputs/Meyer_XLR8_Implementation_Plan.xlsx")
