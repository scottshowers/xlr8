"""
PLAYBOOKS ROUTER - Thin layer using PlaybookFramework
======================================================

This router delegates to the PlaybookFramework for core operations.
Playbook-specific customization comes from registered playbook definitions.

Endpoints:
- GET /playbooks/{type}/structure - Get playbook structure
- GET /playbooks/{type}/progress/{project_id} - Get progress
- POST /playbooks/{type}/progress/{project_id} - Update progress
- POST /playbooks/{type}/scan/{project_id}/{action_id} - Scan for action
- GET /playbooks/{type}/export/{project_id} - Export as XLSX

Author: XLR8 Team
Version: 4.0.0 - Framework Integration
Date: December 2025
"""

from fastapi import APIRouter, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
import logging
import json
import io
import os
import re
import uuid
import asyncio
from datetime import datetime
from pathlib import Path

# Excel generation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# LLMOrchestrator for consistent calls
from utils.llm_orchestrator import LLMOrchestrator
_playbooks_orchestrator = LLMOrchestrator()

router = APIRouter(tags=["playbooks"])

# Job tracking for scan-all background tasks
SCAN_JOBS: Dict[str, Dict[str, Any]] = {}


# =============================================================================
# FRAMEWORK IMPORTS
# =============================================================================

try:
    from backend.utils.playbook_framework import (
        PlaybookEngine,
        PlaybookDefinition,
        PlaybookRegistry,
        PLAYBOOK_REGISTRY,
        PROGRESS_MANAGER,
        ActionStatus,
        LearningHook,
        get_playbook_engine,
    )
    FRAMEWORK_AVAILABLE = True
    logger.info("[PLAYBOOKS] Framework imported successfully")
except ImportError:
    try:
        from utils.playbook_framework import (
            PlaybookEngine,
            PlaybookDefinition,
            PlaybookRegistry,
            PLAYBOOK_REGISTRY,
            PROGRESS_MANAGER,
            ActionStatus,
            LearningHook,
            get_playbook_engine,
        )
        FRAMEWORK_AVAILABLE = True
        logger.info("[PLAYBOOKS] Framework imported (alt path)")
    except ImportError as e:
        FRAMEWORK_AVAILABLE = False
        logger.warning(f"[PLAYBOOKS] Framework not available: {e}")

# Year-End specific config
try:
    from backend.playbooks.year_end_playbook import (
        get_year_end_config,
        register_year_end_playbook,
        YEAR_END_CONSULTATIVE_PROMPTS,
        YEAR_END_DEPENDENT_GUIDANCE,
    )
    YEAR_END_CONFIG = get_year_end_config()
    # Register the playbook on import
    if FRAMEWORK_AVAILABLE:
        register_year_end_playbook()
except ImportError:
    try:
        from playbooks.year_end_playbook import (
            get_year_end_config,
            register_year_end_playbook,
            YEAR_END_CONSULTATIVE_PROMPTS,
            YEAR_END_DEPENDENT_GUIDANCE,
        )
        YEAR_END_CONFIG = get_year_end_config()
        # Register the playbook on import
        if FRAMEWORK_AVAILABLE:
            register_year_end_playbook()
    except ImportError:
        YEAR_END_CONFIG = {}
        logger.warning("[PLAYBOOKS] Year-End config not available")


# =============================================================================
# LEGACY SUPPORT - For backward compatibility during transition
# =============================================================================

# In-memory progress (fallback if framework unavailable)
PLAYBOOK_PROGRESS: Dict[str, Dict[str, Any]] = {}

# Cached structure
PLAYBOOK_CACHE: Dict[str, Any] = {}

# Supabase client
from utils.database.supabase_client import get_supabase


# =============================================================================
# STRUCTURE PARSING - Load playbook from uploaded document
# =============================================================================

async def get_year_end_structure(force_refresh: bool = False) -> Dict[str, Any]:
    """
    Get Year-End playbook structure.
    
    Parses from uploaded Year-End Checklist document.
    """
    cache_key = "year-end-2025"
    
    # Clear cache if force refresh
    if force_refresh and cache_key in PLAYBOOK_CACHE:
        del PLAYBOOK_CACHE[cache_key]
        logger.info("[STRUCTURE] Force refresh - cache cleared")
    
    if cache_key in PLAYBOOK_CACHE:
        return PLAYBOOK_CACHE[cache_key]
    
    try:
        from backend.utils.playbook_parser import parse_year_end_checklist
    except ImportError:
        from utils.playbook_parser import parse_year_end_checklist
    
    structure = parse_year_end_checklist()
    
    if structure:
        PLAYBOOK_CACHE[cache_key] = structure
        logger.info(f"[STRUCTURE] Cached Year-End structure: {len(structure.get('steps', []))} steps")
    
    return structure


def invalidate_year_end_cache():
    """Invalidate cache when new document uploaded."""
    if "year-end-2025" in PLAYBOOK_CACHE:
        del PLAYBOOK_CACHE["year-end-2025"]
        logger.info("[CACHE] Year-End structure cache invalidated")


# =============================================================================
# REQUEST/RESPONSE MODELS
# =============================================================================

class ProgressUpdate(BaseModel):
    status: str
    notes: Optional[str] = None


class FeedbackRequest(BaseModel):
    finding_text: str
    feedback: str  # 'keep', 'discard', 'modify'
    reason: Optional[str] = None


class EntityConfigRequest(BaseModel):
    primary_entity: Optional[str] = None
    fein: Optional[str] = None
    company_name: Optional[str] = None


class QuickSuppressRequest(BaseModel):
    finding_text: str
    reason: str
    suppress_type: str = 'acknowledge'
    fein: Optional[str] = None


# =============================================================================
# CORE ENDPOINTS
# =============================================================================

@router.get("/year-end/structure")
async def get_playbook_structure():
    """Get Year-End playbook structure."""
    try:
        structure = await get_year_end_structure()
        if not structure:
            raise HTTPException(status_code=404, detail="Year-End structure not found")
        return structure
    except Exception as e:
        logger.exception(f"Failed to get structure: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/year-end/refresh-structure")
async def refresh_playbook_structure():
    """Refresh Year-End playbook structure from Reference Library."""
    try:
        logger.info("[PLAYBOOK] Refreshing structure from Reference Library...")
        
        # Force re-parse from DuckDB
        structure = await get_year_end_structure(force_refresh=True)
        
        if not structure:
            return {
                "success": False,
                "error": "Year-End Checklist not found in Reference Library"
            }
        
        # Structure has 'steps' array and 'total_actions' directly
        total_actions = structure.get('total_actions', 0)
        source_file = structure.get('source_file', 'Year-End Checklist')
        
        # If total_actions not in structure, count from steps
        if total_actions == 0 and 'steps' in structure:
            for step in structure['steps']:
                total_actions += len(step.get('actions', []))
        
        logger.info(f"[PLAYBOOK] Refreshed structure: {total_actions} actions from {source_file}")
        
        return {
            "success": True,
            "total_actions": total_actions,
            "source_file": source_file,
            "structure": structure
        }
    except Exception as e:
        logger.exception(f"Failed to refresh structure: {e}")
        return {
            "success": False,
            "error": str(e)
        }


@router.get("/year-end/progress/{project_id}")
async def get_progress(project_id: str):
    """Get progress for a project."""
    try:
        if FRAMEWORK_AVAILABLE:
            progress = PROGRESS_MANAGER.get_raw_progress("year-end", project_id)
        else:
            progress = PLAYBOOK_PROGRESS.get(project_id, {})
        
        return {
            "project_id": project_id,
            "progress": progress
        }
    except Exception as e:
        logger.exception(f"Failed to get progress: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/year-end/progress/{project_id}")
async def update_progress(project_id: str, action_id: str, update: ProgressUpdate):
    """Update progress for an action."""
    try:
        if FRAMEWORK_AVAILABLE:
            engine = get_playbook_engine("year-end")
            if engine:
                result = engine.update_status(
                    project_id, action_id, update.status, update.notes
                )
                return result
        
        # Fallback
        if project_id not in PLAYBOOK_PROGRESS:
            PLAYBOOK_PROGRESS[project_id] = {}
        
        existing = PLAYBOOK_PROGRESS[project_id].get(action_id, {})
        existing["status"] = update.status
        if update.notes is not None:
            existing["notes"] = update.notes
        PLAYBOOK_PROGRESS[project_id][action_id] = existing
        
        return {"success": True, "status": update.status}
        
    except Exception as e:
        logger.exception(f"Failed to update progress: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/year-end/summary/{project_id}")
async def get_ai_summary(project_id: str):
    """Get consolidated AI summary across all scanned actions."""
    if FRAMEWORK_AVAILABLE:
        progress = PROGRESS_MANAGER.get_raw_progress("year-end", project_id)
    else:
        progress = PLAYBOOK_PROGRESS.get(project_id, {})
    
    # Aggregate data
    all_issues = []
    all_recommendations = []
    all_key_values = {}
    all_conflicts = []
    actions_with_flags = []
    high_risk_actions = []
    
    # Keywords to filter out noise
    noise_keywords = [
        'fein', 'ein', 'federal employer', 'multiple feins', 'fein format',
        'formatting issue', 'column alignment', 'garbled', 'misaligned data',
        'conduct complete audit', 'conduct a complete audit', 'ensure consistency',
        'verify consistency', 'verify and update', 'review data', 'reconcile all'
    ]
    
    for action_id, action_progress in progress.items():
        findings = action_progress.get("findings", {})
        if not findings:
            continue
        
        # Collect issues with filtering
        for issue in findings.get("issues", []):
            if not any(kw in issue.lower() for kw in noise_keywords):
                all_issues.append({
                    "action_id": action_id,
                    "issue": issue,
                    "risk_level": findings.get("risk_level", "medium")
                })
        
        # Collect recommendations with filtering
        for rec in findings.get("recommendations", []):
            if not any(kw in rec.lower() for kw in noise_keywords):
                all_recommendations.append({
                    "action_id": action_id,
                    "recommendation": rec
                })
        
        # Collect key values
        for key, value in findings.get("key_values", {}).items():
            if key not in all_key_values:
                all_key_values[key] = {"value": value, "source": action_id}
        
        # Collect conflicts (excluding FEIN)
        for conflict in findings.get("conflicts", []):
            field = conflict.get("field", "").lower()
            if field != "fein" and "fein" not in conflict.get("message", "").lower():
                all_conflicts.append(conflict)
        
        # Track review flags
        if action_progress.get("review_flag"):
            actions_with_flags.append({
                "action_id": action_id,
                "flag": action_progress["review_flag"]
            })
        
        # Track high risk
        if findings.get("risk_level") == "high":
            high_risk_actions.append({
                "action_id": action_id,
                "summary": findings.get("summary", "")
            })
    
    # Sort by risk
    risk_order = {"high": 0, "medium": 1, "low": 2}
    all_issues.sort(key=lambda x: risk_order.get(x["risk_level"], 3))
    
    # Calculate overall risk
    if all_conflicts:
        overall_risk = "high"
    elif len(high_risk_actions) >= 5:
        overall_risk = "medium"
    else:
        overall_risk = "low"
    
    # Generate summary text
    summary_parts = []
    if high_risk_actions:
        summary_parts.append(f"📋 {len(high_risk_actions)} action(s) pending analysis")
    if all_conflicts:
        summary_parts.append(f"❗ {len(all_conflicts)} data conflict(s) detected")
    if actions_with_flags:
        summary_parts.append(f"🔄 {len(actions_with_flags)} action(s) flagged for review")
    
    total_complete = sum(1 for p in progress.values() if p.get("status") == "complete")
    total_in_progress = sum(1 for p in progress.values() if p.get("status") == "in_progress")
    
    return {
        "project_id": project_id,
        "overall_risk": overall_risk,
        "summary_text": " | ".join(summary_parts) if summary_parts else "No critical issues detected",
        "stats": {
            "actions_scanned": len(progress),
            "actions_complete": total_complete,
            "actions_in_progress": total_in_progress,
            "total_issues": len(all_issues),
            "total_recommendations": len(all_recommendations),
            "total_conflicts": len(all_conflicts),
            "actions_flagged": len(actions_with_flags)
        },
        "issues": all_issues[:20],
        "recommendations": all_recommendations[:15],
        "key_values": all_key_values,
        "conflicts": all_conflicts,
        "review_flags": actions_with_flags,
        "high_risk_actions": high_risk_actions
    }


@router.post("/year-end/scan/{project_id}/{action_id}")
async def scan_for_action(project_id: str, action_id: str):
    """
    Scan documents for a specific action.
    
    This is the main scan endpoint - delegates to framework.
    """
    logger.info(f"[SCAN] {action_id} in project {project_id[:8]}")
    
    try:
        if FRAMEWORK_AVAILABLE:
            engine = get_playbook_engine("year-end")
            if engine:
                # Ensure structure is loaded into engine
                structure = await get_year_end_structure()
                if structure:
                    # Update engine's playbook with parsed structure
                    _sync_structure_to_engine(engine, structure)
                
                result = await engine.scan_action(project_id, action_id)
                return result.to_dict()
        
        # Fallback to legacy implementation
        return await _legacy_scan(project_id, action_id)
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception(f"Scan failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _sync_structure_to_engine(engine: PlaybookEngine, structure: Dict):
    """
    Sync parsed structure to engine's playbook definition.
    
    The structure comes from parsing the uploaded document.
    """
    try:
        from backend.utils.playbook_framework import StepDefinition, ActionDefinition, ActionType
    except ImportError:
        from utils.playbook_framework import StepDefinition, ActionDefinition, ActionType
    
    steps = []
    for step_data in structure.get('steps', []):
        actions = []
        for action_data in step_data.get('actions', []):
            action_type_str = action_data.get('action_type', 'recommended')
            try:
                action_type = ActionType(action_type_str)
            except ValueError:
                action_type = ActionType.RECOMMENDED
            
            action = ActionDefinition(
                action_id=action_data['action_id'],
                description=action_data.get('description', ''),
                action_type=action_type,
                reports_needed=action_data.get('reports_needed', []),
                due_date=action_data.get('due_date'),
                quarter_end=action_data.get('quarter_end', False),
                keywords=YEAR_END_CONFIG.get('action_keywords', {}).get(
                    action_data['action_id'], {}
                ).get('keywords', []),
                categories=YEAR_END_CONFIG.get('action_keywords', {}).get(
                    action_data['action_id'], {}
                ).get('categories', []),
                guidance=YEAR_END_CONFIG.get('dependent_guidance', {}).get(
                    action_data['action_id']
                ),
                consultative_prompt=YEAR_END_CONFIG.get('consultative_prompts', {}).get(
                    action_data['action_id']
                ),
            )
            actions.append(action)
        
        step = StepDefinition(
            step_number=step_data['step_number'],
            step_name=step_data.get('step_name', ''),
            phase=step_data.get('phase', 'before_final_payroll'),
            actions=actions
        )
        steps.append(step)
    
    engine.playbook.steps = steps
    # Rebuild dependencies from new structure
    try:
        from backend.utils.playbook_framework import InheritanceManager
    except ImportError:
        from utils.playbook_framework import InheritanceManager
    engine.dependencies = InheritanceManager.build_dependencies(engine.playbook)


async def _legacy_scan(project_id: str, action_id: str) -> Dict:
    """Legacy scan implementation for backward compatibility."""
    # This would contain the old scan logic
    # For now, return a placeholder
    return {
        "found": False,
        "documents": [],
        "findings": None,
        "suggested_status": "not_started",
        "message": "Framework not available, using legacy mode"
    }


# =============================================================================
# FEEDBACK ENDPOINT - Learning integration
# =============================================================================

@router.post("/year-end/feedback/{project_id}/{action_id}")
async def record_feedback(
    project_id: str, 
    action_id: str, 
    request: FeedbackRequest
):
    """Record user feedback on a finding."""
    try:
        if FRAMEWORK_AVAILABLE:
            engine = get_playbook_engine("year-end")
            if engine:
                success = engine.record_feedback(
                    project_id=project_id,
                    action_id=action_id,
                    finding_text=request.finding_text,
                    feedback=request.feedback,
                    reason=request.reason
                )
                return {"success": success}
        
        # Fallback - just acknowledge
        return {"success": True, "message": "Feedback noted (learning not active)"}
        
    except Exception as e:
        logger.error(f"[FEEDBACK] Error: {e}")
        return {"success": False, "message": str(e)}


# =============================================================================
# DOCUMENT CHECKLIST ENDPOINT
# =============================================================================

@router.get("/year-end/document-checklist/{project_id}")
async def get_document_checklist(project_id: str):
    """
    Get the document checklist with real-time upload status.
    Shows which reports are needed per step, matched vs missing.
    
    Uses document_registry as source of truth for uploaded files.
    """
    try:
        from utils.database.models import ProcessingJobModel, DocumentRegistryModel
        from backend.utils.playbook_parser import load_step_documents, match_documents_to_step
    except ImportError:
        try:
            from utils.database.models import ProcessingJobModel, DocumentRegistryModel
            from utils.playbook_parser import load_step_documents, match_documents_to_step
        except ImportError:
            raise HTTPException(status_code=500, detail="Required modules not available")
    
    # =================================================================
    # Get uploaded files from REGISTRY (source of truth)
    # =================================================================
    uploaded_files_list = []
    
    try:
        # Get files for this project (including global/reference library)
        registry_entries = DocumentRegistryModel.get_by_project(project_id, include_global=True)
        
        for entry in registry_entries:
            filename = entry.get('filename')
            if filename:
                # Exclude playbook source documents (the checklist itself)
                usage_type = entry.get('usage_type', '')
                if usage_type == 'playbook_source':
                    logger.debug(f"[DOC-CHECKLIST] Skipping playbook source: {filename}")
                    continue
                
                uploaded_files_list.append(filename)
        
        logger.info(f"[DOC-CHECKLIST] Registry: {len(uploaded_files_list)} files for project {project_id[:8]}")
        
    except Exception as reg_e:
        logger.warning(f"[DOC-CHECKLIST] Registry query failed: {reg_e}")
        # Return empty if registry unavailable
        return {
            "project_id": project_id,
            "has_step_documents": False,
            "uploaded_files": [],
            "step_checklists": [],
            "stats": {"files_in_project": 0, "total_matched": 0, "total_missing": 0, "required_missing": 0},
            "processing_jobs": [],
            "error": f"Registry unavailable: {reg_e}"
        }
    
    uploaded_files_list.sort()
    logger.info(f"[DOC-CHECKLIST] TOTAL: {len(uploaded_files_list)} files")
    
    # Check for active processing jobs
    processing_jobs = []
    try:
        all_jobs = ProcessingJobModel.get_all(limit=20)
        for job in all_jobs:
            job_status = job.get("status", "")
            job_project = job.get("input_data", {}).get("project_id", "")
            if job_status in ["pending", "processing"] and job_project == project_id:
                processing_jobs.append({
                    "filename": job.get("input_data", {}).get("filename", "Unknown"),
                    "progress": job.get("progress", 0),
                    "message": job.get("status_message", "Processing..."),
                    "job_id": job.get("id")
                })
    except Exception as e:
        logger.warning(f"Could not fetch processing jobs: {e}")
    
    # STEP-BASED DOCUMENT CHECKLIST (from Step_Documents sheet)
    step_checklists = []
    has_step_documents = False
    total_matched = 0
    total_missing = 0
    required_missing = 0
    
    try:
        step_documents = load_step_documents()
        
        # Get step names from structure
        step_names_map = {}
        try:
            structure = await get_year_end_structure()
            for step in structure.get('steps', []):
                step_names_map[step['step_number']] = step.get('step_name', f"Step {step['step_number']}")
        except Exception as e:
            logger.warning(f"[DOC-CHECKLIST] Could not load step names: {e}")
        
        if step_documents:
            has_step_documents = True
            logger.info(f"[DOC-CHECKLIST] Found Step_Documents for {len(step_documents)} steps")
            
            for step_num, docs in sorted(step_documents.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999):
                result = match_documents_to_step(docs, uploaded_files_list)
                
                actual_step_name = step_names_map.get(step_num, "")
                
                step_checklists.append({
                    'step_number': step_num,
                    'step_name': actual_step_name,
                    'matched': result['matched'],
                    'missing': result['missing'],
                    'stats': result['stats']
                })
                
                total_matched += result['stats']['matched']
                total_missing += result['stats']['missing']
                required_missing += result['stats']['required_missing']
        else:
            logger.info("[DOC-CHECKLIST] No Step_Documents found - showing uploaded files only")
            
    except Exception as e:
        logger.warning(f"[DOC-CHECKLIST] Could not load Step_Documents: {e}")
    
    return {
        "project_id": project_id,
        "has_step_documents": has_step_documents,
        "uploaded_files": uploaded_files_list,
        "step_checklists": step_checklists,
        "stats": {
            "files_in_project": len(uploaded_files_list),
            "total_matched": total_matched,
            "total_missing": total_missing,
            "required_missing": required_missing
        },
        "processing_jobs": processing_jobs
    }


# =============================================================================
# EXPORT ENDPOINT
# =============================================================================

@router.get("/year-end/export/{project_id}")
async def export_playbook(project_id: str):
    """
    Export playbook progress as comprehensive Excel workbook.
    
    Sheets:
    1. Summary - Overall completion stats
    2. Checklist - All actions with status
    3. Issues - All identified issues across all actions
    4. Recommendations - All recommendations
    5. Analysis Details - Full analysis summaries per action
    """
    try:
        structure = await get_year_end_structure()
        
        if FRAMEWORK_AVAILABLE:
            progress = PROGRESS_MANAGER.get_raw_progress("year-end", project_id)
        else:
            progress = PLAYBOOK_PROGRESS.get(project_id, {})
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        issue_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        complete_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        def style_header(ws, row=1):
            for cell in ws[row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
        
        # Collect all issues and recommendations across actions
        all_issues = []
        all_recommendations = []
        all_analyses = []
        status_counts = {"complete": 0, "in_progress": 0, "not_started": 0, "blocked": 0, "n_a": 0}
        total_actions = 0
        
        for step in structure.get('steps', []):
            for action in step.get('actions', []):
                total_actions += 1
                action_id = action['action_id']
                action_progress = progress.get(action_id, {})
                status = action_progress.get('status', 'not_started')
                
                # Count statuses
                status_key = status.replace(' ', '_').lower()
                if status_key in status_counts:
                    status_counts[status_key] += 1
                
                findings = action_progress.get('findings', {}) or {}
                
                # Collect issues
                for issue in findings.get('issues', []):
                    all_issues.append({
                        "action_id": action_id,
                        "step": f"Step {step['step_number']}",
                        "action_description": action.get('description', '')[:100],
                        "issue": issue,
                        "status": status
                    })
                
                # Collect recommendations
                for rec in findings.get('recommendations', []):
                    all_recommendations.append({
                        "action_id": action_id,
                        "step": f"Step {step['step_number']}",
                        "action_description": action.get('description', '')[:100],
                        "recommendation": rec,
                        "status": status
                    })
                
                # Collect analysis summaries
                summary = findings.get('summary', '')
                if summary:
                    extracted = findings.get('extracted_data', {})
                    extracted_str = ""
                    if extracted:
                        extracted_str = "; ".join([f"{k}: {v}" for k, v in extracted.items()][:5])
                    
                    all_analyses.append({
                        "action_id": action_id,
                        "step": f"Step {step['step_number']}",
                        "description": action.get('description', '')[:150],
                        "summary": summary,
                        "extracted_data": extracted_str,
                        "method": findings.get('analysis_method', ''),
                        "confidence": findings.get('confidence', ''),
                        "sources": ", ".join(action_progress.get('documents_found', [])[:3])
                    })
        
        # =====================================================================
        # Sheet 1: Summary
        # =====================================================================
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        ws_summary['A1'] = "Year-End Playbook Export"
        ws_summary['A1'].font = Font(bold=True, size=16)
        
        ws_summary['A3'] = "Project ID:"
        ws_summary['B3'] = project_id
        ws_summary['A4'] = "Export Date:"
        ws_summary['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws_summary['A5'] = "Total Actions:"
        ws_summary['B5'] = total_actions
        
        ws_summary['A7'] = "Completion Status"
        ws_summary['A7'].font = Font(bold=True)
        ws_summary['A8'] = "Complete:"
        ws_summary['B8'] = status_counts.get('complete', 0)
        ws_summary['A9'] = "In Progress:"
        ws_summary['B9'] = status_counts.get('in_progress', 0)
        ws_summary['A10'] = "Not Started:"
        ws_summary['B10'] = status_counts.get('not_started', 0)
        ws_summary['A11'] = "Blocked:"
        ws_summary['B11'] = status_counts.get('blocked', 0)
        ws_summary['A12'] = "N/A:"
        ws_summary['B12'] = status_counts.get('n_a', 0)
        
        ws_summary['A14'] = "Issues Found:"
        ws_summary['B14'] = len(all_issues)
        ws_summary['A15'] = "Recommendations:"
        ws_summary['B15'] = len(all_recommendations)
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 40
        
        # =====================================================================
        # Sheet 2: Checklist (all actions)
        # =====================================================================
        ws_checklist = wb.create_sheet("Checklist")
        
        headers = [
            "Action ID", "Step", "Type", "Description", "Due Date",
            "Reports Needed", "Docs Found", "Status", "Issues Count", 
            "Analysis Summary", "Notes"
        ]
        ws_checklist.append(headers)
        style_header(ws_checklist)
        
        for step in structure.get('steps', []):
            for action in step.get('actions', []):
                action_id = action['action_id']
                action_progress = progress.get(action_id, {})
                findings = action_progress.get('findings', {}) or {}
                
                status = action_progress.get('status', 'not_started')
                issues_count = len(findings.get('issues', []))
                
                row_data = [
                    action_id,
                    f"Step {step['step_number']}",
                    action.get('action_type', 'Recommended'),
                    action.get('description', '')[:200],
                    action.get('due_date', 'N/A'),
                    ', '.join(action.get('reports_needed', [])),
                    ', '.join(action_progress.get('documents_found', [])[:3]),
                    status.replace('_', ' ').title(),
                    issues_count,
                    (findings.get('summary', '') or '')[:300],
                    action_progress.get('notes', '') or ''
                ]
                ws_checklist.append(row_data)
                
                # Color code by status
                row_num = ws_checklist.max_row
                if status == 'complete':
                    for col in range(1, 12):
                        ws_checklist.cell(row=row_num, column=col).fill = complete_fill
                elif issues_count > 0:
                    for col in range(1, 12):
                        ws_checklist.cell(row=row_num, column=col).fill = issue_fill
        
        # Set column widths
        widths = [10, 10, 12, 60, 12, 40, 40, 12, 10, 60, 30]
        for i, width in enumerate(widths, 1):
            ws_checklist.column_dimensions[get_column_letter(i)].width = width
        
        # =====================================================================
        # Sheet 3: Issues Detail
        # =====================================================================
        if all_issues:
            ws_issues = wb.create_sheet("Issues")
            
            issue_headers = ["Action ID", "Step", "Action", "Issue", "Status"]
            ws_issues.append(issue_headers)
            style_header(ws_issues)
            
            for issue in all_issues:
                ws_issues.append([
                    issue['action_id'],
                    issue['step'],
                    issue['action_description'],
                    issue['issue'],
                    issue['status'].replace('_', ' ').title()
                ])
            
            ws_issues.column_dimensions['A'].width = 10
            ws_issues.column_dimensions['B'].width = 10
            ws_issues.column_dimensions['C'].width = 50
            ws_issues.column_dimensions['D'].width = 80
            ws_issues.column_dimensions['E'].width = 12
            
            # Enable text wrapping for issue column
            for row in ws_issues.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    cell.alignment = wrap_alignment
        
        # =====================================================================
        # Sheet 4: Recommendations
        # =====================================================================
        if all_recommendations:
            ws_recs = wb.create_sheet("Recommendations")
            
            rec_headers = ["Action ID", "Step", "Action", "Recommendation", "Status"]
            ws_recs.append(rec_headers)
            style_header(ws_recs)
            
            for rec in all_recommendations:
                ws_recs.append([
                    rec['action_id'],
                    rec['step'],
                    rec['action_description'],
                    rec['recommendation'],
                    rec['status'].replace('_', ' ').title()
                ])
            
            ws_recs.column_dimensions['A'].width = 10
            ws_recs.column_dimensions['B'].width = 10
            ws_recs.column_dimensions['C'].width = 50
            ws_recs.column_dimensions['D'].width = 80
            ws_recs.column_dimensions['E'].width = 12
            
            for row in ws_recs.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    cell.alignment = wrap_alignment
        
        # =====================================================================
        # Sheet 5: Analysis Details
        # =====================================================================
        if all_analyses:
            ws_analysis = wb.create_sheet("Analysis Details")
            
            analysis_headers = ["Action ID", "Step", "Description", "Analysis Summary", 
                               "Extracted Data", "Method", "Confidence", "Sources"]
            ws_analysis.append(analysis_headers)
            style_header(ws_analysis)
            
            for analysis in all_analyses:
                ws_analysis.append([
                    analysis['action_id'],
                    analysis['step'],
                    analysis['description'],
                    analysis['summary'],
                    analysis['extracted_data'],
                    analysis['method'],
                    str(analysis['confidence']) if analysis['confidence'] else '',
                    analysis['sources']
                ])
            
            ws_analysis.column_dimensions['A'].width = 10
            ws_analysis.column_dimensions['B'].width = 10
            ws_analysis.column_dimensions['C'].width = 50
            ws_analysis.column_dimensions['D'].width = 80
            ws_analysis.column_dimensions['E'].width = 40
            ws_analysis.column_dimensions['F'].width = 15
            ws_analysis.column_dimensions['G'].width = 10
            ws_analysis.column_dimensions['H'].width = 40
            
            for row in ws_analysis.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    cell.alignment = wrap_alignment
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Year_End_Playbook_{project_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.exception(f"Export failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# ENTITY CONFIGURATION ENDPOINTS
# =============================================================================

@router.get("/year-end/entity-config/{project_id}")
async def get_entity_config(project_id: str):
    """Get entity configuration for a project."""
    try:
        from utils.database.models import EntityConfigModel
        config = EntityConfigModel.get(project_id, "year-end")
        return {"config": config}
    except ImportError:
        return {"config": None, "message": "Entity config not available"}
    except Exception as e:
        logger.error(f"[ENTITY] Get config error: {e}")
        return {"config": None, "error": str(e)}


@router.post("/year-end/entity-config/{project_id}")
async def set_entity_config(project_id: str, request: EntityConfigRequest):
    """Set entity configuration for a project."""
    try:
        from utils.database.models import EntityConfigModel
        result = EntityConfigModel.save(
            project_id=project_id,
            playbook_type="year-end",
            primary_entity=request.primary_entity,
            analysis_scope="selected" if request.primary_entity else "all",
            selected_entities=[request.primary_entity] if request.primary_entity else [],
            country_mode="us_only"
        )
        return {"success": True, "config": result}
    except ImportError:
        return {"success": False, "message": "Entity config not available"}
    except Exception as e:
        logger.error(f"[ENTITY] Set config error: {e}")
        return {"success": False, "error": str(e)}


# =============================================================================
# SUPPRESSION ENDPOINTS
# =============================================================================

@router.get("/year-end/suppressions/{project_id}")
async def get_suppressions(project_id: str):
    """Get suppression rules for a project."""
    try:
        from utils.database.models import FindingSuppressionModel
        rules = FindingSuppressionModel.get_by_project(project_id, "year-end")
        stats = FindingSuppressionModel.get_stats(project_id, "year-end")
        return {"rules": rules, "stats": stats}
    except ImportError:
        return {"rules": [], "stats": {}, "message": "Suppression not available"}
    except Exception as e:
        logger.error(f"[SUPPRESS] Get rules error: {e}")
        return {"rules": [], "stats": {}, "error": str(e)}


class SuppressRequest(BaseModel):
    """Request body for suppress endpoint - flexible field names."""
    finding_text: Optional[str] = None
    pattern_text: Optional[str] = None  # Alternative name frontend might use
    text: Optional[str] = None  # Another alternative
    pattern: Optional[str] = None  # Yet another alternative
    reason: Optional[str] = "User suppressed"
    action_id: Optional[str] = None
    suppress_type: str = 'acknowledge'
    suppression_type: Optional[str] = None  # Alternative name
    fein: Optional[str] = None
    
    class Config:
        extra = 'allow'  # Accept any extra fields
    
    def get_finding_text(self) -> str:
        """Get finding text from whichever field was provided."""
        return self.finding_text or self.pattern_text or self.text or self.pattern or ""


@router.post("/year-end/suppress/{project_id}")
async def suppress_finding(project_id: str, request: SuppressRequest):
    """
    Suppress a finding - feeds BOTH suppression model AND learning system.
    
    This is the endpoint the frontend UI calls.
    """
    # Log what we received for debugging
    logger.info(f"[SUPPRESS] Raw request fields: {request.dict()}")
    
    success = False
    rule_id = None
    action_id = request.action_id or "general"
    finding_text = request.get_finding_text()
    reason = request.reason or "User suppressed"
    suppress_type = request.suppression_type or request.suppress_type or 'acknowledge'
    
    if not finding_text:
        logger.warning(f"[SUPPRESS] No finding text found in request: {request.dict()}")
        return {"success": False, "message": "No finding text provided"}
    
    logger.info(f"[SUPPRESS] Received: finding='{finding_text[:50]}...', reason='{reason}'")
    
    # 1. Store in FindingSuppressionModel (existing behavior)
    try:
        from utils.database.models import FindingSuppressionModel
        result = FindingSuppressionModel.create(
            project_id=project_id,
            playbook_type="year-end",
            action_id=action_id,
            suppression_type=suppress_type,
            pattern=finding_text,  # The model uses 'pattern' column
            reason=reason,
            fein_filter=[request.fein] if request.fein else None
        )
        if result:
            success = True
            rule_id = result['id']
            logger.info(f"[SUPPRESS] Created suppression rule {rule_id}")
    except ImportError:
        logger.warning("[SUPPRESS] FindingSuppressionModel not available")
    except Exception as e:
        logger.error(f"[SUPPRESS] Suppress error: {e}")
    
    # 2. ALSO feed into Learning System
    try:
        if FRAMEWORK_AVAILABLE and LearningHook.is_available():
            engine = get_playbook_engine("year-end")
            if engine:
                engine.record_feedback(
                    project_id=project_id,
                    action_id=action_id,
                    finding_text=finding_text,
                    feedback='discard',
                    reason=reason
                )
                logger.info(f"[LEARNING] Recorded discard from suppress UI: {finding_text[:50]}...")
                success = True
    except Exception as e:
        logger.warning(f"[LEARNING] Failed to record in learning system: {e}")
    
    if success:
        return {"success": True, "rule_id": rule_id, "learning_recorded": True}
    return {"success": False, "message": "Both systems unavailable"}


@router.post("/year-end/suppress/quick/{project_id}/{action_id}")
async def quick_suppress(
    project_id: str, 
    action_id: str, 
    request: QuickSuppressRequest
):
    """Quick suppress from UI - feeds BOTH suppression model AND learning system."""
    success = False
    rule_id = None
    
    # 1. Store in FindingSuppressionModel (existing behavior)
    try:
        from utils.database.models import FindingSuppressionModel
        result = FindingSuppressionModel.create(
            project_id=project_id,
            playbook_type="year-end",
            action_id=action_id,
            suppression_type=request.suppress_type,
            pattern=request.finding_text,  # The model uses 'pattern' column
            reason=request.reason,
            fein_filter=[request.fein] if request.fein else None
        )
        if result:
            success = True
            rule_id = result['id']
            logger.info(f"[SUPPRESS] Created suppression rule {rule_id}")
    except ImportError:
        logger.warning("[SUPPRESS] FindingSuppressionModel not available")
    except Exception as e:
        logger.error(f"[SUPPRESS] Quick suppress error: {e}")
    
    # 2. ALSO feed into Learning System (new behavior)
    try:
        if FRAMEWORK_AVAILABLE and LearningHook.is_available():
            engine = get_playbook_engine("year-end")
            if engine:
                engine.record_feedback(
                    project_id=project_id,
                    action_id=action_id,
                    finding_text=request.finding_text,
                    feedback='discard',
                    reason=request.reason
                )
                logger.info(f"[LEARNING] Recorded discard from suppress UI: {request.finding_text[:50]}...")
                success = True  # Mark success even if suppression model failed
    except Exception as e:
        logger.warning(f"[LEARNING] Failed to record in learning system: {e}")
    
    if success:
        return {"success": True, "rule_id": rule_id, "learning_recorded": True}
    return {"success": False, "message": "Both systems unavailable"}


# =============================================================================
# LEARNING ENDPOINTS
# =============================================================================

@router.get("/year-end/learning/stats")
async def get_learning_stats():
    """Get learning system statistics."""
    if not FRAMEWORK_AVAILABLE:
        return {"success": False, "message": "Framework not available"}
    
    if not LearningHook.is_available():
        return {"success": False, "message": "Learning system not active"}
    
    try:
        try:
            from backend.utils.learning_engine import get_learning_system
        except ImportError:
            from utils.learning_engine import get_learning_system
        
        learning = get_learning_system()
        return {
            "success": True,
            "stats": learning.get_stats()
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# SCAN ALL ENDPOINT
# =============================================================================

@router.post("/year-end/scan-all/{project_id}")
async def scan_all_actions(project_id: str, background_tasks: BackgroundTasks):
    """Scan all actions for a project (background job)."""
    try:
        structure = await get_year_end_structure()
        
        # Count total actions
        total_actions = 0
        for step in structure.get('steps', []):
            total_actions += len(step.get('actions', []))
        
        # Create job ID
        job_id = str(uuid.uuid4())
        
        # Initialize job status
        SCAN_JOBS[job_id] = {
            "status": "running",
            "project_id": project_id,
            "total": total_actions,
            "completed": 0,
            "successful": 0,
            "failed": 0,
            "current_action": "Starting...",
            "results": {},
            "started_at": datetime.now().isoformat()
        }
        
        # Run scan in background
        background_tasks.add_task(run_scan_all_background, job_id, project_id, structure)
        
        logger.info(f"[SCAN-ALL] Started background job {job_id} for {project_id} with {total_actions} actions")
        
        return {
            "job_id": job_id,
            "project_id": project_id,
            "total_actions": total_actions,
            "status": "started"
        }
        
    except Exception as e:
        logger.exception(f"Scan all failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def run_scan_all_background(job_id: str, project_id: str, structure: Dict):
    """Background task to scan all actions."""
    try:
        for step in structure.get('steps', []):
            for action in step.get('actions', []):
                action_id = action['action_id']
                action_title = action.get('title', action_id)[:50]
                
                # Update current action
                SCAN_JOBS[job_id]["current_action"] = f"Scanning {action_id}: {action_title}..."
                
                try:
                    result = await scan_for_action(project_id, action_id)
                    SCAN_JOBS[job_id]["results"][action_id] = {
                        "success": True,
                        "status": result.get('suggested_status', 'not_started')
                    }
                    SCAN_JOBS[job_id]["successful"] += 1
                except Exception as e:
                    logger.warning(f"[SCAN-ALL] Action {action_id} failed: {e}")
                    SCAN_JOBS[job_id]["results"][action_id] = {
                        "success": False,
                        "error": str(e)
                    }
                    SCAN_JOBS[job_id]["failed"] += 1
                
                SCAN_JOBS[job_id]["completed"] += 1
                
                # Small delay to prevent overwhelming the system
                await asyncio.sleep(0.1)
        
        # Mark complete
        SCAN_JOBS[job_id]["status"] = "completed"
        SCAN_JOBS[job_id]["current_action"] = "Complete"
        SCAN_JOBS[job_id]["finished_at"] = datetime.now().isoformat()
        
        logger.info(f"[SCAN-ALL] Job {job_id} completed: {SCAN_JOBS[job_id]['successful']} successful, {SCAN_JOBS[job_id]['failed']} failed")
        
    except Exception as e:
        logger.exception(f"[SCAN-ALL] Background job {job_id} failed: {e}")
        SCAN_JOBS[job_id]["status"] = "failed"
        SCAN_JOBS[job_id]["error"] = str(e)


@router.get("/year-end/scan-all/status/{job_id}")
async def get_scan_all_status(job_id: str):
    """Get status of a scan-all background job."""
    if job_id not in SCAN_JOBS:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")
    
    return SCAN_JOBS[job_id]


# =============================================================================
# HEALTH CHECK
# =============================================================================

@router.get("/health")
async def playbooks_health():
    """Health check for playbooks router."""
    return {
        "status": "healthy",
        "framework_available": FRAMEWORK_AVAILABLE,
        "year_end_config_loaded": bool(YEAR_END_CONFIG),
        "learning_available": FRAMEWORK_AVAILABLE and LearningHook.is_available()
    }


# =============================================================================
# LEARNING STATS ENDPOINT
# =============================================================================

@router.get("/learning/stats")
async def get_learning_stats():
    """Get learning system statistics."""
    try:
        try:
            from backend.utils.learning_engine import get_learning_system
        except ImportError:
            from utils.learning_engine import get_learning_system
        
        learning = get_learning_system()
        stats = learning.get_stats()
        
        return {
            "available": True,
            "stats": stats,
            "message": "Learning system active"
        }
        
    except ImportError:
        return {
            "available": False,
            "stats": None,
            "message": "Learning engine not available"
        }
    except Exception as e:
        logger.error(f"[LEARNING] Stats error: {e}")
        return {
            "available": False,
            "stats": None,
            "message": str(e)
        }


@router.get("/learning/patterns/{playbook_id}")
async def get_learning_patterns(playbook_id: str, project_id: str = "global"):
    """Get learned suppression patterns for a playbook."""
    try:
        try:
            from backend.utils.learning_engine import get_learning_system
        except ImportError:
            from utils.learning_engine import get_learning_system
        
        learning = get_learning_system()
        patterns = learning.get_patterns(project_id, playbook_id)
        
        return {
            "playbook_id": playbook_id,
            "project_id": project_id,
            "patterns": patterns
        }
        
    except Exception as e:
        logger.error(f"[LEARNING] Patterns error: {e}")
        return {
            "playbook_id": playbook_id,
            "project_id": project_id,
            "patterns": {"suppressions": [], "preferences": {}},
            "error": str(e)
        }


# =============================================================================
# ENTITY DETECTION ENDPOINT
# =============================================================================

async def get_project_documents_text(project_id: str) -> List[str]:
    """Get all document text for a project (for entity detection)."""
    texts = []
    project_name = None
    
    # Get project NAME from Supabase
    try:
        supabase = get_supabase()
        if supabase:
            result = supabase.table('customers').select('name').eq('id', project_id).execute()
            if result.data and len(result.data) > 0:
                project_name = result.data[0].get('name')
    except Exception as e:
        logger.debug(f"[ENTITIES] Could not get project name: {e}")
    
    # 1. Try DuckDB (structured data)
    if project_name:
        try:
            from backend.utils.playbook_parser import get_duckdb_connection
            conn = get_duckdb_connection()
            if conn:
                try:
                    tables = conn.execute("""
                        SELECT table_name, file_name 
                        FROM _schema_metadata 
                        WHERE LOWER(project) = LOWER(?) AND is_current = TRUE
                    """, [project_name]).fetchall()
                    
                    for table_name, file_name in tables:
                        try:
                            df = conn.execute(f'SELECT * FROM "{table_name}" LIMIT 500').fetchdf()
                            if df is not None and not df.empty:
                                texts.append(f"[Source: {file_name}]\n{df.to_string()}")
                        except Exception:
                            pass
                    
                    # Also check _pdf_tables
                    try:
                        pdf_tables = conn.execute("""
                            SELECT table_name, source_file 
                            FROM _pdf_tables 
                            WHERE project = ? OR project_id = ?
                        """, [project_name, project_id]).fetchall()
                        
                        for table_name, source_file in pdf_tables:
                            try:
                                df = conn.execute(f'SELECT * FROM "{table_name}" LIMIT 500').fetchdf()
                                if df is not None and not df.empty:
                                    texts.append(f"[Source: {source_file}]\n{df.to_string()}")
                            except Exception:
                                pass
                    except Exception:
                        pass
                    
                    conn.close()
                except Exception as e:
                    logger.debug(f"[ENTITIES] DuckDB query error: {e}")
                    try:
                        conn.close()
                    except Exception as e:
                        logger.debug(f"Suppressed: {e}")
        except Exception:
            pass
    
    # 2. Try ChromaDB (unstructured data)
    try:
        from utils.rag_handler import RAGHandler
        rag = RAGHandler()
        results = rag.search(
            collection_name="documents",
            query="company FEIN EIN employer tax federal identification",
            n_results=50,
            project_id=project_id
        )
        if results:
            for result in results:
                doc_text = result.get('document', '')
                if doc_text:
                    texts.append(doc_text)
    except Exception:
        pass
    
    logger.info(f"[ENTITIES] Retrieved {len(texts)} text chunks for project {project_id[:8]}")
    return texts


@router.post("/{playbook_type}/detect-entities/{project_id}")
async def detect_entities(playbook_type: str, project_id: str):
    """Scan project documents for US FEINs and Canada BNs using LLM."""
    logger.info(f"[ENTITIES] Starting entity detection for project {project_id}")
    
    try:
        docs = await get_project_documents_text(project_id)
        
        if not docs:
            return {
                "success": True,
                "entities": {"us": [], "canada": []},
                "summary": {"us_count": 0, "canada_count": 0, "total": 0},
                "warnings": ["No documents found for this project"]
            }
        
        combined_text = "\n\n---\n\n".join(docs)
        if len(combined_text) > 50000:
            combined_text = combined_text[:50000]
        
        logger.info(f"[ENTITIES] Analyzing {len(combined_text)} chars")
        
        prompt = f"""Analyze these documents and extract ALL Federal Employer Identification Numbers (FEINs/EINs).

FEINs are 9-digit numbers, usually formatted as XX-XXXXXXX (like 74-1776312).
They may appear:
- In headers/footers
- In company profile sections
- Embedded in codes (like 036741776312F01 contains 74-1776312)
- Near text like "EIN", "FEIN", "Employer Identification Number", "Tax ID"

Also look for Canada Business Numbers (9 digits + RT/RC/RP/RZ/RR + 4 digits).

Return ONLY a JSON object in this exact format:
{{
  "us": [
    {{"fein": "74-1776312", "company_name": "Company Name if found", "confidence": "high"}}
  ],
  "canada": [
    {{"bn": "123456789 RT 0001", "company_name": "Company Name if found", "confidence": "high"}}
  ]
}}

If no FEINs/BNs found, return: {{"us": [], "canada": []}}

DOCUMENTS:
{combined_text}"""

        system_prompt = "You are a precise data extraction assistant. Extract tax identification numbers and return valid JSON."
        
        # Use orchestrator if available
        global _playbooks_orchestrator
        response_text = None
        
        if _playbooks_orchestrator:
            # Use generate_json for structured output
            result = _playbooks_orchestrator.generate_json(
                prompt=f"{system_prompt}\n\n{prompt}"
            )
            
            if result.get('success') and isinstance(result.get('response'), dict):
                # Already parsed - jump to entity processing
                entities = result['response']
                logger.info(f"[PLAYBOOKS] Entity detection via {result.get('model_used', 'unknown')}")
                
                us_entities = entities.get("us", [])
                ca_entities = entities.get("canada", [])
                
                for i, entity in enumerate(us_entities):
                    entity['id'] = entity.get('fein', f'US-{i}')
                    entity['type'] = 'fein'
                    entity['count'] = 1
                
                for i, entity in enumerate(ca_entities):
                    entity['id'] = entity.get('bn', f'CA-{i}')
                    entity['type'] = 'bn'
                    entity['count'] = 1
                
                warnings = []
                if ca_entities and playbook_type == 'year-end':
                    warnings.append("Canada entities detected - requires Canada Year-End Playbook")
                
                primary = None
                for e in us_entities:
                    if e.get('confidence') == 'high':
                        primary = e['id']
                        break
                
                return {
                    "us_entities": us_entities,
                    "canada_entities": ca_entities,
                    "primary_entity": primary,
                    "warnings": warnings
                }
            elif result.get('response') and isinstance(result['response'], str):
                response_text = result['response'].strip()
        
        # Fallback to parsing text response
        if not response_text:
            logger.warning("[PLAYBOOKS] No LLM response for entity detection, using empty defaults")
            return {
                "us_entities": [],
                "canada_entities": [],
                "primary_entity": None,
                "warnings": ["Entity detection unavailable"]
            }
        
        try:
            if "```json" in response_text:
                response_text = response_text.split("```json")[1].split("```")[0].strip()
            elif "```" in response_text:
                response_text = response_text.split("```")[1].split("```")[0].strip()
            
            entities = json.loads(response_text)
        except json.JSONDecodeError:
            entities = {"us": [], "canada": []}
        
        us_entities = entities.get("us", [])
        ca_entities = entities.get("canada", [])
        
        for i, entity in enumerate(us_entities):
            entity['id'] = entity.get('fein', f'US-{i}')
            entity['type'] = 'fein'
            entity['count'] = 1
        
        for i, entity in enumerate(ca_entities):
            entity['id'] = entity.get('bn', f'CA-{i}')
            entity['type'] = 'bn'
            entity['count'] = 1
        
        warnings = []
        if ca_entities and playbook_type == 'year-end':
            warnings.append("Canada entities detected - requires Canada Year-End Playbook")
        
        primary = None
        for e in us_entities:
            if e.get('confidence') == 'high':
                primary = e['id']
                break
        if not primary and us_entities:
            primary = us_entities[0]['id']
        
        logger.info(f"[ENTITIES] Found {len(us_entities)} US, {len(ca_entities)} Canada entities")
        
        return {
            "success": True,
            "project_id": project_id,
            "playbook_type": playbook_type,
            "entities": {"us": us_entities, "canada": ca_entities},
            "summary": {
                "us_count": len(us_entities),
                "canada_count": len(ca_entities),
                "total": len(us_entities) + len(ca_entities),
                "suggested_primary": primary
            },
            "warnings": warnings
        }
        
    except Exception as e:
        logger.error(f"[ENTITIES] Detection error: {e}")
        return {"success": False, "message": str(e)}
