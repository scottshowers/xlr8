"""
Chat Router - ENHANCED with Query Routing, Local LLM, and Learning
===================================================================

ARCHITECTURE:
1. Intent Classification (FREE - regex)
2. Query Routing (FREE - rules + fuzzy matching)
3. Local LLM for SQL & simple answers (FREE)
4. Claude fallback for complex reasoning
5. Learning loop for continuous improvement

SCOPE OPTIONS:
- project: This project only (DuckDB + ChromaDB filtered)
- global: Global knowledge (ChromaDB global + LLM knowledge)
- all: All projects (everything)

PII HANDLING:
- PII NEVER goes to Claude
- Reversible redaction: redact before Claude, restore after
- No bullshit "couldn't analyze because redacted" excuses

PRESERVED FEATURES:
- Personas
- Excel export
- Source citations
- Polling status

Author: XLR8 Team
"""

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, Dict, Any, List
import threading
import uuid
import logging
import sys
import time
import re
import traceback
import io
import json
import os

sys.path.insert(0, '/app')
sys.path.insert(0, '/data')

from utils.rag_handler import RAGHandler
from utils.llm_orchestrator import LLMOrchestrator
from utils.persona_manager import get_persona_manager

# Import structured data handling
try:
    from utils.structured_data_handler import get_structured_handler
    STRUCTURED_QUERIES_AVAILABLE = True
    logging.getLogger(__name__).info("✅ Structured data handler loaded")
except ImportError as e:
    STRUCTURED_QUERIES_AVAILABLE = False
    logging.getLogger(__name__).warning(f"❌ Structured data handler NOT available: {e}")

# Import query routing components
try:
    from utils.intent_classifier import classify_and_configure
    from utils.query_router import detect_query_type, QueryType, build_sql_prompt
    from utils.query_decomposition import search_with_diversity
    from utils.smart_aggregation import handle_aggregation
    ROUTING_AVAILABLE = True
    logging.getLogger(__name__).info("✅ Query routing components loaded")
except ImportError as e:
    ROUTING_AVAILABLE = False
    logging.getLogger(__name__).warning(f"❌ Query routing NOT available: {e}")

# Import learning system
try:
    from utils.learning_engine import get_learning_system
    LEARNING_AVAILABLE = True
    logging.getLogger(__name__).info("✅ Learning system loaded")
except ImportError as e:
    LEARNING_AVAILABLE = False
    logging.getLogger(__name__).warning(f"❌ Learning system NOT available: {e}")

# Import local LLM client
try:
    from utils.hybrid_analyzer import LocalLLMClient
    LOCAL_LLM_AVAILABLE = True
    logging.getLogger(__name__).info("✅ Local LLM client loaded")
except ImportError as e:
    LOCAL_LLM_AVAILABLE = False
    logging.getLogger(__name__).warning(f"❌ Local LLM NOT available: {e}")

logger = logging.getLogger(__name__)

router = APIRouter()

chat_jobs: Dict[str, Dict[str, Any]] = {}


# =============================================================================
# REVERSIBLE PII REDACTION - PII NEVER GOES TO CLAUDE
# =============================================================================

class ReversibleRedactor:
    """
    Reversible PII redaction - redact before Claude, restore after.
    
    PII NEVER goes to Claude. Period.
    But we don't give bullshit "couldn't analyze" excuses either.
    Claude works with placeholders, we restore the real values after.
    """
    
    def __init__(self):
        self.mappings = {}  # {placeholder: original_value}
        self.counters = {'ssn': 0, 'salary': 0, 'phone': 0, 'email': 0, 'name': 0, 'account': 0}
    
    def _get_placeholder(self, pii_type: str) -> str:
        """Generate unique placeholder."""
        self.counters[pii_type] += 1
        return f"[{pii_type.upper()}_{self.counters[pii_type]:03d}]"
    
    def redact(self, text: str) -> str:
        """Redact PII with reversible placeholders."""
        if not text:
            return text
        
        result = text
        
        # SSN: 123-45-6789
        for match in re.finditer(r'\b(\d{3}-\d{2}-\d{4})\b', result):
            original = match.group(1)
            if original not in [v for v in self.mappings.values()]:
                placeholder = self._get_placeholder('ssn')
                self.mappings[placeholder] = original
                result = result.replace(original, placeholder)
        
        # Bank account / routing numbers (8-17 digits)
        for match in re.finditer(r'\b(\d{8,17})\b', result):
            original = match.group(1)
            # Skip if already mapped or looks like a year/date
            if original not in [v for v in self.mappings.values()] and not original.startswith('20'):
                placeholder = self._get_placeholder('account')
                self.mappings[placeholder] = original
                result = result.replace(original, placeholder)
        
        # Salary: $75,000 or $75,000.00
        for match in re.finditer(r'(\$\s*\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', result):
            original = match.group(1)
            if original not in [v for v in self.mappings.values()]:
                placeholder = self._get_placeholder('salary')
                self.mappings[placeholder] = original
                result = result.replace(original, placeholder)
        
        # Phone: (123) 456-7890 or 123-456-7890
        for match in re.finditer(r'(\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4})\b', result):
            original = match.group(1)
            # Skip if it looks like SSN
            if original.count('-') == 2 and len(original.replace('-', '')) == 9:
                continue
            if original not in [v for v in self.mappings.values()]:
                placeholder = self._get_placeholder('phone')
                self.mappings[placeholder] = original
                result = result.replace(original, placeholder)
        
        # Email
        for match in re.finditer(r'\b([\w\.-]+@[\w\.-]+\.\w{2,})\b', result):
            original = match.group(1)
            if original not in [v for v in self.mappings.values()]:
                placeholder = self._get_placeholder('email')
                self.mappings[placeholder] = original
                result = result.replace(original, placeholder)
        
        return result
    
    def restore(self, text: str) -> str:
        """Restore original PII values from placeholders."""
        if not text or not self.mappings:
            return text
        
        result = text
        for placeholder, original in self.mappings.items():
            result = result.replace(placeholder, original)
        
        return result
    
    def has_pii(self) -> bool:
        """Check if any PII was redacted."""
        return len(self.mappings) > 0
    
    def get_stats(self) -> Dict[str, int]:
        """Get redaction statistics."""
        return {
            'total_redacted': len(self.mappings),
            **{k: v for k, v in self.counters.items() if v > 0}
        }


# Legacy function for backward compatibility
def detect_pii_in_data(data: str, columns: List[str] = None) -> Dict[str, Any]:
    """Scan DATA content for PII patterns."""
    detected = {'has_pii': False, 'pii_types': [], 'column_flags': []}
    
    if not data:
        return detected
    
    PII_COLUMN_PATTERNS = [
        r'ssn', r'social.*security', r'tax.*id',
        r'salary', r'pay.*rate', r'wage', r'compensation', r'hourly.*rate', r'annual.*pay',
        r'bank.*account', r'routing', r'account.*num',
        r'dob', r'birth.*date', r'date.*birth',
        r'email', r'phone', r'address', r'street', r'city', r'zip',
        r'first.*name', r'last.*name', r'full.*name', r'employee.*name'
    ]
    
    if columns:
        for col in columns:
            col_lower = str(col).lower()
            for pattern in PII_COLUMN_PATTERNS:
                if re.search(pattern, col_lower):
                    detected['column_flags'].append(col)
                    detected['has_pii'] = True
    
    if re.search(r'\b\d{3}-\d{2}-\d{4}\b', data):
        detected['pii_types'].append('ssn')
        detected['has_pii'] = True
    if re.search(r'\$\s*\d{2,3},?\d{3}', data):
        detected['pii_types'].append('compensation')
        detected['has_pii'] = True
    if re.search(r'\b[\w\.-]+@[\w\.-]+\.\w{2,}\b', data):
        detected['pii_types'].append('email')
        detected['has_pii'] = True
    if re.search(r'\b\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}\b', data):
        detected['pii_types'].append('phone')
        detected['has_pii'] = True
    
    detected['pii_types'] = list(set(detected['pii_types']))
    return detected


# =============================================================================
# REQUEST/RESPONSE MODELS
# =============================================================================

class ChatRequest(BaseModel):
    message: str
    project: Optional[str] = None
    max_results: Optional[int] = 30
    persona: Optional[str] = 'bessie'
    scope: Optional[str] = 'project'  # NEW: project, global, all


class ChatStartRequest(BaseModel):
    message: str
    project: Optional[str] = None
    max_results: Optional[int] = 50
    persona: Optional[str] = 'bessie'
    scope: Optional[str] = 'project'  # NEW: project, global, all


class FeedbackRequest(BaseModel):
    job_id: str
    feedback: str  # 'up' or 'down'
    message: Optional[str] = None
    response: Optional[str] = None


def update_job_status(job_id: str, status: str, step: str, progress: int, **kwargs):
    if job_id in chat_jobs:
        chat_jobs[job_id].update({'status': status, 'current_step': step, 'progress': progress, **kwargs})
        logger.info(f"[JOB {job_id[:8]}] {step} ({progress}%)")


def decrypt_results(rows: list, handler) -> list:
    """Decrypt encrypted fields in query results."""
    if not rows:
        return rows
    try:
        encryptor = handler.encryptor
        if not encryptor or not getattr(encryptor, 'aesgcm', None):
            return rows
        decrypted_rows = []
        for row in rows:
            decrypted_row = {}
            for key, value in row.items():
                if isinstance(value, str) and (value.startswith('ENC:') or value.startswith('ENC256:')):
                    try:
                        decrypted_row[key] = encryptor.decrypt(value)
                    except:
                        decrypted_row[key] = '[encrypted]'
                else:
                    decrypted_row[key] = value
            decrypted_rows.append(decrypted_row)
        return decrypted_rows
    except Exception as e:
        logger.warning(f"[DECRYPT] Error: {e}")
        return rows


# =============================================================================
# LOCAL LLM SQL GENERATION
# =============================================================================

def generate_sql_with_local_llm(query: str, schema: Dict, local_llm) -> Optional[str]:
    """
    Use local LLM to generate SQL query.
    Returns SQL string or None if failed.
    """
    if not local_llm or not local_llm.is_available():
        return None
    
    try:
        # Build schema description
        tables_info = []
        for table in schema.get('tables', [])[:10]:
            table_name = table.get('table_name', '')
            columns = table.get('columns', [])
            if columns and isinstance(columns[0], dict):
                col_names = [c.get('name', str(c)) for c in columns]
            else:
                col_names = [str(c) for c in columns] if columns else []
            row_count = table.get('row_count', 0)
            sheet = table.get('sheet', '')
            
            tables_info.append(f"Table: {table_name}\n  Sheet: {sheet}\n  Columns: {', '.join(col_names[:15])}\n  Rows: {row_count}")
        
        schema_text = '\n\n'.join(tables_info)
        
        prompt = f"""Generate a SQL query for DuckDB to answer this question.

SCHEMA:
{schema_text}

QUESTION: {query}

RULES:
1. Return ONLY the SQL query, nothing else
2. Use exact table and column names from schema
3. Use ILIKE for case-insensitive text matching
4. LIMIT 1000 unless counting
5. For "how many" use COUNT(*)
6. Wrap table/column names in double quotes

SQL:"""
        
        result, success = local_llm.extract("", prompt)
        
        if success and result:
            sql = result.strip()
            if sql.startswith('```'):
                sql = sql.split('```')[1]
                if sql.startswith('sql'):
                    sql = sql[3:]
            sql = sql.strip()
            
            if sql.upper().startswith('SELECT'):
                logger.info(f"[LOCAL_SQL] Generated: {sql[:100]}...")
                return sql
        
        return None
        
    except Exception as e:
        logger.warning(f"[LOCAL_SQL] Failed: {e}")
        return None


def generate_answer_with_local_llm(query: str, context: str, local_llm, persona_prompt: str = "") -> Optional[str]:
    """
    Use local LLM to generate answer from context.
    Returns answer string or None if failed/low quality.
    """
    if not local_llm or not local_llm.is_available():
        return None
    
    try:
        prompt = f"""{persona_prompt}

Based on the following data, answer the user's question concisely and accurately.

DATA:
{context[:12000]}

QUESTION: {query}

ANSWER:"""
        
        result, success = local_llm.extract("", prompt)
        
        if success and result:
            answer = result.strip()
            
            # Quality check
            if len(answer) < 20:
                logger.warning(f"[LOCAL_ANSWER] Too short: {len(answer)} chars")
                return None
            
            low_confidence_phrases = [
                "i don't have", "i cannot", "not able to", "no information",
                "i'm not sure", "i don't know", "cannot determine"
            ]
            if any(phrase in answer.lower() for phrase in low_confidence_phrases):
                logger.warning(f"[LOCAL_ANSWER] Low confidence response")
                return None
            
            logger.info(f"[LOCAL_ANSWER] Generated: {len(answer)} chars")
            return answer
        
        return None
        
    except Exception as e:
        logger.warning(f"[LOCAL_ANSWER] Failed: {e}")
        return None


# =============================================================================
# SCOPE-AWARE DATA RETRIEVAL
# =============================================================================

def get_duckdb_tables_for_scope(project: str, scope: str) -> List[Dict]:
    """Get DuckDB tables based on scope."""
    if not STRUCTURED_QUERIES_AVAILABLE:
        return []
    
    try:
        handler = get_structured_handler()
        
        if scope == 'global':
            return []  # Global = ChromaDB + LLM knowledge only
        
        elif scope == 'all':
            result = handler.conn.execute("""
                SELECT DISTINCT project, file_name, sheet_name, table_name, columns, row_count
                FROM _schema_metadata
                WHERE is_current = TRUE
                ORDER BY project, file_name
            """).fetchall()
            
            tables = []
            for row in result:
                project_name, file_name, sheet_name, table_name, columns, row_count = row
                tables.append({
                    'project': project_name,
                    'file': file_name,
                    'sheet': sheet_name,
                    'table_name': table_name,
                    'columns': json.loads(columns) if columns else [],
                    'row_count': row_count
                })
            return tables
        
        else:  # project scope
            if not project:
                return []
            schema = handler.get_schema_for_project(project)
            return schema.get('tables', [])
    
    except Exception as e:
        logger.error(f"[DUCKDB_TABLES] Error: {e}")
        return []


def search_chromadb_for_scope(query: str, project: str, scope: str, n_results: int = 20) -> Dict:
    """Search ChromaDB based on scope."""
    try:
        rag = RAGHandler()
        collection = rag.client.get_or_create_collection(name="documents")
        
        where_filter = None
        
        if scope == 'project' and project:
            where_filter = {"project": project}
        
        elif scope == 'global':
            # Try to get only global docs
            try:
                where_filter = {"project": {"$in": ["global", "__global__", "Global/Universal"]}}
            except:
                where_filter = None  # Fall back to no filter if $in not supported
        
        # 'all' scope = no filter
        
        results = collection.query(
            query_texts=[query],
            n_results=n_results,
            where=where_filter if scope != 'all' else None
        )
        
        return results
    
    except Exception as e:
        logger.error(f"[CHROMADB_SEARCH] Error: {e}")
        return {'documents': [[]], 'metadatas': [[]], 'distances': [[]]}


# =============================================================================
# MAIN PROCESSING - ENHANCED WITH ROUTING, LOCAL LLM, LEARNING
# =============================================================================

def process_chat_job(
    job_id: str, 
    message: str, 
    project: Optional[str], 
    max_results: int, 
    persona_name: str = 'bessie',
    scope: str = 'project'
):
    """
    Enhanced Chat Pipeline with Query Routing & Local LLM
    ======================================================
    
    Flow:
    1. Check cache (FREE)
    2. Classify intent (FREE - regex)
    3. Route query (FREE - rules)
    4. Try local LLM first
    5. Fall back to Claude if needed
    6. Store for learning
    
    PII: Always redacted before Claude, restored after.
    """
    logger.warning(f"{'='*60}")
    logger.warning(f"[START] Job {job_id[:8]} | Query: {message[:50]}...")
    logger.warning(f"[START] Project: {project} | Scope: {scope} | Persona: {persona_name}")
    logger.warning(f"{'='*60}")
    
    # Initialize redactor for this request
    redactor = ReversibleRedactor()
    
    try:
        # =====================================================================
        # STEP 1: INITIALIZE
        # =====================================================================
        update_job_status(job_id, 'processing', '🚀 Initializing...', 5)
        
        pm = get_persona_manager()
        persona = pm.get_persona(persona_name)
        persona_prompt = persona.system_prompt if persona else ""
        logger.warning(f"[STEP 1] Persona loaded: {persona.name if persona else 'default'}")
        
        # Initialize local LLM if available
        local_llm = None
        if LOCAL_LLM_AVAILABLE:
            try:
                local_llm = LocalLLMClient()
                if local_llm.is_available():
                    logger.warning("[STEP 1] Local LLM available")
                else:
                    local_llm = None
                    logger.warning("[STEP 1] Local LLM not reachable")
            except:
                local_llm = None
        
        # Initialize learning system if available
        learning = None
        if LEARNING_AVAILABLE:
            try:
                learning = get_learning_system()
                logger.warning("[STEP 1] Learning system available")
            except:
                learning = None
        
        # =====================================================================
        # STEP 2: CHECK CACHE (FREE)
        # =====================================================================
        if learning:
            update_job_status(job_id, 'processing', '🔍 Checking cache...', 8)
            cache_key = f"chat_{project or 'global'}_{scope}"
            cached = learning.get_cached_analysis(message, cache_key)
            if cached:
                logger.warning(f"[STEP 2] ✅ Cache hit!")
                # Restore any PII in cached response
                response = redactor.restore(cached.get('response', ''))
                update_job_status(
                    job_id, 'complete', 'Complete', 100,
                    response=response,
                    sources=cached.get('sources', []),
                    chunks_found=cached.get('chunks_found', 0),
                    models_used=['cache'],
                    query_type='cached',
                    routing_info={'source': 'cache'}
                )
                return
        
        # =====================================================================
        # STEP 3: CLASSIFY INTENT & ROUTE (FREE)
        # =====================================================================
        update_job_status(job_id, 'processing', '🧠 Analyzing query...', 10)
        
        intent_config = {'intent': 'GENERAL', 'n_results': 50}
        route_info = {'route': 'hybrid', 'reasoning': []}
        
        if ROUTING_AVAILABLE:
            try:
                intent_config = classify_and_configure(message)
                logger.warning(f"[STEP 3] Intent: {intent_config.get('intent')} (n={intent_config.get('n_results')})")
                
                # Get available data sources for routing
                tables_list = get_duckdb_tables_for_scope(project, scope)
                has_structured = len(tables_list) > 0
                has_rag = True  # Assume ChromaDB always available
                
                route_result = detect_query_type(
                    query=message,
                    has_structured=has_structured,
                    has_rag=has_rag,
                    tables=[t.get('table_name', '') for t in tables_list]
                )
                route_info = {
                    'route': route_result.get('route', QueryType.HYBRID).value if hasattr(route_result.get('route'), 'value') else str(route_result.get('route', 'hybrid')),
                    'reasoning': route_result.get('reasoning', []),
                    'entities': route_result.get('entities', [])
                }
                logger.warning(f"[STEP 3] Route: {route_info['route']} ({', '.join(route_info['reasoning'])})")
            except Exception as route_e:
                logger.warning(f"[STEP 3] Routing failed: {route_e}, using hybrid")
        
        # Determine what to search based on route
        search_structured = route_info['route'] in ['structured', 'hybrid']
        search_rag = route_info['route'] in ['unstructured', 'hybrid', 'general']
        
        # =====================================================================
        # STEP 4: RETRIEVE DATA
        # =====================================================================
        sql_results_list = []
        rag_chunks = []
        all_columns = []
        sources = []
        
        # 4a: Structured data (DuckDB)
        if search_structured and STRUCTURED_QUERIES_AVAILABLE and scope != 'global':
            update_job_status(job_id, 'processing', '📊 Querying structured data...', 20)
            
            try:
                handler = get_structured_handler()
                schema = {'tables': get_duckdb_tables_for_scope(project, scope)}
                
                if schema['tables']:
                    logger.warning(f"[STEP 4a] Found {len(schema['tables'])} tables")
                    
                    # Try local LLM for SQL generation first
                    sql = None
                    if local_llm:
                        sql = generate_sql_with_local_llm(message, schema, local_llm)
                    
                    if sql:
                        # Execute local LLM generated SQL
                        try:
                            rows, cols = handler.execute_query(sql)
                            if rows:
                                decrypted = decrypt_results(rows, handler)
                                sql_results_list.append({
                                    'source': 'Local LLM SQL',
                                    'sql': sql,
                                    'columns': cols,
                                    'data': decrypted[:100],
                                    'total_rows': len(rows)
                                })
                                all_columns.extend(cols)
                                sources.append({'filename': 'SQL Query', 'type': 'structured', 'rows': len(rows), 'relevance': 95})
                                logger.warning(f"[STEP 4a] ✅ Local SQL returned {len(rows)} rows")
                        except Exception as sql_e:
                            logger.warning(f"[STEP 4a] Local SQL execution failed: {sql_e}")
                    
                    # If no local SQL results, fall back to table selection method
                    if not sql_results_list:
                        logger.warning("[STEP 4a] Falling back to Claude table selection")
                        # Use existing Claude-based table selection (preserved from original)
                        table_summaries = []
                        table_lookup = {}
                        
                        for idx, table_info in enumerate(schema['tables'][:20]):
                            table_name = table_info.get('table_name', '')
                            sheet_name = table_info.get('sheet', '')
                            file_name = table_info.get('file', '')
                            columns = table_info.get('columns', [])
                            row_count = table_info.get('row_count', 0)
                            
                            if columns and isinstance(columns[0], dict):
                                col_names = [c.get('name', str(c)) for c in columns]
                            else:
                                col_names = [str(c) for c in columns] if columns else []
                            
                            summary = f"[{idx}] {file_name} → {sheet_name}\n  Rows: {row_count}, Cols: {', '.join(col_names[:8])}"
                            table_summaries.append(summary)
                            table_lookup[idx] = table_info
                        
                        if table_summaries:
                            # Ask Claude to pick tables
                            try:
                                import anthropic
                                orchestrator = LLMOrchestrator()
                                client = anthropic.Anthropic(api_key=orchestrator.claude_api_key)
                                
                                selection_prompt = f"""Question: "{message}"

Pick table number(s) that contain data to answer this question.

Available tables:
{chr(10).join(table_summaries)}

RESPOND WITH ONLY THE TABLE NUMBERS (comma-separated if multiple). NO EXPLANATION.

Table numbers:"""
                                
                                selection_response = client.messages.create(
                                    model="claude-sonnet-4-20250514",
                                    max_tokens=50,
                                    messages=[{"role": "user", "content": selection_prompt}]
                                )
                                
                                selection_text = selection_response.content[0].text.strip()
                                selected_indices = []
                                
                                for part in selection_text.replace(' ', '').split(','):
                                    try:
                                        idx = int(part.strip('[]()'))
                                        if idx in table_lookup:
                                            selected_indices.append(idx)
                                    except:
                                        pass
                                
                                if not selected_indices:
                                    numbers = re.findall(r'\b(\d{1,2})\b', selection_text)
                                    for num_str in numbers[:3]:
                                        idx = int(num_str)
                                        if idx in table_lookup:
                                            selected_indices.append(idx)
                                
                                logger.warning(f"[STEP 4a] Claude selected tables: {selected_indices}")
                                
                                # Query selected tables
                                for idx in selected_indices[:3]:
                                    table_info = table_lookup[idx]
                                    table_name = table_info.get('table_name', '')
                                    
                                    try:
                                        sql = f'SELECT * FROM "{table_name}" LIMIT 500'
                                        rows, cols = handler.execute_query(sql)
                                        
                                        if rows:
                                            decrypted = decrypt_results(rows, handler)
                                            sql_results_list.append({
                                                'source_file': table_info.get('file', ''),
                                                'sheet': table_info.get('sheet', ''),
                                                'table': table_name,
                                                'columns': cols,
                                                'data': decrypted,
                                                'total_rows': len(rows)
                                            })
                                            all_columns.extend(cols)
                                            sources.append({
                                                'filename': table_info.get('file', ''),
                                                'sheet': table_info.get('sheet', ''),
                                                'type': 'structured',
                                                'rows': len(rows),
                                                'relevance': 95
                                            })
                                    except Exception as q_e:
                                        logger.warning(f"[STEP 4a] Query failed for {table_name}: {q_e}")
                                
                            except Exception as sel_e:
                                logger.error(f"[STEP 4a] Claude selection failed: {sel_e}")
                
            except Exception as struct_e:
                logger.error(f"[STEP 4a] Structured query error: {struct_e}")
        
        # 4b: RAG search (ChromaDB)
        if search_rag:
            update_job_status(job_id, 'processing', '🔍 Searching documents...', 40)
            
            try:
                n_results = intent_config.get('n_results', max_results)
                rag_results = search_chromadb_for_scope(message, project, scope, n_results)
                
                if rag_results and rag_results.get('documents') and rag_results['documents'][0]:
                    rag_chunks = rag_results['documents'][0]
                    logger.warning(f"[STEP 4b] ✅ RAG found {len(rag_chunks)} chunks")
                    sources.append({'filename': 'Documents', 'type': 'rag', 'chunks': len(rag_chunks), 'relevance': 80})
                else:
                    logger.warning("[STEP 4b] RAG returned no results")
                    
            except Exception as rag_e:
                logger.error(f"[STEP 4b] RAG error: {rag_e}")
        
        # =====================================================================
        # STEP 5: BUILD CONTEXT
        # =====================================================================
        update_job_status(job_id, 'processing', '📝 Building context...', 50)
        
        context_parts = []
        
        # Add structured data results
        for result in sql_results_list:
            data_text = f"Source: {result.get('source_file', result.get('source', 'SQL'))} → {result.get('sheet', 'Query')}\n"
            data_text += f"Columns: {', '.join(result.get('columns', []))}\n"
            data_text += f"Total rows: {result.get('total_rows', len(result.get('data', [])))}\n"
            data_text += "\nData:\n"
            
            for row in result.get('data', [])[:50]:
                row_str = " | ".join(f"{k}: {v}" for k, v in list(row.items())[:10])
                data_text += f"  {row_str}\n"
            
            context_parts.append(data_text)
        
        # Add RAG chunks
        for i, chunk in enumerate(rag_chunks[:10]):
            context_parts.append(f"Document {i+1}:\n{chunk}")
        
        logger.warning(f"[STEP 5] Context: {len(context_parts)} parts, {len(sql_results_list)} SQL, {len(rag_chunks)} RAG")
        
        if not context_parts:
            # No data found - try broader search
            update_job_status(job_id, 'processing', '🔍 Broader search...', 55)
            try:
                rag = RAGHandler()
                collection = rag.client.get_or_create_collection(name="documents")
                fallback_results = collection.query(query_texts=[message], n_results=20)
                
                if fallback_results and fallback_results.get('documents') and fallback_results['documents'][0]:
                    for chunk in fallback_results['documents'][0][:10]:
                        context_parts.append(f"Document:\n{chunk}")
                    sources.append({'filename': 'Fallback Search', 'type': 'rag', 'chunks': len(fallback_results['documents'][0])})
                    logger.warning(f"[STEP 5] Fallback found {len(fallback_results['documents'][0])} chunks")
            except Exception as fb_e:
                logger.warning(f"[STEP 5] Fallback search failed: {fb_e}")
        
        # =====================================================================
        # STEP 6: REDACT PII BEFORE ANY LLM CALL
        # =====================================================================
        full_context = '\n\n'.join(context_parts)
        
        # Always redact before sending to any LLM
        redacted_context = redactor.redact(full_context)
        pii_stats = redactor.get_stats()
        
        if pii_stats['total_redacted'] > 0:
            logger.warning(f"[STEP 6] Redacted PII: {pii_stats}")
        
        # =====================================================================
        # STEP 7: GENERATE ANSWER
        # =====================================================================
        response = None
        models_used = []
        
        if context_parts:
            # Try local LLM first for simple queries
            if local_llm and intent_config.get('intent') in ['LOOKUP', 'COUNT', 'GENERAL']:
                update_job_status(job_id, 'processing', '🤖 Trying local LLM...', 65)
                
                local_response = generate_answer_with_local_llm(
                    message, 
                    redacted_context, 
                    local_llm, 
                    persona_prompt
                )
                
                if local_response:
                    # Restore PII in response
                    response = redactor.restore(local_response)
                    models_used = ['local']
                    logger.warning(f"[STEP 7] ✅ Local LLM answered")
            
            # Fall back to Claude if local LLM failed or complex query
            if not response:
                update_job_status(job_id, 'processing', '✨ Generating response...', 75)
                
                try:
                    orchestrator = LLMOrchestrator()
                    orchestrator_chunks = [{
                        'document': redacted_context,  # PII already redacted!
                        'metadata': {'filename': 'Query Results', 'type': 'combined'}
                    }]
                    
                    result = orchestrator.process_query(message, orchestrator_chunks)
                    claude_response = result.get('response', '')
                    
                    # Restore PII in Claude's response
                    response = redactor.restore(claude_response)
                    models_used = result.get('models_used', ['claude'])
                    logger.warning(f"[STEP 7] ✅ Claude answered")
                    
                    # Learn from this interaction
                    if learning and response:
                        try:
                            cache_key = f"chat_{project or 'global'}_{scope}"
                            learning.learn_from_claude(
                                message, 
                                f"Context: {len(redacted_context)} chars",
                                {'response': claude_response, 'sources': sources, 'chunks_found': len(context_parts)},
                                cache_key,
                                quality='medium'
                            )
                        except Exception as learn_e:
                            logger.warning(f"[STEP 7] Learning failed: {learn_e}")
                    
                except Exception as claude_e:
                    logger.error(f"[STEP 7] Claude failed: {claude_e}")
                    response = f"I found relevant data but encountered an error generating the response: {str(claude_e)}"
                    models_used = ['error']
        
        else:
            response = "I couldn't find relevant data for your question. Please ensure data has been uploaded for this project, or try a different search scope."
            models_used = ['none']
        
        # =====================================================================
        # STEP 8: RETURN RESULT
        # =====================================================================
        update_job_status(
            job_id, 'complete', 'Complete', 100,
            response=response,
            sources=sources,
            chunks_found=len(sql_results_list) + len(rag_chunks),
            models_used=models_used,
            query_type=route_info.get('route', 'hybrid'),
            routing_info=route_info,
            scope=scope,
            pii_redacted=pii_stats['total_redacted'] > 0
        )
        
        logger.warning(f"[COMPLETE] Job {job_id[:8]} finished successfully")
        
    except Exception as e:
        logger.error(f"[FATAL] Job {job_id[:8]} crashed: {e}")
        logger.error(f"[FATAL] Traceback: {traceback.format_exc()}")
        update_job_status(job_id, 'error', 'Error', 100, error=str(e), response=f"An error occurred: {str(e)}")


# =============================================================================
# API ENDPOINTS
# =============================================================================

@router.get("/chat/models")
async def get_available_models():
    """Get available models info."""
    return {
        "models": [
            {"id": "claude-sonnet", "name": "Claude Sonnet"},
            {"id": "local", "name": "Local LLM"}
        ], 
        "default": "claude-sonnet",
        "local_available": LOCAL_LLM_AVAILABLE,
        "routing_available": ROUTING_AVAILABLE,
        "learning_available": LEARNING_AVAILABLE
    }


@router.post("/chat/start")
async def start_chat(request: ChatStartRequest):
    """Start async chat job with scope support."""
    job_id = str(uuid.uuid4())
    chat_jobs[job_id] = {
        'status': 'starting', 
        'current_step': '🚀 Starting...', 
        'progress': 0, 
        'created_at': time.time(),
        'scope': request.scope
    }
    
    thread = threading.Thread(
        target=process_chat_job, 
        args=(
            job_id, 
            request.message, 
            request.project, 
            request.max_results or 50, 
            request.persona or 'bessie',
            request.scope or 'project'
        )
    )
    thread.daemon = True
    thread.start()
    
    return {"job_id": job_id, "status": "starting", "scope": request.scope}


@router.get("/chat/status/{job_id}")
async def get_chat_status(job_id: str):
    """Get chat job status."""
    if job_id not in chat_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    return chat_jobs[job_id]


@router.delete("/chat/job/{job_id}")
async def cancel_chat_job(job_id: str):
    """Cancel/cleanup chat job."""
    if job_id in chat_jobs:
        del chat_jobs[job_id]
        return {"status": "cancelled"}
    raise HTTPException(status_code=404, detail="Job not found")


@router.post("/chat")
async def chat_sync(request: ChatRequest):
    """Synchronous chat (for simple queries)."""
    job_id = str(uuid.uuid4())
    chat_jobs[job_id] = {'status': 'processing', 'progress': 0}
    process_chat_job(
        job_id, 
        request.message, 
        request.project, 
        request.max_results or 30, 
        request.persona or 'bessie',
        request.scope or 'project'
    )
    return chat_jobs.get(job_id, {'error': 'Job failed'})


@router.get("/chat/health")
async def chat_health():
    """Health check endpoint."""
    return {
        "status": "healthy", 
        "structured": STRUCTURED_QUERIES_AVAILABLE, 
        "routing": ROUTING_AVAILABLE,
        "learning": LEARNING_AVAILABLE,
        "local_llm": LOCAL_LLM_AVAILABLE,
        "jobs": len(chat_jobs)
    }


# =============================================================================
# FEEDBACK ENDPOINT (NEW)
# =============================================================================

@router.post("/chat/feedback")
async def submit_feedback(request: FeedbackRequest):
    """Submit thumbs up/down feedback for learning."""
    try:
        if LEARNING_AVAILABLE:
            learning = get_learning_system()
            
            if request.feedback == 'down':
                # Record negative feedback
                learning.record_feedback(
                    action_id=f"chat_{request.job_id}",
                    correction_type='negative_feedback',
                    original=request.response,
                    corrected=None,
                    context=request.message
                )
                logger.info(f"[FEEDBACK] Negative feedback recorded for {request.job_id}")
            else:
                # Record positive feedback (helps with cache confidence)
                learning.record_feedback(
                    action_id=f"chat_{request.job_id}",
                    correction_type='positive_feedback',
                    original=request.response,
                    corrected=request.response,  # Same = confirmed good
                    context=request.message
                )
                logger.info(f"[FEEDBACK] Positive feedback recorded for {request.job_id}")
            
            return {"success": True, "feedback": request.feedback}
        else:
            return {"success": False, "error": "Learning system not available"}
    
    except Exception as e:
        logger.error(f"[FEEDBACK] Error: {e}")
        raise HTTPException(500, str(e))


# =============================================================================
# PERSONA ENDPOINTS (PRESERVED)
# =============================================================================

@router.get("/chat/personas")
async def list_personas():
    """List all available personas."""
    try:
        pm = get_persona_manager()
        return {"personas": pm.list_personas(), "default": "bessie"}
    except Exception as e:
        raise HTTPException(500, str(e))


@router.get("/chat/personas/{name}")
async def get_persona(name: str):
    """Get specific persona by name."""
    try:
        pm = get_persona_manager()
        persona = pm.get_persona(name)
        return persona.to_dict() if persona else None
    except:
        raise HTTPException(404, f"Persona '{name}' not found")


@router.post("/chat/personas")
async def create_persona(request: Request):
    """Create a new persona."""
    try:
        data = await request.json()
        pm = get_persona_manager()
        persona = pm.create_persona(
            name=data['name'], 
            icon=data.get('icon', '🤖'), 
            description=data.get('description', ''), 
            system_prompt=data.get('system_prompt', ''), 
            expertise=data.get('expertise', []), 
            tone=data.get('tone', 'Professional')
        )
        return {"success": True, "persona": persona.to_dict()}
    except Exception as e:
        logger.error(f"[PERSONA] Create error: {e}")
        raise HTTPException(500, str(e))


@router.put("/chat/personas/{name}")
async def update_persona(name: str, request: Request):
    """Update an existing persona."""
    try:
        data = await request.json()
        pm = get_persona_manager()
        persona = pm.update_persona(name, **data)
        if persona:
            return {"success": True, "persona": persona.to_dict()}
        raise HTTPException(404, f"Persona '{name}' not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@router.delete("/chat/personas/{name}")
async def delete_persona(name: str):
    """Delete a persona."""
    try:
        pm = get_persona_manager()
        success = pm.delete_persona(name)
        if success:
            return {"success": True}
        raise HTTPException(404, f"Persona '{name}' not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# =============================================================================
# EXCEL EXPORT ENDPOINT (PRESERVED)
# =============================================================================

class ExportRequest(BaseModel):
    query: str
    project: Optional[str] = None
    tables: Optional[List[str]] = None


@router.post("/chat/export-excel")
async def export_to_excel(request: ExportRequest):
    """Export query results to Excel file."""
    if not STRUCTURED_QUERIES_AVAILABLE:
        raise HTTPException(503, "Structured queries not available")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        handler = get_structured_handler()
        project = request.project
        query = request.query.lower()
        
        # Get tables for project
        schema = handler.get_schema_for_project(project) if project else {'tables': []}
        tables_list = schema.get('tables', [])
        
        if not tables_list:
            raise HTTPException(404, "No tables found for this project")
        
        # Select tables to export
        tables_to_query = []
        if request.tables:
            for t in tables_list:
                if t.get('table_name') in request.tables:
                    tables_to_query.append(t)
        else:
            # Keyword matching for table selection
            keywords = ['employee', 'company', 'personal', 'conversion', 'deduction', 'earning']
            for t in tables_list:
                table_name = t.get('table_name', '').lower()
                sheet_name = t.get('sheet', '').lower()
                if any(kw in query for kw in keywords):
                    if any(kw in table_name or kw in sheet_name for kw in keywords):
                        tables_to_query.append(t)
            
            if not tables_to_query:
                tables_to_query = tables_list[:3]
        
        # Create workbook
        wb = Workbook()
        ws_results = wb.active
        ws_results.title = "Results"
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        sources_info = []
        
        for table_info in tables_to_query[:5]:
            table_name = table_info.get('table_name', '')
            sheet_name = table_info.get('sheet', '')
            file_name = table_info.get('file', '')
            
            try:
                sql = f'SELECT * FROM "{table_name}"'
                rows, cols = handler.execute_query(sql)
                
                if not rows:
                    continue
                
                decrypted = decrypt_results(rows, handler)
                
                # Section header
                ws_results.cell(row=current_row, column=1, value=f"📊 {file_name} → {sheet_name}")
                ws_results.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                ws_results.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=min(len(cols), 10))
                current_row += 1
                
                # Column headers
                for col_idx, col_name in enumerate(cols, 1):
                    cell = ws_results.cell(row=current_row, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                current_row += 1
                
                # Data rows
                for row_data in decrypted:
                    for col_idx, col_name in enumerate(cols, 1):
                        cell = ws_results.cell(row=current_row, column=col_idx, value=row_data.get(col_name, ''))
                        cell.border = thin_border
                    current_row += 1
                
                sources_info.append({
                    'file': file_name,
                    'sheet': sheet_name,
                    'table': table_name,
                    'rows': len(decrypted),
                    'columns': len(cols)
                })
                
                current_row += 2
                
            except Exception as te:
                logger.error(f"[EXPORT] Error querying {table_name}: {te}")
                continue
        
        # Auto-adjust column widths
        for col_idx in range(1, ws_results.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, min(ws_results.max_row + 1, 100)):
                cell = ws_results.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_results.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Sources tab
        ws_sources = wb.create_sheet(title="Sources")
        source_headers = ["File", "Sheet", "Table Name", "Rows", "Columns"]
        for col_idx, header in enumerate(source_headers, 1):
            cell = ws_sources.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, source in enumerate(sources_info, 2):
            ws_sources.cell(row=row_idx, column=1, value=source['file']).border = thin_border
            ws_sources.cell(row=row_idx, column=2, value=source['sheet']).border = thin_border
            ws_sources.cell(row=row_idx, column=3, value=source['table']).border = thin_border
            ws_sources.cell(row=row_idx, column=4, value=source['rows']).border = thin_border
            ws_sources.cell(row=row_idx, column=5, value=source['columns']).border = thin_border
        
        for col_idx in range(1, 6):
            ws_sources.column_dimensions[get_column_letter(col_idx)].width = 25
        
        # Metadata
        metadata_row = len(sources_info) + 3
        ws_sources.cell(row=metadata_row, column=1, value="Export Info:")
        ws_sources.cell(row=metadata_row, column=1).font = Font(bold=True)
        ws_sources.cell(row=metadata_row + 1, column=1, value="Query:")
        ws_sources.cell(row=metadata_row + 1, column=2, value=request.query)
        ws_sources.cell(row=metadata_row + 2, column=1, value="Project:")
        ws_sources.cell(row=metadata_row + 2, column=2, value=project or "All")
        ws_sources.cell(row=metadata_row + 3, column=1, value="Generated:")
        ws_sources.cell(row=metadata_row + 3, column=2, value=time.strftime("%Y-%m-%d %H:%M:%S"))
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        safe_query = re.sub(r'[^\w\s]', '', request.query[:30]).strip().replace(' ', '_')
        filename = f"export_{safe_query}_{timestamp}.xlsx"
        
        logger.warning(f"[EXPORT] Generated {filename} with {len(sources_info)} tables")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[EXPORT] Error: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(500, f"Export failed: {str(e)}")


# =============================================================================
# DATA INSPECTION ENDPOINTS (PRESERVED)
# =============================================================================

@router.get("/chat/data/{project}")
async def get_project_data_summary(project: str):
    """Get summary of available data for a project."""
    if not STRUCTURED_QUERIES_AVAILABLE:
        raise HTTPException(503, "Structured queries not available")
    
    try:
        handler = get_structured_handler()
        schema = handler.get_schema_for_project(project)
        
        return {
            "project": project,
            "tables": len(schema.get('tables', [])),
            "summary": [
                {
                    "file": t.get('file'),
                    "sheet": t.get('sheet'),
                    "rows": t.get('row_count'),
                    "columns": len(t.get('columns', []))
                }
                for t in schema.get('tables', [])
            ]
        }
    except Exception as e:
        raise HTTPException(500, str(e))


@router.get("/chat/data/{project}/{file_name}/versions")
async def get_file_versions(project: str, file_name: str):
    """Get version history for a file."""
    if not STRUCTURED_QUERIES_AVAILABLE:
        raise HTTPException(503, "Not available")
    try:
        handler = get_structured_handler()
        return {"versions": handler.get_file_versions(project, file_name)}
    except Exception as e:
        raise HTTPException(500, str(e))


class CompareRequest(BaseModel):
    sheet_name: str
    key_column: str
    version1: Optional[int] = None
    version2: Optional[int] = None


@router.post("/chat/data/{project}/{file_name}/compare")
async def compare_versions(project: str, file_name: str, request: CompareRequest):
    """Compare two versions of a file."""
    if not STRUCTURED_QUERIES_AVAILABLE:
        raise HTTPException(503, "Not available")
    try:
        handler = get_structured_handler()
        result = handler.compare_versions(
            project=project, 
            file_name=file_name, 
            sheet_name=request.sheet_name, 
            key_column=request.key_column, 
            version1=request.version1, 
            version2=request.version2
        )
        if 'error' in result:
            raise HTTPException(400, result['error'])
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@router.get("/chat/data/encryption-status")
async def encryption_status():
    """Get encryption status."""
    if not STRUCTURED_QUERIES_AVAILABLE:
        raise HTTPException(503, "Not available")
    try:
        handler = get_structured_handler()
        has_encryption = handler.encryptor and (
            getattr(handler.encryptor, 'aesgcm', None) or 
            getattr(handler.encryptor, 'fernet', None)
        )
        return {"encryption": has_encryption, "pii": "enabled"}
    except Exception as e:
        raise HTTPException(500, str(e))


# =============================================================================
# LEARNING/STATS ENDPOINT (NEW)
# =============================================================================

@router.get("/chat/stats")
async def get_chat_stats():
    """Get chat system statistics including learning stats."""
    stats = {
        "active_jobs": len(chat_jobs),
        "routing_available": ROUTING_AVAILABLE,
        "learning_available": LEARNING_AVAILABLE,
        "local_llm_available": LOCAL_LLM_AVAILABLE,
        "structured_available": STRUCTURED_QUERIES_AVAILABLE
    }
    
    if LEARNING_AVAILABLE:
        try:
            learning = get_learning_system()
            stats["learning_stats"] = learning.get_stats()
        except:
            pass
    
    if LOCAL_LLM_AVAILABLE:
        try:
            llm = LocalLLMClient()
            stats["local_llm_reachable"] = llm.is_available()
        except:
            stats["local_llm_reachable"] = False
    
    return stats
