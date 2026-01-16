"""
UNIFIED CHAT ROUTER - THE REVOLUTIONARY HEART OF XLR8
======================================================

This is it. The single, unified chat system that transforms XLR8 from 
"a tool that queries data" into "a world-class consultant in a box."

REVOLUTIONARY FEATURES:
-----------------------
1. DATA MODEL INTELLIGENCE
   - Auto-detect reference tables (location codes, dept codes, pay groups)
   - Build lookup dictionaries on upload
   - Transform "LOC001" → "Houston, TX (LOC001)" automatically
   
2. PROACTIVE DATA QUALITY ALERTS
   - Don't just answer - NOTICE things
   - Terminated employees with active status? Flag it.
   - Missing hire dates? Surface it.
   - Duplicate SSNs? Alert immediately.
   
3. SUGGESTED FOLLOW-UPS
   - Guide the conversation like a real consultant
   - "Break this down by department"
   - "Show me the trend over time"
   - "Compare to benchmarks"
   
4. CITATION & AUDIT TRAIL
   - Every claim backed by data
   - Full SQL transparency
   - Source table attribution
   - Defensible findings

ARCHITECTURE:
-------------
┌─────────────────────────────────────────────────────────────┐
│                    POST /api/chat/unified                    │
└─────────────────────────────┬───────────────────────────────┘
                              │
        ┌─────────────────────┼─────────────────────┐
        ▼                     ▼                     ▼
┌──────────────┐    ┌──────────────┐    ┌──────────────┐
│ DataModel    │    │ Intelligence │    │ DataQuality  │
│ Service      │    │ Engine       │    │ Service      │
│ (Lookups)    │    │ (Core Logic) │    │ (Alerts)     │
└──────────────┘    └──────────────┘    └──────────────┘
        │                     │                     │
        └─────────────────────┼─────────────────────┘
                              ▼
                    ┌──────────────────┐
                    │ Citation Builder │
                    │ + Follow-up Gen  │
                    └──────────────────┘
                              │
                              ▼
                    ┌──────────────────┐
                    │ PII Redaction    │
                    │ (if Claude call) │
                    └──────────────────┘
                              │
                              ▼
                    ┌──────────────────┐
                    │ RESPONSE         │
                    │ Revolutionary.   │
                    └──────────────────┘

DEPLOYMENT:
-----------
1. Copy to backend/routers/unified_chat.py
2. Add to main.py:
   from routers import unified_chat
   app.include_router(unified_chat.router, prefix="/api")
3. Update frontend to call /api/chat/unified

Author: XLR8 Team
Version: 1.1.0 - Phase 3.5 Intelligence Consumer
Date: December 2025
"""

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Dict, List, Optional, Any, Tuple
import logging
import uuid
import json
import re
import time
import io
import traceback

logger = logging.getLogger(__name__)

router = APIRouter(tags=["unified-chat"])


# =============================================================================
# IMPORTS - Graceful degradation for all dependencies
# =============================================================================

# Intelligence Engine - V2 ONLY (modular architecture)
# The monolith (intelligence_engine.py) is deprecated and archived.
import os

# Modular engine (V2) - THE engine going forward
from backend.utils.intelligence import IntelligenceEngineV2, IntelligenceMode
ENGINE_V2_AVAILABLE = True
logger.info("[UNIFIED] IntelligenceEngineV2 available")

# Legacy alias for compatibility (points to V2)
IntelligenceEngine = IntelligenceEngineV2
INTELLIGENCE_AVAILABLE = ENGINE_V2_AVAILABLE

# Learning Module
from backend.utils.learning import get_learning_module
LEARNING_AVAILABLE = True

# Structured Data Handler - use read handler for queries (no write conflicts)
from utils.structured_data_handler import get_structured_handler
STRUCTURED_AVAILABLE = True

# RAG Handler
from utils.rag_handler import RAGHandler
RAG_AVAILABLE = True

# LLM Orchestrator
from utils.llm_orchestrator import LLMOrchestrator
LLM_AVAILABLE = True

# Persona Manager
from utils.persona_manager import get_persona_manager
PERSONAS_AVAILABLE = True

# Expert Context & Domain Inference (Phase 4: Intelligent Context Selection)
from backend.utils.expert_context_registry import (
    select_expert_context,
    record_expert_feedback,
    get_expert_selector,
    get_expert_registry,
)
EXPERT_CONTEXT_AVAILABLE = True

from backend.utils.domain_inference_engine import (
    get_domain_engine,
    infer_project_domains,
    get_primary_domain,
)
DOMAIN_INFERENCE_AVAILABLE = True

# Supabase
from utils.database.supabase_client import get_supabase
SUPABASE_AVAILABLE = True

# Metrics Service (for query tracking)
from backend.utils.metrics_service import MetricsService
METRICS_AVAILABLE = True

# Project Intelligence Service (Phase 3)
from backend.utils.project_intelligence import ProjectIntelligenceService, get_project_intelligence
PROJECT_INTELLIGENCE_AVAILABLE = True

# Chat Services (extracted from this file for modularity)
from backend.utils.chat_services import (
    ReversibleRedactor,
    DataModelService,
    DataQualityService,
    FollowUpGenerator,
    CitationBuilder
)
CHAT_SERVICES_AVAILABLE = True


# =============================================================================
# REQUEST/RESPONSE MODELS
# =============================================================================

class UnifiedChatRequest(BaseModel):
    """
    Unified chat request - supports all chat modes.
    
    Attributes:
        message: The user's question or command
        project: Project identifier (e.g., "TEA1000")
        project_id: Alias for project (for backward compatibility)
        persona: Persona to use for response style (default: None - professional)
        scope: Data scope - "project", "global", or "all"
        mode: Force specific intelligence mode (optional)
        clarifications: Answers to clarification questions
        session_id: Session ID for conversation continuity
        include_quality_alerts: Whether to run data quality checks
        include_follow_ups: Whether to suggest follow-up questions
        include_citations: Whether to include full audit trail
    """
    message: str
    project: Optional[str] = None
    project_id: Optional[str] = None  # Alias for project (backward compat)
    persona: Optional[str] = None
    scope: Optional[str] = 'project'
    mode: Optional[str] = None
    clarifications: Optional[Dict[str, Any]] = None
    session_id: Optional[str] = None
    
    # Revolutionary feature flags
    include_quality_alerts: Optional[bool] = True
    include_follow_ups: Optional[bool] = True
    include_citations: Optional[bool] = True
    
    # Deprecated - V2 is now the only engine (kept for backward compatibility)
    use_engine_v2: Optional[bool] = True  # Ignored, always uses V2


class ClarificationAnswer(BaseModel):
    """User's answers to clarification questions."""
    session_id: str
    original_question: str
    answers: Dict[str, Any]


class FeedbackRequest(BaseModel):
    """Feedback on chat response."""
    job_id: Optional[str] = None
    session_id: Optional[str] = None
    feedback: str  # 'up' or 'down'
    message: Optional[str] = None
    response: Optional[str] = None
    correction: Optional[str] = None


class ResetPreferencesRequest(BaseModel):
    """Request to reset user preferences."""
    session_id: Optional[str] = None
    project: Optional[str] = None
    reset_type: str = "session"  # "session" or "learned"


class ExportRequest(BaseModel):
    """Request to export data to Excel."""
    query: str
    project: Optional[str] = None
    tables: Optional[List[str]] = None
    include_summary: Optional[bool] = True
    include_quality_report: Optional[bool] = True



# =============================================================================
# CHAT SERVICE CLASSES (Imported from backend.utils.chat_services)
# =============================================================================
# The following classes are imported at the top of this file:
# - ReversibleRedactor: PII redaction before LLM calls
# - DataModelService: Code-to-description lookups  
# - DataQualityService: Proactive data quality alerts
# - FollowUpGenerator: Suggested follow-up questions
# - CitationBuilder: Audit trail and source attribution
#
# See: backend/utils/chat_services/ for implementations

# =============================================================================
# SESSION MANAGEMENT
# =============================================================================

# In-memory session storage (in production, use Redis)
unified_sessions: Dict[str, Dict] = {}


def get_or_create_session(session_id: str, project: str) -> Tuple[str, Dict]:
    """Get existing session or create new one."""
    if not session_id:
        session_id = f"session_{uuid.uuid4().hex[:8]}"
    
    if session_id not in unified_sessions:
        unified_sessions[session_id] = {
            'engine': None,
            'project': project,
            'created_at': time.time(),
            'last_sql': None,
            'last_result': None,
            'last_question': None,
            'skip_learning': False,
            'data_model': None,
            'conversation_history': [],
            'interaction_count': 0  # Track interactions for learning control
        }
    
    return session_id, unified_sessions[session_id]


def cleanup_old_sessions(max_sessions: int = 100) -> None:
    """Remove oldest sessions if over limit."""
    if len(unified_sessions) > max_sessions:
        # Sort by created_at and remove oldest
        sorted_sessions = sorted(
            unified_sessions.items(),
            key=lambda x: x[1].get('created_at', 0)
        )
        for session_id, _ in sorted_sessions[:len(sorted_sessions) - max_sessions]:
            del unified_sessions[session_id]


# =============================================================================
# SCHEMA RETRIEVAL - Registry-Filtered (v3.0)
# =============================================================================

def _get_valid_files_from_registry(project: str) -> set:
    """
    Get valid filenames from DocumentRegistry for this project.
    Returns lowercase filenames for case-insensitive matching.
    Returns None if registry unavailable (allows fallback to unfiltered).
    
    REGISTRY IS SOURCE OF TRUTH - if a file isn't registered, 
    its data should not be visible to chat.
    """
    try:
        from utils.database.models import DocumentRegistryModel, ProjectModel
        
        # Get project_id from project name
        project_id = None
        if project:
            proj_record = ProjectModel.get_by_name(project)
            if proj_record:
                project_id = proj_record.get('id')
        
        # Get registered files for this project
        if project_id:
            entries = DocumentRegistryModel.get_by_project(project_id, include_global=True)
        else:
            entries = DocumentRegistryModel.get_all()
        
        valid_files = set()
        for entry in entries:
            filename = entry.get('filename', '')
            if filename:
                valid_files.add(filename.lower())
        
        logger.info(f"[SCHEMA] Registry has {len(valid_files)} valid files for project '{project}'")
        return valid_files
        
    except Exception as e:
        logger.warning(f"[SCHEMA] Registry lookup failed: {e} - proceeding without filter")
        return None  # Return None to indicate fallback (not empty set)


async def get_project_schema(project: str, scope: str, handler) -> Dict:
    """
    Get comprehensive schema for project including column profiles.
    
    v3.0 FIX (2025-12-21): Now filters through DocumentRegistry first.
    Registry is the SOURCE OF TRUTH - if a file isn't in registry,
    its tables should not be returned to chat.
    
    v2.0 FIX (2025-12-17): Uses handler.get_tables() which queries 
    _schema_metadata with exact project match.
    """
    tables = []
    filter_candidates = {}
    
    if not handler or not handler.conn:
        logger.error("[SCHEMA] No handler connection")
        return {'tables': [], 'filter_candidates': {}}
    
    try:
        # ================================================================
        # STEP 1: Get valid files from Registry (SOURCE OF TRUTH)
        # ================================================================
        valid_files = _get_valid_files_from_registry(project)
        registry_available = valid_files is not None
        
        if registry_available and not valid_files:
            # Registry is available but has no files for this project
            logger.warning(f"[SCHEMA] No files in registry for project '{project}'")
            return {'tables': [], 'filter_candidates': {}}
        
        # ================================================================
        # STEP 2: Get tables from metadata
        # ================================================================
        metadata_tables = handler.get_tables(project)
        
        logger.warning(f"[SCHEMA] get_tables('{project}') returned {len(metadata_tables)} tables")
        
        # If no tables found via metadata, try fallback to SHOW TABLES
        if not metadata_tables:
            logger.warning(f"[SCHEMA] No tables in _schema_metadata for '{project}', trying SHOW TABLES fallback")
            
            all_tables = handler.conn.execute("SHOW TABLES").fetchall()
            logger.warning(f"[SCHEMA] DuckDB SHOW TABLES returned {len(all_tables)} total tables")
            
            # Build project prefix for filtering
            project_clean = (project or '').strip()
            sanitized_project = re.sub(r'[^\w\s]', '', project_clean)
            sanitized_project = re.sub(r'\s+', '_', sanitized_project.strip()).lower()
            
            project_prefixes = [
                sanitized_project,
                project_clean.lower(),
                project_clean.lower().replace(' ', '_'),
                project_clean.lower().replace(' ', '_').replace('-', '_'),
            ]
            project_prefixes = list(dict.fromkeys([p for p in project_prefixes if p]))
            
            matched_tables = []
            
            for (table_name,) in all_tables:
                if table_name.startswith('_'):
                    continue
                
                table_lower = table_name.lower()
                matches_project = any(
                    table_lower.startswith(prefix) 
                    for prefix in project_prefixes
                )
                
                if matches_project:
                    matched_tables.append(table_name)
            
            tables_to_process = matched_tables if matched_tables else [t[0] for t in all_tables if not t[0].startswith('_')]
            
            for table_name in tables_to_process:
                try:
                    # ================================================
                    # REGISTRY FILTER: Skip tables for unregistered files
                    # ================================================
                    if registry_available and valid_files:
                        source_file = None
                        try:
                            file_result = handler.conn.execute("""
                                SELECT file_name FROM _schema_metadata WHERE table_name = ? LIMIT 1
                            """, [table_name]).fetchone()
                            if file_result:
                                source_file = file_result[0]
                        except Exception:
                            pass
                        
                        if not source_file:
                            try:
                                file_result = handler.conn.execute("""
                                    SELECT source_file FROM _pdf_tables WHERE table_name = ? LIMIT 1
                                """, [table_name]).fetchone()
                                if file_result:
                                    source_file = file_result[0]
                            except Exception:
                                pass
                        
                        if source_file and source_file.lower() not in valid_files:
                            logger.debug(f"[SCHEMA] Skipping {table_name} - file '{source_file}' not in registry")
                            continue
                    # ================================================
                    
                    # Get columns
                    columns = []
                    try:
                        col_result = handler.conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
                        columns = [row[1] for row in col_result]
                    except Exception:
                        try:
                            col_result = handler.conn.execute(f'DESCRIBE "{table_name}"').fetchall()
                            columns = [row[0] for row in col_result]
                        except Exception:
                            result = handler.conn.execute(f'SELECT * FROM "{table_name}" LIMIT 0')
                            columns = [desc[0] for desc in result.description]
                    
                    if not columns:
                        continue
                    
                    # Get row count
                    try:
                        count_result = handler.conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()
                        row_count = count_result[0] if count_result else 0
                    except Exception:
                        row_count = 0
                    
                    # Try to get display_name from _schema_metadata
                    display_name = ''
                    try:
                        dn_result = handler.conn.execute("""
                            SELECT display_name FROM _schema_metadata WHERE table_name = ? LIMIT 1
                        """, [table_name]).fetchone()
                        if dn_result and dn_result[0]:
                            display_name = dn_result[0]
                    except Exception:
                        pass
                    
                    tables.append({
                        'table_name': table_name,
                        'display_name': display_name,
                        'project': project or 'unknown',
                        'columns': columns,
                        'row_count': row_count,
                        'column_profiles': {},
                        'categorical_columns': []
                    })
                    
                except Exception as col_e:
                    logger.warning(f"[SCHEMA] Error processing {table_name}: {col_e}")
        
        else:
            # ================================================================
            # PRIMARY PATH: Use metadata_tables from get_tables()
            # ================================================================
            for table_info in metadata_tables:
                table_name = table_info.get('table_name', '')
                if not table_name:
                    continue
                
                # ================================================
                # REGISTRY FILTER: Skip tables for unregistered files
                # ================================================
                if registry_available and valid_files:
                    source_file = table_info.get('file_name', table_info.get('source_file', ''))
                    if source_file and source_file.lower() not in valid_files:
                        logger.debug(f"[SCHEMA] Skipping {table_name} - file '{source_file}' not in registry")
                        continue
                # ================================================
                
                # Get columns - use metadata columns or query
                columns = table_info.get('columns', [])
                if columns and isinstance(columns[0], dict):
                    columns = [c.get('name', str(c)) for c in columns]
                
                if not columns:
                    try:
                        col_result = handler.conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
                        columns = [row[1] for row in col_result]
                    except Exception:
                        try:
                            result = handler.conn.execute(f'SELECT * FROM "{table_name}" LIMIT 0')
                            columns = [desc[0] for desc in result.description]
                        except Exception:
                            columns = []
                
                row_count = table_info.get('row_count', 0)
                
                # Get column profiles (Phase 2: Data-driven clarification)
                categorical_columns = []
                column_profiles = {}
                
                try:
                    profile_result = handler.conn.execute("""
                        SELECT column_name, inferred_type, distinct_count, 
                               distinct_values, value_distribution, is_categorical,
                               min_value, max_value, filter_category
                        FROM _column_profiles 
                        WHERE table_name = ?
                    """, [table_name]).fetchall()
                    
                    for prow in profile_result:
                        col_name = prow[0]
                        profile = {
                            'inferred_type': prow[1],
                            'distinct_count': prow[2],
                            'is_categorical': prow[5],
                            'filter_category': prow[8]
                        }
                        
                        if prow[3]:  # distinct_values
                            try:
                                profile['distinct_values'] = json.loads(prow[3])
                            except Exception:
                                pass
                        
                        if prow[4]:  # value_distribution
                            try:
                                profile['value_distribution'] = json.loads(prow[4])
                            except Exception:
                                pass
                        
                        if prow[1] == 'numeric':
                            profile['min_value'] = prow[6]
                            profile['max_value'] = prow[7]
                        
                        column_profiles[col_name] = profile
                        
                        if prow[5] and prow[2] and prow[2] <= 20:  # is_categorical and small distinct count
                            categorical_columns.append({
                                'column': col_name,
                                'values': profile.get('distinct_values', []),
                                'distribution': profile.get('value_distribution', {}),
                                'filter_category': prow[8]
                            })
                            
                except Exception as profile_e:
                    logger.debug(f"[SCHEMA] No profiles for {table_name}: {profile_e}")
                
                tables.append({
                    'table_name': table_name,
                    'display_name': table_info.get('display_name', ''),
                    'file_name': table_info.get('file_name', ''),
                    'project': project or 'unknown',
                    'columns': columns,
                    'row_count': row_count,
                    'column_profiles': column_profiles,
                    'categorical_columns': categorical_columns
                })
        
        # Get filter candidates
        try:
            filter_candidates = handler.get_filter_candidates(project)
        except Exception:
            pass
        
        logger.warning(f"[SCHEMA] Returning {len(tables)} tables for project '{project}' (registry_filtered={registry_available})")
        
    except Exception as e:
        logger.error(f"[SCHEMA] Failed: {e}")
        import traceback
        logger.error(traceback.format_exc())
    
    return {'tables': tables, 'filter_candidates': filter_candidates}


# =============================================================================
# ANSWER GENERATION WITH INTELLIGENT CONTEXT SELECTION
# =============================================================================

def _get_project_domains(project: str, handler=None) -> Tuple[Optional[str], List[Dict]]:
    """
    Get detected domains for a project.
    
    Returns:
        (project_id, domains_list)
    """
    if not DOMAIN_INFERENCE_AVAILABLE:
        return None, []
    
    try:
        # First try to get project_id and cached domains
        project_id = None
        if SUPABASE_AVAILABLE:
            try:
                supabase = get_supabase()
                result = supabase.table('projects').select('id, metadata').eq('name', project).execute()
                if result.data:
                    project_id = result.data[0].get('id')
                    # Check if domains already computed
                    metadata = result.data[0].get('metadata', {}) or {}
                    detected = metadata.get('detected_domains', {})
                    if detected and detected.get('domains'):
                        logger.info(f"[UNIFIED] Using cached domains for {project}")
                        return project_id, detected.get('domains', [])
            except Exception as e:
                logger.warning(f"[UNIFIED] Project lookup error: {e}")
        
        # If no cached domains, infer them now
        engine = get_domain_engine(handler)
        result = engine.infer_domains(project, project_id)
        if result:
            logger.info(f"[UNIFIED] Inferred domains for {project}: primary={result.primary_domain}")
            return project_id, [d.to_dict() for d in result.domains]
        
        return project_id, []
        
    except Exception as e:
        logger.warning(f"[UNIFIED] Domain inference error: {e}")
        return None, []


async def generate_synthesized_answer(
    question: str,
    context: str,
    persona: str,
    insights: List,
    conflicts: List,
    citations: CitationBuilder,
    quality_alerts: DataQualityService,
    follow_ups: List[str],
    redactor: ReversibleRedactor,
    project: str = None,
    project_domains: List[Dict] = None,
) -> Tuple[str, Optional[str]]:
    """
    Generate the final synthesized answer using LOCAL LLMs FIRST.
    
    Flow:
    1. Select expert context based on question + domains
    2. Try Mistral via Ollama for synthesis
    3. Fall back to Claude ONLY if Mistral fails
    
    Returns:
        (answer_text, expert_context_id_used)
    """
    expert_context_id = None
    
    try:
        if not LLM_AVAILABLE:
            logger.warning("[UNIFIED] LLM not available, returning raw context")
            return context[:3000], None
        
        orchestrator = LLMOrchestrator()
        
        # Redact PII before sending to any LLM
        redacted_context = redactor.redact(context)
        
        # ===================================================================
        # INTELLIGENT CONTEXT SELECTION
        # ===================================================================
        expert_prompt = None
        
        # STEP 1: Try expert context auto-selection
        if EXPERT_CONTEXT_AVAILABLE:
            try:
                selector = get_expert_selector()
                matches = selector.select(question, project_domains, top_k=1)
                
                if matches and matches[0].match_score > 0.35:
                    expert_prompt = matches[0].context.prompt_template
                    expert_context_id = matches[0].context.id
                    logger.info(f"[UNIFIED] Using expert context: {matches[0].context.name} "
                               f"(score={matches[0].match_score:.2f})")
            except Exception as e:
                logger.warning(f"[UNIFIED] Expert context selection failed: {e}")
        
        # STEP 2: Fall back to persona if no expert match
        if not expert_prompt and PERSONAS_AVAILABLE and persona:
            try:
                pm = get_persona_manager()
                persona_obj = pm.get_persona(persona)
                if persona_obj and persona_obj.system_prompt:
                    expert_prompt = persona_obj.system_prompt
                    logger.info(f"[UNIFIED] Using persona: {persona_obj.name}")
            except Exception as e:
                logger.warning(f"[UNIFIED] Persona load failed: {e}")
        
        # STEP 3: No expert prompt - orchestrator will use default
        if not expert_prompt:
            logger.info("[UNIFIED] Using default prompt (no expert match or persona)")
        
        # ===================================================================
        # SYNTHESIS VIA ORCHESTRATOR - Mistral first, Claude fallback
        # ===================================================================
        result = orchestrator.synthesize_answer(
            question=question,
            context=redacted_context,
            expert_prompt=expert_prompt,
            use_claude_fallback=True  # Only if Mistral fails
        )
        
        if result.get('success'):
            answer = result.get('response', '')
            model_used = result.get('model_used', 'unknown')
            logger.info(f"[UNIFIED] Synthesis complete via {model_used}")
            
            # Restore any PII in the response
            answer = redactor.restore(answer)
            
            return answer, expert_context_id
        else:
            error = result.get('error', 'Unknown error')
            logger.warning(f"[UNIFIED] Synthesis failed: {error}")
            return context[:2000], None
        
    except Exception as e:
        logger.error(f"[UNIFIED] Synthesis error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return context[:2000], None


# =============================================================================
# MAIN ENDPOINT
# =============================================================================

@router.post("/chat/unified")
async def unified_chat(request: UnifiedChatRequest):
    """
    UNIFIED CHAT ENDPOINT - The Revolutionary Heart of XLR8
    
    This endpoint combines:
    - Intelligent clarification (from Phase 2)
    - Data model intelligence (reference lookups)
    - Proactive data quality alerts
    - Suggested follow-ups
    - Full citation/audit trail
    - PII protection
    
    Returns a comprehensive response suitable for rich frontend display.
    """
    query_start_time = time.time()  # For metrics tracking
    
    # Use project_id as fallback for project (backward compatibility)
    project = request.project or request.project_id
    message = request.message
    session_id, session = get_or_create_session(request.session_id, project)
    
    logger.warning(f"[UNIFIED] ===== NEW REQUEST =====")
    logger.warning(f"[UNIFIED] Message: {message[:100]}...")
    logger.warning(f"[UNIFIED] Project: {project}, Session: {session_id}")
    
    # Initialize services
    redactor = ReversibleRedactor()
    citation_builder = CitationBuilder()
    
    try:
        # Get project_id (UUID) and project_code from project name for DuckDB/RAG operations
        project_id = None
        project_code = None  # The code used in DuckDB term index
        product_id = None  # Phase 5F: Multi-product support
        if project and SUPABASE_AVAILABLE:
            try:
                supabase = get_supabase()
                result = supabase.table('projects').select('id, code, metadata').eq('name', project).limit(1).execute()
                if result.data:
                    project_id = result.data[0].get('id')
                    project_code = result.data[0].get('code')  # Get the project code for DuckDB
                    # Phase 5F: Extract product from metadata
                    metadata = result.data[0].get('metadata', {}) or {}
                    product_id = metadata.get('product')
                    logger.info(f"[UNIFIED] Resolved project_id: {project_id}, project_code: {project_code}, product: {product_id}")
            except Exception as e:
                logger.warning(f"[UNIFIED] Could not resolve project_id: {e}")
        
        # Use project_code for DuckDB operations if available, otherwise fall back to project name
        project_for_duckdb = project_code or project
        logger.warning(f"[UNIFIED] Using project_for_duckdb: {project_for_duckdb}")
        
        # Get or create intelligence engine (V2 ONLY)
        if session['engine']:
            engine = session['engine']
            # Ensure project_id is set on existing engines
            if project_id and not getattr(engine, 'project_id', None):
                engine.project_id = project_id
            # Phase 5F: Load product schema if changed
            if product_id and getattr(engine, 'product_id', None) != product_id:
                engine._load_product_schema(product_id)
        elif ENGINE_V2_AVAILABLE:
            logger.warning("[UNIFIED] Creating IntelligenceEngineV2 (modular)")
            # Use project_code for the engine so term index lookups work
            engine = IntelligenceEngineV2(project_for_duckdb or 'default', project_id=project_id, product_id=product_id)
            session['engine'] = engine
            session['engine_type'] = 'v2'
        else:
            return {
                "session_id": session_id,
                "answer": "Intelligence engine not available. Please check deployment.",
                "needs_clarification": False,
                "confidence": 0.0,
                "success": False
            }
        
        # Load structured data handler and schema
        handler = None
        schema = {'tables': [], 'filter_candidates': {}}
        data_model = None
        quality_service = None
        
        # Load RAG handler FIRST (before load_context creates gatherers)
        rag = None
        if RAG_AVAILABLE:
            try:
                rag = RAGHandler()
                logger.info("[UNIFIED] RAG handler created")
            except Exception as e:
                logger.warning(f"[UNIFIED] RAG handler error: {e}")
        
        if STRUCTURED_AVAILABLE:
            try:
                handler = get_structured_handler()
                if handler and handler.conn:
                    schema = await get_project_schema(project, request.scope, handler)
                    
                    # Always initialize engine - ChromaDB works even without DuckDB tables
                    engine.load_context(structured_handler=handler, schema=schema, rag_handler=rag)
                    
                    if schema['tables']:
                        logger.info(f"[UNIFIED] Loaded {len(schema['tables'])} tables")
                        
                        # Initialize Data Model Service
                        data_model = DataModelService(project)
                        data_model.load_lookups(handler)
                        session['data_model'] = data_model
                        
                        # Initialize Quality Service
                        quality_service = DataQualityService(project)
                    else:
                        logger.warning("[UNIFIED] No DuckDB tables, but engine initialized for ChromaDB")
                        
            except Exception as e:
                logger.error(f"[UNIFIED] Structured handler error: {e}")
        
        # Load relationships
        if SUPABASE_AVAILABLE:
            try:
                supabase = get_supabase()
                result = supabase.table('project_relationships').select('*').eq(
                    'project_name', project
                ).in_('status', ['confirmed', 'auto_confirmed']).execute()
                
                if result.data:
                    engine.relationships = result.data
            except Exception as e:
                logger.warning(f"[UNIFIED] Relationships error: {e}")
        
        # Apply clarifications if provided
        if request.clarifications:
            logger.info(f"[UNIFIED] Applying clarifications: {request.clarifications}")
            engine.confirmed_facts.update(request.clarifications)
            session['skip_learning'] = False
            
            # v4.0: If we have a pending question from clarification, use it
            if session.get('pending_clarification_question'):
                original_question = session.pop('pending_clarification_question')
                logger.warning(f"[UNIFIED] Using pending question: {original_question[:50]}...")
                message = original_question  # Use original question, not the clarification answer
            
            # Record to learning module
        
        # v4.0: Auto-detect scope clarification responses
        # If we have a pending question and the message looks like a scope value, auto-apply it
        elif session.get('pending_clarification_question'):
            pending_q = session.get('pending_clarification_question')
            # Check if current message could be a scope value (short, no spaces, uppercase company code)
            if len(message) < 30 and ' ' not in message.strip():
                logger.warning(f"[UNIFIED] Auto-detecting scope clarification: '{message}' for pending question")
                # Treat this as a scope clarification
                # Try to detect the dimension from the message
                # Common patterns: company code, country code, etc.
                scope_value = message.strip()
                
                # Apply as scope clarification
                engine.confirmed_facts['scope'] = f"company:{scope_value}"  # Default to company
                logger.warning(f"[UNIFIED] Auto-applied scope: {engine.confirmed_facts['scope']}")
                
                # Use the original question
                message = session.pop('pending_clarification_question')
                logger.warning(f"[UNIFIED] Using pending question: {message[:50]}...")
            if LEARNING_AVAILABLE:
                try:
                    learning = get_learning_module()
                    for q_id, choice in request.clarifications.items():
                        if not q_id.startswith('_'):
                            learning.record_clarification_choice(
                                question_id=q_id,
                                chosen_option=str(choice),
                                project=project
                            )
                except Exception as e:
                    logger.warning(f"[UNIFIED] Learning record error: {e}")
        
        # Pass conversation context to engine
        if session.get('last_sql') or session.get('last_result'):
            engine.conversation_context = {
                'last_sql': session.get('last_sql'),
                'last_result': session.get('last_result'),
                'last_question': session.get('last_question')
            }
        
        # Determine mode
        mode = None
        if request.mode:
            try:
                mode = IntelligenceMode(request.mode)
            except Exception:
                pass
        
        # Check learning for similar queries
        learned_sql = None
        auto_applied_facts = {}
        
        if LEARNING_AVAILABLE and not request.clarifications and not session.get('skip_learning'):
            try:
                learning = get_learning_module()
                
                # Check for learned SQL patterns
                similar = learning.find_similar_query(
                    question=message,
                    intent=mode.value if mode else None,
                    project=project
                )
                if similar and similar.get('successful_sql'):
                    learned_sql = similar['successful_sql']
                    logger.info(f"[UNIFIED] Found learned SQL pattern")
                    
            except Exception as e:
                logger.warning(f"[UNIFIED] Learning check error: {e}")
        
        # ASK THE INTELLIGENCE ENGINE
        answer = engine.ask(message, mode=mode, context={'learned_sql': learned_sql} if learned_sql else None)
        
        # Check if we can skip clarification using learning
        # BUT only after the first interaction in a session (let users see clarification first)
        if answer.structured_output and answer.structured_output.get('type') == 'clarification_needed':
            session_interactions = session.get('interaction_count', 0)
            
            # Only auto-apply learning if user has already interacted this session
            # First question should show clarification so user knows the system asks
            if LEARNING_AVAILABLE and not session.get('skip_learning') and session_interactions > 0:
                try:
                    learning = get_learning_module()
                    questions = answer.structured_output.get('questions', [])
                    detected_domain = answer.structured_output.get('detected_domains', ['general'])[0]
                    
                    can_skip, learned_answers = learning.should_skip_clarification(
                        questions=questions,
                        domain=detected_domain,
                        project=project
                    )
                    
                    if can_skip and learned_answers:
                        logger.info(f"[UNIFIED] Auto-applying learned answers: {learned_answers}")
                        engine.confirmed_facts.update(learned_answers)
                        auto_applied_facts = learned_answers.copy()
                        answer = engine.ask(message, mode=mode)
                        
                except Exception as e:
                    logger.warning(f"[UNIFIED] Skip clarification error: {e}")
        
        # Build auto-applied note
        auto_applied_note = ""
        if auto_applied_facts:
            notes = []
            for key, value in auto_applied_facts.items():
                if key in ['employee_status', 'status']:
                    if value == 'active':
                        notes.append("Active employees only")
                    elif value == 'termed':
                        notes.append("Terminated employees only")
                    else:
                        notes.append(f"Status: {value}")
                else:
                    notes.append(f"{key}: {value}")
            if notes:
                auto_applied_note = "📌 *Remembered: " + ", ".join(notes) + "*"
        
        # Build response
        response = {
            "session_id": session_id,
            "question": answer.question,
            "confidence": answer.confidence,
            "reasoning": answer.reasoning,
            
            # Auto-applied preferences
            "auto_applied_note": auto_applied_note,
            "auto_applied_facts": auto_applied_facts,
            "can_reset_preferences": bool(auto_applied_facts),
            
            # Clarification
            "needs_clarification": (
                answer.structured_output and 
                answer.structured_output.get('type') == 'clarification_needed'
            ),
            "clarification_questions": (
                answer.structured_output.get('questions', []) 
                if answer.structured_output else []
            ),
            
            # Five Truths
            "from_reality": [],
            "from_intent": [],
            "from_configuration": [],
            "from_reference": [],
            "from_regulatory": [],
            "from_compliance": [],
            
            # Core results
            "conflicts": [],
            "insights": [],
            "structured_output": answer.structured_output,
            
            # Revolutionary features
            "quality_alerts": None,
            "follow_up_suggestions": [],
            "citations": None,
            
            # v3.2: Consultative metadata (excel_spec, proactive_offers, hcmpact_hook)
            "question_type": None,
            "question_category": None,
            "excel_spec": [],
            "proactive_offers": [],
            "hcmpact_hook": None,
            "synthesis_method": None,
            
            # v4: Provenance / Show Our Math
            "provenance": {
                "resolution_path": [],      # Step-by-step lookup chain
                "sql_executed": None,       # Actual SQL run
                "source_tables": [],        # Tables used
                "methodology": None         # Human-readable explanation
            },
            
            # Learning
            "used_learning": learned_sql is not None,
            
            # The answer
            "answer": None,
            "success": True
        }
        
        # v3.2: Extract consultative metadata if available
        if hasattr(answer, 'consultative_metadata') and answer.consultative_metadata:
            cm = answer.consultative_metadata
            response["question_type"] = cm.get('question_type')
            response["question_category"] = cm.get('question_category')
            response["excel_spec"] = cm.get('excel_spec', [])
            response["proactive_offers"] = cm.get('proactive_offers', [])
            response["hcmpact_hook"] = cm.get('hcmpact_hook')
            response["synthesis_method"] = cm.get('synthesis_method')
            logger.warning(f"[UNIFIED] Consultative metadata: type={cm.get('question_type')}, "
                          f"offers={len(cm.get('proactive_offers', []))}, "
                          f"excel_sheets={len(cm.get('excel_spec', []))}")
        
        # Serialize truths (Five Truths)
        for truth in answer.from_reality:
            response["from_reality"].append(_serialize_truth(truth, data_model))
        for truth in answer.from_intent:
            response["from_intent"].append(_serialize_truth(truth, data_model))
        for truth in answer.from_configuration:
            response["from_configuration"].append(_serialize_truth(truth, data_model))
        for truth in answer.from_reference:
            response["from_reference"].append(_serialize_truth(truth, data_model))
        for truth in answer.from_regulatory:
            response["from_regulatory"].append(_serialize_truth(truth, data_model))
        for truth in answer.from_compliance:
            response["from_compliance"].append(_serialize_truth(truth, data_model))
        
        # v4: Populate provenance from reality truths
        for truth in answer.from_reality:
            if isinstance(truth.content, dict):
                # SQL executed
                if truth.content.get('sql') and not response["provenance"]["sql_executed"]:
                    response["provenance"]["sql_executed"] = truth.content.get('sql')
                
                # Resolution path (step-by-step)
                if truth.content.get('resolution_path'):
                    response["provenance"]["resolution_path"] = truth.content.get('resolution_path')
                
                # Source table
                table = truth.content.get('table') or truth.content.get('display_name')
                if table and table not in response["provenance"]["source_tables"]:
                    response["provenance"]["source_tables"].append(table)
                
                # Reality context has methodology info
                reality_ctx = truth.content.get('reality_context', {})
                if reality_ctx and reality_ctx.get('answer_label'):
                    response["provenance"]["methodology"] = reality_ctx.get('answer_label')
        
        # Also add source tables from other truths
        for truth in answer.from_configuration + answer.from_reference + answer.from_regulatory:
            if truth.source_name and truth.source_name not in response["provenance"]["source_tables"]:
                response["provenance"]["source_tables"].append(truth.source_name)
        
        # Serialize conflicts
        for conflict in answer.conflicts:
            response["conflicts"].append({
                "description": conflict.description,
                "severity": conflict.severity,
                "recommendation": conflict.recommendation
            })
        
        # Serialize insights
        for insight in answer.insights:
            response["insights"].append({
                "type": insight.type,
                "title": insight.title,
                "description": insight.description,
                "severity": insight.severity,
                "action_required": insight.action_required
            })
        
        # Generate answer if not clarification
        if not response["needs_clarification"] and answer.answer:
            
            # Check for export request first
            if answer.structured_output and answer.structured_output.get('type') == 'export_ready':
                logger.warning(f"[UNIFIED] Export request - generating Excel file")
                export_data = answer.structured_output.get('export_data', {})
                
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Border, Side
                    import io
                    import base64
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Validation Results"
                    
                    # Styling
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    
                    rows = export_data.get('rows', [])
                    columns = export_data.get('columns', [])
                    findings = export_data.get('findings', [])
                    
                    if rows and columns:
                        # Headers
                        for col_idx, col_name in enumerate(columns, 1):
                            cell = ws.cell(row=1, column=col_idx, value=col_name)
                            cell.font = header_font
                            cell.fill = header_fill
                        
                        # Data
                        for row_idx, row in enumerate(rows, 2):
                            for col_idx, col_name in enumerate(columns, 1):
                                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ''))
                        
                        # Auto-width columns
                        for col_idx, col_name in enumerate(columns, 1):
                            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = max(12, len(col_name) + 2)
                    
                    # Findings sheet
                    if findings:
                        ws_findings = wb.create_sheet(title="Findings")
                        ws_findings.cell(row=1, column=1, value="Finding").font = header_font
                        ws_findings.cell(row=1, column=1).fill = header_fill
                        ws_findings.cell(row=1, column=2, value="Severity").font = header_font
                        ws_findings.cell(row=1, column=2).fill = header_fill
                        ws_findings.cell(row=1, column=3, value="Message").font = header_font
                        ws_findings.cell(row=1, column=3).fill = header_fill
                        ws_findings.cell(row=1, column=4, value="Action").font = header_font
                        ws_findings.cell(row=1, column=4).fill = header_fill
                        
                        for f_idx, finding in enumerate(findings, 2):
                            ws_findings.cell(row=f_idx, column=1, value=finding.get('title', ''))
                            ws_findings.cell(row=f_idx, column=2, value=finding.get('severity', ''))
                            ws_findings.cell(row=f_idx, column=3, value=finding.get('message', ''))
                            ws_findings.cell(row=f_idx, column=4, value=finding.get('action', ''))
                            
                            # Color by severity
                            if finding.get('severity') == 'high':
                                for col in range(1, 5):
                                    ws_findings.cell(row=f_idx, column=col).fill = error_fill
                            elif finding.get('severity') == 'medium':
                                for col in range(1, 5):
                                    ws_findings.cell(row=f_idx, column=col).fill = warning_fill
                    
                    # Save to bytes
                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    excel_base64 = base64.b64encode(excel_buffer.read()).decode('utf-8')
                    
                    filename = answer.structured_output.get('filename_suggestion', 'validation_export.xlsx')
                    
                    response["answer"] = f"📥 **Export Ready**\n\nI've prepared {export_data.get('total_records', 0)} records with {len(findings)} findings for download."
                    response["export"] = {
                        "filename": filename,
                        "data": excel_base64,
                        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    }
                    logger.warning(f"[UNIFIED] Export generated: {filename}")
                    
                except Exception as e:
                    logger.error(f"[UNIFIED] Export failed: {e}")
                    response["answer"] = f"Export generation failed: {str(e)}"
                
                return response
            
            # Check for simple answer (count, sum, etc.)
            # USE THE ENGINE'S CONSULTATIVE ANSWER DIRECTLY
            # The intelligence_engine._generate_consultative_response() already
            # builds the Three Truths response - don't override it with simple_answer
            engine_answer = answer.answer if answer.answer else None
            
            # Check if this is an analytical question that REQUIRES expert interpretation
            analytical_keywords = ['correct', 'valid', 'issue', 'problem', 'check', 'verify', 
                                   'audit', 'review', 'configure', 'setup', 'missing', 'wrong',
                                   'should', 'compliance', 'accurate', 'quality', 'rate', 'rates',
                                   'properly', 'right', 'okay', 'ok', 'good', 'bad', 'error']
            is_analytical = any(kw in message.lower() for kw in analytical_keywords)
            
            # DEBUG - always log this to see what's happening
            logger.warning(f"[UNIFIED] EXPERT CHECK: is_analytical={is_analytical}, EXPERT_AVAILABLE={EXPERT_CONTEXT_AVAILABLE}, msg='{message[:50]}'")
            
            # Check if engine already produced a validation analysis (has our OLD consultant formatting)
            # v4.0: Don't trigger on hybrid responses (those have the data listing first)
            # Only trigger on OLD-style responses that start with emojis
            is_validation_response = engine_answer and (
                # Original emoji-based detection - must START with these indicators
                engine_answer.strip().startswith(('🔴', '🟡', '✅', '**Rates', '**Rate', '**Zero'))
                # OR contains the old validation markers
                or any(marker in engine_answer for marker in ['Rates Valid', 'Rate Issues', 'Zero Rates', 'Years Old'])
            )
            # v4.0: Hybrid responses contain "🔍 Consultant Analysis" but should NOT trigger this bypass
            # because they already have the template + analysis concatenated
            if '🔍 Consultant Analysis' in engine_answer:
                is_validation_response = False
                logger.warning("[UNIFIED] Hybrid response detected - not bypassing")
            
            # Detect garbage SQL responses (just literals, no real data)
            garbage_indicators = [
                'configured correctly',  # LLM returned fake confirmation
                'no issues found',
                'everything looks',
                'appears to be',
            ]
            is_garbage_response = engine_answer and any(
                indicator in engine_answer.lower() for indicator in garbage_indicators
            )
            
            # Detect "no data" placeholder messages
            no_data_indicators = ['no data found', 'no matching', 'couldn\'t find', 'no results']
            is_no_data_placeholder = engine_answer and any(
                indicator in engine_answer.lower() for indicator in no_data_indicators
            )
            
            if is_validation_response:
                # Engine already did the consultant analysis - use it directly
                logger.warning(f"[UNIFIED] Using engine's validation analysis directly (has consultant formatting)")
                # Limit answer size to prevent browser choking
                safe_answer = engine_answer[:15000] if engine_answer else ""
                if auto_applied_note:
                    response["answer"] = auto_applied_note + "\n\n" + safe_answer
                else:
                    response["answer"] = safe_answer
                # Clear clarification flag - we have a real answer now
                response["needs_clarification"] = False
                response["clarification_questions"] = []
                # Keep sources intact but mark to skip heavy quality_alerts
                response["structured_output"] = {"type": "validation_complete"}
                response["_skip_quality_alerts"] = True  # Flag to skip quality check population
            elif is_analytical and EXPERT_CONTEXT_AVAILABLE:
                logger.info(f"[UNIFIED] Analytical question detected - using expert context")
                logger.info(f"[UNIFIED] from_reality count: {len(answer.from_reality) if answer.from_reality else 0}")
                logger.info(f"[UNIFIED] from_intent count: {len(answer.from_intent) if answer.from_intent else 0}")
                logger.info(f"[UNIFIED] is_garbage: {is_garbage_response}, is_no_data: {is_no_data_placeholder}")
                
                try:
                    project_id, project_domains = _get_project_domains(project, handler)
                    
                    # Build context from what we have
                    context_parts = [f"User Question: {message}", ""]
                    
                    # Include any real data we found from DuckDB (from_reality)
                    has_real_data = False
                    if answer.from_reality:
                        for truth in answer.from_reality:
                            if isinstance(truth.content, dict):
                                rows = truth.content.get('rows', [])
                                if rows:
                                    has_real_data = True
                                    context_parts.append(f"Data found in {truth.source_name}:")
                                    context_parts.append(f"Columns: {truth.content.get('columns', [])}")
                                    context_parts.append(f"Total rows: {len(rows)}")
                                    context_parts.append("")
                                    # Format rows as context - show all if < 100, else sample
                                    rows_to_show = rows if len(rows) < 100 else rows[:50]
                                    for row in rows_to_show:
                                        context_parts.append(f"  {row}")
                                    if len(rows) > len(rows_to_show):
                                        context_parts.append(f"  ... and {len(rows) - len(rows_to_show)} more rows")
                                    logger.info(f"[UNIFIED] Added {len(rows)} rows from {truth.source_name}")
                    
                    # =========================================================
                    # FALLBACK TO RAG (from_intent) when no DuckDB data
                    # This handles PDFs and documents stored in ChromaDB
                    # =========================================================
                    if not has_real_data and answer.from_intent:
                        logger.info(f"[UNIFIED] No DuckDB data - falling back to RAG documents ({len(answer.from_intent)} results)")
                        context_parts.append("Information found in uploaded documents:")
                        context_parts.append("")
                        
                        for truth in answer.from_intent[:5]:  # Limit to top 5 most relevant
                            source_name = getattr(truth, 'source_name', 'Document')
                            content = getattr(truth, 'content', '')
                            confidence = getattr(truth, 'confidence', 0.5)
                            
                            if content:
                                has_real_data = True
                                context_parts.append(f"From {source_name} (relevance: {confidence:.0%}):")
                                # Truncate very long content
                                content_preview = content[:1500] if len(content) > 1500 else content
                                context_parts.append(content_preview)
                                context_parts.append("")
                                logger.info(f"[UNIFIED] Added RAG content from {source_name} ({len(content)} chars)")
                    
                    if not has_real_data:
                        context_parts.append("No relevant data was found in the available tables or documents.")
                        context_parts.append("This could mean: the data hasn't been uploaded yet, or it's stored with different names.")
                        logger.warning(f"[UNIFIED] No real data found for analytical question")
                    
                    # Add table context
                    if schema.get('tables'):
                        table_names = [t.get('table_name', '').split('__')[-1][:40] for t in schema['tables'][:20]]
                        context_parts.append(f"\nAvailable tables: {', '.join(table_names)}")
                    
                    context = "\n".join(context_parts)
                    
                    synthesized, expert_context_used = await generate_synthesized_answer(
                        question=message,
                        context=context,
                        persona=request.persona,
                        insights=answer.insights,
                        conflicts=answer.conflicts,
                        citations=citation_builder,
                        quality_alerts=quality_service,
                        follow_ups=[],
                        redactor=redactor,
                        project=project,
                        project_domains=project_domains,
                    )
                    
                    if synthesized and len(synthesized) > 50:
                        if auto_applied_note:
                            response["answer"] = auto_applied_note + "\n\n" + synthesized
                        else:
                            response["answer"] = synthesized
                        session['last_expert_context'] = expert_context_used
                        if expert_context_used:
                            response["expert_context_id"] = expert_context_used
                        logger.info(f"[UNIFIED] Expert context: {expert_context_used}")
                    else:
                        # Expert context failed, fall through to normal flow
                        logger.warning(f"[UNIFIED] Expert synthesis too short, using engine answer")
                        if auto_applied_note:
                            response["answer"] = auto_applied_note + "\n\n" + engine_answer
                        else:
                            response["answer"] = engine_answer
                except Exception as e:
                    logger.warning(f"[UNIFIED] Expert context failed: {e}")
                    import traceback
                    logger.warning(traceback.format_exc())
                    if auto_applied_note:
                        response["answer"] = auto_applied_note + "\n\n" + engine_answer
                    else:
                        response["answer"] = engine_answer
                        
            elif engine_answer and len(engine_answer) > 50:
                # Engine gave us a proper consultative response
                if auto_applied_note:
                    response["answer"] = auto_applied_note + "\n\n" + engine_answer
                else:
                    response["answer"] = engine_answer
            else:
                # Fallback for edge cases - try simple extraction
                simple_answer = _try_simple_answer(answer, data_model)
                if simple_answer:
                    if auto_applied_note:
                        simple_answer = auto_applied_note + "\n\n" + simple_answer
                    response["answer"] = simple_answer
                else:
                    # Last resort - use Claude synthesis WITH INTELLIGENT CONTEXT
                    # Get project domains for intelligent context selection
                    project_id, project_domains = _get_project_domains(project, handler)
                    
                    synthesized, expert_context_used = await generate_synthesized_answer(
                        question=message,
                        context=answer.answer or "",
                        persona=request.persona,
                        insights=answer.insights,
                        conflicts=answer.conflicts,
                        citations=citation_builder,
                        quality_alerts=quality_service,
                        follow_ups=[],
                        redactor=redactor,
                        project=project,
                        project_domains=project_domains,
                    )
                    
                    # Track expert context for feedback loop
                    session['last_expert_context'] = expert_context_used
                    if expert_context_used:
                        response["expert_context_id"] = expert_context_used
                    
                    if auto_applied_note and synthesized:
                        response["answer"] = auto_applied_note + "\n\n" + synthesized
                    else:
                        response["answer"] = synthesized
            
            # Add citations
            if request.include_citations and answer.from_reality:
                for truth in answer.from_reality:
                    if isinstance(truth.content, dict):
                        citation_builder.add_citation(
                            claim=str(truth.content.get('rows', [])[:1]),
                            source_table=truth.source_name,
                            sql=truth.content.get('sql'),
                            row_count=len(truth.content.get('rows', []))
                        )
                response["citations"] = citation_builder.build_audit_trail()
            
            # Run quality checks (skip if validation response already handled it)
            if request.include_quality_alerts and quality_service and schema['tables'] and not response.get("_skip_quality_alerts"):
                quality_service.run_checks(handler, schema['tables'][:5])
                summary = quality_service.get_summary()
                # Limit findings to prevent frontend from choking
                if summary and 'findings' in summary:
                    summary['findings'] = summary['findings'][:50]  # Max 50 findings
                    summary['total_findings'] = summary.get('total_findings', len(summary['findings']))
                response["quality_alerts"] = summary
            
            # Remove internal flag before returning
            response.pop("_skip_quality_alerts", None)
            
            # Generate follow-ups
            if request.include_follow_ups:
                follow_up_gen = FollowUpGenerator(schema)
                query_type = _detect_query_type(message, answer)
                response["follow_up_suggestions"] = follow_up_gen.generate(
                    query_type=query_type,
                    question=message,
                    result={'answer': response["answer"]}
                )
        
        # Handle no answer case
        if response["answer"] is None and not response["needs_clarification"]:
            # Check if this is an analytical question that deserves expert guidance
            analytical_keywords = ['correct', 'valid', 'issue', 'problem', 'check', 'verify', 
                                   'audit', 'review', 'configure', 'setup', 'missing', 'wrong',
                                   'should', 'compliance', 'accurate', 'quality']
            is_analytical = any(kw in message.lower() for kw in analytical_keywords)
            
            if is_analytical and EXPERT_CONTEXT_AVAILABLE:
                # Even without data, try to provide expert guidance
                try:
                    project_id, project_domains = _get_project_domains(project, handler)
                    
                    # Build context from what we know
                    context_parts = []
                    if answer.from_reality:
                        context_parts.append("Available data tables were searched but no matching records found.")
                    if schema.get('tables'):
                        table_names = [t.get('table_name', '').split('__')[-1] for t in schema['tables'][:10]]
                        context_parts.append(f"Available tables include: {', '.join(table_names)}")
                    
                    context = "\n".join(context_parts) if context_parts else "No relevant data found in the project."
                    
                    synthesized, expert_context_used = await generate_synthesized_answer(
                        question=message,
                        context=context,
                        persona=request.persona,
                        insights=answer.insights if answer else [],
                        conflicts=answer.conflicts if answer else [],
                        citations=citation_builder,
                        quality_alerts=quality_service,
                        follow_ups=[],
                        redactor=redactor,
                        project=project,
                        project_domains=project_domains,
                    )
                    
                    if synthesized and len(synthesized) > 50:
                        response["answer"] = synthesized
                        response["confidence"] = 0.5
                        if expert_context_used:
                            response["expert_context_id"] = expert_context_used
                        session['last_expert_context'] = expert_context_used
                        logger.info(f"[UNIFIED] Used expert context for no-data analytical question")
                except Exception as e:
                    logger.warning(f"[UNIFIED] Expert guidance failed: {e}")
            
            # Fallback if expert guidance didn't work
            if response["answer"] is None:
                if answer and (answer.from_reality or answer.from_intent or answer.from_reference):
                    response["answer"] = "I found some related information but couldn't generate a complete answer. Please try rephrasing your question."
                else:
                    response["answer"] = "I couldn't find any data matching your query. Please check that data has been uploaded for this project."
                response["confidence"] = 0.3
        
        # Clarification message
        if response["needs_clarification"]:
            questions = response.get("clarification_questions", [])
            if questions:
                q_text = questions[0].get('question', 'I need more information')
                response["answer"] = f"Before I can answer, {q_text.lower()}"
            
            # v4.0: Store the original question so we can use it after clarification
            session['pending_clarification_question'] = message
            logger.warning(f"[UNIFIED] Stored pending question for clarification: {message[:50]}...")
        
        # Update session
        actual_sql = getattr(answer, 'executed_sql', None) or learned_sql
        session['last_sql'] = actual_sql
        session['last_result'] = response["answer"][:1000] if response.get("answer") else None
        session['last_question'] = message
        session['interaction_count'] = session.get('interaction_count', 0) + 1
        
        # Track lineage: sources → response
        try:
            from utils.database.models import LineageModel, ProjectModel
            import uuid
            
            # Only track if we have actual sources cited
            sources_tracked = 0
            response_id = session.get('session_id', str(uuid.uuid4())[:8]) + f"_{session.get('interaction_count', 0)}"
            
            # Get project_id
            project_record = ProjectModel.get_by_name(project)
            project_id = project_record.get('id') if project_record else None
            
            # Track Reality sources (DuckDB tables)
            if answer and answer.from_reality:
                for truth in answer.from_reality:
                    source_table = truth.source_name if hasattr(truth, 'source_name') else None
                    if source_table:
                        LineageModel.track(
                            source_type=LineageModel.NODE_TABLE,
                            source_id=source_table,
                            target_type=LineageModel.NODE_RESPONSE,
                            target_id=response_id,
                            relationship=LineageModel.REL_CITED,
                            project_id=project_id,
                            metadata={'query': message[:100], 'confidence': response.get('confidence', 0)}
                        )
                        sources_tracked += 1
            
            # Track Intent sources (ChromaDB chunks)
            if answer and answer.from_intent:
                for truth in answer.from_intent:
                    source_doc = truth.source_name if hasattr(truth, 'source_name') else None
                    if source_doc:
                        LineageModel.track(
                            source_type=LineageModel.NODE_CHUNK,
                            source_id=f"{source_doc}:chunks",
                            target_type=LineageModel.NODE_RESPONSE,
                            target_id=response_id,
                            relationship=LineageModel.REL_CITED,
                            project_id=project_id,
                            metadata={'query': message[:100], 'truth_type': 'intent'}
                        )
                        sources_tracked += 1
            
            # Track Reference sources
            if answer and answer.from_reference:
                for truth in answer.from_reference:
                    source_doc = truth.source_name if hasattr(truth, 'source_name') else None
                    if source_doc:
                        LineageModel.track(
                            source_type=LineageModel.NODE_CHUNK,
                            source_id=f"{source_doc}:chunks",
                            target_type=LineageModel.NODE_RESPONSE,
                            target_id=response_id,
                            relationship=LineageModel.REL_CITED,
                            project_id=project_id,
                            metadata={'query': message[:100], 'truth_type': 'reference'}
                        )
                        sources_tracked += 1
            
            # Track Regulatory sources
            if answer and answer.from_regulatory:
                for truth in answer.from_regulatory:
                    source_doc = truth.source_name if hasattr(truth, 'source_name') else None
                    if source_doc:
                        LineageModel.track(
                            source_type=LineageModel.NODE_CHUNK,
                            source_id=f"{source_doc}:chunks",
                            target_type=LineageModel.NODE_RESPONSE,
                            target_id=response_id,
                            relationship=LineageModel.REL_CITED,
                            project_id=project_id,
                            metadata={'query': message[:100], 'truth_type': 'regulatory'}
                        )
                        sources_tracked += 1
            
            if sources_tracked > 0:
                logger.debug(f"[UNIFIED] Tracked lineage: {sources_tracked} sources -> response_{response_id}")
                
        except Exception as lineage_error:
            logger.debug(f"[UNIFIED] Lineage tracking skipped: {lineage_error}")
        
        # Record query metrics
        query_duration_ms = int((time.time() - query_start_time) * 1000)
        if METRICS_AVAILABLE:
            try:
                # Count rows if we have data results
                rows_processed = 0
                if response.get("data") and isinstance(response["data"], dict):
                    rows_data = response["data"].get("rows", [])
                    rows_processed = len(rows_data) if isinstance(rows_data, list) else 0
                
                MetricsService.record_query(
                    query_type='chat',
                    duration_ms=query_duration_ms,
                    success=True,
                    project_id=project_id,
                    rows_processed=rows_processed
                )
                logger.debug(f"[UNIFIED] Recorded query metric: {query_duration_ms}ms, {rows_processed} rows")
            except Exception as metrics_err:
                logger.debug(f"[UNIFIED] Query metrics failed: {metrics_err}")
        
        # Cleanup old sessions
        cleanup_old_sessions()
        
        return response
        
    except Exception as e:
        # Record failed query
        query_duration_ms = int((time.time() - query_start_time) * 1000)
        if METRICS_AVAILABLE:
            try:
                MetricsService.record_query(
                    query_type='chat',
                    duration_ms=query_duration_ms,
                    success=False,
                    project_id=project_id if 'project_id' in dir() else None,
                    error_message=str(e)[:200]
                )
            except Exception:
                pass
        
        logger.error(f"[UNIFIED] Error: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(500, f"Chat error: {str(e)}")


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _serialize_truth(truth, data_model: Optional[DataModelService] = None) -> Dict:
    """Serialize a Truth object for JSON response."""
    content = truth.content
    
    # Handle data content (rows from DuckDB)
    if isinstance(content, dict) and 'rows' in content:
        rows = content.get('rows', [])
        
        # Enrich with data model if available
        if data_model and rows:
            columns = content.get('columns', [])
            rows = data_model.enrich_results(rows, columns)
        
        content = {
            'columns': content.get('columns', []),
            'rows': rows,  # Return ALL rows - no arbitrary limit
            'total': content.get('total', len(rows)),
            'sql': content.get('sql'),
            'query_type': content.get('query_type'),
            'resolution_path': content.get('resolution_path'),  # Provenance
            'reality_context': content.get('reality_context')   # Breakdowns
        }
    elif not isinstance(content, (dict, list, str, int, float, bool, type(None))):
        content = str(content)[:2000]
    
    return {
        "source_type": truth.source_type,
        "source_name": truth.source_name,
        "content": content,
        "confidence": truth.confidence,
        "location": truth.location
    }


def _try_simple_answer(answer, data_model: Optional[DataModelService] = None) -> Optional[str]:
    """Try to generate a simple direct answer without Claude."""
    if not answer.from_reality:
        return None
    
    for truth in answer.from_reality:
        if not isinstance(truth.content, dict):
            continue
        
        query_type = truth.content.get('query_type', '')
        rows = truth.content.get('rows', [])
        sql = truth.content.get('sql', '')
        cols = truth.content.get('columns', [])
        
        if query_type == 'count' and rows:
            count_val = list(rows[0].values())[0] if rows[0] else 0
            simple_answer = f"**{count_val:,}** employees match your criteria."
            if sql:
                simple_answer += f"\n\n*Query executed:* `{sql[:300]}`"
            return simple_answer
        
        elif query_type in ['sum', 'average'] and rows:
            result_val = list(rows[0].values())[0] if rows[0] else 0
            label = "Total" if query_type == 'sum' else "Average"
            simple_answer = f"**{label}: {result_val:,.2f}**"
            if sql:
                simple_answer += f"\n\n*Query executed:* `{sql[:300]}`"
            return simple_answer
        
        elif query_type == 'list' and rows and cols:
            # Enrich rows if data model available
            if data_model:
                rows = data_model.enrich_results(rows, cols)
            
            # Format as markdown table
            table_lines = []
            display_cols = cols[:6]
            table_lines.append("| " + " | ".join(display_cols) + " |")
            table_lines.append("|" + "|".join(["---"] * len(display_cols)) + "|")
            
            for row in rows[:20]:
                vals = [str(row.get(c, ''))[:30] for c in display_cols]
                table_lines.append("| " + " | ".join(vals) + " |")
            
            simple_answer = f"**Found {len(rows)} results:**\n\n" + "\n".join(table_lines)
            
            if len(rows) > 20:
                simple_answer += f"\n\n*Showing first 20 of {len(rows)} results*"
            if sql:
                simple_answer += f"\n\n*Query:* `{sql[:200]}`"
            
            return simple_answer
    
    return None


def _detect_query_type(question: str, answer) -> str:
    """Detect the type of query for follow-up generation."""
    question_lower = question.lower()
    
    if any(word in question_lower for word in ['how many', 'count', 'total number']):
        return 'count'
    elif any(word in question_lower for word in ['list', 'show me', 'who are', 'which']):
        return 'list'
    elif any(word in question_lower for word in ['sum', 'total', 'add up']):
        return 'sum'
    elif any(word in question_lower for word in ['analyze', 'compare', 'trend', 'pattern']):
        return 'analysis'
    else:
        return 'general'


def _decrypt_results(rows: List[Dict], handler) -> List[Dict]:
    """Decrypt encrypted fields in query results."""
    if not rows:
        return rows
    
    try:
        encryptor = getattr(handler, 'encryptor', None)
        if not encryptor:
            return rows
        
        decrypted_rows = []
        for row in rows:
            decrypted_row = {}
            for key, value in row.items():
                if isinstance(value, str) and (value.startswith('ENC:') or value.startswith('ENC256:')):
                    try:
                        decrypted_row[key] = encryptor.decrypt(value)
                    except Exception:
                        decrypted_row[key] = '[encrypted]'
                else:
                    decrypted_row[key] = value
            decrypted_rows.append(decrypted_row)
        
        return decrypted_rows
        
    except Exception as e:
        logger.warning(f"[DECRYPT] Error: {e}")
        return rows


# =============================================================================
# SESSION ENDPOINTS
# =============================================================================

@router.get("/chat/unified/session/{session_id}")
async def get_session(session_id: str):
    """Get current state of a unified chat session."""
    if session_id not in unified_sessions:
        raise HTTPException(404, "Session not found")
    
    session = unified_sessions[session_id]
    engine = session.get('engine')
    
    return {
        "session_id": session_id,
        "project": session.get('project'),
        "confirmed_facts": dict(engine.confirmed_facts) if engine else {},
        "conversation_length": len(session.get('conversation_history', [])),
        "last_question": session.get('last_question'),
        "created_at": session.get('created_at')
    }


@router.delete("/chat/unified/session/{session_id}")
async def end_session(session_id: str):
    """End a unified chat session."""
    if session_id in unified_sessions:
        del unified_sessions[session_id]
    return {"success": True}


@router.post("/chat/unified/clarify")
async def submit_clarification(request: ClarificationAnswer):
    """Submit answers to clarification questions."""
    if request.session_id not in unified_sessions:
        raise HTTPException(404, "Session not found")
    
    session = unified_sessions[request.session_id]
    engine = session.get('engine')
    
    if not engine:
        raise HTTPException(400, "Session has no active engine")
    
    # Store confirmed facts
    engine.confirmed_facts.update(request.answers)
    
    # Re-ask the original question
    answer = engine.ask(request.original_question)
    
    return {
        "session_id": request.session_id,
        "question": answer.question,
        "confidence": answer.confidence,
        "answer": answer.answer,
        "needs_clarification": False
    }


# =============================================================================
# PREFERENCES ENDPOINTS
# =============================================================================

@router.post("/chat/unified/reset-preferences")
async def reset_preferences(request: ResetPreferencesRequest):
    """Reset user preferences/filters."""
    result = {
        "success": False,
        "reset_type": request.reset_type,
        "message": ""
    }
    
    try:
        if request.reset_type == "session":
            if request.session_id and request.session_id in unified_sessions:
                session = unified_sessions[request.session_id]
                engine = session.get('engine')
                if engine:
                    old_facts = dict(engine.confirmed_facts)
                    engine.confirmed_facts.clear()
                    session['skip_learning'] = True
                    result["success"] = True
                    result["message"] = f"Cleared session facts: {old_facts}"
                    result["cleared_facts"] = old_facts
                else:
                    result["message"] = "Session found but no engine"
            else:
                result["message"] = "Session not found"
        
        elif request.reset_type == "learned":
            if LEARNING_AVAILABLE and SUPABASE_AVAILABLE:
                try:
                    supabase = get_supabase()
                    supabase.table('clarification_choices').delete().eq(
                        'project', request.project
                    ).execute()
                    result["success"] = True
                    result["message"] = "Cleared learned preferences"
                except Exception as e:
                    result["message"] = f"Database error: {e}"
            else:
                result["message"] = "Learning/Supabase not available"
        
        else:
            result["message"] = f"Unknown reset_type: {request.reset_type}"
            
    except Exception as e:
        result["message"] = f"Error: {e}"
        logger.error(f"[RESET] Error: {e}")
    
    return result


@router.get("/chat/unified/preferences")
async def get_preferences(session_id: str = None, project: str = None):
    """Get current preferences for debugging."""
    result = {
        "session_facts": {},
        "learned_preferences": {},
        "session_exists": False
    }
    
    if session_id and session_id in unified_sessions:
        session = unified_sessions[session_id]
        engine = session.get('engine')
        if engine:
            result["session_facts"] = dict(engine.confirmed_facts)
            result["session_exists"] = True
    
    if LEARNING_AVAILABLE and SUPABASE_AVAILABLE and project:
        try:
            supabase = get_supabase()
            choices = supabase.table('clarification_choices').select('*').eq(
                'project', project
            ).order('created_at', desc=True).limit(20).execute()
            
            if choices.data:
                result["learned_preferences"] = {
                    c['question_id']: c['chosen_option']
                    for c in choices.data
                }
        except Exception as e:
            result["learning_error"] = str(e)
    
    return result


# =============================================================================
# FEEDBACK ENDPOINT
# =============================================================================

@router.post("/chat/unified/feedback")
async def submit_feedback(request: FeedbackRequest):
    """Submit feedback for learning."""
    try:
        if not LEARNING_AVAILABLE:
            return {"success": False, "error": "Learning not available"}
        
        learning = get_learning_module()
        
        # Record feedback
        if hasattr(learning, 'record_feedback'):
            learning.record_feedback(
                question=request.message,
                response=request.response,
                feedback=request.feedback,
                correction=request.correction
            )
        
        return {"success": True, "feedback": request.feedback}
        
    except Exception as e:
        logger.error(f"[FEEDBACK] Error: {e}")
        return {"success": False, "error": str(e)}


@router.post("/chat/unified/expert-feedback")
async def submit_expert_feedback(request: Request):
    """
    Submit feedback on expert context selection.
    
    This closes the learning loop - good/bad selections improve future matching.
    Call this when user indicates the response quality (thumbs up/down).
    """
    if not EXPERT_CONTEXT_AVAILABLE:
        return {"success": False, "error": "Expert context system not available"}
    
    try:
        data = await request.json()
        context_id = data.get('context_id')
        feedback = data.get('feedback')  # 'positive' or 'negative'
        
        if not context_id:
            return {"success": False, "error": "context_id required"}
        
        if feedback not in ['positive', 'negative']:
            return {"success": False, "error": "feedback must be 'positive' or 'negative'"}
        
        record_expert_feedback(context_id, feedback)
        
        logger.info(f"[UNIFIED] Expert feedback recorded: {context_id} = {feedback}")
        return {"success": True, "context_id": context_id, "feedback": feedback}
        
    except Exception as e:
        logger.error(f"[UNIFIED] Expert feedback error: {e}")
        return {"success": False, "error": str(e)}


@router.get("/chat/unified/expert-contexts")
async def list_expert_contexts():
    """
    List all available expert contexts.
    
    Useful for debugging and understanding what contexts are available.
    """
    if not EXPERT_CONTEXT_AVAILABLE:
        return {"contexts": [], "available": False}
    
    try:
        registry = get_expert_registry()
        contexts = registry.get_all()
        return {
            "contexts": [c.to_dict() for c in contexts],
            "available": True,
            "count": len(contexts)
        }
    except Exception as e:
        logger.error(f"[UNIFIED] List expert contexts error: {e}")
        return {"contexts": [], "available": True, "error": str(e)}


# =============================================================================
# PERSONA ENDPOINTS
# =============================================================================

@router.get("/chat/unified/personas")
async def list_personas():
    """List all available personas."""
    if not PERSONAS_AVAILABLE:
        return {"personas": [], "default": None}
    
    try:
        pm = get_persona_manager()
        return {"personas": pm.list_personas(), "default": None}
    except Exception as e:
        logger.error(f"[PERSONAS] List error: {e}")
        return {"personas": [], "default": None, "error": str(e)}


@router.get("/chat/unified/personas/{name}")
async def get_persona(name: str):
    """Get specific persona."""
    if not PERSONAS_AVAILABLE:
        raise HTTPException(503, "Personas not available")
    
    try:
        pm = get_persona_manager()
        persona = pm.get_persona(name)
        if persona:
            return persona.to_dict()
        raise HTTPException(404, f"Persona '{name}' not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@router.post("/chat/unified/personas")
async def create_persona(request: Request):
    """Create a new persona."""
    if not PERSONAS_AVAILABLE:
        raise HTTPException(503, "Personas not available")
    
    try:
        data = await request.json()
        pm = get_persona_manager()
        persona = pm.create_persona(
            name=data['name'],
            icon=data.get('icon', '🤖'),
            description=data.get('description', ''),
            system_prompt=data.get('system_prompt', ''),
            expertise=data.get('expertise', []),
            tone=data.get('tone', 'Professional')
        )
        return {"success": True, "persona": persona.to_dict()}
    except Exception as e:
        logger.error(f"[PERSONAS] Create error: {e}")
        raise HTTPException(500, str(e))


@router.put("/chat/unified/personas/{name}")
async def update_persona(name: str, request: Request):
    """Update an existing persona."""
    if not PERSONAS_AVAILABLE:
        raise HTTPException(503, "Personas not available")
    
    try:
        data = await request.json()
        pm = get_persona_manager()
        persona = pm.update_persona(name, **data)
        if persona:
            return {"success": True, "persona": persona.to_dict()}
        raise HTTPException(404, f"Persona '{name}' not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@router.delete("/chat/unified/personas/{name}")
async def delete_persona(name: str):
    """Delete a persona."""
    if not PERSONAS_AVAILABLE:
        raise HTTPException(503, "Personas not available")
    
    try:
        pm = get_persona_manager()
        success = pm.delete_persona(name)
        if success:
            return {"success": True}
        raise HTTPException(404, f"Persona '{name}' not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# =============================================================================
# EXCEL EXPORT ENDPOINT
# =============================================================================

@router.post("/chat/unified/export-excel")
async def export_to_excel(request: ExportRequest):
    """Export query results to Excel with quality report."""
    if not STRUCTURED_AVAILABLE:
        raise HTTPException(503, "Structured queries not available")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        handler = get_structured_handler()
        project = request.project
        
        # Get tables for project
        tables_list = handler.get_tables(project) if project else []
        
        if not tables_list:
            raise HTTPException(404, "No tables found for this project")
        
        # Select tables to export
        tables_to_query = []
        if request.tables:
            for t in tables_list:
                if t.get('table_name') in request.tables:
                    tables_to_query.append(t)
        else:
            # Use first 3 tables
            tables_to_query = tables_list[:3]
        
        # Create workbook
        wb = Workbook()
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary tab (if requested)
        if request.include_summary:
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            ws_summary.cell(row=1, column=1, value="XLR8 Data Export").font = Font(bold=True, size=14)
            ws_summary.cell(row=2, column=1, value=f"Project: {project or 'All'}")
            ws_summary.cell(row=3, column=1, value=f"Query: {request.query}")
            ws_summary.cell(row=4, column=1, value=f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}")
            ws_summary.cell(row=6, column=1, value="Tables Included:").font = Font(bold=True)
            
            for i, t in enumerate(tables_to_query, 1):
                ws_summary.cell(row=6+i, column=1, value=f"• {t.get('table_name', '')}")
        else:
            wb.remove(wb.active)
        
        # Data tabs
        sources_info = []
        for table_info in tables_to_query[:5]:
            table_name = table_info.get('table_name', '')
            sheet_name = table_info.get('sheet', table_name)[:31]  # Excel limit
            
            try:
                sql = f'SELECT * FROM "{table_name}"'
                rows = handler.query(sql)
                cols = list(rows[0].keys()) if rows else []
                
                if not rows:
                    continue
                
                decrypted = _decrypt_results(rows, handler)
                
                ws = wb.create_sheet(title=sheet_name)
                
                # Headers
                for col_idx, col_name in enumerate(cols, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                
                # Data
                for row_idx, row_data in enumerate(decrypted[:5000], 2):
                    for col_idx, col_name in enumerate(cols, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
                        cell.border = thin_border
                
                # Auto-width
                for col_idx in range(1, min(len(cols) + 1, 20)):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
                
                sources_info.append({
                    'table': table_name,
                    'rows': len(decrypted),
                    'columns': len(cols)
                })
                
            except Exception as e:
                logger.error(f"[EXPORT] Error on {table_name}: {e}")
        
        # Quality report tab (if requested)
        if request.include_quality_report and project:
            try:
                quality_service = DataQualityService(project)
                quality_service.run_checks(handler, tables_to_query)
                alerts = quality_service.alerts
                
                if alerts:
                    ws_quality = wb.create_sheet(title="Quality Report")
                    
                    headers = ["Severity", "Issue", "Table", "Count", "Description"]
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws_quality.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                    
                    for row_idx, alert in enumerate(alerts, 2):
                        ws_quality.cell(row=row_idx, column=1, value=alert['severity'].upper())
                        ws_quality.cell(row=row_idx, column=2, value=alert['name'])
                        ws_quality.cell(row=row_idx, column=3, value=alert['table'])
                        ws_quality.cell(row=row_idx, column=4, value=alert['count'])
                        ws_quality.cell(row=row_idx, column=5, value=alert['description'])
                        
            except Exception as qe:
                logger.warning(f"[EXPORT] Quality report error: {qe}")
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        safe_query = re.sub(r'[^\w\s]', '', request.query[:20]).strip().replace(' ', '_')
        filename = f"xlr8_export_{safe_query}_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[EXPORT] Error: {e}")
        raise HTTPException(500, f"Export failed: {str(e)}")


# =============================================================================
# DATA INSPECTION ENDPOINTS
# =============================================================================

@router.get("/chat/unified/data/{project}")
async def get_project_data_summary(project: str):
    """Get summary of available data for a project."""
    if not STRUCTURED_AVAILABLE:
        raise HTTPException(503, "Structured queries not available")
    
    try:
        handler = get_structured_handler()
        tables = handler.get_tables(project)
        
        return {
            "project": project,
            "tables": len(tables),
            "summary": [
                {
                    "file": t.get('file'),
                    "sheet": t.get('sheet_name'),
                    "rows": t.get('row_count'),
                    "columns": len(t.get('columns', []))
                }
                for t in tables
            ]
        }
    except Exception as e:
        raise HTTPException(500, str(e))


@router.get("/chat/unified/data/{project}/{file_name}/versions")
async def get_file_versions(project: str, file_name: str):
    """Get version history for a file."""
    if not STRUCTURED_AVAILABLE:
        raise HTTPException(503, "Not available")
    
    try:
        handler = get_structured_handler()
        return {"versions": handler.get_file_versions(project, file_name)}
    except Exception as e:
        raise HTTPException(500, str(e))


class CompareRequest(BaseModel):
    sheet_name: str
    key_column: str
    version1: Optional[int] = None
    version2: Optional[int] = None


@router.post("/chat/unified/data/{project}/{file_name}/compare")
async def compare_versions(project: str, file_name: str, request: CompareRequest):
    """Compare two versions of a file."""
    if not STRUCTURED_AVAILABLE:
        raise HTTPException(503, "Not available")
    
    try:
        handler = get_structured_handler()
        result = handler.compare_versions(
            project=project,
            file_name=file_name,
            sheet_name=request.sheet_name,
            key_column=request.key_column,
            version1=request.version1,
            version2=request.version2
        )
        if 'error' in result:
            raise HTTPException(400, result['error'])
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# =============================================================================
# DIAGNOSTICS ENDPOINTS
# =============================================================================

@router.get("/chat/unified/diagnostics")
async def get_diagnostics(project: str = None):
    """Get diagnostic information about the chat system."""
    result = {
        "intelligence_available": INTELLIGENCE_AVAILABLE,
        "learning_available": LEARNING_AVAILABLE,
        "structured_available": STRUCTURED_AVAILABLE,
        "rag_available": RAG_AVAILABLE,
        "llm_available": LLM_AVAILABLE,
        "personas_available": PERSONAS_AVAILABLE,
        "expert_context_available": EXPERT_CONTEXT_AVAILABLE,
        "domain_inference_available": DOMAIN_INFERENCE_AVAILABLE,
        "active_sessions": len(unified_sessions),
        "profile_status": None,
        "sample_profiles": [],
        "detected_domains": None
    }
    
    if STRUCTURED_AVAILABLE:
        try:
            handler = get_structured_handler()
            
            # Check for _column_profiles table
            try:
                count = handler.conn.execute("SELECT COUNT(*) FROM _column_profiles").fetchone()[0]
                result["profile_status"] = f"✅ {count} column profiles"
            except Exception:
                result["profile_status"] = "❌ No column profiles - run backfill"
            
            # Sample profiles
            if project:
                try:
                    samples = handler.conn.execute("""
                        SELECT table_name, column_name, inferred_type, distinct_count, is_categorical, filter_category
                        FROM _column_profiles 
                        WHERE table_name LIKE ?
                        LIMIT 15
                    """, [f"{project}%"]).fetchall()
                    
                    result["sample_profiles"] = [
                        {
                            "table": s[0].split('__')[-1],
                            "column": s[1],
                            "type": s[2],
                            "distinct_count": s[3],
                            "is_categorical": s[4],
                            "filter_category": s[5]
                        }
                        for s in samples
                    ]
                except Exception:
                    pass
                
                # Get detected domains for project
                if project and DOMAIN_INFERENCE_AVAILABLE:
                    try:
                        project_id, domains = _get_project_domains(project, handler)
                        if domains:
                            result["detected_domains"] = {
                                "project_id": project_id,
                                "domains": domains[:5],  # Top 5
                                "primary": domains[0].get('domain') if domains else None
                            }
                    except Exception as de:
                        result["domain_error"] = str(de)
                    
        except Exception as e:
            result["diagnostics_error"] = str(e)
    
    return result


@router.post("/chat/unified/diagnostics/backfill")
async def run_backfill(project: str = None):
    """Run profile backfill for existing tables."""
    if not STRUCTURED_AVAILABLE:
        raise HTTPException(503, "Structured queries not available")
    
    try:
        handler = get_structured_handler()
        
        if not hasattr(handler, 'backfill_profiles'):
            return {"error": "backfill_profiles method not found - deploy latest structured_data_handler.py"}
        
        result = handler.backfill_profiles(project)
        return result
        
    except Exception as e:
        raise HTTPException(500, str(e))


# =============================================================================
# HEALTH & STATS
# =============================================================================

@router.get("/chat/unified/health")
async def health_check():
    """Health check endpoint."""
    return {
        "status": "healthy",
        "version": "1.1.0",  # Phase 3.5
        "intelligence": INTELLIGENCE_AVAILABLE,
        "project_intelligence": PROJECT_INTELLIGENCE_AVAILABLE,
        "learning": LEARNING_AVAILABLE,
        "structured": STRUCTURED_AVAILABLE,
        "rag": RAG_AVAILABLE,
        "sessions": len(unified_sessions)
    }


@router.get("/chat/unified/stats")
async def get_stats():
    """Get chat system statistics."""
    stats = {
        "active_sessions": len(unified_sessions),
        "intelligence_available": INTELLIGENCE_AVAILABLE,
        "project_intelligence_available": PROJECT_INTELLIGENCE_AVAILABLE,
        "learning_available": LEARNING_AVAILABLE,
        "structured_available": STRUCTURED_AVAILABLE,
        "rag_available": RAG_AVAILABLE
    }
    
    if LEARNING_AVAILABLE:
        try:
            learning = get_learning_module()
            if hasattr(learning, 'get_stats'):
                stats["learning_stats"] = learning.get_stats()
        except Exception:
            pass
    
    return stats


@router.get("/chat/intelligent/learning/stats")
async def get_learning_stats():
    """
    Get learning system statistics.
    
    This endpoint is called by the Admin UI Learning page.
    Provides counts for learned queries, feedback, preferences, and patterns.
    """
    if not LEARNING_AVAILABLE:
        return {
            'available': False,
            'learned_queries': 0,
            'feedback_records': 0,
            'user_preferences': 0,
            'clarification_patterns': 0
        }
    
    try:
        learning = get_learning_module()
        return learning.get_learning_stats()
    except Exception as e:
        logger.error(f"[LEARNING] Stats error: {e}")
        return {
            'available': False,
            'error': str(e)
        }


# =============================================================================
# BACKWARD COMPATIBILITY REDIRECTS
# =============================================================================

@router.get("/chat/models")
async def get_chat_models():
    """
    Get available chat models and their status.
    Called by frontend Chat component on load.
    """
    models = {
        "primary": "mistral:7b",
        "fallback": "claude-sonnet-4",
        "sql": "deepseek-coder:6.7b",
        "embeddings": "nomic-embed-text"
    }
    
    status = {
        "ollama_available": False,
        "claude_available": False,
        "models": models
    }
    
    # Check Ollama
    try:
        import os
        ollama_url = os.getenv("LLM_ENDPOINT", "")
        if ollama_url:
            import httpx
            async with httpx.AsyncClient(timeout=2.0) as client:
                r = await client.get(f"{ollama_url.rstrip('/')}/api/tags")
                if r.status_code == 200:
                    status["ollama_available"] = True
                    data = r.json()
                    status["ollama_models"] = [m.get("name") for m in data.get("models", [])]
    except Exception:
        pass
    
    # Check Claude
    try:
        import os
        if os.getenv("CLAUDE_API_KEY") or os.getenv("ANTHROPIC_API_KEY"):
            status["claude_available"] = True
    except Exception:
        pass
    
    return status


@router.post("/chat/intelligent")
async def intelligent_chat_redirect(request: UnifiedChatRequest):
    """Redirect old /chat/intelligent to unified endpoint."""
    logger.info("[REDIRECT] /chat/intelligent → /chat/unified")
    return await unified_chat(request)


# =============================================================================
# END OF UNIFIED CHAT ROUTER
# =============================================================================
