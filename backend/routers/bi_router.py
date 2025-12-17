"""
XLR8 BI ROUTER - Natural Language Analytics
=============================================

Deploy to: backend/routers/bi_router.py

FEATURES:
- Natural language → SQL via IntelligenceEngine
- Smart chart type recommendation based on result shape
- Export with user-friendly transforms
- Suggestions from filter_candidates + intelligence findings
- Reusable by playbooks and other components

ENDPOINTS:
- POST /api/bi/query - NL query → SQL → results + chart
- GET /api/bi/suggestions/{project} - Smart query suggestions
- POST /api/bi/export - Export with transforms
- GET /api/bi/schema/{project} - Schema for project
- GET /api/bi/saved/{project} - Saved queries/reports

Author: XLR8 Team
Version: 1.0.0
"""

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Dict, List, Optional, Any
import logging
import json
import re
import time
import io

logger = logging.getLogger(__name__)

router = APIRouter(tags=["bi"])


# =============================================================================
# IMPORTS - Graceful degradation
# =============================================================================

try:
    from utils.intelligence_engine import IntelligenceEngine, IntelligenceMode
    INTELLIGENCE_AVAILABLE = True
except ImportError:
    try:
        from backend.utils.intelligence_engine import IntelligenceEngine, IntelligenceMode
        INTELLIGENCE_AVAILABLE = True
    except ImportError:
        INTELLIGENCE_AVAILABLE = False
        logger.warning("[BI] Intelligence engine not available")

try:
    from utils.structured_data_handler import get_structured_handler
    STRUCTURED_AVAILABLE = True
except ImportError:
    try:
        from backend.utils.structured_data_handler import get_structured_handler
        STRUCTURED_AVAILABLE = True
    except ImportError:
        STRUCTURED_AVAILABLE = False
        logger.warning("[BI] Structured data handler not available")

try:
    from utils.rag_handler import RAGHandler
    RAG_AVAILABLE = True
except ImportError:
    try:
        from backend.utils.rag_handler import RAGHandler
        RAG_AVAILABLE = True
    except ImportError:
        RAG_AVAILABLE = False

try:
    from utils.database.supabase_client import get_supabase
    SUPABASE_AVAILABLE = True
except ImportError:
    try:
        from backend.utils.database.supabase_client import get_supabase
        SUPABASE_AVAILABLE = True
    except ImportError:
        SUPABASE_AVAILABLE = False


# =============================================================================
# REQUEST/RESPONSE MODELS
# =============================================================================

class BIQueryRequest(BaseModel):
    """Natural language BI query request."""
    query: str
    project: str
    filters: Optional[Dict[str, Any]] = None
    session_id: Optional[str] = None


class TransformOperation(BaseModel):
    """A single transform operation."""
    type: str  # rename, map_lookup, combine, split, format_currency, etc.
    column: str
    params: Optional[Dict[str, Any]] = None


class BIExportRequest(BaseModel):
    """Export request with transforms."""
    query: str
    project: str
    sql: Optional[str] = None  # Pre-generated SQL (skip NL processing)
    transforms: Optional[List[TransformOperation]] = []
    format: str = "xlsx"  # xlsx, csv
    include_metadata: bool = True


class SavedQueryRequest(BaseModel):
    """Save a query for reuse."""
    name: str
    query: str
    project: str
    sql: str
    chart_type: Optional[str] = None
    transforms: Optional[List[Dict]] = []


# =============================================================================
# CHART TYPE RECOMMENDATION
# =============================================================================

def recommend_chart_type(columns: List[str], data: List[Dict], sql: str) -> Dict[str, Any]:
    """
    Recommend chart type based on query results.
    
    Returns:
        {
            'recommended': 'bar',
            'alternatives': ['pie', 'table'],
            'config': {
                'xAxis': 'department',
                'yAxis': 'count',
                'series': []
            }
        }
    """
    if not data or not columns:
        return {'recommended': 'table', 'alternatives': [], 'config': {}}
    
    sql_upper = sql.upper() if sql else ''
    num_rows = len(data)
    num_cols = len(columns)
    
    # Detect column types from first row
    numeric_cols = []
    categorical_cols = []
    date_cols = []
    
    sample = data[0] if data else {}
    for col in columns:
        val = sample.get(col)
        if val is None:
            continue
        if isinstance(val, (int, float)):
            numeric_cols.append(col)
        elif isinstance(val, str):
            # Check for date patterns
            if re.match(r'\d{4}-\d{2}-\d{2}', str(val)):
                date_cols.append(col)
            else:
                categorical_cols.append(col)
    
    # Determine chart type
    config = {}
    
    # COUNT/SUM with GROUP BY → Bar or Pie
    if ('COUNT(' in sql_upper or 'SUM(' in sql_upper) and 'GROUP BY' in sql_upper:
        if len(categorical_cols) >= 1 and len(numeric_cols) >= 1:
            config['xAxis'] = categorical_cols[0]
            config['yAxis'] = numeric_cols[0]
            
            if num_rows <= 6:
                return {'recommended': 'pie', 'alternatives': ['bar', 'table'], 'config': config}
            elif num_rows <= 20:
                return {'recommended': 'bar', 'alternatives': ['pie', 'table'], 'config': config}
            else:
                return {'recommended': 'bar', 'alternatives': ['table'], 'config': config}
    
    # Time series (date column + numeric)
    if date_cols and numeric_cols:
        config['xAxis'] = date_cols[0]
        config['yAxis'] = numeric_cols[0]
        return {'recommended': 'line', 'alternatives': ['area', 'bar', 'table'], 'config': config}
    
    # Two numeric columns → Scatter
    if len(numeric_cols) >= 2:
        config['xAxis'] = numeric_cols[0]
        config['yAxis'] = numeric_cols[1]
        return {'recommended': 'scatter', 'alternatives': ['table'], 'config': config}
    
    # Single aggregate (COUNT(*) without GROUP BY)
    if num_rows == 1 and num_cols == 1:
        return {'recommended': 'metric', 'alternatives': ['table'], 'config': {'value': columns[0]}}
    
    # Default to table for complex/list results
    return {'recommended': 'table', 'alternatives': ['bar'], 'config': config}


# =============================================================================
# TRANSFORM ENGINE
# =============================================================================

# US State code → name mapping (for state_names transform)
STATE_NAMES = {
    'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas',
    'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
    'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii', 'ID': 'Idaho',
    'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa', 'KS': 'Kansas',
    'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
    'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi',
    'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada',
    'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NY': 'New York',
    'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio', 'OK': 'Oklahoma',
    'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
    'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah',
    'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia',
    'WI': 'Wisconsin', 'WY': 'Wyoming', 'DC': 'District of Columbia'
}


def apply_transforms(data: List[Dict], columns: List[str], transforms: List[TransformOperation], handler=None, project: str = None) -> tuple:
    """
    Apply user-defined transforms to data.
    
    Returns: (transformed_data, new_columns, column_renames)
    """
    if not transforms:
        return data, columns, {}
    
    result = [row.copy() for row in data]
    new_columns = list(columns)
    column_renames = {}  # original -> display name
    
    # Load lookups if needed
    lookups = {}
    if handler and project:
        try:
            # Get lookup tables from schema
            schema = _build_bi_schema(handler, project) if handler else {'tables': []}
            for table in schema.get('tables', []):
                table_name = table.get('table_name', '').lower()
                if any(p in table_name for p in ['lookup', 'code', 'ref', '_lkp']):
                    # Load as lookup
                    cols = table.get('columns', [])
                    if len(cols) >= 2:
                        code_col = cols[0]
                        desc_col = cols[1] if len(cols) > 1 else cols[0]
                        try:
                            sql = f'SELECT "{code_col}", "{desc_col}" FROM "{table.get("table_name")}"'
                            rows = handler.query(sql)
                            lookups[table_name] = {str(r[code_col]): str(r[desc_col]) for r in rows}
                        except:
                            pass
        except Exception as e:
            logger.warning(f"[BI] Could not load lookups: {e}")
    
    for transform in transforms:
        t_type = transform.type
        col = transform.column
        params = transform.params or {}
        
        if t_type == 'rename':
            # Rename column header
            new_name = params.get('new_name', col)
            column_renames[col] = new_name
        
        elif t_type == 'state_names':
            # Convert state codes to full names
            for row in result:
                if col in row and row[col]:
                    code = str(row[col]).upper().strip()
                    row[col] = STATE_NAMES.get(code, row[col])
        
        elif t_type == 'format_currency':
            # Format as $X,XXX.XX
            for row in result:
                if col in row and row[col] is not None:
                    try:
                        val = float(row[col])
                        row[col] = f"${val:,.2f}"
                    except (ValueError, TypeError):
                        pass
        
        elif t_type == 'format_percent':
            # Format as XX.X%
            for row in result:
                if col in row and row[col] is not None:
                    try:
                        val = float(row[col])
                        row[col] = f"{val * 100:.1f}%"
                    except (ValueError, TypeError):
                        pass
        
        elif t_type == 'uppercase':
            for row in result:
                if col in row and row[col]:
                    row[col] = str(row[col]).upper()
        
        elif t_type == 'titlecase':
            for row in result:
                if col in row and row[col]:
                    row[col] = str(row[col]).title()
        
        elif t_type == 'map_lookup':
            # Map codes to descriptions using lookup table
            lookup_table = params.get('lookup_table', '').lower()
            if lookup_table in lookups:
                lookup = lookups[lookup_table]
                for row in result:
                    if col in row and row[col]:
                        code = str(row[col])
                        row[col] = lookup.get(code, code)
        
        elif t_type == 'combine':
            # Combine multiple columns
            source_cols = params.get('columns', [])
            separator = params.get('separator', ' ')
            new_col = params.get('new_column', f"{col}_combined")
            
            for row in result:
                parts = [str(row.get(c, '')) for c in source_cols if row.get(c)]
                row[new_col] = separator.join(parts)
            
            if new_col not in new_columns:
                new_columns.append(new_col)
        
        elif t_type == 'split':
            # Split column by delimiter
            delimiter = params.get('delimiter', ',')
            new_col_names = params.get('new_columns', [f"{col}_1", f"{col}_2"])
            
            for row in result:
                if col in row and row[col]:
                    parts = str(row[col]).split(delimiter)
                    for i, new_col in enumerate(new_col_names):
                        row[new_col] = parts[i].strip() if i < len(parts) else ''
            
            for nc in new_col_names:
                if nc not in new_columns:
                    new_columns.append(nc)
        
        elif t_type == 'calculated':
            # Add calculated column
            formula = params.get('formula', '')
            new_col = params.get('new_column', 'calculated')
            
            # Simple formula parsing (col1 + col2, col1 * 100, etc.)
            for row in result:
                try:
                    # Replace column names with values
                    expr = formula
                    for c in columns:
                        if c in expr:
                            val = row.get(c, 0)
                            if val is None:
                                val = 0
                            expr = expr.replace(c, str(float(val)))
                    row[new_col] = eval(expr)  # Safe because we control the formula
                except:
                    row[new_col] = None
            
            if new_col not in new_columns:
                new_columns.append(new_col)
        
        elif t_type == 'filter':
            # Filter rows
            filter_col = params.get('column', col)
            operator = params.get('operator', '=')
            value = params.get('value')
            
            def matches(row_val):
                if operator == '=':
                    return str(row_val) == str(value)
                elif operator == '!=':
                    return str(row_val) != str(value)
                elif operator == '>':
                    return float(row_val) > float(value)
                elif operator == '<':
                    return float(row_val) < float(value)
                elif operator == 'contains':
                    return str(value).lower() in str(row_val).lower()
                return True
            
            result = [row for row in result if filter_col in row and matches(row[filter_col])]
    
    return result, new_columns, column_renames


# =============================================================================
# SCHEMA HELPER (mirrors unified_chat pattern)
# =============================================================================

def _build_bi_schema(handler, project: str) -> Dict[str, Any]:
    """
    Get schema for BI queries - compatible with IntelligenceEngine.
    Returns {'tables': [...], 'filter_candidates': {...}}
    """
    tables = []
    filter_candidates = {}
    
    if not handler or not handler.conn:
        return {'tables': [], 'filter_candidates': {}}
    
    try:
        # Get all tables from DuckDB
        all_tables = handler.conn.execute("SHOW TABLES").fetchall()
        
        # Build project prefix for filtering
        project_clean = (project or '').strip()
        project_prefixes = [
            project_clean.lower(),
            project_clean.lower().replace(' ', '_'),
            project_clean.lower().replace('-', '_'),
        ]
        
        for (table_name,) in all_tables:
            if table_name.startswith('_'):
                continue
            
            table_lower = table_name.lower()
            matches_project = any(
                table_lower.startswith(prefix.lower()) 
                for prefix in project_prefixes if prefix
            )
            
            if not matches_project and project:
                continue
            
            try:
                # Get columns
                columns = []
                try:
                    col_result = handler.conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
                    columns = [row[1] for row in col_result]
                except:
                    result = handler.conn.execute(f'SELECT * FROM "{table_name}" LIMIT 0')
                    columns = [desc[0] for desc in result.description]
                
                if not columns:
                    continue
                
                # Get row count
                try:
                    count_result = handler.conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()
                    row_count = count_result[0] if count_result else 0
                except:
                    row_count = 0
                
                tables.append({
                    'table_name': table_name,
                    'project': project,
                    'columns': columns,
                    'row_count': row_count
                })
                
            except Exception as e:
                logger.warning(f"[BI] Error processing table {table_name}: {e}")
        
        # Get filter candidates
        try:
            filter_candidates = handler.get_filter_candidates(project)
        except:
            pass
        
    except Exception as e:
        logger.error(f"[BI] Schema error: {e}")
    
    return {'tables': tables, 'filter_candidates': filter_candidates}


# =============================================================================
# MAIN QUERY ENDPOINT
# =============================================================================

@router.post("/bi/query")
async def execute_bi_query(request: BIQueryRequest):
    """
    Execute a natural language BI query.
    
    Uses IntelligenceEngine for:
    - SQL generation via LLM
    - Filter handling (status, location, etc.)
    - Schema awareness
    
    Returns:
    - SQL generated
    - Query results
    - Chart recommendation
    - Available transforms
    """
    if not INTELLIGENCE_AVAILABLE or not STRUCTURED_AVAILABLE:
        raise HTTPException(503, "BI service not available")
    
    try:
        handler = get_structured_handler()
        
        # Get schema for project
        schema = _build_bi_schema(handler, request.project)
        if not schema or not schema.get('tables'):
            raise HTTPException(404, f"No data found for project: {request.project}")
        
        # Get relationships
        relationships = []
        try:
            relationships = handler.get_relationships(request.project)
        except:
            pass
        
        # Initialize intelligence engine
        engine = IntelligenceEngine(request.project)
        engine.load_context(
            structured_handler=handler,
            schema=schema,
            relationships=relationships
        )
        
        # Apply any pre-set filters
        if request.filters:
            engine.confirmed_facts.update(request.filters)
        
        # Execute query via intelligence engine
        start_time = time.time()
        answer = engine.ask(request.query)
        execution_time = time.time() - start_time
        
        # Check if clarification needed
        if answer.structured_output and answer.structured_output.get('type') == 'clarification_needed':
            return {
                "success": True,
                "needs_clarification": True,
                "clarification": answer.structured_output,
                "session_id": request.session_id
            }
        
        # Extract SQL and results from answer
        sql = None
        data = []
        columns = []
        
        # Get the SQL that was executed
        if hasattr(engine, 'last_executed_sql') and engine.last_executed_sql:
            sql = engine.last_executed_sql
        
        # Try to get structured output data
        if answer.structured_output:
            if 'data' in answer.structured_output:
                data = answer.structured_output['data']
            if 'columns' in answer.structured_output:
                columns = answer.structured_output['columns']
            if 'sql' in answer.structured_output:
                sql = answer.structured_output['sql']
        
        # If no data in structured_output, try to re-execute SQL
        if sql and not data:
            try:
                result_rows = handler.query(sql)
                data = result_rows
                columns = list(result_rows[0].keys()) if result_rows else []
            except Exception as e:
                logger.warning(f"[BI] Could not execute SQL: {e}")
        
        # Get chart recommendation
        chart_rec = recommend_chart_type(columns, data, sql or '')
        
        # Get available transforms based on columns
        available_transforms = _get_available_transforms(columns, data, schema)
        
        # Post-process: Convert month numbers to names for better display
        data = _convert_month_numbers_to_names(data, columns)
        
        return {
            "success": True,
            "needs_clarification": False,
            "query": request.query,
            "sql": sql,
            "data": data[:1000],  # Limit for response size
            "columns": columns,
            "total_rows": len(data),
            "truncated": len(data) > 1000,
            "chart": chart_rec,
            "available_transforms": available_transforms,
            "execution_time": round(execution_time, 3),
            "answer_text": answer.answer,  # Natural language answer
            "confidence": answer.confidence,
            "filters_applied": dict(engine.confirmed_facts)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[BI] Query error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(500, f"Query failed: {str(e)}")


def _convert_month_numbers_to_names(data: List[Dict], columns: List[str]) -> List[Dict]:
    """
    Convert month columns to readable format.
    Handles:
    - YYYY-MM format (2024-01) → January 2024
    - Month numbers (1-12) → January, February, etc. (less useful without year)
    """
    if not data or not columns:
        return data
    
    MONTH_NAMES = {
        1: 'January', 2: 'February', 3: 'March', 4: 'April',
        5: 'May', 6: 'June', 7: 'July', 8: 'August',
        9: 'September', 10: 'October', 11: 'November', 12: 'December',
        '01': 'January', '02': 'February', '03': 'March', '04': 'April',
        '05': 'May', '06': 'June', '07': 'July', '08': 'August',
        '09': 'September', '10': 'October', '11': 'November', '12': 'December'
    }
    
    # Find month columns
    month_cols = []
    for col in columns:
        col_lower = col.lower()
        if 'month' in col_lower or col_lower == 'month':
            month_cols.append(col)
    
    if not month_cols:
        return data
    
    # Convert month values to readable format
    converted = []
    for row in data:
        new_row = dict(row)
        for col in month_cols:
            val = row.get(col)
            if val is not None:
                val_str = str(val)
                
                # Handle YYYY-MM format (e.g., "2024-01")
                if '-' in val_str and len(val_str) >= 7:
                    parts = val_str.split('-')
                    if len(parts) >= 2:
                        year = parts[0]
                        month = parts[1]
                        month_name = MONTH_NAMES.get(month) or MONTH_NAMES.get(int(month)) if month.isdigit() else month
                        if month_name:
                            new_row[col] = f"{month_name} {year}"
                            continue
                
                # Handle plain month number (1-12)
                try:
                    month_num = int(val)
                    if 1 <= month_num <= 12:
                        new_row[col] = MONTH_NAMES[month_num]
                except (ValueError, TypeError):
                    pass
        converted.append(new_row)
    
    return converted


def _get_available_transforms(columns: List[str], data: List[Dict], schema: Dict) -> List[Dict]:
    """Get available transforms based on column types and data."""
    transforms = []
    
    if not columns or not data:
        return transforms
    
    sample = data[0] if data else {}
    
    # Get column profiles from schema
    profiles = {}
    for table in schema.get('tables', []):
        for col_profile in table.get('column_profiles', {}).items() if isinstance(table.get('column_profiles'), dict) else []:
            col_name, profile = col_profile
            profiles[col_name] = profile
    
    for col in columns:
        val = sample.get(col)
        col_lower = col.lower()
        
        # Always offer rename
        transforms.append({
            'column': col,
            'type': 'rename',
            'label': 'Rename Column',
            'icon': '✏️'
        })
        
        # State columns → state_names
        if any(p in col_lower for p in ['state', 'province', 'st_', '_st']):
            if val and isinstance(val, str) and len(val) == 2:
                transforms.append({
                    'column': col,
                    'type': 'state_names',
                    'label': 'Expand State Codes',
                    'icon': '🗺️'
                })
        
        # Numeric columns → currency/percent
        if isinstance(val, (int, float)):
            if any(p in col_lower for p in ['amount', 'salary', 'pay', 'rate', 'price', 'cost', 'earning']):
                transforms.append({
                    'column': col,
                    'type': 'format_currency',
                    'label': 'Format as Currency',
                    'icon': '💰'
                })
            if any(p in col_lower for p in ['percent', 'pct', 'ratio', 'rate']):
                transforms.append({
                    'column': col,
                    'type': 'format_percent',
                    'label': 'Format as Percent',
                    'icon': '%'
                })
        
        # Text columns → case transforms
        if isinstance(val, str) and len(val) > 1:
            if val == val.lower() or val == val.upper():
                transforms.append({
                    'column': col,
                    'type': 'titlecase',
                    'label': 'Title Case',
                    'icon': 'Aa'
                })
        
        # Code columns → lookup mapping
        if any(p in col_lower for p in ['code', '_cd', 'type', 'category']):
            transforms.append({
                'column': col,
                'type': 'map_lookup',
                'label': 'Map to Description',
                'icon': '🔗'
            })
    
    return transforms


# =============================================================================
# SUGGESTIONS ENDPOINT
# =============================================================================

@router.get("/bi/suggestions/{project}")
async def get_suggestions(project: str):
    """
    Get smart query suggestions for a project.
    
    Sources:
    - filter_candidates (what dimensions exist)
    - Intelligence findings (data quality issues)
    - Common query patterns
    - Recently used queries
    """
    if not STRUCTURED_AVAILABLE:
        raise HTTPException(503, "BI service not available")
    
    try:
        handler = get_structured_handler()
        schema = _build_bi_schema(handler, project)
        
        if not schema or not schema.get('tables'):
            return {"suggestions": [], "categories": []}
        
        suggestions = []
        categories = set()
        
        # 1. Build suggestions from filter_candidates
        filter_candidates = schema.get('filter_candidates', {})
        
        for category, candidates in filter_candidates.items():
            categories.add(category)
            
            if category == 'status':
                suggestions.append({
                    'text': 'How many active employees?',
                    'type': 'common',
                    'category': 'employees'
                })
                suggestions.append({
                    'text': 'Show terminated employees by month',
                    'type': 'common',
                    'category': 'employees'
                })
            
            elif category == 'location':
                if candidates:
                    # Get a sample location value
                    sample_values = candidates[0].get('values', [])[:3] if candidates else []
                    if sample_values:
                        suggestions.append({
                            'text': f'Employees by {candidates[0].get("column", "location")}',
                            'type': 'common',
                            'category': 'location'
                        })
            
            elif category == 'company':
                suggestions.append({
                    'text': 'Headcount by company',
                    'type': 'common',
                    'category': 'organization'
                })
            
            elif category == 'organization':
                suggestions.append({
                    'text': 'Employee count by department',
                    'type': 'common',
                    'category': 'organization'
                })
        
        # 2. Add suggestions from table structure
        for table in schema.get('tables', []):
            table_name = table.get('table_name', '').lower()
            short_name = table_name.split('__')[-1] if '__' in table_name else table_name
            
            if 'earning' in short_name:
                suggestions.append({
                    'text': 'Total earnings by pay type',
                    'type': 'common',
                    'category': 'compensation'
                })
                suggestions.append({
                    'text': 'Average salary by department',
                    'type': 'common',
                    'category': 'compensation'
                })
                categories.add('compensation')
            
            if 'deduction' in short_name:
                suggestions.append({
                    'text': 'Deductions by type',
                    'type': 'common',
                    'category': 'benefits'
                })
                categories.add('benefits')
            
            if 'job' in short_name or 'position' in short_name:
                suggestions.append({
                    'text': 'Employees by job title',
                    'type': 'common',
                    'category': 'organization'
                })
        
        # 3. Try to get intelligence findings as suggestions
        try:
            if INTELLIGENCE_AVAILABLE:
                engine = IntelligenceEngine(project)
                engine.load_context(structured_handler=handler, schema=schema)
                
                # Run quick analysis for findings
                # (In production, cache these or run async)
                # For now, add placeholder for common data quality issues
                
                # Check for location mismatches
                if 'location' in filter_candidates:
                    suggestions.append({
                        'text': 'Employees where home state differs from work state',
                        'type': 'finding',
                        'category': 'data_quality',
                        'badge': 'Potential issue'
                    })
        except Exception as e:
            logger.warning(f"[BI] Could not get intelligence findings: {e}")
        
        # 4. Add saved queries (from Supabase)
        if SUPABASE_AVAILABLE:
            try:
                supabase = get_supabase()
                saved = supabase.table('bi_saved_queries').select('name, query').eq(
                    'project', project
                ).limit(5).execute()
                
                for item in saved.data or []:
                    suggestions.append({
                        'text': item['query'],
                        'type': 'saved',
                        'name': item['name'],
                        'category': 'saved'
                    })
                    categories.add('saved')
            except:
                pass
        
        # Deduplicate by text
        seen = set()
        unique_suggestions = []
        for s in suggestions:
            if s['text'] not in seen:
                seen.add(s['text'])
                unique_suggestions.append(s)
        
        return {
            "suggestions": unique_suggestions,
            "categories": list(categories),
            "filter_candidates": filter_candidates,
            "table_count": len(schema.get('tables', []))
        }
        
    except Exception as e:
        logger.error(f"[BI] Suggestions error: {e}")
        raise HTTPException(500, str(e))


# =============================================================================
# SCHEMA ENDPOINT
# =============================================================================

@router.get("/bi/schema/{project}")
async def get_bi_schema(project: str):
    """Get schema info for BI builder."""
    if not STRUCTURED_AVAILABLE:
        raise HTTPException(503, "Not available")
    
    try:
        handler = get_structured_handler()
        schema = _build_bi_schema(handler, project)
        
        # Simplify for frontend
        tables = []
        for t in schema.get('tables', []):
            tables.append({
                'name': t.get('table_name', '').split('__')[-1],
                'full_name': t.get('table_name'),
                'rows': t.get('row_count', 0),
                'columns': [c.get('name') if isinstance(c, dict) else c for c in t.get('columns', [])],
                'file': t.get('file'),
                'sheet': t.get('sheet')
            })
        
        return {
            "project": project,
            "tables": tables,
            "filter_candidates": schema.get('filter_candidates', {}),
            "relationships": schema.get('relationships', [])
        }
        
    except Exception as e:
        raise HTTPException(500, str(e))


# =============================================================================
# EXECUTE RAW SQL ENDPOINT
# =============================================================================

class BIExecuteRequest(BaseModel):
    """Raw SQL execution request."""
    sql: str
    project: str


@router.post("/bi/execute")
async def execute_raw_sql(request: BIExecuteRequest):
    """
    Execute raw SQL query against project data.
    
    Used by the visual Query Builder for direct SQL execution.
    Returns data, columns, row count, and execution time.
    """
    if not STRUCTURED_AVAILABLE:
        raise HTTPException(503, "Data handler not available")
    
    import time
    start_time = time.time()
    
    try:
        handler = get_structured_handler()
        
        # Execute the SQL
        logger.warning(f"[BI-EXECUTE] Running SQL for project {request.project}")
        logger.warning(f"[BI-EXECUTE] SQL: {request.sql[:200]}...")
        
        result = handler.query(request.sql)
        
        execution_time = int((time.time() - start_time) * 1000)
        
        if result.get('success') and result.get('data'):
            data = result['data']
            columns = result.get('columns', list(data[0].keys()) if data else [])
            
            logger.warning(f"[BI-EXECUTE] Success: {len(data)} rows in {execution_time}ms")
            
            return {
                "success": True,
                "data": data,
                "columns": columns,
                "row_count": len(data),
                "execution_time": execution_time,
                "sql": request.sql
            }
        else:
            error = result.get('error', 'Query returned no results')
            logger.warning(f"[BI-EXECUTE] Failed: {error}")
            raise HTTPException(400, error)
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[BI-EXECUTE] Error: {e}")
        raise HTTPException(500, str(e))


# =============================================================================
# EXPORT ENDPOINT
# =============================================================================

@router.post("/bi/export")
async def export_bi_data(request: BIExportRequest):
    """
    Export query results with transforms.
    
    Can either:
    - Use provided SQL directly
    - Generate SQL from natural language query
    
    Applies transforms before export.
    """
    if not STRUCTURED_AVAILABLE:
        raise HTTPException(503, "Export not available")
    
    try:
        handler = get_structured_handler()
        
        # Get SQL
        sql = request.sql
        if not sql:
            # Generate from NL query
            schema = _build_bi_schema(handler, request.project)
            engine = IntelligenceEngine(request.project)
            engine.load_context(structured_handler=handler, schema=schema)
            answer = engine.ask(request.query)
            
            if hasattr(engine, 'last_executed_sql'):
                sql = engine.last_executed_sql
            elif answer.structured_output and 'sql' in answer.structured_output:
                sql = answer.structured_output['sql']
        
        if not sql:
            raise HTTPException(400, "Could not generate SQL for query")
        
        # Execute query
        rows = handler.query(sql)
        columns = list(rows[0].keys()) if rows else []
        
        if not rows:
            raise HTTPException(404, "No data returned")
        
        # Apply transforms
        transformed_data, new_columns, renames = apply_transforms(
            rows, columns, 
            [TransformOperation(**t) if isinstance(t, dict) else t for t in request.transforms or []],
            handler, request.project
        )
        
        # Apply column renames
        final_columns = [renames.get(c, c) for c in new_columns]
        
        if request.format == 'csv':
            # CSV export
            output = io.StringIO()
            output.write(','.join(final_columns) + '\n')
            for row in transformed_data:
                values = [str(row.get(c, '')) for c in new_columns]
                output.write(','.join(values) + '\n')
            
            output.seek(0)
            filename = f"xlr8_export_{int(time.time())}.csv"
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
        else:
            # Excel export
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # Styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4A7C59", end_color="4A7C59", fill_type="solid")  # Brand green
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            for col_idx, col_name in enumerate(final_columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_idx, row_data in enumerate(transformed_data, 2):
                for col_idx, col_name in enumerate(new_columns, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ''))
                    cell.border = thin_border
            
            # Auto-width
            for col_idx, col_name in enumerate(final_columns, 1):
                max_len = len(str(col_name))
                for row in transformed_data[:100]:
                    val = row.get(new_columns[col_idx-1], '')
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
            
            # Metadata sheet
            if request.include_metadata:
                ws_meta = wb.create_sheet(title="Metadata")
                ws_meta['A1'] = "XLR8 BI Export"
                ws_meta['A1'].font = Font(bold=True, size=14)
                ws_meta['A2'] = f"Query: {request.query}"
                ws_meta['A3'] = f"Project: {request.project}"
                ws_meta['A4'] = f"Rows: {len(transformed_data)}"
                ws_meta['A5'] = f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}"
                
                if request.transforms:
                    ws_meta['A7'] = "Transforms Applied:"
                    ws_meta['A7'].font = Font(bold=True)
                    for i, t in enumerate(request.transforms, 8):
                        t_dict = t if isinstance(t, dict) else t.dict()
                        ws_meta[f'A{i}'] = f"• {t_dict.get('type', '')} on {t_dict.get('column', '')}"
            
            # Save to buffer
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            safe_query = re.sub(r'[^\w\s]', '', request.query[:20]).strip().replace(' ', '_')
            filename = f"xlr8_{safe_query}_{timestamp}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[BI] Export error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(500, f"Export failed: {str(e)}")


# =============================================================================
# SAVED QUERIES ENDPOINTS
# =============================================================================

@router.get("/bi/saved/{project}")
async def get_saved_queries(project: str):
    """Get saved queries for a project."""
    if not SUPABASE_AVAILABLE:
        return {"queries": []}
    
    try:
        supabase = get_supabase()
        result = supabase.table('bi_saved_queries').select('*').eq(
            'project', project
        ).order('created_at', desc=True).execute()
        
        return {"queries": result.data or []}
    except Exception as e:
        logger.error(f"[BI] Error loading saved queries: {e}")
        return {"queries": []}


@router.post("/bi/saved")
async def save_query(request: SavedQueryRequest):
    """Save a query for reuse."""
    if not SUPABASE_AVAILABLE:
        raise HTTPException(503, "Save not available")
    
    try:
        supabase = get_supabase()
        result = supabase.table('bi_saved_queries').insert({
            'name': request.name,
            'query': request.query,
            'project': request.project,
            'sql': request.sql,
            'chart_type': request.chart_type,
            'transforms': request.transforms
        }).execute()
        
        return {"success": True, "id": result.data[0]['id'] if result.data else None}
    except Exception as e:
        logger.error(f"[BI] Error saving query: {e}")
        raise HTTPException(500, str(e))


@router.delete("/bi/saved/{query_id}")
async def delete_saved_query(query_id: str):
    """Delete a saved query."""
    if not SUPABASE_AVAILABLE:
        raise HTTPException(503, "Not available")
    
    try:
        supabase = get_supabase()
        supabase.table('bi_saved_queries').delete().eq('id', query_id).execute()
        return {"success": True}
    except Exception as e:
        raise HTTPException(500, str(e))
