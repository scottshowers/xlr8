"""
Analysis & Templates Page - XLR8 Analysis Engine
COMPLETE VERSION - Question Browser + RAG Analysis + LLM Synthesis
BUILD 20251118-2330
"""

import streamlit as st
import json
from pathlib import Path
import logging
from typing import List, Dict, Any
import os
import sys
import time
from datetime import datetime

# Import RAG handler for document querying
try:
    from utils.rag_handler import RAGHandler
    RAG_AVAILABLE = True
except ImportError:
    RAG_AVAILABLE = False
    st.error("RAG Handler not available")

# Import LLM synthesizer for answer generation
try:
    from utils.ai.enhanced_llm_synthesizer import get_enhanced_synthesizer
    LLM_AVAILABLE = True
except ImportError:
    try:
        from utils.ai.llm_synthesizer import get_synthesizer
        LLM_AVAILABLE = True
        USING_ENHANCED = False
    except ImportError:
        LLM_AVAILABLE = False
        logger = logging.getLogger(__name__)
        logger.warning("LLM Synthesizer not available - will use raw chunks")
    else:
        USING_ENHANCED = False
else:
    USING_ENHANCED = True

logger = logging.getLogger(__name__)

# Print to logs
print("=" * 80)
print("🚀 ANALYSIS_ENGINE.PY LOADING - MISTRAL MODEL - BUILD 2344")
print("=" * 80)


def find_questions_database() -> Path:
    """Find analysis_questions.json by checking multiple possible locations."""
    possible_paths = [
        Path("/data/analysis_questions.json"),  # Primary location (matches upload tool)
        Path(__file__).parent.parent.parent / "data" / "analysis_questions.json",
        Path(__file__).parent / "analysis_questions.json",
        Path.cwd() / "data" / "analysis_questions.json",
        Path("/app/data/analysis_questions.json"),
    ]
    
    for path in possible_paths:
        if path.exists():
            logger.info(f"Found questions database at: {path}")
            return path
    
    logger.warning("Questions database not found")
    return None


def load_questions() -> Dict[str, Any]:
    """Load questions database from JSON file."""
    try:
        db_path = find_questions_database()
        
        if db_path is None:
            return {
                'metadata': {'total_questions': 0, 'error': 'analysis_questions.json not found'},
                'questions': []
            }
        
        with open(db_path, 'r') as f:
            data = json.load(f)
            
            # Handle BOTH formats:
            # NEW format (from analysis_questions.py): {"version": "1.0", "questions": [...]}
            # OLD format: {"metadata": {...}, "questions": [...]}
            
            if 'metadata' not in data:
                # NEW format - build metadata from questions
                questions = data.get('questions', [])
                data = {
                    'metadata': {
                        'total_questions': len(questions),
                        'version': data.get('version', '1.0'),
                        'updated_at': data.get('updated_at', '')
                    },
                    'questions': questions
                }
            
            logger.info(f"Loaded {data['metadata']['total_questions']} questions")
            return data
            
    except Exception as e:
        logger.error(f"Error loading questions: {e}")
        return {
            'metadata': {'total_questions': 0, 'error': str(e)},
            'questions': []
        }


def save_questions(questions_data: Dict[str, Any]) -> bool:
    """Save updated questions back to JSON file."""
    try:
        db_path = find_questions_database()
        if db_path is None:
            logger.error("Cannot save - database path not found")
            return False
        
        with open(db_path, 'w') as f:
            json.dump(questions_data, f, indent=2)
        
        logger.info("Questions database saved successfully")
        return True
    except Exception as e:
        logger.error(f"Error saving questions: {e}")
        return False


def analyze_single_question(question: Dict[str, Any], rag_handler: RAGHandler) -> Dict[str, Any]:
    """
    Analyze a single question using RAG to query customer documents + LLM synthesis.
    
    Args:
        question: Question dictionary with 'question', 'keywords', etc.
        rag_handler: RAG handler instance for querying documents
    
    Returns:
        Dictionary with 'answer', 'sources', 'confidence', 'status'
    """
    try:
        # Build search query from question and keywords
        query_text = question['question']
        
        # Add keywords to enhance search
        if question.get('keywords'):
            query_text += " " + " ".join(question['keywords'][:5])
        
        logger.info(f"Analyzing question {question['id']}: {question['question'][:50]}...")
        
        # Get the collection
        collection = rag_handler.client.get_collection(name="hcmpact_docs")
        
        # Generate embedding for query
        query_embedding = rag_handler.get_embedding(query_text)
        
        if query_embedding is None:
            return {
                'answer': 'Error: Could not generate embedding for query.',
                'sources': [],
                'confidence': 0.0,
                'status': 'pending'
            }
        
        # Query collection directly
        results = collection.query(
            query_embeddings=[query_embedding],
            n_results=5,
            include=['documents', 'metadatas', 'distances']
        )
        
        if not results or not results.get('documents'):
            return {
                'answer': 'No relevant information found in uploaded documents.',
                'sources': [],
                'confidence': 0.0,
                'status': 'analyzed'
            }
        
        # Extract documents and metadata
        documents = results['documents'][0] if results['documents'] else []
        metadatas = results['metadatas'][0] if results.get('metadatas') else []
        distances = results['distances'][0] if results.get('distances') else []
        
        if not documents:
            return {
                'answer': 'No relevant information found.',
                'sources': [],
                'confidence': 0.0,
                'status': 'analyzed'
            }
        
        # Filter out template documents by category (much more robust!)
        filtered_results = []
        for doc, meta, dist in zip(documents, metadatas, distances):
            category = meta.get('category', '')
            source = meta.get('source', '')
            
            # ONLY include customer documents - exclude ALL templates
            # Templates have category: "UKG Templates" or similar
            # Customer docs have category: "Customer Documents" or other non-template categories
            if category not in ['UKG Templates', 'Templates', 'UKG Pro', 'WFM', 'Implementation Guide']:
                # Also double-check filename as backup
                if 'Analysis_Workbook' not in source and 'BRIT' not in source and 'Items_to_Gather' not in source:
                    filtered_results.append((doc, meta, dist))
        
        # If we filtered out everything, inform user (don't fall back to templates!)
        if not filtered_results:
            return {
                'answer': 'No relevant information found in customer documents. All results were from templates. Please ensure customer documents are uploaded with category "Customer Documents".',
                'sources': [],
                'confidence': 0.0,
                'status': 'analyzed'
            }
        
        # Extract chunks and sources for LLM
        chunks = []
        sources = []
        
        for doc, meta, dist in filtered_results[:3]:  # Top 3 chunks
            source_name = meta.get('source', 'Unknown')
            if source_name not in sources:
                sources.append(source_name)
            chunks.append(doc.strip())
        
        # USE ENHANCED LLM TO SYNTHESIZE ANSWER (if available)
        if LLM_AVAILABLE:
            try:
                if USING_ENHANCED:
                    # Use enhanced synthesizer with full context
                    synthesizer = get_enhanced_synthesizer()
                    result = synthesizer.synthesize_answer(
                        question=question['question'],
                        chunks=chunks,
                        sources=sources,
                        reason=question.get('reason', ''),
                        category=question.get('category', ''),
                        required=question.get('required', False)
                    )
                else:
                    # Use basic synthesizer
                    synthesizer = get_synthesizer()
                    result = synthesizer.synthesize_answer(
                        question=question['question'],
                        chunks=chunks,
                        sources=sources,
                        reason=question.get('reason', '')
                    )
                
                return {
                    'answer': result['answer'],
                    'sources': sources,
                    'confidence': result['confidence'],
                    'status': 'analyzed',
                    'reasoning': result.get('reasoning', '')
                }
                
            except Exception as e:
                logger.error(f"LLM synthesis failed, falling back to raw chunks: {e}")
                # Fall through to raw chunk method below
        
        # FALLBACK: Build answer from raw chunks (if LLM not available or failed)
        answer_parts = []
        
        for chunk, source in zip(chunks, sources):
            if chunk and len(chunk) > 20:
                answer_parts.append(f"**From {source}:**\n{chunk[:300]}...")
        
        # Combine answer parts
        if answer_parts:
            answer = "\n\n".join(answer_parts)
        else:
            answer = "Information found but could not extract meaningful content."
        
        # Calculate confidence from distances
        if distances:
            avg_distance = sum(distances[:len(filtered_results)]) / len(filtered_results)
            confidence = max(0.0, min(1.0, 1.0 - (avg_distance / 2.0)))
        else:
            confidence = 0.5
        
        return {
            'answer': answer,
            'sources': sources,
            'confidence': confidence,
            'status': 'analyzed'
        }
        
    except Exception as e:
        logger.error(f"Error analyzing question: {e}", exc_info=True)
        return {
            'answer': f'Error during analysis: {str(e)}',
            'sources': [],
            'confidence': 0.0,
            'status': 'pending'
        }


def render_analysis_page():
    """Main analysis page render function."""
    
    st.title("📊 UKG Analysis Engine")
    st.markdown("""
    <div style='color: #6B7280; font-size: 0.9rem; margin-bottom: 1.5rem;'>
    Automated analysis question answering using RAG-powered document intelligence.
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize RAG handler
    rag_handler = None
    if RAG_AVAILABLE:
        try:
            rag_handler = RAGHandler()
            # Test connection
            count = rag_handler.get_collection_count("hcmpact_docs")
            st.success(f"✅ Connected to RAG system: {count:,} chunks available")
        except Exception as e:
            st.error(f"⚠️ RAG system error: {e}")
            rag_handler = None
    
    # Load questions
    questions_data = load_questions()
    questions = questions_data.get('questions', [])
    total_questions = questions_data.get('metadata', {}).get('total_questions', 0)
    error = questions_data.get('metadata', {}).get('error')
    
    # Show error if questions failed to load
    if error:
        st.error(f"⚠️ Error loading questions database: {error}")
        st.info("**File should be at:** `data/questions_database.json` (root of repo)")
        return
    
    if total_questions == 0:
        st.warning("⚠️ No questions loaded yet.")
        return
    
    # Create tabs
    tab1, tab2, tab3 = st.tabs([
        "📋 Question Browser",
        "⚡ Batch Analysis",
        "📤 Export & Review"
    ])
    
    # TAB 1: QUESTION BROWSER
    with tab1:
        render_question_browser(questions, questions_data, rag_handler)
    
    # TAB 2: BATCH ANALYSIS
    with tab2:
        render_batch_analysis(questions, questions_data, rag_handler)
    
    # TAB 3: EXPORT & REVIEW
    with tab3:
        render_export_review(questions)


def render_question_browser(questions: List[Dict[str, Any]], questions_data: Dict[str, Any], rag_handler):
    """Render the question browser interface with analysis capability."""
    
    st.subheader("📋 Question Browser")
    
    # Summary metrics
    col1, col2, col3, col4 = st.columns(4)
    
    analyzed_count = sum(1 for q in questions if q.get('status') == 'analyzed')
    pending_count = sum(1 for q in questions if q.get('status') == 'pending')
    required_count = sum(1 for q in questions if q.get('required'))
    
    with col1:
        st.metric("Total Questions", len(questions))
    with col2:
        st.metric("Analyzed", analyzed_count, delta=f"{analyzed_count/len(questions)*100:.0f}%")
    with col3:
        st.metric("Pending", pending_count)
    with col4:
        st.metric("Required", required_count)
    
    st.markdown("---")
    
    # Get unique categories
    categories = sorted(list(set(q['category'] for q in questions)))
    
    # Filters
    col1, col2, col3 = st.columns([2, 2, 2])
    
    with col1:
        selected_category = st.selectbox(
            "Filter by Category",
            ["All Categories"] + categories,
            key="question_browser_category"
        )
    
    with col2:
        status_filter = st.selectbox(
            "Filter by Status",
            ["All Status", "Pending", "Analyzed", "Reviewed"],
            key="question_browser_status"
        )
    
    with col3:
        required_filter = st.selectbox(
            "Required Only?",
            ["All Questions", "Required Only", "Optional Only"],
            key="question_browser_required"
        )
    
    # Apply filters
    filtered_questions = questions
    
    if selected_category != "All Categories":
        filtered_questions = [q for q in filtered_questions if q['category'] == selected_category]
    
    if status_filter != "All Status":
        filtered_questions = [q for q in filtered_questions if q['status'] == status_filter.lower()]
    
    if required_filter == "Required Only":
        filtered_questions = [q for q in filtered_questions if q.get('required', False)]
    elif required_filter == "Optional Only":
        filtered_questions = [q for q in filtered_questions if not q.get('required', False)]
    
    # Display count
    st.metric("Questions Matching Filters", len(filtered_questions))
    
    # Display questions as expandable cards
    st.markdown("---")
    
    if not filtered_questions:
        st.info("No questions match the current filters.")
        return
    
    # Pagination
    questions_per_page = 10
    total_pages = max(1, (len(filtered_questions) + questions_per_page - 1) // questions_per_page)
    
    # Use unique key to avoid conflicts with other parts of the app
    page_key = 'analysis_questions_page'
    
    # Initialize or validate current page
    if page_key not in st.session_state:
        st.session_state[page_key] = 0
    
    # Ensure current_page is valid integer within range
    try:
        current_page = int(st.session_state[page_key])
        if current_page >= total_pages or current_page < 0:
            current_page = 0
            st.session_state[page_key] = 0
    except (ValueError, TypeError):
        current_page = 0
        st.session_state[page_key] = 0
    
    # Page controls
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        page = st.selectbox(
            f"Page (showing {questions_per_page} per page)",
            options=list(range(total_pages)),
            index=current_page,
            format_func=lambda x: f"Page {x+1} of {total_pages}",
            key="analysis_page_selector"
        )
        st.session_state[page_key] = int(page)
    
    # Get questions for current page
    start_idx = page * questions_per_page
    end_idx = start_idx + questions_per_page
    page_questions = filtered_questions[start_idx:end_idx]
    
    # Display questions
    for q in page_questions:
        with st.expander(
            f"**{q['id']}** - {q['question'][:80]}{'...' if len(q['question']) > 80 else ''}", 
            expanded=False
        ):
            # Question details
            col1, col2 = st.columns([3, 1])
            
            with col1:
                st.markdown(f"**Category:** {q['category']}")
                if q.get('section'):
                    st.markdown(f"**Section:** {q['section']}")
                st.markdown(f"**Required:** {'✅ Yes' if q.get('required') else '❌ No'}")
            
            with col2:
                status_color = {
                    'pending': '🟡',
                    'analyzed': '🔵',
                    'reviewed': '🟢'
                }.get(q.get('status', 'pending'), '⚪')
                st.markdown(f"**Status:** {status_color} {q.get('status', 'pending').capitalize()}")
            
            st.markdown("---")
            st.markdown(f"**Question:**  \n{q['question']}")
            
            if q.get('reason'):
                st.markdown(f"**Why we ask this:**  \n{q['reason']}")
            
            # Show existing answer if available
            if q.get('answer'):
                st.success(f"**Answer:**  \n{q['answer']}")
                
                if q.get('sources'):
                    st.markdown("**Sources:**")
                    for source in q['sources']:
                        st.caption(f"📄 {source}")
                
                if q.get('confidence'):
                    confidence = q['confidence']
                    confidence_color = "🟢" if confidence > 0.8 else "🟡" if confidence > 0.6 else "🔴"
                    st.markdown(f"**Confidence:** {confidence_color} {confidence*100:.0f}%")
                
                # Show LLM reasoning if available (helps understand the answer)
                if q.get('reasoning'):
                    with st.expander("🧠 Analysis Reasoning"):
                        st.info(q['reasoning'])
            
            # Action buttons
            st.markdown("---")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if rag_handler and st.button("🔍 Analyze This Question", key=f"analyze_{q['id']}"):
                    with st.spinner("Analyzing..."):
                        # Analyze the question
                        result = analyze_single_question(q, rag_handler)
                        
                        # Update question with results
                        q['answer'] = result['answer']
                        q['sources'] = result['sources']
                        q['confidence'] = result['confidence']
                        q['status'] = result['status']
                        q['reasoning'] = result.get('reasoning', '')
                        
                        # Save updated questions
                        if save_questions(questions_data):
                            st.success("✅ Analysis complete!")
                            st.rerun()
                        else:
                            st.error("⚠️ Could not save results")
            
            with col2:
                if q.get('answer') and st.button("✏️ Edit Answer", key=f"edit_{q['id']}"):
                    st.info("💡 Edit feature: Open the answer in a text area for manual editing (coming soon)")
            
            with col3:
                if q.get('answer') and st.button("✅ Mark Reviewed", key=f"review_{q['id']}"):
                    q['status'] = 'reviewed'
                    if save_questions(questions_data):
                        st.success("✅ Marked as reviewed!")
                        st.rerun()


def render_batch_analysis(questions: List[Dict[str, Any]], questions_data: Dict[str, Any], rag_handler):
    """Render the batch analysis interface with working batch processing."""
    
    st.subheader("⚡ Batch Analysis")
    
    if not rag_handler:
        st.error("⚠️ RAG system not available. Cannot perform batch analysis.")
        return
    
    # Calculate stats
    pending = [q for q in questions if q.get('status') == 'pending']
    analyzed = [q for q in questions if q.get('status') in ['analyzed', 'reviewed']]
    
    # Summary metrics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Questions", len(questions))
    with col2:
        st.metric("Analyzed", len(analyzed), delta=f"{len(analyzed)/len(questions)*100:.0f}%")
    with col3:
        st.metric("Pending", len(pending))
    with col4:
        avg_conf = sum(q.get('confidence', 0) for q in analyzed) / len(analyzed) if analyzed else 0
        st.metric("Avg Confidence", f"{avg_conf*100:.0f}%")
    
    st.markdown("---")
    
    if len(pending) == 0:
        st.success("✅ All questions have been analyzed!")
        st.info("👉 Go to Export & Review tab to download results")
        return
    
    # Batch processing controls
    st.markdown("### 🚀 Batch Processing")
    
    # Options
    col1, col2 = st.columns(2)
    with col1:
        batch_size = st.selectbox(
            "Batch Size",
            [5, 10, 20, 50],
            index=1,
            help="Process this many questions before auto-saving"
        )
    
    with col2:
        process_mode = st.selectbox(
            "Process",
            ["All Pending", "By Category", "Required Only"],
            help="Choose which questions to process"
        )
    
    # Filter questions based on mode
    if process_mode == "By Category":
        categories = sorted(list(set(q['category'] for q in pending)))
        selected_category = st.selectbox("Select Category", categories)
        questions_to_process = [q for q in pending if q['category'] == selected_category]
    elif process_mode == "Required Only":
        questions_to_process = [q for q in pending if q.get('required')]
    else:
        questions_to_process = pending
    
    st.info(f"📊 **{len(questions_to_process)} questions** will be processed")
    
    # Estimate time
    time_per_question = 12  # seconds (with LLM synthesis)
    estimated_time = (len(questions_to_process) * time_per_question) / 60
    st.caption(f"⏱️ Estimated time: ~{estimated_time:.0f} minutes")
    
    # Start button
    if st.button("🚀 Start Batch Analysis", type="primary", disabled=len(questions_to_process)==0):
        st.session_state.batch_processing = True
        st.session_state.batch_questions = questions_to_process
        st.session_state.batch_size = batch_size
        st.rerun()
    
    # Process if triggered
    if st.session_state.get('batch_processing'):
        run_batch_analysis(
            questions_to_process=st.session_state.batch_questions,
            questions_data=questions_data,
            rag_handler=rag_handler,
            batch_size=st.session_state.batch_size
        )


def run_batch_analysis(
    questions_to_process: List[Dict[str, Any]],
    questions_data: Dict[str, Any],
    rag_handler: RAGHandler,
    batch_size: int = 10
):
    """Run batch analysis with progress tracking and auto-save."""
    
    st.markdown("---")
    st.markdown("### 📊 Processing in Progress...")
    
    total = len(questions_to_process)
    
    # Progress tracking
    progress_bar = st.progress(0.0)
    status_text = st.empty()
    stats_container = st.empty()
    
    # Track results
    successful = 0
    failed = 0
    start_time = time.time()
    
    # Process questions
    for idx, question in enumerate(questions_to_process):
        # Update progress
        progress = (idx + 1) / total
        progress_bar.progress(progress)
        
        # Update status
        status_text.info(f"🔍 Processing question {idx+1}/{total}: **{question['id']}** - {question['question'][:60]}...")
        
        # Analyze question
        try:
            result = analyze_single_question(question, rag_handler)
            
            # Update question with results
            question['answer'] = result['answer']
            question['sources'] = result['sources']
            question['confidence'] = result['confidence']
            question['status'] = result['status']
            question['reasoning'] = result.get('reasoning', '')
            
            successful += 1
            
        except Exception as e:
            logger.error(f"Error analyzing {question['id']}: {e}")
            question['answer'] = f"Error during analysis: {str(e)}"
            question['status'] = 'pending'
            failed += 1
        
        # Auto-save at batch intervals
        if (idx + 1) % batch_size == 0 or (idx + 1) == total:
            if save_questions(questions_data):
                status_text.success(f"💾 Auto-saved progress at {idx+1}/{total} questions")
            else:
                status_text.warning(f"⚠️ Failed to auto-save at {idx+1}/{total}")
        
        # Update stats
        elapsed = time.time() - start_time
        avg_time = elapsed / (idx + 1)
        remaining = (total - (idx + 1)) * avg_time
        
        with stats_container.container():
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Processed", f"{idx+1}/{total}")
            with col2:
                st.metric("Successful", successful)
            with col3:
                st.metric("Failed", failed)
            with col4:
                st.metric("Time Remaining", f"{remaining/60:.0f}m")
    
    # Completion
    progress_bar.progress(1.0)
    status_text.success(f"✅ Batch analysis complete!")
    
    # Final save
    if save_questions(questions_data):
        st.success("💾 Final results saved successfully!")
    
    # Summary
    st.markdown("---")
    st.markdown("### 📊 Batch Analysis Summary")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Processed", total)
    with col2:
        st.metric("Successful", successful, delta="✅")
    with col3:
        st.metric("Failed", failed, delta="❌" if failed > 0 else None)
    
    total_time = time.time() - start_time
    st.info(f"⏱️ Total time: {total_time/60:.1f} minutes ({total_time/total:.1f}s per question)")
    
    # Clear state
    if st.button("✅ Done - Return to Analysis"):
        st.session_state.pop('batch_processing', None)
        st.session_state.pop('batch_questions', None)
        st.session_state.pop('batch_size', None)
        st.rerun()


def render_export_review(questions: List[Dict[str, Any]]):
    """Render the export and review interface with Excel export."""
    
    st.subheader("📤 Export & Review")
    
    # Show summary of analyzed questions
    analyzed = [q for q in questions if q.get('answer')]
    pending = [q for q in questions if q.get('status') == 'pending']
    
    if not analyzed:
        st.warning("⚠️ No analyzed questions to export yet.")
        st.info("👉 Go to Question Browser or Batch Analysis to analyze questions first")
        return
    
    st.success(f"✅ {len(analyzed)} questions have been analyzed and are ready for export!")
    
    # Confidence breakdown
    high_conf = [q for q in analyzed if q.get('confidence', 0) > 0.8]
    med_conf = [q for q in analyzed if 0.6 < q.get('confidence', 0) <= 0.8]
    low_conf = [q for q in analyzed if q.get('confidence', 0) <= 0.6]
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("High Confidence", len(high_conf), help=">80%", delta="🟢")
    with col2:
        st.metric("Medium Confidence", len(med_conf), help="60-80%", delta="🟡")
    with col3:
        st.metric("Low Confidence", len(low_conf), help="<60%", delta="🔴")
    with col4:
        st.metric("Still Pending", len(pending))
    
    st.markdown("---")
    
    # Export options
    st.markdown("### 📥 Export Options")
    
    export_format = st.radio(
        "Export Format",
        ["Excel (XLSX)", "CSV", "JSON"],
        help="Choose the export format"
    )
    
    include_options = st.multiselect(
        "Include in Export",
        ["All Questions", "Analyzed Only", "High Confidence Only", "Include Sources", "Include Confidence Scores"],
        default=["All Questions", "Include Sources", "Include Confidence Scores"]
    )
    
    # Export button
    if st.button("📥 Generate Export File", type="primary"):
        # Filter questions based on options
        if "Analyzed Only" in include_options:
            export_questions = analyzed
        elif "High Confidence Only" in include_options:
            export_questions = high_conf
        else:
            export_questions = questions
        
        # Generate export based on format
        if export_format == "Excel (XLSX)":
            export_data = generate_excel_export(
                export_questions,
                include_sources="Include Sources" in include_options,
                include_confidence="Include Confidence Scores" in include_options
            )
            
            if export_data:
                st.download_button(
                    label="📥 Download Analysis_Workbook_Completed.xlsx",
                    data=export_data,
                    file_name=f"Analysis_Workbook_Completed_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel"
                )
                st.success("✅ Excel file ready for download!")
            else:
                st.error("❌ Failed to generate Excel file")
        
        elif export_format == "CSV":
            export_data = generate_csv_export(export_questions)
            st.download_button(
                label="📥 Download analysis_results.csv",
                data=export_data,
                file_name=f"analysis_results_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                key="download_csv"
            )
            st.success("✅ CSV file ready for download!")
        
        elif export_format == "JSON":
            export_data = json.dumps({
                'exported_at': datetime.now().isoformat(),
                'total_questions': len(export_questions),
                'questions': export_questions
            }, indent=2)
            st.download_button(
                label="📥 Download analysis_results.json",
                data=export_data,
                file_name=f"analysis_results_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                mime="application/json",
                key="download_json"
            )
            st.success("✅ JSON file ready for download!")
    
    # Preview section
    st.markdown("---")
    st.markdown("### 👁️ Preview Export Data")
    
    with st.expander("Show Export Preview (First 5 Questions)"):
        import pandas as pd
        
        preview_data = []
        for q in analyzed[:5]:
            preview_data.append({
                'ID': q['id'],
                'Category': q['category'],
                'Question': q['question'][:60] + '...',
                'Answer': q.get('answer', 'N/A')[:100] + '...' if q.get('answer') else 'N/A',
                'Confidence': f"{q.get('confidence', 0)*100:.0f}%",
                'Sources': ', '.join(q.get('sources', [])[:2])
            })
        
        df = pd.DataFrame(preview_data)
        st.dataframe(df, use_container_width=True)


def generate_excel_export(
    questions: List[Dict[str, Any]],
    include_sources: bool = True,
    include_confidence: bool = True
) -> bytes:
    """Generate Excel export file."""
    
    try:
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis Results"
        
        # Headers
        headers = ['ID', 'Category', 'Question', 'Answer', 'Status']
        if include_sources:
            headers.append('Sources')
        if include_confidence:
            headers.append('Confidence')
        headers.append('Required')
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, q in enumerate(questions, 2):
            col_idx = 1
            
            # ID
            ws.cell(row=row_idx, column=col_idx, value=q['id'])
            col_idx += 1
            
            # Category
            ws.cell(row=row_idx, column=col_idx, value=q['category'])
            col_idx += 1
            
            # Question
            ws.cell(row=row_idx, column=col_idx, value=q['question'])
            col_idx += 1
            
            # Answer
            answer_cell = ws.cell(row=row_idx, column=col_idx, value=q.get('answer', 'Not analyzed'))
            
            # Color-code by confidence
            if include_confidence and q.get('confidence'):
                conf = q['confidence']
                if conf > 0.8:
                    answer_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                elif conf > 0.6:
                    answer_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                else:
                    answer_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
            
            col_idx += 1
            
            # Status
            ws.cell(row=row_idx, column=col_idx, value=q.get('status', 'pending').capitalize())
            col_idx += 1
            
            # Sources
            if include_sources:
                sources = ', '.join(q.get('sources', []))
                ws.cell(row=row_idx, column=col_idx, value=sources)
                col_idx += 1
            
            # Confidence
            if include_confidence:
                conf_val = f"{q.get('confidence', 0)*100:.0f}%" if q.get('confidence') else 'N/A'
                ws.cell(row=row_idx, column=col_idx, value=conf_val)
                col_idx += 1
            
            # Required
            ws.cell(row=row_idx, column=col_idx, value="Yes" if q.get('required') else "No")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating Excel export: {e}")
        return None


def generate_csv_export(questions: List[Dict[str, Any]]) -> str:
    """Generate CSV export."""
    
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow(['ID', 'Category', 'Question', 'Answer', 'Confidence', 'Sources', 'Status', 'Required'])
    
    # Data
    for q in questions:
        writer.writerow([
            q['id'],
            q['category'],
            q['question'],
            q.get('answer', 'Not analyzed'),
            f"{q.get('confidence', 0)*100:.0f}%",
            ', '.join(q.get('sources', [])),
            q.get('status', 'pending'),
            'Yes' if q.get('required') else 'No'
        ])
    
    return output.getvalue()
    
    if not analyzed:
        st.warning("No analyzed questions to export yet.")
        return
    
    st.success(f"✅ {len(analyzed)} questions have been analyzed!")
    
    # Confidence breakdown
    high_conf = sum(1 for q in analyzed if q.get('confidence', 0) > 0.8)
    med_conf = sum(1 for q in analyzed if 0.6 < q.get('confidence', 0) <= 0.8)
    low_conf = sum(1 for q in analyzed if q.get('confidence', 0) <= 0.6)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("High Confidence", high_conf, help=">80%", delta="✅")
    with col2:
        st.metric("Medium Confidence", med_conf, help="60-80%", delta="⚠️")
    with col3:
        st.metric("Low Confidence", low_conf, help="<60%", delta="⚠️")
    
    st.markdown("---")
    st.markdown("### Coming Features:")
    st.markdown("""
    - **Excel Export:** Export to Analysis_Workbook.xlsx format with answers populated
    - **Confidence Color Coding:** Green/Yellow/Red based on confidence
    - **Source Citations:** Include all sources in export
    - **Review Workflow:** Filter and review low-confidence answers
    - **Bulk Operations:** Approve multiple answers at once
    """)


# Entry point
if __name__ == "__main__":
    print("⚠️ analysis_engine.py running in standalone mode - BUILD 2315")
    render_analysis_page()
